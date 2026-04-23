"""Excel import service — parse and import contacts, invoices, payments from Excel files."""

from __future__ import annotations

from sqlalchemy.ext.asyncio import AsyncSession

from backend.services.excel_import._comparison import (
    _gestion_bank_comparison_signature,
    _gestion_cash_comparison_signature,
    _gestion_payment_comparison_signature,
)
from backend.services.excel_import._comparison_loaders import (
    _load_existing_bank_comparison_signatures,
    _load_existing_cash_comparison_signatures,
    _load_existing_client_payment_comparison_signatures,
)
from backend.services.excel_import._constants import (
    _CLIENT_INVOICE_CLARIFIED_MESSAGE,
    _load_existing_generated_accounting_group_signatures,
)
from backend.services.excel_import._entry_groups import (
    _build_entry_row_groups,
    _invoice_row_from_supplier_candidate,
    _matching_existing_salary_entry_group,
    _normalized_entry_group_signature,
    _supplier_invoice_candidate_from_bank_row,
    _supplier_invoice_candidate_from_cash_row,
)
from backend.services.excel_import._invoices import (
    _client_settlement_account_from_method,
    _find_clarifiable_existing_client_invoice,
    _load_existing_client_invoices_by_number,
    _load_existing_client_payment_reference_signatures,
    _load_existing_supplier_payment_reference_signatures,
    _matching_existing_client_invoice_reference,
    _matching_existing_client_payment_reference,
    _matching_existing_supplier_invoice_payment_reference,
    _merge_existing_client_invoice_entry_groups,
    _single_client_invoice_reference,
)
from backend.services.excel_import._loaders import (
    _load_existing_salary_group_signatures,
    _load_existing_salary_keys,
)
from backend.services.excel_import._salary import _salary_employee_key
from backend.services.excel_import._sheet_wrappers import (
    _parse_bank_sheet,
    _parse_cash_sheet,
    _parse_contact_sheet,
    _parse_entries_sheet,
    _parse_invoice_sheet,
    _parse_payment_sheet,
    _parse_salary_sheet,
)
from backend.services.excel_import_classification import (
    classify_comptabilite_sheet as _classify_comptabilite_sheet,
)
from backend.services.excel_import_classification import (
    classify_gestion_sheet as _classify_gestion_sheet,
)
from backend.services.excel_import_parsing import (
    normalize_text as _normalize_text,
)
from backend.services.excel_import_payment_matching import (
    PaymentMatchCandidate,
)
from backend.services.excel_import_payment_matching import (
    make_workbook_invoice_candidate as _make_workbook_invoice_candidate,
)
from backend.services.excel_import_payment_matching import (
    resolve_payment_match_with_database as _resolve_payment_match,
)
from backend.services.excel_import_policy import (
    ENTRY_COVERED_BY_SOLDE_MESSAGE,
    ENTRY_EXISTING_MESSAGE,
    ENTRY_NEAR_EXISTING_MANUAL_MESSAGE,
    EXISTING_GESTION_ROW_MESSAGE,
    EXISTING_SALARY_MESSAGE,
    filter_duplicate_contact_rows,
    filter_duplicate_invoice_rows,
    find_ambiguous_invoice_contact_issues,
    find_existing_contact_issues,
    find_existing_invoice_issues,
    make_existing_invoice_issue,
    make_payment_resolution_issue,
    resolve_invoice_contact_match,
)
from backend.services.excel_import_preview_helpers import (
    append_preview_blocked_issue as _append_preview_blocked_issue,
)
from backend.services.excel_import_preview_helpers import (
    append_preview_ignored_issue as _append_preview_ignored_issue,
)
from backend.services.excel_import_preview_helpers import (
    append_preview_warning_issue as _append_preview_warning_issue,
)
from backend.services.excel_import_preview_helpers import (
    contact_preview_key as _contact_preview_key,
)
from backend.services.excel_import_preview_helpers import (
    find_sheet_preview as _find_sheet_preview,
)
from backend.services.excel_import_preview_helpers import (
    recompute_preview_can_import as _recompute_preview_can_import,
)
from backend.services.excel_import_results import PreviewResult
from backend.services.excel_import_state import (
    accounting_entry_line_signature as _accounting_entry_line_signature,
)
from backend.services.excel_import_state import (
    accounting_entry_signature as _accounting_entry_signature,
)
from backend.services.excel_import_state import (
    load_existing_accounting_entry_signatures as _load_existing_accounting_entry_signatures,
)
from backend.services.excel_import_state import (
    load_existing_contacts_by_preview_key as _load_existing_contacts_by_preview_key,
)
from backend.services.excel_import_state import (
    load_existing_invoice_numbers as _load_existing_invoice_numbers,
)
from backend.services.excel_import_state import (
    load_existing_manual_accounting_line_signatures,
)
from backend.services.excel_import_types import (
    NormalizedPaymentRow,
    RowIgnoredIssue,
    RowWarningIssue,
)


async def _add_gestion_existing_rows_preview(
    db: AsyncSession, file_bytes: bytes, preview: PreviewResult
) -> None:
    from io import BytesIO  # noqa: PLC0415

    import openpyxl  # noqa: PLC0415

    existing_contacts_by_preview_key = await _load_existing_contacts_by_preview_key(db)
    existing_contact_keys = set(existing_contacts_by_preview_key)
    existing_invoice_numbers = await _load_existing_invoice_numbers(db)
    existing_salary_keys = await _load_existing_salary_keys(db)

    wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    workbook_candidates: list[PaymentMatchCandidate] = []
    payment_years: set[int] = set()
    bank_years: set[int] = set()
    cash_years: set[int] = set()
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        kind, _ = _classify_gestion_sheet(sheet_name)
        if kind == "invoices":
            parsed_sheet, invoice_rows, _, _ = _parse_invoice_sheet(ws)
            if parsed_sheet is None or parsed_sheet.missing_columns:
                continue
            invoice_rows, _ = filter_duplicate_invoice_rows(
                invoice_rows,
                normalize_text=_normalize_text,
            )
            workbook_candidates.extend(
                _make_workbook_invoice_candidate(invoice_row) for invoice_row in invoice_rows
            )
        elif kind == "payments":
            parsed_sheet, payment_rows, _ = _parse_payment_sheet(ws)
            if parsed_sheet is None or parsed_sheet.missing_columns:
                continue
            payment_years.update(payment_row.payment_date.year for payment_row in payment_rows)
        elif kind == "bank":
            parsed_sheet, bank_rows, _, _ = _parse_bank_sheet(ws)
            if parsed_sheet is None or parsed_sheet.missing_columns:
                continue
            bank_years.update(bank_row.entry_date.year for bank_row in bank_rows)
        elif kind == "cash":
            parsed_sheet, cash_rows, _, _ = _parse_cash_sheet(ws)
            if parsed_sheet is None or parsed_sheet.missing_columns:
                continue
            cash_years.update(cash_row.entry_date.year for cash_row in cash_rows)

    existing_payment_signatures = await _load_existing_client_payment_comparison_signatures(
        db,
        payment_years,
    )
    existing_bank_signatures = await _load_existing_bank_comparison_signatures(db, bank_years)
    existing_cash_signatures = await _load_existing_cash_comparison_signatures(db, cash_years)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        kind, _ = _classify_gestion_sheet(sheet_name)
        sheet_preview = _find_sheet_preview(preview, sheet_name)
        if kind is None or sheet_preview is None or sheet_preview["status"] != "recognized":
            continue

        if kind == "contacts":
            parsed_sheet, contact_rows, _ = _parse_contact_sheet(ws)
            if parsed_sheet is None or parsed_sheet.missing_columns:
                continue
            contact_rows, _ = filter_duplicate_contact_rows(
                contact_rows,
                normalize_text=_normalize_text,
            )
            ignored_issues = find_existing_contact_issues(
                contact_rows,
                existing_contact_keys,
                contact_preview_key=_contact_preview_key,
            )
            for ignored_issue in ignored_issues:
                _append_preview_ignored_issue(
                    preview,
                    sheet_preview,
                    ignored_issue,
                )
            if ignored_count := len(ignored_issues):
                sheet_preview["rows"] = max(0, sheet_preview["rows"] - ignored_count)

        elif kind == "invoices":
            if (invoice_parsed := _parse_invoice_sheet(ws))[0] is None or invoice_parsed[
                0
            ].missing_columns:
                continue
            _, invoice_rows, _, _ = invoice_parsed
            invoice_rows, _ = filter_duplicate_invoice_rows(
                invoice_rows,
                normalize_text=_normalize_text,
            )
            ignored_issues = find_existing_invoice_issues(
                invoice_rows,
                existing_invoice_numbers,
                normalize_text=_normalize_text,
            )
            blocked_issues = find_ambiguous_invoice_contact_issues(
                invoice_rows,
                existing_contacts_by_preview_key,
                normalize_text=_normalize_text,
            )
            for ignored_issue in ignored_issues:
                _append_preview_ignored_issue(
                    preview,
                    sheet_preview,
                    ignored_issue,
                )
            for blocked_issue in blocked_issues:
                _append_preview_blocked_issue(
                    preview,
                    sheet_preview,
                    blocked_issue,
                )
            ignored_count = len(ignored_issues)
            blocked_count = len(blocked_issues)
            if ignored_count:
                sheet_preview["rows"] = max(0, sheet_preview["rows"] - ignored_count)
                preview.estimated_invoices = max(0, preview.estimated_invoices - ignored_count)
            if blocked_count:
                sheet_preview["rows"] = max(0, sheet_preview["rows"] - blocked_count)
                preview.estimated_invoices = max(0, preview.estimated_invoices - blocked_count)

        elif kind == "payments":
            parsed_sheet, payment_rows, _ = _parse_payment_sheet(ws)
            if parsed_sheet is None or parsed_sheet.missing_columns:
                continue
            for payment_row in payment_rows:
                resolution = await _resolve_payment_match(
                    db,
                    payment_row,
                    workbook_candidates,
                )
                if resolution.status != "matched" or resolution.candidate is None:
                    continue
                if not resolution.candidate.invoice_number:
                    continue
                payment_signature = _gestion_payment_comparison_signature(
                    reference=resolution.candidate.invoice_number,
                    payment_date=payment_row.payment_date,
                    amount=payment_row.amount,
                    settlement_account=_client_settlement_account_from_method(payment_row.method),
                    cheque_number=payment_row.cheque_number,
                )
                if payment_signature not in existing_payment_signatures:
                    continue
                _append_preview_ignored_issue(
                    preview,
                    sheet_preview,
                    RowIgnoredIssue(
                        source_row_number=payment_row.source_row_number,
                        message=EXISTING_GESTION_ROW_MESSAGE,
                    ),
                )
                sheet_preview["rows"] = max(0, sheet_preview["rows"] - 1)
                preview.estimated_payments = max(0, preview.estimated_payments - 1)

        elif kind == "bank":
            parsed_sheet, bank_rows, _, _ = _parse_bank_sheet(ws)
            if parsed_sheet is None or parsed_sheet.missing_columns:
                continue
            for bank_row in bank_rows:
                bank_signature = _gestion_bank_comparison_signature(
                    entry_date=bank_row.entry_date,
                    amount=bank_row.amount,
                    description=bank_row.description,
                    reference=bank_row.reference,
                )
                candidate = _supplier_invoice_candidate_from_bank_row(bank_row)
                if bank_signature in existing_bank_signatures:
                    _append_preview_ignored_issue(
                        preview,
                        sheet_preview,
                        RowIgnoredIssue(
                            source_row_number=bank_row.source_row_number,
                            message=EXISTING_GESTION_ROW_MESSAGE,
                        ),
                    )
                    sheet_preview["rows"] = max(0, sheet_preview["rows"] - 1)
                    if candidate is not None:
                        preview.estimated_invoices = max(0, preview.estimated_invoices - 1)
                        preview.estimated_payments = max(0, preview.estimated_payments - 1)
                    elif _single_client_invoice_reference(
                        bank_row.reference,
                        bank_row.description,
                    ):
                        preview.estimated_payments = max(0, preview.estimated_payments - 1)
                    continue
                if candidate is None:
                    continue
                if _normalize_text(candidate.invoice_number) in existing_invoice_numbers:
                    _append_preview_ignored_issue(
                        preview,
                        sheet_preview,
                        make_existing_invoice_issue(candidate.source_row_number),
                    )
                    sheet_preview["rows"] = max(0, sheet_preview["rows"] - 1)
                    preview.estimated_invoices = max(0, preview.estimated_invoices - 1)
                    preview.estimated_payments = max(0, preview.estimated_payments - 1)
                    continue
                _, supplier_blocked_issue = resolve_invoice_contact_match(
                    _invoice_row_from_supplier_candidate(candidate),
                    existing_contacts_by_preview_key,
                    normalize_text=_normalize_text,
                )
                if supplier_blocked_issue is not None:
                    _append_preview_blocked_issue(
                        preview,
                        sheet_preview,
                        supplier_blocked_issue,
                    )
                    preview.estimated_invoices = max(0, preview.estimated_invoices - 1)
                    preview.estimated_payments = max(0, preview.estimated_payments - 1)

        elif kind == "cash":
            parsed_sheet, cash_rows, _, _ = _parse_cash_sheet(ws)
            if parsed_sheet is None or parsed_sheet.missing_columns:
                continue
            for cash_row in cash_rows:
                cash_signature = _gestion_cash_comparison_signature(
                    entry_date=cash_row.entry_date,
                    movement_type=cash_row.movement_type,
                    amount=cash_row.amount,
                    description=cash_row.description,
                    reference=cash_row.reference,
                )
                candidate = _supplier_invoice_candidate_from_cash_row(cash_row)
                if cash_signature in existing_cash_signatures:
                    _append_preview_ignored_issue(
                        preview,
                        sheet_preview,
                        RowIgnoredIssue(
                            source_row_number=cash_row.source_row_number,
                            message=EXISTING_GESTION_ROW_MESSAGE,
                        ),
                    )
                    sheet_preview["rows"] = max(0, sheet_preview["rows"] - 1)
                    if candidate is not None:
                        preview.estimated_invoices = max(0, preview.estimated_invoices - 1)
                        preview.estimated_payments = max(0, preview.estimated_payments - 1)
                    continue
                if candidate is None:
                    continue
                if _normalize_text(candidate.invoice_number) in existing_invoice_numbers:
                    _append_preview_ignored_issue(
                        preview,
                        sheet_preview,
                        make_existing_invoice_issue(candidate.source_row_number),
                    )
                    sheet_preview["rows"] = max(0, sheet_preview["rows"] - 1)
                    preview.estimated_invoices = max(0, preview.estimated_invoices - 1)
                    preview.estimated_payments = max(0, preview.estimated_payments - 1)
                    continue
                _, supplier_blocked_issue = resolve_invoice_contact_match(
                    _invoice_row_from_supplier_candidate(candidate),
                    existing_contacts_by_preview_key,
                    normalize_text=_normalize_text,
                )
                if supplier_blocked_issue is not None:
                    _append_preview_blocked_issue(
                        preview,
                        sheet_preview,
                        supplier_blocked_issue,
                    )
                    preview.estimated_invoices = max(0, preview.estimated_invoices - 1)
                    preview.estimated_payments = max(0, preview.estimated_payments - 1)

        elif kind == "salaries":
            parsed_sheet, salary_rows, _ = _parse_salary_sheet(ws)
            if parsed_sheet is None or parsed_sheet.missing_columns:
                continue
            salary_ignored_issues: list[RowIgnoredIssue] = []
            seen_salary_keys = set(existing_salary_keys)
            for salary_row in salary_rows:
                salary_key = (salary_row.month, _salary_employee_key(salary_row.employee_name))
                if salary_key in seen_salary_keys:
                    salary_ignored_issues.append(
                        RowIgnoredIssue(
                            source_row_number=salary_row.source_row_number,
                            message=EXISTING_SALARY_MESSAGE,
                        )
                    )
                    continue
                seen_salary_keys.add(salary_key)
            for ignored_issue in salary_ignored_issues:
                _append_preview_ignored_issue(
                    preview,
                    sheet_preview,
                    ignored_issue,
                )
            if ignored_count := len(salary_ignored_issues):
                sheet_preview["rows"] = max(0, sheet_preview["rows"] - ignored_count)
                preview.estimated_salaries = max(0, preview.estimated_salaries - ignored_count)

    preview._candidate_contacts.difference_update(existing_contact_keys)
    preview.estimated_contacts = len(preview._candidate_contacts)
    _recompute_preview_can_import(preview)


async def _add_comptabilite_existing_rows_preview(
    db: AsyncSession, file_bytes: bytes, preview: PreviewResult
) -> None:
    from io import BytesIO  # noqa: PLC0415

    import openpyxl  # noqa: PLC0415

    existing_entry_signatures = await _load_existing_accounting_entry_signatures(db)
    existing_invoice_numbers = await _load_existing_invoice_numbers(db)
    existing_client_payment_signatures = await _load_existing_client_payment_reference_signatures(
        db
    )
    existing_supplier_payment_signatures = (
        await _load_existing_supplier_payment_reference_signatures(db)
    )
    existing_client_invoices_by_number = await _load_existing_client_invoices_by_number(db)
    existing_salary_group_signatures = await _load_existing_salary_group_signatures(db)
    existing_manual_line_signatures = await load_existing_manual_accounting_line_signatures(db)
    existing_generated_group_signatures = (
        await _load_existing_generated_accounting_group_signatures(db)
    )
    if (
        not existing_entry_signatures
        and not existing_generated_group_signatures
        and not existing_invoice_numbers
    ):
        return

    wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        kind, _ = _classify_comptabilite_sheet(sheet_name)
        sheet_preview = _find_sheet_preview(preview, sheet_name)
        if kind != "entries" or sheet_preview is None or sheet_preview["status"] != "recognized":
            continue

        parsed_sheet, normalized_rows, _, _ = _parse_entries_sheet(ws)
        if parsed_sheet is None or parsed_sheet.missing_columns:
            continue

        entry_groups = _build_entry_row_groups(normalized_rows)
        index = 0
        while index < len(entry_groups):
            entry_group, next_index = _merge_existing_client_invoice_entry_groups(
                entry_groups,
                index,
                existing_invoice_numbers,
            )
            existing_client_invoice_reference = _matching_existing_client_invoice_reference(
                entry_group,
                existing_invoice_numbers,
            )
            if existing_client_invoice_reference is not None:
                clarified_invoice = await _find_clarifiable_existing_client_invoice(
                    db,
                    entry_group,
                    existing_client_invoices_by_number,
                )
                for entry_row in entry_group:
                    _append_preview_ignored_issue(
                        preview,
                        sheet_preview,
                        RowIgnoredIssue(
                            source_row_number=entry_row.source_row_number,
                            message=(
                                _CLIENT_INVOICE_CLARIFIED_MESSAGE
                                if clarified_invoice is not None
                                else ENTRY_COVERED_BY_SOLDE_MESSAGE
                            ),
                        ),
                    )
                    sheet_preview["rows"] = max(0, sheet_preview["rows"] - 1)
                    preview.estimated_entries = max(0, preview.estimated_entries - 1)
                index = next_index
                continue

            if _matching_existing_client_payment_reference(
                entry_group,
                existing_client_payment_signatures,
            ):
                for entry_row in entry_group:
                    _append_preview_ignored_issue(
                        preview,
                        sheet_preview,
                        RowIgnoredIssue(
                            source_row_number=entry_row.source_row_number,
                            message=ENTRY_COVERED_BY_SOLDE_MESSAGE,
                        ),
                    )
                    sheet_preview["rows"] = max(0, sheet_preview["rows"] - 1)
                    preview.estimated_entries = max(0, preview.estimated_entries - 1)
                index = next_index
                continue

            if _matching_existing_supplier_invoice_payment_reference(
                entry_group,
                existing_supplier_payment_signatures,
            ):
                for entry_row in entry_group:
                    _append_preview_ignored_issue(
                        preview,
                        sheet_preview,
                        RowIgnoredIssue(
                            source_row_number=entry_row.source_row_number,
                            message=ENTRY_COVERED_BY_SOLDE_MESSAGE,
                        ),
                    )
                    sheet_preview["rows"] = max(0, sheet_preview["rows"] - 1)
                    preview.estimated_entries = max(0, preview.estimated_entries - 1)
                index = next_index
                continue

            if _matching_existing_salary_entry_group(
                entry_group,
                existing_salary_group_signatures,
            ):
                for entry_row in entry_group:
                    _append_preview_ignored_issue(
                        preview,
                        sheet_preview,
                        RowIgnoredIssue(
                            source_row_number=entry_row.source_row_number,
                            message=ENTRY_COVERED_BY_SOLDE_MESSAGE,
                        ),
                    )
                    sheet_preview["rows"] = max(0, sheet_preview["rows"] - 1)
                    preview.estimated_entries = max(0, preview.estimated_entries - 1)
                index = next_index
                continue

            if (
                _normalized_entry_group_signature(entry_group)
                in existing_generated_group_signatures
            ):
                for entry_row in entry_group:
                    _append_preview_ignored_issue(
                        preview,
                        sheet_preview,
                        RowIgnoredIssue(
                            source_row_number=entry_row.source_row_number,
                            message=ENTRY_COVERED_BY_SOLDE_MESSAGE,
                        ),
                    )
                    sheet_preview["rows"] = max(0, sheet_preview["rows"] - 1)
                    preview.estimated_entries = max(0, preview.estimated_entries - 1)
                index = next_index
                continue

            for entry_row in entry_group:
                signature = _accounting_entry_signature(
                    entry_date=entry_row.entry_date,
                    account_number=entry_row.account_number,
                    label=entry_row.label,
                    debit=entry_row.debit,
                    credit=entry_row.credit,
                )
                if signature not in existing_entry_signatures:
                    line_signature = _accounting_entry_line_signature(
                        entry_date=entry_row.entry_date,
                        account_number=entry_row.account_number,
                        debit=entry_row.debit,
                        credit=entry_row.credit,
                    )
                    if line_signature in existing_manual_line_signatures:
                        _append_preview_warning_issue(
                            preview,
                            sheet_preview,
                            RowWarningIssue(
                                source_row_number=entry_row.source_row_number,
                                message=ENTRY_NEAR_EXISTING_MANUAL_MESSAGE,
                            ),
                        )
                    continue
                _append_preview_ignored_issue(
                    preview,
                    sheet_preview,
                    RowIgnoredIssue(
                        source_row_number=entry_row.source_row_number,
                        message=ENTRY_EXISTING_MESSAGE,
                    ),
                )
                sheet_preview["rows"] = max(0, sheet_preview["rows"] - 1)
                preview.estimated_entries = max(0, preview.estimated_entries - 1)
            index = next_index

    _recompute_preview_can_import(preview)


async def _add_gestion_payment_validation(
    db: AsyncSession, file_bytes: bytes, preview: PreviewResult
) -> None:
    from io import BytesIO  # noqa: PLC0415

    import openpyxl  # noqa: PLC0415

    workbook_candidates: list[PaymentMatchCandidate] = []
    payment_rows_by_sheet: dict[str, list[NormalizedPaymentRow]] = {}

    wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        kind, _ = _classify_gestion_sheet(sheet_name)
        if kind == "invoices":
            parsed_sheet, invoice_rows, _, _ = _parse_invoice_sheet(ws)
            if parsed_sheet is None or parsed_sheet.missing_columns:
                continue
            workbook_candidates.extend(
                _make_workbook_invoice_candidate(invoice_row) for invoice_row in invoice_rows
            )
        elif kind == "payments":
            parsed_sheet, payment_rows, _ = _parse_payment_sheet(ws)
            if parsed_sheet is None or parsed_sheet.missing_columns:
                continue
            payment_rows_by_sheet[sheet_name] = payment_rows

    for sheet_name, payment_rows in payment_rows_by_sheet.items():
        sheet_preview = _find_sheet_preview(preview, sheet_name)
        if sheet_preview is None:
            continue

        blocked_count = 0
        for payment_row in payment_rows:
            resolution = await _resolve_payment_match(
                db,
                payment_row,
                workbook_candidates,
            )
            blocking_issue = make_payment_resolution_issue(
                source_row_number=payment_row.source_row_number,
                status=resolution.status,
                candidate=resolution.candidate,
                message=resolution.message,
                require_persistable_candidate=False,
            )
            if blocking_issue is not None:
                _append_preview_blocked_issue(preview, sheet_preview, blocking_issue)
                blocked_count += 1

        if blocked_count:
            sheet_preview["rows"] = max(0, sheet_preview["rows"] - blocked_count)
            preview.estimated_payments = max(0, preview.estimated_payments - blocked_count)
