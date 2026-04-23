"""Excel import service — parse and import contacts, invoices, payments from Excel files."""

from __future__ import annotations

from typing import Any

from sqlalchemy.ext.asyncio import AsyncSession

import backend.services.excel_import as _pkg
from backend.services.excel_import._constants import (
    _GESTION_IMPORT_ORDER,
    _ImportSheetFailure,
    logger,
)
from backend.services.excel_import._exceptions import ImportFileOpenError
from backend.services.excel_import_classification import (
    classify_comptabilite_sheet as _classify_comptabilite_sheet,
)
from backend.services.excel_import_classification import (
    classify_gestion_sheet as _classify_gestion_sheet,
)
from backend.services.excel_import_results import ImportResult
from backend.services.excel_import_state import (
    compute_file_hash as _compute_file_hash,
)
from backend.services.excel_import_state import (
    record_import_log as _record_import_log,
)


async def import_gestion_file(
    db: AsyncSession, file_bytes: bytes, file_name: str | None = None
) -> ImportResult:
    """Import contacts, invoices and payments from a 'Gestion YYYY.xlsx' file.

    Expected sheets:
    - 'Factures' or 'Clients' : contact, date, montant, type, statut
    - 'Paiements' : date, contact, montant, mode, N° chèque

    The parser blocks recognized sheets containing invalid rows or unresolved payments.
    """
    import openpyxl  # noqa: PLC0415

    result = ImportResult()

    file_hash = _compute_file_hash(file_bytes)
    preview = await _pkg.preview_gestion_file(db, file_bytes, file_name)
    if preview.errors:
        result.absorb_preview(preview)
        await _record_import_log(
            db,
            import_type="gestion",
            status="blocked",
            file_hash=file_hash,
            file_name=file_name,
            result=result,
        )
        return result

    try:
        from io import BytesIO  # noqa: PLC0415

        wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        open_err = ImportFileOpenError(str(exc), file_name=file_name)
        result.add_open_file_error(open_err)
        await _record_import_log(
            db,
            import_type="gestion",
            status="failed",
            file_hash=file_hash,
            file_name=file_name,
            result=result,
        )
        return result

    sheets_by_kind: dict[str, list[Any]] = {kind: [] for kind in _GESTION_IMPORT_ORDER}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        kind, _ = _classify_gestion_sheet(sheet_name)
        if kind in sheets_by_kind:
            sheets_by_kind[kind].append(ws)

    for kind in _GESTION_IMPORT_ORDER:
        try:
            for ws in sheets_by_kind[kind]:
                if kind == "invoices":
                    await _pkg._import_invoices_sheet(db, ws, result)
                elif kind == "payments":
                    await _pkg._import_payments_sheet(db, ws, result)
                elif kind == "contacts":
                    await _pkg._import_contacts_sheet(db, ws, result)
                elif kind == "salaries":
                    await _pkg._import_salaries_sheet(db, ws, result)
                elif kind == "cash":
                    await _pkg._import_cash_sheet(db, ws, result)
                elif kind == "bank":
                    await _pkg._import_bank_sheet(db, ws, result)
        except _ImportSheetFailure:
            logger.error("Gestion import aborted during %s (sheet failure)", kind, exc_info=True)
            await db.rollback()
            result.reset_persisted_counts()
            await _record_import_log(
                db,
                import_type="gestion",
                status="failed",
                file_hash=file_hash,
                file_name=file_name,
                result=result,
            )
            return result
        except Exception as exc:
            logger.error("Gestion import aborted during %s: %s", kind, exc, exc_info=True)
            await db.rollback()
            result.reset_persisted_counts()
            result.add_import_error("gestion", exc)
            await _record_import_log(
                db,
                import_type="gestion",
                status="failed",
                file_hash=file_hash,
                file_name=file_name,
                result=result,
            )
            return result

    await _record_import_log(
        db,
        import_type="gestion",
        status="success",
        file_hash=file_hash,
        file_name=file_name,
        result=result,
    )
    return result


async def import_comptabilite_file(
    db: AsyncSession, file_bytes: bytes, file_name: str | None = None
) -> ImportResult:
    """Import accounting entries from a 'Comptabilité YYYY.xlsx' file.

    Expected sheet columns (flexible column detection):
    date | N° pièce | compte | libellé | débit | crédit
    """
    import openpyxl  # noqa: PLC0415

    result = ImportResult()

    file_hash = _compute_file_hash(file_bytes)
    preview = await _pkg.preview_comptabilite_file(db, file_bytes)
    if preview.errors:
        result.absorb_preview(preview)
        await _record_import_log(
            db,
            import_type="comptabilite",
            status="blocked",
            file_hash=file_hash,
            file_name=file_name,
            result=result,
        )
        return result

    try:
        from io import BytesIO  # noqa: PLC0415

        wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        open_err = ImportFileOpenError(str(exc), file_name=file_name)
        result.add_open_file_error(open_err)
        await _record_import_log(
            db,
            import_type="comptabilite",
            status="failed",
            file_hash=file_hash,
            file_name=file_name,
            result=result,
        )
        return result

    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            kind, _ = _classify_comptabilite_sheet(sheet_name)
            if kind == "entries":
                await _pkg._import_entries_sheet(db, ws, result)
    except _ImportSheetFailure:
        logger.error("Comptabilite import aborted (sheet failure)", exc_info=True)
        await db.rollback()
        result.reset_persisted_counts()
        await _record_import_log(
            db,
            import_type="comptabilite",
            status="failed",
            file_hash=file_hash,
            file_name=file_name,
            result=result,
        )
        return result
    except Exception as exc:
        logger.error("Comptabilite import aborted: %s", exc, exc_info=True)
        await db.rollback()
        result.reset_persisted_counts()
        result.add_import_error("comptabilite", exc)
        await _record_import_log(
            db,
            import_type="comptabilite",
            status="failed",
            file_hash=file_hash,
            file_name=file_name,
            result=result,
        )
        return result

    await _record_import_log(
        db,
        import_type="comptabilite",
        status="success",
        file_hash=file_hash,
        file_name=file_name,
        result=result,
    )
    return result
