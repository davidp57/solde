"""Excel import service — parse and import contacts, invoices, payments from Excel files."""

from __future__ import annotations

from datetime import date
from typing import Any

from sqlalchemy.ext.asyncio import AsyncSession

from backend.services.excel_import._comparison import _resolve_comparison_window
from backend.services.excel_import._comparison_domains import (
    _build_comptabilite_preview_comparison_domains,
    _build_gestion_preview_comparison_domains,
    _collect_comptabilite_extra_in_solde_by_kind,
)
from backend.services.excel_import._entry_groups import (
    _supplier_invoice_candidate_from_bank_row,
    _supplier_invoice_candidate_from_cash_row,
)
from backend.services.excel_import._exceptions import ImportFileOpenError
from backend.services.excel_import._invoices import _single_client_invoice_reference
from backend.services.excel_import._preview_existing import (
    _add_comptabilite_existing_rows_preview,
    _add_gestion_existing_rows_preview,
    _add_gestion_payment_validation,
)
from backend.services.excel_import._sheet_wrappers import (
    _parse_bank_sheet,
    _parse_cash_sheet,
    _parse_contact_sheet,
    _parse_entries_sheet,
    _parse_invoice_sheet,
    _parse_payment_sheet,
    _parse_salary_sheet,
)
from backend.services.excel_import_classification import (
    classify_comptabilite_sheet as _classify_comptabilite_sheet,
)
from backend.services.excel_import_classification import (
    classify_gestion_sheet as _classify_gestion_sheet,
)
from backend.services.excel_import_classification import (
    sheet_has_content as _sheet_has_content,
)
from backend.services.excel_import_parsing import (
    format_contact_display_name as _format_contact_display_name,
)
from backend.services.excel_import_parsing import (
    normalize_text as _normalize_text,
)
from backend.services.excel_import_parsing import (
    parse_str as _parse_str,
)
from backend.services.excel_import_policy import (
    COMPTABILITE_UNKNOWN_STRUCTURE_MESSAGE,
    GESTION_UNKNOWN_STRUCTURE_MESSAGE,
    detect_gestion_preview_header,
    filter_duplicate_contact_rows,
    filter_duplicate_invoice_rows,
    preview_warning_for_comptabilite_reason,
    preview_warning_for_gestion_reason,
)
from backend.services.excel_import_preview_helpers import (
    append_finalized_sheet_preview as _append_finalized_sheet_preview,
)
from backend.services.excel_import_preview_helpers import (
    append_ignored_issues as _append_ignored_issues,
)
from backend.services.excel_import_preview_helpers import (
    append_preview_open_error as _append_preview_open_error,
)
from backend.services.excel_import_preview_helpers import (
    append_reasoned_ignored_sheet_preview as _append_reasoned_ignored_sheet_preview,
)
from backend.services.excel_import_preview_helpers import (
    append_row_issues as _append_row_issues,
)
from backend.services.excel_import_preview_helpers import (
    append_unknown_structure_sheet_preview as _append_unknown_structure_sheet_preview,
)
from backend.services.excel_import_preview_helpers import (
    finalize_preview_can_import as _finalize_preview_can_import,
)
from backend.services.excel_import_preview_helpers import (
    register_preview_contact as _register_preview_contact,
)
from backend.services.excel_import_results import PreviewResult
from backend.services.excel_import_sheet_helpers import (
    detect_header_row as _detect_header_row,
)
from backend.services.excel_import_state import (
    add_comptabilite_coexistence_validation as _add_comptabilite_coexistence_validation,
)
from backend.services.excel_import_state import (
    compute_file_hash as _compute_file_hash,
)
from backend.services.excel_import_state import (
    find_successful_import_log as _find_successful_import_log,
)
from backend.services.excel_import_state import (
    mark_preview_as_already_imported as _mark_preview_as_already_imported,
)


def _collect_sample_rows(
    ws: Any, header_row: int, col_map: dict[str, int], *, limit: int = 3
) -> list[dict[str, str]]:
    samples: list[dict[str, str]] = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if all(cell is None for cell in row):
            continue
        samples.append(
            {key: _parse_str(row[idx] if idx < len(row) else None) for key, idx in col_map.items()}
        )
        if len(samples) >= limit:
            break
    return samples


def _preview_sheet_gestion(ws: Any, sheet_name: str, preview: PreviewResult) -> None:
    """Count parseable rows and report diagnostics for gestion files."""
    kind, reason = _classify_gestion_sheet(sheet_name)
    if kind is None:
        _append_reasoned_ignored_sheet_preview(
            preview,
            sheet_name=sheet_name,
            has_content=_sheet_has_content(ws),
            warning=preview_warning_for_gestion_reason(reason),
        )
        return

    header_info = detect_gestion_preview_header(
        ws,
        kind,
        detect_header_row=_detect_header_row,
    )

    if header_info is None:
        _append_unknown_structure_sheet_preview(
            preview,
            sheet_name=sheet_name,
            kind=kind,
            warning=GESTION_UNKNOWN_STRUCTURE_MESSAGE,
        )
        return

    header_row, col_map = header_info
    detected_columns = list(col_map.keys())
    count = 0
    missing_columns: list[str] = []
    ignored_rows = 0
    errors: list[str] = []
    warnings: list[str] = []

    if kind == "contacts":
        parsed_sheet, contact_rows, row_issues = _parse_contact_sheet(ws)
        if parsed_sheet is None:
            _append_unknown_structure_sheet_preview(
                preview,
                sheet_name=sheet_name,
                kind=kind,
                warning=GESTION_UNKNOWN_STRUCTURE_MESSAGE,
            )
            return
        header_row = parsed_sheet.header_row
        detected_columns = list(parsed_sheet.col_map.keys())
        missing_columns = parsed_sheet.missing_columns
        contact_rows, ignored_issues = filter_duplicate_contact_rows(
            contact_rows,
            normalize_text=_normalize_text,
        )
        ignored_rows = len(ignored_issues)
        count = len(contact_rows)
        _append_row_issues(errors, row_issues)
        _append_ignored_issues(warnings, ignored_issues)
        if not missing_columns:
            for contact_row in contact_rows:
                _register_preview_contact(
                    preview,
                    _format_contact_display_name(contact_row.nom, contact_row.prenom),
                )
    elif kind == "invoices":
        parsed_sheet, invoice_rows, row_issues, ignored_issues = _parse_invoice_sheet(ws)
        if parsed_sheet is None:
            _append_unknown_structure_sheet_preview(
                preview,
                sheet_name=sheet_name,
                kind=kind,
                warning=GESTION_UNKNOWN_STRUCTURE_MESSAGE,
            )
            return
        header_row = parsed_sheet.header_row
        detected_columns = list(parsed_sheet.col_map.keys())
        missing_columns = parsed_sheet.missing_columns
        invoice_rows, duplicate_ignored_issues = filter_duplicate_invoice_rows(
            invoice_rows,
            normalize_text=_normalize_text,
        )
        ignored_rows = len(ignored_issues) + len(duplicate_ignored_issues)
        count = len(invoice_rows)
        _append_row_issues(errors, row_issues)
        _append_ignored_issues(warnings, ignored_issues)
        _append_ignored_issues(warnings, duplicate_ignored_issues)
        if not missing_columns:
            for invoice_row in invoice_rows:
                _register_preview_contact(preview, invoice_row.contact_name)
            preview.estimated_invoices += count
    elif kind == "payments":
        parsed_sheet, payment_rows, row_issues = _parse_payment_sheet(ws)
        if parsed_sheet is None:
            _append_unknown_structure_sheet_preview(
                preview,
                sheet_name=sheet_name,
                kind=kind,
                warning=GESTION_UNKNOWN_STRUCTURE_MESSAGE,
            )
            return
        header_row = parsed_sheet.header_row
        detected_columns = list(parsed_sheet.col_map.keys())
        missing_columns = parsed_sheet.missing_columns
        count = len(payment_rows)
        _append_row_issues(errors, row_issues)
        if not missing_columns:
            for payment_row in payment_rows:
                if payment_row.contact_name:
                    _register_preview_contact(preview, payment_row.contact_name)
            preview.estimated_payments += count
    elif kind == "salaries":
        parsed_sheet, salary_rows, row_issues = _parse_salary_sheet(ws)
        if parsed_sheet is None:
            _append_unknown_structure_sheet_preview(
                preview,
                sheet_name=sheet_name,
                kind=kind,
                warning=GESTION_UNKNOWN_STRUCTURE_MESSAGE,
            )
            return
        header_row = parsed_sheet.header_row
        detected_columns = list(parsed_sheet.col_map.keys())
        missing_columns = parsed_sheet.missing_columns
        count = len(salary_rows)
        _append_row_issues(errors, row_issues)
        if not missing_columns:
            for salary_row in salary_rows:
                _register_preview_contact(preview, salary_row.employee_name)
            preview.estimated_salaries += count
    elif kind == "bank":
        parsed_sheet, bank_rows, row_issues, ignored_issues = _parse_bank_sheet(ws)
        if parsed_sheet is None:
            _append_unknown_structure_sheet_preview(
                preview,
                sheet_name=sheet_name,
                kind=kind,
                warning=GESTION_UNKNOWN_STRUCTURE_MESSAGE,
            )
            return
        header_row = parsed_sheet.header_row
        detected_columns = list(parsed_sheet.col_map.keys())
        missing_columns = parsed_sheet.missing_columns
        count = len(bank_rows)
        _append_row_issues(errors, row_issues)
        ignored_rows = len(ignored_issues)
        _append_ignored_issues(warnings, ignored_issues)
        if not missing_columns:
            for bank_row in bank_rows:
                candidate = _supplier_invoice_candidate_from_bank_row(bank_row)
                if candidate is not None:
                    preview.estimated_invoices += 1
                    preview.estimated_payments += 1
                    _register_preview_contact(preview, candidate.contact_name)
                    continue
                if _single_client_invoice_reference(bank_row.reference, bank_row.description):
                    preview.estimated_payments += 1
    elif kind == "cash":
        parsed_sheet, cash_rows, row_issues, ignored_issues = _parse_cash_sheet(ws)
        if parsed_sheet is None:
            _append_unknown_structure_sheet_preview(
                preview,
                sheet_name=sheet_name,
                kind=kind,
                warning=GESTION_UNKNOWN_STRUCTURE_MESSAGE,
            )
            return
        header_row = parsed_sheet.header_row
        detected_columns = list(parsed_sheet.col_map.keys())
        missing_columns = parsed_sheet.missing_columns
        count = len(cash_rows)
        _append_row_issues(errors, row_issues)
        ignored_rows = len(ignored_issues)
        _append_ignored_issues(warnings, ignored_issues)
        if not missing_columns:
            for cash_row in cash_rows:
                candidate = _supplier_invoice_candidate_from_cash_row(cash_row)
                if candidate is None:
                    continue
                preview.estimated_invoices += 1
                preview.estimated_payments += 1
                _register_preview_contact(preview, candidate.contact_name)

    assert parsed_sheet is not None
    col_map = parsed_sheet.col_map
    _append_finalized_sheet_preview(
        preview,
        sheet_name=sheet_name,
        kind=kind,
        header_row=header_row,
        rows=count,
        detected_columns=detected_columns,
        missing_columns=missing_columns,
        ignored_rows=ignored_rows,
        sample_rows=_collect_sample_rows(ws, header_row, col_map),
        warnings=warnings,
        errors=errors,
    )


def _preview_sheet_comptabilite(ws: Any, sheet_name: str, preview: PreviewResult) -> None:
    """Count parseable accounting entry rows and report diagnostics."""
    kind, reason = _classify_comptabilite_sheet(sheet_name)
    if kind is None:
        _append_reasoned_ignored_sheet_preview(
            preview,
            sheet_name=sheet_name,
            has_content=_sheet_has_content(ws),
            warning=preview_warning_for_comptabilite_reason(reason),
        )
        return

    parsed_sheet, normalized_rows, row_issues, ignored_issues = _parse_entries_sheet(ws)
    if parsed_sheet is None:
        _append_unknown_structure_sheet_preview(
            preview,
            sheet_name=sheet_name,
            kind=kind,
            warning=COMPTABILITE_UNKNOWN_STRUCTURE_MESSAGE,
        )
        return

    header_row = parsed_sheet.header_row
    detected_columns = list(parsed_sheet.col_map.keys())
    missing_columns = parsed_sheet.missing_columns
    count = len(normalized_rows)
    warnings: list[str] = []
    errors: list[str] = []
    _append_row_issues(errors, row_issues)
    _append_ignored_issues(warnings, ignored_issues)
    if not missing_columns:
        preview.estimated_entries += count

    _append_finalized_sheet_preview(
        preview,
        sheet_name=sheet_name,
        kind=kind,
        header_row=header_row,
        rows=count,
        detected_columns=detected_columns,
        missing_columns=missing_columns,
        ignored_rows=len(ignored_issues),
        sample_rows=_collect_sample_rows(ws, header_row, parsed_sheet.col_map),
        warnings=warnings,
        errors=errors,
    )


def _preview_gestion_file(file_bytes: bytes) -> PreviewResult:
    """Dry-run parse of a Gestion file using file-only validation."""
    from io import BytesIO  # noqa: PLC0415

    import openpyxl  # noqa: PLC0415

    preview = PreviewResult()
    try:
        wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        open_err = ImportFileOpenError(str(exc))
        _append_preview_open_error(preview, open_err)
        return preview

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        _preview_sheet_gestion(ws, sheet_name, preview)

    return preview


async def preview_gestion_file(
    db: AsyncSession,
    file_bytes: bytes,
    file_name: str | None = None,
    comparison_start_date: date | None = None,
    comparison_end_date: date | None = None,
) -> PreviewResult:
    """Dry-run parse of a Gestion file with shared business validation."""
    preview = _preview_gestion_file(file_bytes)
    preview.comparison_mode = "gestion-excel-to-solde"
    if preview.errors:
        preview.can_import = False
        return preview

    comparison_window = _resolve_comparison_window(
        file_name,
        comparison_start_date,
        comparison_end_date,
    )
    preview.comparison_context["domains"] = await _build_gestion_preview_comparison_domains(
        db,
        file_bytes,
        file_name,
        comparison_window,
    )

    await _add_gestion_existing_rows_preview(db, file_bytes, preview)

    import_log = await _find_successful_import_log(db, "gestion", _compute_file_hash(file_bytes))
    if import_log is not None:
        _mark_preview_as_already_imported(preview, "gestion", import_log)
        return preview

    await _add_gestion_payment_validation(db, file_bytes, preview)
    _finalize_preview_can_import(preview)
    return preview


async def preview_comptabilite_file(
    db: AsyncSession,
    file_bytes: bytes,
    file_name: str | None = None,
    comparison_start_date: date | None = None,
    comparison_end_date: date | None = None,
) -> PreviewResult:
    """Dry-run parse of a Comptabilité file — no DB writes."""
    from io import BytesIO  # noqa: PLC0415

    import openpyxl  # noqa: PLC0415

    preview = PreviewResult()
    try:
        wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        open_err = ImportFileOpenError(str(exc))
        _append_preview_open_error(preview, open_err)
        return preview

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        _preview_sheet_comptabilite(ws, sheet_name, preview)

    comparison_window = _resolve_comparison_window(
        file_name,
        comparison_start_date,
        comparison_end_date,
    )
    preview.comparison_context["domains"] = await _build_comptabilite_preview_comparison_domains(
        db,
        file_bytes,
        file_name,
        comparison_window,
    )

    await _add_comptabilite_coexistence_validation(db, preview)
    await _add_comptabilite_existing_rows_preview(db, file_bytes, preview)
    preview.comparison_mode = "global-convergence"
    preview.comparison_context[
        "extra_in_solde_by_kind"
    ] = await _collect_comptabilite_extra_in_solde_by_kind(db, file_bytes, preview)
    if preview.errors:
        preview.can_import = False
        return preview

    import_log = await _find_successful_import_log(
        db, "comptabilite", _compute_file_hash(file_bytes)
    )
    if import_log is not None:
        _mark_preview_as_already_imported(preview, "comptabilite", import_log)
        return preview

    _finalize_preview_can_import(preview)

    return preview
