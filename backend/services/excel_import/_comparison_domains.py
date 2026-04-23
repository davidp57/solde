"""Excel import service — parse and import contacts, invoices, payments from Excel files."""

from __future__ import annotations

from datetime import date
from typing import Any

from sqlalchemy.ext.asyncio import AsyncSession

from backend.services.excel_import._comparison import (
    _build_salary_row_months,
    _build_sheet_row_dates,
    _comparison_years_within_bounds,
    _expand_date_bounds,
    _filter_date_issues_in_window,
    _filter_salary_issues_in_window,
    _gestion_bank_comparison_signature,
    _gestion_cash_comparison_signature,
    _gestion_file_fiscal_year_bounds,
    _gestion_payment_comparison_signature,
    _gestion_salary_comparison_signature,
    _is_salary_month_within_comparison_window,
    _is_within_comparison_window,
    _is_within_date_bounds,
    _PreviewFilterWindow,
)
from backend.services.excel_import._comparison_loaders import (
    _load_existing_bank_comparison_items,
    _load_existing_bank_comparison_signatures,
    _load_existing_cash_comparison_items,
    _load_existing_cash_comparison_signatures,
    _load_existing_client_invoice_comparison_items,
    _load_existing_client_invoice_comparison_signatures,
    _load_existing_client_payment_comparison_items,
    _load_existing_client_payment_comparison_signatures,
    _load_existing_salary_comparison_items,
    _load_existing_salary_comparison_signatures,
)
from backend.services.excel_import._constants import (
    _load_existing_generated_accounting_group_signatures,
)
from backend.services.excel_import._entry_groups import (
    _build_entry_row_groups,
    _invoice_row_from_supplier_candidate,
    _matching_existing_salary_entry_group,
    _normalized_entry_group_signature,
    _supplier_invoice_candidate_from_bank_row,
    _supplier_invoice_candidate_from_cash_row,
)
from backend.services.excel_import._invoices import (
    _client_settlement_account_from_method,
    _load_existing_client_payment_reference_signatures,
    _load_existing_supplier_payment_reference_signatures,
    _matching_existing_client_invoice_reference,
    _matching_existing_client_payment_reference,
    _matching_existing_supplier_invoice_payment_reference,
    _merge_existing_client_invoice_entry_groups,
)
from backend.services.excel_import._loaders import _load_existing_salary_group_signatures
from backend.services.excel_import._sheet_wrappers import (
    _parse_bank_sheet,
    _parse_cash_sheet,
    _parse_entries_sheet,
    _parse_invoice_sheet,
    _parse_payment_sheet,
    _parse_salary_sheet,
)
from backend.services.excel_import_classification import (
    classify_comptabilite_sheet as _classify_comptabilite_sheet,
)
from backend.services.excel_import_classification import (
    classify_gestion_sheet as _classify_gestion_sheet,
)
from backend.services.excel_import_parsing import (
    normalize_text as _normalize_text,
)
from backend.services.excel_import_payment_matching import (
    PaymentMatchCandidate,
)
from backend.services.excel_import_payment_matching import (
    make_workbook_invoice_candidate as _make_workbook_invoice_candidate,
)
from backend.services.excel_import_payment_matching import (
    resolve_payment_match_with_database as _resolve_payment_match,
)
from backend.services.excel_import_policy import (
    filter_duplicate_invoice_rows,
    find_ambiguous_invoice_contact_issues,
    find_existing_invoice_issues,
    make_payment_resolution_issue,
    resolve_invoice_contact_match,
)
from backend.services.excel_import_preview_helpers import (
    find_sheet_preview as _find_sheet_preview,
)
from backend.services.excel_import_results import PreviewResult
from backend.services.excel_import_state import (
    accounting_entry_signature as _accounting_entry_signature,
)
from backend.services.excel_import_state import (
    load_existing_accounting_entry_signatures as _load_existing_accounting_entry_signatures,
)
from backend.services.excel_import_state import (
    load_existing_contacts_by_preview_key as _load_existing_contacts_by_preview_key,
)
from backend.services.excel_import_state import (
    load_existing_invoice_numbers as _load_existing_invoice_numbers,
)
from backend.services.excel_import_types import (
    NormalizedPaymentRow,
)


def _make_preview_comparison_domain(
    *,
    kind: str,
    file_rows: int,
    already_in_solde: int,
    missing_in_solde: int,
    extra_in_solde: int,
    ignored_by_policy: int,
    blocked: int,
    extra_in_solde_details: list[dict[str, str]] | None = None,
) -> dict[str, Any]:
    domain = {
        "kind": kind,
        "file_rows": file_rows,
        "already_in_solde": already_in_solde,
        "missing_in_solde": missing_in_solde,
        "extra_in_solde": extra_in_solde,
        "ignored_by_policy": ignored_by_policy,
        "blocked": blocked,
    }
    if extra_in_solde_details:
        domain["extra_in_solde_details"] = extra_in_solde_details
    return domain


async def _build_gestion_preview_comparison_domains(
    db: AsyncSession,
    file_bytes: bytes,
    file_name: str | None,
    comparison_window: _PreviewFilterWindow | None,
) -> list[dict[str, Any]]:
    from io import BytesIO  # noqa: PLC0415

    import openpyxl  # noqa: PLC0415

    existing_contacts_by_preview_key = await _load_existing_contacts_by_preview_key(db)
    default_extra_bounds = _gestion_file_fiscal_year_bounds(file_name)
    default_extra_start_date = default_extra_bounds[0] if default_extra_bounds else None
    default_extra_end_date = default_extra_bounds[1] if default_extra_bounds else None
    workbook_candidates: list[PaymentMatchCandidate] = []

    wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        kind, _ = _classify_gestion_sheet(sheet_name)
        if kind != "invoices":
            continue
        parsed_sheet, invoice_rows, _, _ = _parse_invoice_sheet(ws)
        if parsed_sheet is None or parsed_sheet.missing_columns:
            continue
        row_dates = {
            invoice_row.source_row_number: invoice_row.invoice_date for invoice_row in invoice_rows
        }
        deduped_invoice_rows, _ = filter_duplicate_invoice_rows(
            invoice_rows,
            normalize_text=_normalize_text,
        )
        workbook_candidates.extend(
            _make_workbook_invoice_candidate(invoice_row) for invoice_row in deduped_invoice_rows
        )

    domains: list[dict[str, Any]] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        kind, _ = _classify_gestion_sheet(sheet_name)
        if kind is None or kind == "contacts":
            continue

        if kind == "invoices":
            parsed_sheet, invoice_rows, row_issues, ignored_issues = _parse_invoice_sheet(ws)
            if parsed_sheet is None or parsed_sheet.missing_columns:
                continue
            row_dates = {
                invoice_row.source_row_number: invoice_row.invoice_date
                for invoice_row in invoice_rows
            }
            deduped_invoice_rows, duplicate_ignored_issues = filter_duplicate_invoice_rows(
                invoice_rows,
                normalize_text=_normalize_text,
            )
            filtered_invoice_rows = (
                [
                    invoice_row
                    for invoice_row in deduped_invoice_rows
                    if _is_within_comparison_window(invoice_row.invoice_date, comparison_window)
                ]
                if comparison_window is not None
                else deduped_invoice_rows
            )
            filtered_ignored_issues = (
                _filter_date_issues_in_window(
                    ignored_issues + duplicate_ignored_issues,
                    row_dates,
                    comparison_window,
                )
                if comparison_window is not None
                else ignored_issues + duplicate_ignored_issues
            )
            filtered_blocked_issues = (
                _filter_date_issues_in_window(
                    row_issues,
                    _build_sheet_row_dates(ws, parsed_sheet),
                    comparison_window,
                )
                if comparison_window is not None
                else row_issues
            )

            invoice_years = {invoice_row.invoice_date.year for invoice_row in filtered_invoice_rows}
            invoice_start_date = (
                comparison_window.start_date
                if comparison_window is not None
                else default_extra_start_date
            )
            invoice_end_date = (
                comparison_window.end_date
                if comparison_window is not None
                else default_extra_end_date
            )
            invoice_comparison_years = _comparison_years_within_bounds(
                invoice_years,
                invoice_start_date,
                invoice_end_date,
            )
            existing_invoice_numbers = await _load_existing_client_invoice_comparison_signatures(
                db,
                invoice_comparison_years,
                start_date=invoice_start_date,
                end_date=invoice_end_date,
            )
            existing_invoice_items = await _load_existing_client_invoice_comparison_items(
                db,
                invoice_comparison_years,
                start_date=invoice_start_date,
                end_date=invoice_end_date,
            )
            existing_invoice_issues = find_existing_invoice_issues(
                filtered_invoice_rows,
                existing_invoice_numbers,
                normalize_text=_normalize_text,
            )
            ambiguous_invoice_issues = find_ambiguous_invoice_contact_issues(
                filtered_invoice_rows,
                existing_contacts_by_preview_key,
                normalize_text=_normalize_text,
            )
            workbook_invoice_numbers = {
                _normalize_text(invoice_row.invoice_number)
                for invoice_row in filtered_invoice_rows
                if invoice_row.invoice_number
            }
            extra_invoice_numbers = sorted(existing_invoice_numbers - workbook_invoice_numbers)
            domains.append(
                _make_preview_comparison_domain(
                    kind="invoices",
                    file_rows=len(filtered_invoice_rows)
                    + len(filtered_ignored_issues)
                    + len(filtered_blocked_issues),
                    already_in_solde=len(existing_invoice_issues),
                    missing_in_solde=max(
                        0,
                        len(filtered_invoice_rows)
                        - len(existing_invoice_issues)
                        - len(ambiguous_invoice_issues),
                    ),
                    extra_in_solde=len(extra_invoice_numbers),
                    ignored_by_policy=len(filtered_ignored_issues),
                    blocked=len(filtered_blocked_issues) + len(ambiguous_invoice_issues),
                    extra_in_solde_details=[
                        existing_invoice_items[number] for number in extra_invoice_numbers
                    ],
                )
            )
            continue

        if kind == "payments":
            parsed_sheet, payment_rows, row_issues = _parse_payment_sheet(ws)
            if parsed_sheet is None or parsed_sheet.missing_columns:
                continue
            filtered_payment_rows = (
                [
                    payment_row
                    for payment_row in payment_rows
                    if _is_within_comparison_window(payment_row.payment_date, comparison_window)
                ]
                if comparison_window is not None
                else payment_rows
            )
            filtered_blocked_issues = (
                _filter_date_issues_in_window(
                    row_issues,
                    _build_sheet_row_dates(ws, parsed_sheet),
                    comparison_window,
                )
                if comparison_window is not None
                else row_issues
            )
            payment_start_date = (
                comparison_window.start_date
                if comparison_window is not None
                else default_extra_start_date
            )
            payment_end_date = (
                comparison_window.end_date
                if comparison_window is not None
                else default_extra_end_date
            )
            payment_comparison_years = _comparison_years_within_bounds(
                {payment_row.payment_date.year for payment_row in filtered_payment_rows},
                payment_start_date,
                payment_end_date,
            )
            existing_payment_signatures = await _load_existing_client_payment_comparison_signatures(
                db,
                payment_comparison_years,
                start_date=payment_start_date,
                end_date=payment_end_date,
            )
            existing_payment_items = await _load_existing_client_payment_comparison_items(
                db,
                payment_comparison_years,
                start_date=payment_start_date,
                end_date=payment_end_date,
            )
            already_in_solde = 0
            blocked_validation = 0
            workbook_payment_signatures: set[tuple[str, str, str, str, str]] = set()
            for payment_row in filtered_payment_rows:
                resolution = await _resolve_payment_match(
                    db,
                    payment_row,
                    workbook_candidates,
                )
                blocking_issue = make_payment_resolution_issue(
                    source_row_number=payment_row.source_row_number,
                    status=resolution.status,
                    candidate=resolution.candidate,
                    message=resolution.message,
                    require_persistable_candidate=False,
                )
                if blocking_issue is not None:
                    blocked_validation += 1
                    continue
                if resolution.candidate is None or not resolution.candidate.invoice_number:
                    continue
                payment_signature = _gestion_payment_comparison_signature(
                    reference=resolution.candidate.invoice_number,
                    payment_date=payment_row.payment_date,
                    amount=payment_row.amount,
                    settlement_account=_client_settlement_account_from_method(payment_row.method),
                    cheque_number=payment_row.cheque_number,
                )
                workbook_payment_signatures.add(payment_signature)
                if payment_signature in existing_payment_signatures:
                    already_in_solde += 1
            extra_payment_signatures = sorted(
                existing_payment_signatures - workbook_payment_signatures
            )
            domains.append(
                _make_preview_comparison_domain(
                    kind="payments",
                    file_rows=len(filtered_payment_rows) + len(filtered_blocked_issues),
                    already_in_solde=already_in_solde,
                    missing_in_solde=max(
                        0,
                        len(filtered_payment_rows) - already_in_solde - blocked_validation,
                    ),
                    extra_in_solde=len(extra_payment_signatures),
                    ignored_by_policy=0,
                    blocked=len(filtered_blocked_issues) + blocked_validation,
                    extra_in_solde_details=[
                        existing_payment_items[signature] for signature in extra_payment_signatures
                    ],
                )
            )
            continue

        if kind == "salaries":
            parsed_sheet, salary_rows, row_issues = _parse_salary_sheet(ws)
            if parsed_sheet is None:
                continue
            filtered_salary_rows = (
                [
                    salary_row
                    for salary_row in salary_rows
                    if _is_salary_month_within_comparison_window(
                        salary_row.month, comparison_window
                    )
                ]
                if comparison_window is not None
                else salary_rows
            )
            filtered_blocked_issues = (
                _filter_salary_issues_in_window(
                    row_issues,
                    _build_salary_row_months(ws),
                    comparison_window,
                )
                if comparison_window is not None
                else row_issues
            )
            salary_months = {salary_row.month for salary_row in filtered_salary_rows}
            existing_salary_signatures = await _load_existing_salary_comparison_signatures(
                db,
                salary_months,
            )
            existing_salary_items = await _load_existing_salary_comparison_items(
                db,
                salary_months,
            )
            workbook_salary_signatures = {
                _gestion_salary_comparison_signature(
                    month=salary_row.month,
                    employee_name=salary_row.employee_name,
                    gross=salary_row.gross,
                    net_pay=salary_row.net_pay,
                )
                for salary_row in filtered_salary_rows
            }
            already_in_solde = sum(
                1
                for signature in workbook_salary_signatures
                if signature in existing_salary_signatures
            )
            extra_salary_signatures = sorted(
                existing_salary_signatures - workbook_salary_signatures
            )
            domains.append(
                _make_preview_comparison_domain(
                    kind="salaries",
                    file_rows=len(filtered_salary_rows) + len(filtered_blocked_issues),
                    already_in_solde=already_in_solde,
                    missing_in_solde=max(0, len(filtered_salary_rows) - already_in_solde),
                    extra_in_solde=len(extra_salary_signatures),
                    ignored_by_policy=0,
                    blocked=len(filtered_blocked_issues),
                    extra_in_solde_details=[
                        existing_salary_items[signature] for signature in extra_salary_signatures
                    ],
                )
            )
            continue

        if kind == "bank":
            parsed_sheet, bank_rows, row_issues, ignored_issues = _parse_bank_sheet(ws)
            if parsed_sheet is None or parsed_sheet.missing_columns:
                continue
            filtered_bank_rows = (
                [
                    bank_row
                    for bank_row in bank_rows
                    if _is_within_comparison_window(bank_row.entry_date, comparison_window)
                ]
                if comparison_window is not None
                else bank_rows
            )
            filtered_ignored_issues = (
                _filter_date_issues_in_window(
                    ignored_issues,
                    _build_sheet_row_dates(ws, parsed_sheet),
                    comparison_window,
                )
                if comparison_window is not None
                else ignored_issues
            )
            filtered_blocked_issues = (
                _filter_date_issues_in_window(
                    row_issues,
                    _build_sheet_row_dates(ws, parsed_sheet),
                    comparison_window,
                )
                if comparison_window is not None
                else row_issues
            )
            bank_start_date = (
                comparison_window.start_date if comparison_window is not None else None
            )
            bank_end_date = comparison_window.end_date if comparison_window is not None else None
            if comparison_window is None:
                for bank_row in filtered_bank_rows:
                    bank_start_date, bank_end_date = _expand_date_bounds(
                        bank_start_date,
                        bank_end_date,
                        bank_row.entry_date,
                    )
            bank_comparison_years = _comparison_years_within_bounds(
                {bank_row.entry_date.year for bank_row in filtered_bank_rows},
                bank_start_date,
                bank_end_date,
            )
            existing_bank_signatures = await _load_existing_bank_comparison_signatures(
                db,
                bank_comparison_years,
                start_date=bank_start_date,
                end_date=bank_end_date,
            )
            existing_bank_items = await _load_existing_bank_comparison_items(
                db,
                bank_comparison_years,
                start_date=bank_start_date,
                end_date=bank_end_date,
            )
            filtered_invoice_numbers = await _load_existing_client_invoice_comparison_signatures(
                db,
                _comparison_years_within_bounds(set(), bank_start_date, bank_end_date),
                start_date=bank_start_date,
                end_date=bank_end_date,
            )
            already_in_solde = 0
            blocked_supplier = 0
            workbook_bank_signatures: set[tuple[str, str, str, str]] = set()
            for bank_row in filtered_bank_rows:
                bank_signature = _gestion_bank_comparison_signature(
                    entry_date=bank_row.entry_date,
                    amount=bank_row.amount,
                    description=bank_row.description,
                    reference=bank_row.reference,
                )
                workbook_bank_signatures.add(bank_signature)
                candidate = _supplier_invoice_candidate_from_bank_row(bank_row)
                if bank_signature in existing_bank_signatures:
                    already_in_solde += 1
                    continue
                if candidate is None:
                    continue
                if _normalize_text(candidate.invoice_number) in filtered_invoice_numbers:
                    already_in_solde += 1
                    continue
                _, supplier_blocked_issue = resolve_invoice_contact_match(
                    _invoice_row_from_supplier_candidate(candidate),
                    existing_contacts_by_preview_key,
                    normalize_text=_normalize_text,
                )
                if supplier_blocked_issue is not None:
                    blocked_supplier += 1
            extra_bank_signatures = sorted(existing_bank_signatures - workbook_bank_signatures)
            domains.append(
                _make_preview_comparison_domain(
                    kind="bank",
                    file_rows=len(filtered_bank_rows)
                    + len(filtered_ignored_issues)
                    + len(filtered_blocked_issues),
                    already_in_solde=already_in_solde,
                    missing_in_solde=max(
                        0,
                        len(filtered_bank_rows) - already_in_solde - blocked_supplier,
                    ),
                    extra_in_solde=len(extra_bank_signatures),
                    ignored_by_policy=len(filtered_ignored_issues),
                    blocked=len(filtered_blocked_issues) + blocked_supplier,
                    extra_in_solde_details=[
                        existing_bank_items[signature] for signature in extra_bank_signatures
                    ],
                )
            )
            continue

        if kind == "cash":
            parsed_sheet, cash_rows, row_issues, ignored_issues = _parse_cash_sheet(ws)
            if parsed_sheet is None or parsed_sheet.missing_columns:
                continue
            filtered_cash_rows = (
                [
                    cash_row
                    for cash_row in cash_rows
                    if _is_within_comparison_window(cash_row.entry_date, comparison_window)
                ]
                if comparison_window is not None
                else cash_rows
            )
            filtered_ignored_issues = (
                _filter_date_issues_in_window(
                    ignored_issues,
                    _build_sheet_row_dates(ws, parsed_sheet),
                    comparison_window,
                )
                if comparison_window is not None
                else ignored_issues
            )
            filtered_blocked_issues = (
                _filter_date_issues_in_window(
                    row_issues,
                    _build_sheet_row_dates(ws, parsed_sheet),
                    comparison_window,
                )
                if comparison_window is not None
                else row_issues
            )
            cash_start_date = (
                comparison_window.start_date if comparison_window is not None else None
            )
            cash_end_date = comparison_window.end_date if comparison_window is not None else None
            if comparison_window is None:
                for cash_row in filtered_cash_rows:
                    cash_start_date, cash_end_date = _expand_date_bounds(
                        cash_start_date,
                        cash_end_date,
                        cash_row.entry_date,
                    )
            cash_comparison_years = _comparison_years_within_bounds(
                {cash_row.entry_date.year for cash_row in filtered_cash_rows},
                cash_start_date,
                cash_end_date,
            )
            existing_cash_signatures = await _load_existing_cash_comparison_signatures(
                db,
                cash_comparison_years,
                start_date=cash_start_date,
                end_date=cash_end_date,
            )
            existing_cash_items = await _load_existing_cash_comparison_items(
                db,
                cash_comparison_years,
                start_date=cash_start_date,
                end_date=cash_end_date,
            )
            filtered_invoice_numbers = await _load_existing_client_invoice_comparison_signatures(
                db,
                _comparison_years_within_bounds(set(), cash_start_date, cash_end_date),
                start_date=cash_start_date,
                end_date=cash_end_date,
            )
            already_in_solde = 0
            blocked_supplier = 0
            workbook_cash_signatures: set[tuple[str, str, str, str, str]] = set()
            for cash_row in filtered_cash_rows:
                cash_signature = _gestion_cash_comparison_signature(
                    entry_date=cash_row.entry_date,
                    movement_type=cash_row.movement_type,
                    amount=cash_row.amount,
                    description=cash_row.description,
                    reference=cash_row.reference,
                )
                workbook_cash_signatures.add(cash_signature)
                candidate = _supplier_invoice_candidate_from_cash_row(cash_row)
                if cash_signature in existing_cash_signatures:
                    already_in_solde += 1
                    continue
                if candidate is None:
                    continue
                if _normalize_text(candidate.invoice_number) in filtered_invoice_numbers:
                    already_in_solde += 1
                    continue
                _, supplier_blocked_issue = resolve_invoice_contact_match(
                    _invoice_row_from_supplier_candidate(candidate),
                    existing_contacts_by_preview_key,
                    normalize_text=_normalize_text,
                )
                if supplier_blocked_issue is not None:
                    blocked_supplier += 1
            extra_cash_signatures = sorted(existing_cash_signatures - workbook_cash_signatures)
            domains.append(
                _make_preview_comparison_domain(
                    kind="cash",
                    file_rows=len(filtered_cash_rows)
                    + len(filtered_ignored_issues)
                    + len(filtered_blocked_issues),
                    already_in_solde=already_in_solde,
                    missing_in_solde=max(
                        0,
                        len(filtered_cash_rows) - already_in_solde - blocked_supplier,
                    ),
                    extra_in_solde=len(extra_cash_signatures),
                    ignored_by_policy=len(filtered_ignored_issues),
                    blocked=len(filtered_blocked_issues) + blocked_supplier,
                    extra_in_solde_details=[
                        existing_cash_items[signature] for signature in extra_cash_signatures
                    ],
                )
            )

    return domains


async def _build_comptabilite_preview_comparison_domains(
    db: AsyncSession,
    file_bytes: bytes,
    file_name: str | None,
    comparison_window: _PreviewFilterWindow | None,
) -> list[dict[str, Any]]:
    from io import BytesIO  # noqa: PLC0415

    import openpyxl  # noqa: PLC0415

    existing_entry_signatures = await _load_existing_accounting_entry_signatures(db)
    existing_invoice_numbers = await _load_existing_invoice_numbers(db)
    existing_client_payment_signatures = await _load_existing_client_payment_reference_signatures(
        db
    )
    existing_supplier_payment_signatures = (
        await _load_existing_supplier_payment_reference_signatures(db)
    )
    existing_salary_group_signatures = await _load_existing_salary_group_signatures(db)
    existing_generated_group_signatures = (
        await _load_existing_generated_accounting_group_signatures(db)
    )

    wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    domains: list[dict[str, Any]] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        kind, _ = _classify_comptabilite_sheet(sheet_name)
        if kind != "entries":
            continue
        parsed_sheet, normalized_rows, row_issues, ignored_issues = _parse_entries_sheet(ws)
        if parsed_sheet is None or parsed_sheet.missing_columns:
            continue

        filtered_rows = [
            entry_row
            for entry_row in normalized_rows
            if _is_within_comparison_window(entry_row.entry_date, comparison_window)
        ]
        filtered_ignored_issues = _filter_date_issues_in_window(
            ignored_issues,
            _build_sheet_row_dates(ws, parsed_sheet),
            comparison_window,
        )
        filtered_blocked_issues = _filter_date_issues_in_window(
            row_issues,
            _build_sheet_row_dates(ws, parsed_sheet),
            comparison_window,
        )

        entry_groups = _build_entry_row_groups(filtered_rows)
        already_in_solde = 0
        index = 0
        while index < len(entry_groups):
            entry_group, next_index = _merge_existing_client_invoice_entry_groups(
                entry_groups,
                index,
                existing_invoice_numbers,
            )
            if _matching_existing_client_invoice_reference(entry_group, existing_invoice_numbers):
                already_in_solde += len(entry_group)
                index = next_index
                continue
            if _matching_existing_client_payment_reference(
                entry_group,
                existing_client_payment_signatures,
            ):
                already_in_solde += len(entry_group)
                index = next_index
                continue
            if _matching_existing_supplier_invoice_payment_reference(
                entry_group,
                existing_supplier_payment_signatures,
            ):
                already_in_solde += len(entry_group)
                index = next_index
                continue
            if _matching_existing_salary_entry_group(
                entry_group,
                existing_salary_group_signatures,
            ):
                already_in_solde += len(entry_group)
                index = next_index
                continue
            if (
                _normalized_entry_group_signature(entry_group)
                in existing_generated_group_signatures
            ):
                already_in_solde += len(entry_group)
                index = next_index
                continue

            for entry_row in entry_group:
                signature = _accounting_entry_signature(
                    entry_date=entry_row.entry_date,
                    account_number=entry_row.account_number,
                    label=entry_row.label,
                    debit=entry_row.debit,
                    credit=entry_row.credit,
                )
                if signature in existing_entry_signatures:
                    already_in_solde += 1
            index = next_index

        comparison_years = _comparison_years_within_bounds(
            {entry_row.entry_date.year for entry_row in filtered_rows},
            comparison_window.start_date if comparison_window else None,
            comparison_window.end_date if comparison_window else None,
        )
        existing_comparison_signatures = (
            await _load_existing_accounting_entry_comparison_signatures(
                db,
                comparison_years,
                start_date=comparison_window.start_date if comparison_window else None,
                end_date=comparison_window.end_date if comparison_window else None,
            )
        )
        workbook_entry_signatures = {
            _accounting_entry_signature(
                entry_date=entry_row.entry_date,
                account_number=entry_row.account_number,
                label=entry_row.label,
                debit=entry_row.debit,
                credit=entry_row.credit,
            )
            for entry_row in filtered_rows
        }
        domains.append(
            _make_preview_comparison_domain(
                kind="entries",
                file_rows=(
                    len(filtered_rows) + len(filtered_ignored_issues) + len(filtered_blocked_issues)
                ),
                already_in_solde=already_in_solde,
                missing_in_solde=max(0, len(filtered_rows) - already_in_solde),
                extra_in_solde=len(existing_comparison_signatures - workbook_entry_signatures),
                ignored_by_policy=len(filtered_ignored_issues),
                blocked=len(filtered_blocked_issues),
            )
        )

    return domains


async def _collect_gestion_extra_in_solde_by_kind(
    db: AsyncSession,
    file_bytes: bytes,
    preview: PreviewResult,
    file_name: str | None = None,
) -> dict[str, Any]:
    from io import BytesIO  # noqa: PLC0415

    import openpyxl  # noqa: PLC0415

    workbook_invoice_numbers: set[str] = set()
    workbook_payment_signatures: set[tuple[str, str, str, str, str]] = set()
    workbook_salary_signatures: set[tuple[str, str, str, str]] = set()
    workbook_bank_signatures: set[tuple[str, str, str, str]] = set()
    workbook_cash_signatures: set[tuple[str, str, str, str, str]] = set()
    invoice_years: set[int] = set()
    payment_years: set[int] = set()
    bank_years: set[int] = set()
    cash_years: set[int] = set()
    salary_months: set[str] = set()
    gestion_exercise_bounds = _gestion_file_fiscal_year_bounds(file_name)
    invoice_exercise_bounds = gestion_exercise_bounds
    invoice_start_date: date | None = (
        invoice_exercise_bounds[0] if invoice_exercise_bounds else None
    )
    invoice_end_date: date | None = invoice_exercise_bounds[1] if invoice_exercise_bounds else None
    payment_start_date: date | None = (
        gestion_exercise_bounds[0] if gestion_exercise_bounds else None
    )
    payment_end_date: date | None = gestion_exercise_bounds[1] if gestion_exercise_bounds else None
    bank_start_date: date | None = None
    bank_end_date: date | None = None
    cash_start_date: date | None = None
    cash_end_date: date | None = None
    workbook_candidates: list[PaymentMatchCandidate] = []
    payment_rows: list[NormalizedPaymentRow] = []

    wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        kind, _ = _classify_gestion_sheet(sheet_name)
        sheet_preview = _find_sheet_preview(preview, sheet_name)
        if kind is None or sheet_preview is None or sheet_preview["status"] != "recognized":
            continue

        if kind == "invoices":
            parsed_sheet, invoice_rows, _, _ = _parse_invoice_sheet(ws)
            if parsed_sheet is None or parsed_sheet.missing_columns:
                continue
            invoice_rows, _ = filter_duplicate_invoice_rows(
                invoice_rows,
                normalize_text=_normalize_text,
            )
            for invoice_row in invoice_rows:
                if invoice_row.invoice_number:
                    workbook_invoice_numbers.add(_normalize_text(invoice_row.invoice_number))
                invoice_years.add(invoice_row.invoice_date.year)
                if invoice_exercise_bounds is None:
                    invoice_start_date, invoice_end_date = _expand_date_bounds(
                        invoice_start_date,
                        invoice_end_date,
                        invoice_row.invoice_date,
                    )
                workbook_candidates.append(_make_workbook_invoice_candidate(invoice_row))

        elif kind == "payments":
            parsed_sheet, parsed_payment_rows, _ = _parse_payment_sheet(ws)
            if parsed_sheet is None or parsed_sheet.missing_columns:
                continue
            payment_rows.extend(parsed_payment_rows)
            for payment_row in parsed_payment_rows:
                payment_years.add(payment_row.payment_date.year)
                if gestion_exercise_bounds is None:
                    payment_start_date, payment_end_date = _expand_date_bounds(
                        payment_start_date,
                        payment_end_date,
                        payment_row.payment_date,
                    )

        elif kind == "salaries":
            parsed_sheet, salary_rows, _ = _parse_salary_sheet(ws)
            if parsed_sheet is None or parsed_sheet.missing_columns:
                continue
            for salary_row in salary_rows:
                salary_months.add(salary_row.month)
                workbook_salary_signatures.add(
                    _gestion_salary_comparison_signature(
                        month=salary_row.month,
                        employee_name=salary_row.employee_name,
                        gross=salary_row.gross,
                        net_pay=salary_row.net_pay,
                    )
                )

        elif kind == "bank":
            parsed_sheet, bank_rows, _, _ = _parse_bank_sheet(ws)
            if parsed_sheet is None or parsed_sheet.missing_columns:
                continue
            for bank_row in bank_rows:
                bank_years.add(bank_row.entry_date.year)
                bank_start_date, bank_end_date = _expand_date_bounds(
                    bank_start_date,
                    bank_end_date,
                    bank_row.entry_date,
                )
                workbook_bank_signatures.add(
                    _gestion_bank_comparison_signature(
                        entry_date=bank_row.entry_date,
                        amount=bank_row.amount,
                        description=bank_row.description,
                        reference=bank_row.reference,
                    )
                )

        elif kind == "cash":
            parsed_sheet, cash_rows, _, _ = _parse_cash_sheet(ws)
            if parsed_sheet is None or parsed_sheet.missing_columns:
                continue
            for cash_row in cash_rows:
                cash_years.add(cash_row.entry_date.year)
                cash_start_date, cash_end_date = _expand_date_bounds(
                    cash_start_date,
                    cash_end_date,
                    cash_row.entry_date,
                )
                workbook_cash_signatures.add(
                    _gestion_cash_comparison_signature(
                        entry_date=cash_row.entry_date,
                        movement_type=cash_row.movement_type,
                        amount=cash_row.amount,
                        description=cash_row.description,
                        reference=cash_row.reference,
                    )
                )

    for payment_row in payment_rows:
        resolution = await _resolve_payment_match(
            db,
            payment_row,
            workbook_candidates=workbook_candidates,
        )
        if resolution.status != "matched" or resolution.candidate is None:
            continue
        if not resolution.candidate.invoice_number:
            continue
        workbook_payment_signatures.add(
            _gestion_payment_comparison_signature(
                reference=resolution.candidate.invoice_number,
                payment_date=payment_row.payment_date,
                amount=payment_row.amount,
                settlement_account=_client_settlement_account_from_method(payment_row.method),
                cheque_number=payment_row.cheque_number,
            )
        )

    extra_in_solde_by_kind: dict[str, int] = {}
    extra_in_solde_details_by_kind: dict[str, list[dict[str, str]]] = {}
    if invoice_years:
        invoice_comparison_years = _comparison_years_within_bounds(
            invoice_years,
            invoice_start_date,
            invoice_end_date,
        )
        existing_invoice_numbers = await _load_existing_client_invoice_comparison_signatures(
            db,
            invoice_comparison_years,
            start_date=invoice_start_date,
            end_date=invoice_end_date,
        )
        existing_invoice_items = await _load_existing_client_invoice_comparison_items(
            db,
            invoice_comparison_years,
            start_date=invoice_start_date,
            end_date=invoice_end_date,
        )
        extra_invoice_numbers = sorted(existing_invoice_numbers - workbook_invoice_numbers)
        extra_in_solde_by_kind["invoices"] = len(extra_invoice_numbers)
        if extra_invoice_numbers:
            extra_in_solde_details_by_kind["invoices"] = [
                existing_invoice_items[number] for number in extra_invoice_numbers
            ]
    if payment_years:
        payment_comparison_years = _comparison_years_within_bounds(
            payment_years,
            payment_start_date,
            payment_end_date,
        )
        existing_payment_signatures = await _load_existing_client_payment_comparison_signatures(
            db,
            payment_comparison_years,
            start_date=payment_start_date,
            end_date=payment_end_date,
        )
        existing_payment_items = await _load_existing_client_payment_comparison_items(
            db,
            payment_comparison_years,
            start_date=payment_start_date,
            end_date=payment_end_date,
        )
        extra_payment_signatures = sorted(existing_payment_signatures - workbook_payment_signatures)
        extra_in_solde_by_kind["payments"] = len(extra_payment_signatures)
        if extra_payment_signatures:
            extra_in_solde_details_by_kind["payments"] = [
                existing_payment_items[signature] for signature in extra_payment_signatures
            ]
    if salary_months:
        existing_salary_signatures = await _load_existing_salary_comparison_signatures(
            db,
            salary_months,
        )
        existing_salary_items = await _load_existing_salary_comparison_items(
            db,
            salary_months,
        )
        extra_salary_signatures = sorted(existing_salary_signatures - workbook_salary_signatures)
        extra_in_solde_by_kind["salaries"] = len(extra_salary_signatures)
        if extra_salary_signatures:
            extra_in_solde_details_by_kind["salaries"] = [
                existing_salary_items[signature] for signature in extra_salary_signatures
            ]
    if bank_years:
        bank_comparison_years = _comparison_years_within_bounds(
            bank_years,
            bank_start_date,
            bank_end_date,
        )
        existing_bank_signatures = await _load_existing_bank_comparison_signatures(
            db,
            bank_comparison_years,
            start_date=bank_start_date,
            end_date=bank_end_date,
        )
        existing_bank_items = await _load_existing_bank_comparison_items(
            db,
            bank_comparison_years,
            start_date=bank_start_date,
            end_date=bank_end_date,
        )
        extra_bank_signatures = sorted(existing_bank_signatures - workbook_bank_signatures)
        extra_in_solde_by_kind["bank"] = len(extra_bank_signatures)
        if extra_bank_signatures:
            extra_in_solde_details_by_kind["bank"] = [
                existing_bank_items[signature] for signature in extra_bank_signatures
            ]
    if cash_years:
        cash_comparison_years = _comparison_years_within_bounds(
            cash_years,
            cash_start_date,
            cash_end_date,
        )
        existing_cash_signatures = await _load_existing_cash_comparison_signatures(
            db,
            cash_comparison_years,
            start_date=cash_start_date,
            end_date=cash_end_date,
        )
        existing_cash_items = await _load_existing_cash_comparison_items(
            db,
            cash_comparison_years,
            start_date=cash_start_date,
            end_date=cash_end_date,
        )
        extra_cash_signatures = sorted(existing_cash_signatures - workbook_cash_signatures)
        extra_in_solde_by_kind["cash"] = len(extra_cash_signatures)
        if extra_cash_signatures:
            extra_in_solde_details_by_kind["cash"] = [
                existing_cash_items[signature] for signature in extra_cash_signatures
            ]

    return {
        "counts": extra_in_solde_by_kind,
        "details": extra_in_solde_details_by_kind,
    }


async def _load_existing_accounting_entry_comparison_signatures(
    db: AsyncSession,
    years: set[int],
    start_date: date | None = None,
    end_date: date | None = None,
) -> set[str]:
    if not years:
        return set()

    from sqlalchemy import select  # noqa: PLC0415

    from backend.models.accounting_entry import AccountingEntry  # noqa: PLC0415

    result = await db.execute(
        select(
            AccountingEntry.date,
            AccountingEntry.account_number,
            AccountingEntry.label,
            AccountingEntry.debit,
            AccountingEntry.credit,
        )
    )
    return {
        _accounting_entry_signature(
            entry_date=entry_date,
            account_number=account_number,
            label=label,
            debit=debit,
            credit=credit,
        )
        for entry_date, account_number, label, debit, credit in result.all()
        if entry_date.year in years and _is_within_date_bounds(entry_date, start_date, end_date)
    }


async def _collect_comptabilite_extra_in_solde_by_kind(
    db: AsyncSession,
    file_bytes: bytes,
    preview: PreviewResult,
) -> dict[str, int]:
    from io import BytesIO  # noqa: PLC0415

    import openpyxl  # noqa: PLC0415

    workbook_entry_signatures: set[str] = set()
    workbook_years: set[int] = set()

    wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        kind, _ = _classify_comptabilite_sheet(sheet_name)
        sheet_preview = _find_sheet_preview(preview, sheet_name)
        if kind != "entries" or sheet_preview is None or sheet_preview["status"] != "recognized":
            continue

        parsed_sheet, normalized_rows, _, _ = _parse_entries_sheet(ws)
        if parsed_sheet is None or parsed_sheet.missing_columns:
            continue

        for entry_row in normalized_rows:
            workbook_years.add(entry_row.entry_date.year)
            workbook_entry_signatures.add(
                _accounting_entry_signature(
                    entry_date=entry_row.entry_date,
                    account_number=entry_row.account_number,
                    label=entry_row.label,
                    debit=entry_row.debit,
                    credit=entry_row.credit,
                )
            )

    if not workbook_years:
        return {}

    existing_entry_signatures = await _load_existing_accounting_entry_comparison_signatures(
        db,
        workbook_years,
    )
    return {"entries": len(existing_entry_signatures - workbook_entry_signatures)}
