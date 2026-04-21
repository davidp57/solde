"""Excel import service — parse and import contacts, invoices, payments from Excel files."""

from __future__ import annotations

import logging
import re
from dataclasses import dataclass
from datetime import date, timedelta
from decimal import Decimal
from typing import Any, TypeVar
from uuid import uuid4

from sqlalchemy.ext.asyncio import AsyncSession

from backend.config import get_settings
from backend.services import settings as settings_service
from backend.services.excel_import_classification import (
    classify_comptabilite_sheet as _classify_comptabilite_sheet,
)
from backend.services.excel_import_classification import (
    classify_gestion_sheet as _classify_gestion_sheet,
)
from backend.services.excel_import_classification import (
    sheet_has_content as _sheet_has_content,
)
from backend.services.excel_import_parsers import (
    parse_bank_sheet as _parse_bank_sheet_impl,
)
from backend.services.excel_import_parsers import (
    parse_cash_sheet as _parse_cash_sheet_impl,
)
from backend.services.excel_import_parsers import (
    parse_contact_sheet as _parse_contact_sheet_impl,
)
from backend.services.excel_import_parsers import (
    parse_entries_sheet as _parse_entries_sheet_impl,
)
from backend.services.excel_import_parsers import (
    parse_invoice_sheet as _parse_invoice_sheet_impl,
)
from backend.services.excel_import_parsers import (
    parse_payment_sheet as _parse_payment_sheet_impl,
)
from backend.services.excel_import_parsers import (
    parse_salary_sheet as _parse_salary_sheet_impl,
)
from backend.services.excel_import_parsing import (
    extract_supplier_invoice_reference as _extract_supplier_invoice_reference,
)
from backend.services.excel_import_parsing import (
    format_contact_display_name as _format_contact_display_name,
)
from backend.services.excel_import_parsing import (
    infer_supplier_contact_name as _infer_supplier_contact_name,
)
from backend.services.excel_import_parsing import (
    is_supplier_subcontracting_description as _is_supplier_subcontracting_description,
)
from backend.services.excel_import_parsing import (
    normalize_text as _normalize_text,
)
from backend.services.excel_import_parsing import (
    parse_date as _parse_date,
)
from backend.services.excel_import_parsing import (
    parse_str as _parse_str,
)
from backend.services.excel_import_parsing import (
    split_contact_full_name as _split_contact_full_name,
)
from backend.services.excel_import_parsing import (
    strip_supplier_invoice_reference as _strip_supplier_invoice_reference,
)
from backend.services.excel_import_payment_matching import (
    PaymentMatchCandidate,
)
from backend.services.excel_import_payment_matching import (
    make_workbook_invoice_candidate as _make_workbook_invoice_candidate,
)
from backend.services.excel_import_payment_matching import (
    resolve_payment_match_with_database as _resolve_payment_match,
)
from backend.services.excel_import_policy import (
    COMPTABILITE_UNKNOWN_STRUCTURE_MESSAGE,
    ENTRY_COVERED_BY_SOLDE_MESSAGE,
    ENTRY_EXISTING_MESSAGE,
    ENTRY_NEAR_EXISTING_MANUAL_MESSAGE,
    EXISTING_GESTION_ROW_MESSAGE,
    EXISTING_SALARY_MESSAGE,
    GESTION_UNKNOWN_STRUCTURE_MESSAGE,
    detect_gestion_preview_header,
    filter_duplicate_contact_rows,
    filter_duplicate_invoice_rows,
    find_ambiguous_invoice_contact_issues,
    find_existing_contact_issues,
    find_existing_invoice_issues,
    format_row_issue,
    make_existing_contact_issue,
    make_existing_invoice_issue,
    make_payment_resolution_issue,
    preview_warning_for_comptabilite_reason,
    preview_warning_for_gestion_reason,
    resolve_invoice_contact_match,
    should_ignore_cash_pending_deposit_forecast,
)
from backend.services.excel_import_preview_helpers import (
    append_finalized_sheet_preview as _append_finalized_sheet_preview,
)
from backend.services.excel_import_preview_helpers import (
    append_ignored_issues as _append_ignored_issues,
)
from backend.services.excel_import_preview_helpers import (
    append_preview_blocked_issue as _append_preview_blocked_issue,
)
from backend.services.excel_import_preview_helpers import (
    append_preview_ignored_issue as _append_preview_ignored_issue,
)
from backend.services.excel_import_preview_helpers import (
    append_preview_open_error as _append_preview_open_error,
)
from backend.services.excel_import_preview_helpers import (
    append_preview_warning_issue as _append_preview_warning_issue,
)
from backend.services.excel_import_preview_helpers import (
    append_reasoned_ignored_sheet_preview as _append_reasoned_ignored_sheet_preview,
)
from backend.services.excel_import_preview_helpers import (
    append_row_issues as _append_row_issues,
)
from backend.services.excel_import_preview_helpers import (
    append_unknown_structure_sheet_preview as _append_unknown_structure_sheet_preview,
)
from backend.services.excel_import_preview_helpers import (
    contact_preview_key as _contact_preview_key,
)
from backend.services.excel_import_preview_helpers import (
    finalize_preview_can_import as _finalize_preview_can_import,
)
from backend.services.excel_import_preview_helpers import (
    find_sheet_preview as _find_sheet_preview,
)
from backend.services.excel_import_preview_helpers import (
    recompute_preview_can_import as _recompute_preview_can_import,
)
from backend.services.excel_import_preview_helpers import (
    register_preview_contact as _register_preview_contact,
)
from backend.services.excel_import_results import ImportResult, PreviewResult
from backend.services.excel_import_sheet_helpers import (
    detect_header_row as _detect_header_row,
)
from backend.services.excel_import_sheet_helpers import (
    get_row_value as _get_row_value,
)
from backend.services.excel_import_state import (
    accounting_entry_group_signature as _accounting_entry_group_signature,
)
from backend.services.excel_import_state import (
    accounting_entry_line_signature as _accounting_entry_line_signature,
)
from backend.services.excel_import_state import (
    accounting_entry_signature as _accounting_entry_signature,
)
from backend.services.excel_import_state import (
    add_comptabilite_coexistence_validation as _add_comptabilite_coexistence_validation,
)
from backend.services.excel_import_state import (
    compute_file_hash as _compute_file_hash,
)
from backend.services.excel_import_state import (
    find_successful_import_log as _find_successful_import_log,
)
from backend.services.excel_import_state import (
    load_existing_accounting_entry_signatures as _load_existing_accounting_entry_signatures,
)
from backend.services.excel_import_state import (
    load_existing_contacts_by_preview_key as _load_existing_contacts_by_preview_key,
)
from backend.services.excel_import_state import (
    load_existing_generated_accounting_group_signatures,
    load_existing_manual_accounting_line_signatures,
)
from backend.services.excel_import_state import (
    load_existing_invoice_numbers as _load_existing_invoice_numbers,
)
from backend.services.excel_import_state import (
    mark_preview_as_already_imported as _mark_preview_as_already_imported,
)
from backend.services.excel_import_state import (
    record_import_log as _record_import_log,
)
from backend.services.excel_import_types import (
    NormalizedBankRow,
    NormalizedCashRow,
    NormalizedContactRow,
    NormalizedEntryRow,
    NormalizedInvoiceRow,
    NormalizedPaymentRow,
    NormalizedSalaryRow,
    ParsedSheet,
    RowIgnoredIssue,
    RowValidationIssue,
    RowWarningIssue,
)
from backend.services.invoice import apply_default_due_date

logger = logging.getLogger(__name__)

_load_existing_generated_accounting_group_signatures = (
    load_existing_generated_accounting_group_signatures
)

_GESTION_IMPORT_ORDER = ("contacts", "invoices", "payments", "salaries", "cash", "bank")
_GESTION_FILE_YEAR_RE = re.compile(r"gestion\D*(20\d{2})", re.IGNORECASE)
_COMPTABILITE_FILE_YEAR_RE = re.compile(r"comptabilite\D*(20\d{2})", re.IGNORECASE)
_CLIENT_INVOICE_REFERENCE_RE = re.compile(r"\b\d{4}-\d{4}\b")
_SALARY_MONTH_RE = re.compile(r"\b(20\d{2})[.-](\d{2})\b")
_SALARY_TRAILING_INITIAL_RE = re.compile(r"\s+[a-z]$")
_SALARY_ACCRUAL_ACCOUNT_PREFIXES = ("421", "431", "443", "641", "645")
_SALARY_PAYMENT_ACCOUNT_PREFIXES = ("421", "512")
_CLIENT_INVOICE_CLARIFIED_MESSAGE = (
    "Facture client existante clarifiee a partir des ecritures comptables"
)


class _ImportSheetFailure(RuntimeError):
    """Internal marker used to abort the global import after a sheet-local failure."""


@dataclass(slots=True)
class _SupplierInvoiceCandidate:
    source_row_number: int
    invoice_date: date
    amount: Decimal
    invoice_number: str
    contact_name: str
    description: str
    label: str
    payment_method: str


_IssueT = TypeVar("_IssueT", RowValidationIssue, RowIgnoredIssue)


def _salary_month_label(month: str) -> str:
    return month.replace("-", ".")


def _salary_entry_date(month: str) -> date:
    year_value, month_value = month.split("-")
    return date(int(year_value), int(month_value), 1)


def _salary_employee_key(employee_name: str) -> str:
    key = _normalize_text(employee_name).replace(".", " ").strip()
    key = re.sub(r"\s+", " ", key)
    return _SALARY_TRAILING_INITIAL_RE.sub("", key)


def _extract_salary_month(*values: Any) -> str | None:
    matches: set[str] = set()
    for value in values:
        text = _parse_str(value, max_len=500)
        if not text:
            continue
        for year_part, month_part in _SALARY_MONTH_RE.findall(text):
            matches.add(f"{year_part}-{month_part}")
    if len(matches) != 1:
        return None
    return next(iter(matches))


def _accounting_amount_signature(
    *,
    account_number: str,
    debit: Decimal,
    credit: Decimal,
) -> str:
    return "|".join(
        (
            _normalize_text(account_number),
            _normalize_decimal_text(debit),
            _normalize_decimal_text(credit),
        )
    )


def _salary_group_amount_signature(
    lines: list[tuple[str, Decimal, Decimal]],
) -> tuple[str, ...]:
    return tuple(
        sorted(
            _accounting_amount_signature(
                account_number=account_number,
                debit=debit,
                credit=credit,
            )
            for account_number, debit, credit in lines
            if debit > 0 or credit > 0
        )
    )


def _entry_group_amount_signature(
    entry_rows: list[NormalizedEntryRow],
) -> tuple[str, ...]:
    return tuple(
        sorted(
            _accounting_amount_signature(
                account_number=entry_row.account_number,
                debit=entry_row.debit,
                credit=entry_row.credit,
            )
            for entry_row in entry_rows
            if entry_row.debit > 0 or entry_row.credit > 0
        )
    )


def _is_salary_related_label(label: str) -> bool:
    return (
        "salaire" in label
        or "charges salariales" in label
        or "charges patronales" in label
        or "impot sur le revenu" in label
    )


def _is_salary_accrual_like_entry_group(entry_rows: list[NormalizedEntryRow]) -> bool:
    account_numbers = [
        _normalize_text(entry_row.account_number)
        for entry_row in entry_rows
        if entry_row.debit > 0 or entry_row.credit > 0
    ]
    if not account_numbers:
        return False
    if any(account_number.startswith(("512", "53")) for account_number in account_numbers):
        return False
    return all(
        account_number.startswith(_SALARY_ACCRUAL_ACCOUNT_PREFIXES)
        for account_number in account_numbers
    )


def _is_salary_payment_like_entry_group(entry_rows: list[NormalizedEntryRow]) -> bool:
    account_numbers = [
        _normalize_text(entry_row.account_number)
        for entry_row in entry_rows
        if entry_row.debit > 0 or entry_row.credit > 0
    ]
    if not account_numbers:
        return False
    if not all(
        account_number.startswith(_SALARY_PAYMENT_ACCOUNT_PREFIXES)
        for account_number in account_numbers
    ):
        return False
    has_net_salary_debit = any(
        _normalize_text(entry_row.account_number).startswith("421") and entry_row.debit > 0
        for entry_row in entry_rows
    )
    has_salary_credit = any(
        _normalize_text(entry_row.account_number).startswith("512") and entry_row.credit > 0
        for entry_row in entry_rows
    )
    return has_net_salary_debit and has_salary_credit


def _salary_month_group_lines(
    month: str,
    salary_rows: list[NormalizedSalaryRow],
) -> list[list[tuple[str, str, Decimal, Decimal]]]:
    month_label = _salary_month_label(month)
    gross_total = sum((salary_row.gross for salary_row in salary_rows), Decimal("0"))
    employee_total = sum((salary_row.employee_charges for salary_row in salary_rows), Decimal("0"))
    employer_total = sum((salary_row.employer_charges for salary_row in salary_rows), Decimal("0"))
    tax_total = sum((salary_row.tax for salary_row in salary_rows), Decimal("0"))
    net_total = sum((salary_row.net_pay for salary_row in salary_rows), Decimal("0"))

    accrual_lines = [
        ("645100", f"Charges patronales {month_label}", employer_total, Decimal("0")),
        ("641000", f"Salaires bruts {month_label}", gross_total, Decimal("0")),
        ("421000", f"Salaires nets {month_label}", Decimal("0"), net_total),
        ("431100", f"Charges patronales {month_label}", Decimal("0"), employer_total),
        ("431100", f"Charges salariales {month_label}", Decimal("0"), employee_total),
    ]
    if tax_total > 0:
        accrual_lines.append(
            ("443000", f"Impôt sur le revenu {month_label}", Decimal("0"), tax_total)
        )

    payment_lines = [("421000", f"Paiement salaires {month_label}", net_total, Decimal("0"))]
    for salary_row in sorted(
        salary_rows, key=lambda row: (_salary_employee_key(row.employee_name), row.employee_name)
    ):
        if salary_row.net_pay <= 0:
            continue
        payment_lines.append(
            (
                "512100",
                f"Salaire {salary_row.employee_name} {month_label}",
                Decimal("0"),
                salary_row.net_pay,
            )
        )

    return [accrual_lines, payment_lines]


def _salary_month_group_signatures(
    month: str,
    salary_rows: list[NormalizedSalaryRow],
) -> set[tuple[str, ...]]:
    return {
        _salary_group_amount_signature(
            [(account_number, debit, credit) for account_number, _, debit, credit in group_lines]
        )
        for group_lines in _salary_month_group_lines(month, salary_rows)
    }


def _salary_month_standalone_tax_signatures(
    month: str,
    salary_rows: list[NormalizedSalaryRow],
) -> set[tuple[str, ...]]:
    tax_total = sum((salary_row.tax for salary_row in salary_rows), Decimal("0"))
    if tax_total <= 0:
        return set()

    signatures: set[tuple[str, ...]] = set()
    for account_number in ("431100", "443000"):
        signatures.add(
            (
                _accounting_amount_signature(
                    account_number=account_number,
                    debit=Decimal("0"),
                    credit=tax_total,
                ),
            )
        )
    return signatures


def _extract_client_invoice_references(*values: Any) -> list[str]:
    references: list[str] = []
    seen: set[str] = set()
    for value in values:
        text = _parse_str(value, max_len=500)
        if not text:
            continue
        for reference in _CLIENT_INVOICE_REFERENCE_RE.findall(text):
            normalized_reference = _normalize_text(reference)
            if normalized_reference in seen:
                continue
            seen.add(normalized_reference)
            references.append(reference)
    return references


def _single_client_invoice_reference(*values: Any) -> str | None:
    references = _extract_client_invoice_references(*values)
    if len(references) != 1:
        return None
    return references[0]


def _normalize_decimal_text(value: Decimal) -> str:
    return format(value.quantize(Decimal("0.01")).normalize(), "f")


def _client_settlement_account_from_method(method: str) -> str:
    normalized_method = _normalize_text(method)
    if normalized_method == "especes":
        return "531000"
    if normalized_method == "cheque":
        return "511200"
    return "512100"


def _supplier_settlement_account_from_method(method: str) -> str:
    if _normalize_text(method) == "especes":
        return "531000"
    return "512100"


def _is_client_invoice_entry_group(entry_rows: list[NormalizedEntryRow]) -> bool:
    has_receivable_debit = any(
        _normalize_text(entry_row.account_number).startswith("411") and entry_row.debit > 0
        for entry_row in entry_rows
    )
    has_revenue_credit = any(
        _normalize_text(entry_row.account_number).startswith(("70", "75")) and entry_row.credit > 0
        for entry_row in entry_rows
    )
    has_bank_or_cash_line = any(
        _normalize_text(entry_row.account_number).startswith(("512", "53"))
        for entry_row in entry_rows
    )
    return has_receivable_debit and has_revenue_credit and not has_bank_or_cash_line


def _matching_existing_client_invoice_reference(
    entry_rows: list[NormalizedEntryRow],
    existing_invoice_numbers: set[str],
) -> str | None:
    reference = _single_client_invoice_reference(*(entry_row.label for entry_row in entry_rows))
    if reference is None:
        return None
    if _normalize_text(reference) not in existing_invoice_numbers:
        return None
    if not _is_client_invoice_entry_group(entry_rows):
        return None
    return reference


def _client_invoice_line_type_from_account_number(account_number: str) -> Any | None:
    from backend.models.invoice import InvoiceLineType  # noqa: PLC0415

    normalized_account_number = _normalize_text(account_number)
    if normalized_account_number.startswith("706"):
        return InvoiceLineType.COURSE
    if normalized_account_number.startswith("756"):
        return InvoiceLineType.ADHESION
    if normalized_account_number.startswith("758"):
        return InvoiceLineType.OTHER
    return None


def _client_invoice_breakdown_from_entry_group(
    entry_rows: list[NormalizedEntryRow],
) -> dict[Any, Decimal] | None:
    from backend.models.invoice import InvoiceLineType  # noqa: PLC0415

    if not _is_client_invoice_entry_group(entry_rows):
        return None

    breakdown: dict[Any, Decimal] = {
        InvoiceLineType.COURSE: Decimal("0"),
        InvoiceLineType.ADHESION: Decimal("0"),
        InvoiceLineType.OTHER: Decimal("0"),
    }
    receivable_total = Decimal("0")
    revenue_total = Decimal("0")

    for entry_row in entry_rows:
        normalized_account_number = _normalize_text(entry_row.account_number)
        if normalized_account_number.startswith("411"):
            receivable_total += entry_row.debit
            continue
        if entry_row.credit <= 0:
            continue
        if normalized_account_number.startswith(("70", "75")):
            line_type = _client_invoice_line_type_from_account_number(entry_row.account_number)
            if line_type is None:
                return None
            breakdown[line_type] += entry_row.credit
            revenue_total += entry_row.credit

    if revenue_total <= 0 or receivable_total != revenue_total:
        return None

    return {line_type: amount for line_type, amount in breakdown.items() if amount != Decimal("0")}


def _current_client_invoice_breakdown(invoice: Any) -> dict[Any, Decimal]:
    from backend.models.invoice import infer_client_line_type  # noqa: PLC0415

    breakdown: dict[Any, Decimal] = {}
    for line in invoice.lines:
        line_type = line.line_type or infer_client_line_type(line.description, invoice.label)
        breakdown[line_type] = breakdown.get(line_type, Decimal("0")) + Decimal(str(line.amount))
    return {line_type: amount for line_type, amount in breakdown.items() if amount != Decimal("0")}


def _can_clarify_existing_client_invoice(invoice: Any) -> bool:
    from backend.models.invoice import InvoiceLineType, infer_client_line_type  # noqa: PLC0415

    if not invoice.lines:
        return False
    current_line_types = {
        line.line_type or infer_client_line_type(line.description, invoice.label)
        for line in invoice.lines
    }
    return current_line_types == {InvoiceLineType.OTHER}


async def _find_clarifiable_existing_client_invoice(
    db: AsyncSession,
    entry_rows: list[NormalizedEntryRow],
    existing_client_invoices_by_number: dict[str, Any] | None = None,
) -> Any | None:
    from sqlalchemy import select  # noqa: PLC0415
    from sqlalchemy.orm import selectinload  # noqa: PLC0415

    from backend.models.invoice import (  # noqa: PLC0415
        Invoice,
        InvoiceType,
    )

    reference = _single_client_invoice_reference(*(entry_row.label for entry_row in entry_rows))
    if reference is None:
        return None

    breakdown = _client_invoice_breakdown_from_entry_group(entry_rows)
    if breakdown is None:
        return None

    if existing_client_invoices_by_number is None:
        invoice_result = await db.execute(
            select(Invoice)
            .where(Invoice.type == InvoiceType.CLIENT, Invoice.number == reference)
            .options(selectinload(Invoice.lines))
        )
        invoice = invoice_result.scalar_one_or_none()
    else:
        invoice = existing_client_invoices_by_number.get(_normalize_text(reference))

    if invoice is None or not _can_clarify_existing_client_invoice(invoice):
        return None

    current_breakdown = _current_client_invoice_breakdown(invoice)
    if current_breakdown == breakdown:
        return None

    total_amount = sum(breakdown.values(), Decimal("0"))
    if total_amount != Decimal(str(invoice.total_amount)):
        return None

    return invoice


async def _clarify_existing_client_invoice_from_entries(
    db: AsyncSession,
    entry_rows: list[NormalizedEntryRow],
    existing_client_invoices_by_number: dict[str, Any] | None = None,
) -> Any | None:
    from sqlalchemy import delete  # noqa: PLC0415

    from backend.models.accounting_entry import AccountingEntry, EntrySourceType  # noqa: PLC0415
    from backend.models.invoice import (  # noqa: PLC0415
        InvoiceLabel,
        InvoiceLine,
        default_client_line_description,
        derive_client_invoice_label,
    )
    from backend.services.accounting_engine import generate_entries_for_invoice  # noqa: PLC0415

    invoice = await _find_clarifiable_existing_client_invoice(
        db,
        entry_rows,
        existing_client_invoices_by_number,
    )
    if invoice is None:
        return None

    breakdown = _client_invoice_breakdown_from_entry_group(entry_rows)
    if breakdown is None:
        return None

    await db.execute(
        delete(AccountingEntry).where(
            AccountingEntry.source_type == EntrySourceType.INVOICE,
            AccountingEntry.source_id == invoice.id,
        )
    )

    invoice.lines.clear()
    await db.flush()

    positive_line_types = {line_type for line_type, amount in breakdown.items() if amount > 0}
    invoice.label = derive_client_invoice_label(positive_line_types)
    invoice.has_explicit_breakdown = len(breakdown) > 1 or invoice.label != InvoiceLabel.GENERAL
    invoice.lines.extend(
        InvoiceLine(
            invoice_id=invoice.id,
            description=default_client_line_description(line_type),
            line_type=line_type,
            quantity=Decimal("1"),
            unit_price=amount,
            amount=amount,
        )
        for line_type, amount in sorted(breakdown.items(), key=lambda item: item[0].value)
    )

    await db.flush()
    await generate_entries_for_invoice(db, invoice)
    await db.flush()
    return invoice


async def _load_existing_client_invoices_by_number(db: AsyncSession) -> dict[str, Any]:
    from sqlalchemy import select  # noqa: PLC0415
    from sqlalchemy.orm import selectinload  # noqa: PLC0415

    from backend.models.invoice import Invoice, InvoiceType  # noqa: PLC0415

    result = await db.execute(
        select(Invoice)
        .where(Invoice.type == InvoiceType.CLIENT)
        .options(selectinload(Invoice.lines))
    )
    return {
        _normalize_text(invoice.number): invoice
        for invoice in result.scalars().all()
        if invoice.number
    }


def _is_client_payment_entry_group(entry_rows: list[NormalizedEntryRow]) -> bool:
    has_receivable_credit = any(
        _normalize_text(entry_row.account_number).startswith("411") and entry_row.credit > 0
        for entry_row in entry_rows
    )
    settlement_accounts = {
        _normalize_text(entry_row.account_number)
        for entry_row in entry_rows
        if entry_row.debit > 0
        and _normalize_text(entry_row.account_number).startswith(("511", "512", "53"))
    }
    has_revenue_line = any(
        _normalize_text(entry_row.account_number).startswith(("70", "75"))
        for entry_row in entry_rows
    )
    has_supplier_line = any(
        _normalize_text(entry_row.account_number).startswith("401") for entry_row in entry_rows
    )
    return (
        has_receivable_credit
        and len(settlement_accounts) == 1
        and not has_revenue_line
        and not has_supplier_line
    )


async def _load_existing_client_payment_reference_signatures(
    db: AsyncSession,
) -> set[tuple[str, str, str]]:
    from sqlalchemy import select  # noqa: PLC0415

    from backend.models.invoice import Invoice, InvoiceType  # noqa: PLC0415
    from backend.models.payment import Payment  # noqa: PLC0415

    result = await db.execute(
        select(Invoice.number, Invoice.reference, Payment.amount, Payment.method)
        .join(Payment, Payment.invoice_id == Invoice.id)
        .where(Invoice.type == InvoiceType.CLIENT)
    )

    signatures: set[tuple[str, str, str]] = set()
    for number, reference, amount, method in result.all():
        references = {number}
        if reference:
            references.add(reference)
        for client_reference in references:
            signatures.add(
                (
                    _normalize_text(client_reference),
                    _normalize_decimal_text(amount),
                    _client_settlement_account_from_method(str(method)),
                )
            )
    return signatures


def _build_client_invoice_lines_from_import_row(
    invoice_row: NormalizedInvoiceRow,
) -> list[dict[str, Any]]:
    from backend.models.invoice import (  # noqa: PLC0415
        InvoiceLabel,
        InvoiceLineType,
        default_client_line_description,
    )

    if invoice_row.course_amount is not None and invoice_row.adhesion_amount is not None:
        return [
            {
                "description": default_client_line_description(InvoiceLineType.COURSE),
                "line_type": InvoiceLineType.COURSE,
                "amount": invoice_row.course_amount,
            },
            {
                "description": default_client_line_description(InvoiceLineType.ADHESION),
                "line_type": InvoiceLineType.ADHESION,
                "amount": invoice_row.adhesion_amount,
            },
        ]

    label = InvoiceLabel(invoice_row.label)
    if label == InvoiceLabel.CS:
        line_type = InvoiceLineType.COURSE
    elif label == InvoiceLabel.ADHESION:
        line_type = InvoiceLineType.ADHESION
    else:
        line_type = InvoiceLineType.OTHER

    return [
        {
            "description": default_client_line_description(line_type),
            "line_type": line_type,
            "amount": invoice_row.amount,
        }
    ]


def _merge_existing_client_invoice_entry_groups(
    entry_groups: list[list[NormalizedEntryRow]],
    start_index: int,
    existing_invoice_numbers: set[str],
) -> tuple[list[NormalizedEntryRow], int]:
    merged_group = list(entry_groups[start_index])
    reference = _matching_existing_client_invoice_reference(merged_group, existing_invoice_numbers)
    if reference is None:
        return merged_group, start_index + 1

    next_index = start_index + 1
    while next_index < len(entry_groups):
        next_group = entry_groups[next_index]
        next_reference = _matching_existing_client_invoice_reference(
            next_group,
            existing_invoice_numbers,
        )
        if next_reference != reference:
            break
        merged_group.extend(next_group)
        next_index += 1

    return merged_group, next_index


def _matching_existing_client_payment_reference(
    entry_rows: list[NormalizedEntryRow],
    existing_client_payment_signatures: set[tuple[str, str, str]],
) -> str | None:
    reference = _single_client_invoice_reference(*(entry_row.label for entry_row in entry_rows))
    if reference is None:
        return None
    if not _is_client_payment_entry_group(entry_rows):
        return None

    settlement_accounts = {
        _normalize_text(entry_row.account_number)
        for entry_row in entry_rows
        if entry_row.debit > 0
        and _normalize_text(entry_row.account_number).startswith(("511", "512", "53"))
    }
    if len(settlement_accounts) != 1:
        return None

    amount = sum(
        (
            entry_row.debit
            for entry_row in entry_rows
            if entry_row.debit > 0
            and _normalize_text(entry_row.account_number).startswith(("511", "512", "53"))
        ),
        start=Decimal("0"),
    )
    if amount <= 0:
        return None

    signature = (
        _normalize_text(reference),
        _normalize_decimal_text(amount),
        next(iter(settlement_accounts)),
    )
    if signature not in existing_client_payment_signatures:
        return None
    return reference


def _is_supplier_direct_settlement_entry_group(entry_rows: list[NormalizedEntryRow]) -> bool:
    has_expense_debit = any(
        _normalize_text(entry_row.account_number).startswith(("6", "60", "61", "62"))
        and entry_row.debit > 0
        for entry_row in entry_rows
    )
    settlement_accounts = {
        _normalize_text(entry_row.account_number)
        for entry_row in entry_rows
        if entry_row.credit > 0
        and _normalize_text(entry_row.account_number).startswith(("512", "53"))
    }
    has_supplier_account = any(
        _normalize_text(entry_row.account_number).startswith("401") for entry_row in entry_rows
    )
    return has_expense_debit and len(settlement_accounts) == 1 and not has_supplier_account


async def _load_existing_supplier_payment_reference_signatures(
    db: AsyncSession,
) -> set[tuple[str, str, str, str]]:
    from sqlalchemy import select  # noqa: PLC0415

    from backend.models.invoice import Invoice, InvoiceType  # noqa: PLC0415
    from backend.models.payment import Payment  # noqa: PLC0415

    result = await db.execute(
        select(Invoice.number, Invoice.reference, Payment.date, Payment.amount, Payment.method)
        .join(Payment, Payment.invoice_id == Invoice.id)
        .where(Invoice.type == InvoiceType.FOURNISSEUR)
    )
    signatures: set[tuple[str, str, str, str]] = set()
    for number, reference, payment_date, amount, method in result.all():
        supplier_reference = reference or number
        if not supplier_reference:
            continue
        signatures.add(
            (
                _normalize_text(supplier_reference),
                payment_date.isoformat(),
                _normalize_decimal_text(amount),
                _supplier_settlement_account_from_method(str(method)),
            )
        )
    return signatures


def _matching_existing_supplier_invoice_payment_reference(
    entry_rows: list[NormalizedEntryRow],
    existing_supplier_payment_signatures: set[tuple[str, str, str, str]],
) -> str | None:
    reference = _extract_supplier_invoice_reference(*(entry_row.label for entry_row in entry_rows))
    if reference is None:
        return None
    if not _is_supplier_direct_settlement_entry_group(entry_rows):
        return None

    dates = {entry_row.entry_date for entry_row in entry_rows}
    if len(dates) != 1:
        return None

    amount = sum((entry_row.debit for entry_row in entry_rows), start=Decimal("0"))
    settlement_accounts = {
        _normalize_text(entry_row.account_number)
        for entry_row in entry_rows
        if entry_row.credit > 0
        and _normalize_text(entry_row.account_number).startswith(("512", "53"))
    }
    if len(settlement_accounts) != 1 or amount <= 0:
        return None

    signature = (
        _normalize_text(reference),
        next(iter(dates)).isoformat(),
        _normalize_decimal_text(amount),
        next(iter(settlement_accounts)),
    )
    if signature not in existing_supplier_payment_signatures:
        return None
    return reference


def _direct_bank_trigger_from_row(bank_row: NormalizedBankRow) -> tuple[Any, Decimal] | None:
    from backend.models.accounting_rule import TriggerType  # noqa: PLC0415

    description = _normalize_text(bank_row.description)
    reference = _normalize_text(bank_row.reference)
    combined_text = " ".join(part for part in (description, reference) if part)
    if "remise cheq" in combined_text and bank_row.amount > 0:
        return TriggerType.DEPOSIT_CHEQUES, bank_row.amount
    if "remise espe" in combined_text and bank_row.amount > 0:
        return TriggerType.DEPOSIT_ESPECES, bank_row.amount
    if "frais banc" in combined_text and bank_row.amount < 0:
        return TriggerType.BANK_FEES, abs(bank_row.amount)
    if "maif" in combined_text and bank_row.amount < 0:
        return TriggerType.BANK_INSURANCE, abs(bank_row.amount)
    if "interets" in combined_text and "livret bleu" in combined_text and bank_row.amount > 0:
        return TriggerType.BANK_SAVINGS_INTEREST, bank_row.amount
    if "virement interne" in combined_text:
        if bank_row.amount > 0:
            return TriggerType.BANK_INTERNAL_TRANSFER_FROM_SAVINGS, bank_row.amount
        if bank_row.amount < 0:
            return TriggerType.BANK_INTERNAL_TRANSFER_TO_SAVINGS, abs(bank_row.amount)
    if "urssaf" in combined_text and bank_row.amount < 0:
        return TriggerType.BANK_SOCIAL_CHARGES, abs(bank_row.amount)
    if "subvention" in combined_text and bank_row.amount > 0:
        return TriggerType.SUBSIDY_RECEIVED, bank_row.amount
    return None


async def _generate_direct_bank_entries(
    db: AsyncSession,
    *,
    bank_row: NormalizedBankRow,
    bank_entry: Any,
) -> list[Any]:
    from backend.models.accounting_entry import EntrySourceType  # noqa: PLC0415
    from backend.services.accounting_engine import generate_entries_for_trigger  # noqa: PLC0415

    direct_trigger = _direct_bank_trigger_from_row(bank_row)
    if direct_trigger is None:
        return []

    trigger, amount = direct_trigger
    return await generate_entries_for_trigger(
        db,
        trigger,
        amount,
        bank_row.entry_date,
        {
            "label": bank_row.description,
            "date": str(bank_row.entry_date),
        },
        source_type=EntrySourceType.GESTION,
        source_id=bank_entry.id,
    )


def _payment_signature(
    *,
    invoice_id: int,
    payment_date: date,
    amount: Decimal,
    method: str,
    cheque_number: str | None = None,
) -> tuple[int, str, str, str, str]:
    return (
        invoice_id,
        payment_date.isoformat(),
        format(amount.normalize(), "f"),
        _normalize_text(method),
        _normalize_text(cheque_number or "") if _normalize_text(method) == "cheque" else "",
    )


def _payment_row_from_bank_row(
    row: NormalizedBankRow,
    *,
    invoice_reference: str,
) -> NormalizedPaymentRow:
    return NormalizedPaymentRow(
        source_row_number=row.source_row_number,
        payment_date=row.entry_date,
        amount=row.amount,
        method="virement",
        invoice_ref=invoice_reference,
        contact_name="",
        cheque_number=None,
        deposited=True,
        deposit_date=row.entry_date,
    )


async def _load_existing_payment_signatures(
    db: AsyncSession,
) -> set[tuple[int, str, str, str, str]]:
    from sqlalchemy import select  # noqa: PLC0415

    from backend.models.payment import Payment  # noqa: PLC0415

    result = await db.execute(
        select(
            Payment.invoice_id,
            Payment.date,
            Payment.amount,
            Payment.method,
            Payment.cheque_number,
        )
    )
    return {
        _payment_signature(
            invoice_id=invoice_id,
            payment_date=payment_date,
            amount=amount,
            method=str(method),
            cheque_number=cheque_number,
        )
        for invoice_id, payment_date, amount, method, cheque_number in result.all()
    }


async def _load_existing_contacts_by_salary_key(db: AsyncSession) -> dict[str, Any]:
    from sqlalchemy import select  # noqa: PLC0415

    from backend.models.contact import Contact  # noqa: PLC0415

    result = await db.execute(select(Contact))
    contacts_by_key: dict[str, Any] = {}
    for contact in result.scalars().all():
        display_name = _format_contact_display_name(contact.nom, contact.prenom) or contact.nom
        contacts_by_key.setdefault(_salary_employee_key(display_name), contact)
        contacts_by_key.setdefault(_salary_employee_key(contact.nom), contact)
    return contacts_by_key


async def _load_existing_salary_keys(db: AsyncSession) -> set[tuple[str, str]]:
    from sqlalchemy import select  # noqa: PLC0415

    from backend.models.contact import Contact  # noqa: PLC0415
    from backend.models.salary import Salary  # noqa: PLC0415

    result = await db.execute(select(Salary.month, Contact.nom, Contact.prenom).join(Contact))
    return {
        (month, _salary_employee_key(_format_contact_display_name(nom, prenom) or nom))
        for month, nom, prenom in result.all()
    }


async def _load_existing_salary_group_signatures(
    db: AsyncSession,
) -> dict[str, set[tuple[str, ...]]]:
    from sqlalchemy import select  # noqa: PLC0415

    from backend.models.contact import Contact  # noqa: PLC0415
    from backend.models.salary import Salary  # noqa: PLC0415

    result = await db.execute(
        select(
            Salary.month,
            Contact.nom,
            Contact.prenom,
            Salary.hours,
            Salary.gross,
            Salary.employee_charges,
            Salary.employer_charges,
            Salary.tax,
            Salary.net_pay,
        ).join(Contact)
    )

    salaries_by_month: dict[str, list[NormalizedSalaryRow]] = {}
    for (
        month,
        nom,
        prenom,
        hours,
        gross,
        employee_charges,
        employer_charges,
        tax,
        net_pay,
    ) in result.all():
        salaries_by_month.setdefault(month, []).append(
            NormalizedSalaryRow(
                source_row_number=0,
                month=month,
                employee_name=_format_contact_display_name(nom, prenom) or nom,
                hours=Decimal(str(hours)),
                gross=Decimal(str(gross)),
                employee_charges=Decimal(str(employee_charges)),
                employer_charges=Decimal(str(employer_charges)),
                tax=Decimal(str(tax)),
                net_pay=Decimal(str(net_pay)),
            )
        )

    return {
        month: _salary_month_group_signatures(month, salary_rows)
        | _salary_month_standalone_tax_signatures(month, salary_rows)
        for month, salary_rows in salaries_by_month.items()
    }


def _gestion_payment_comparison_signature(
    *,
    reference: str,
    payment_date: date,
    amount: Decimal,
    settlement_account: str,
    cheque_number: str | None = None,
) -> tuple[str, str, str, str, str]:
    return (
        _normalize_text(reference),
        payment_date.isoformat(),
        _normalize_decimal_text(amount),
        _normalize_text(settlement_account),
        (
            _normalize_text(cheque_number or "")
            if _normalize_text(settlement_account) == "511200"
            else ""
        ),
    )


def _gestion_salary_comparison_signature(
    *,
    month: str,
    employee_name: str,
    gross: Decimal,
    net_pay: Decimal,
) -> tuple[str, str, str, str]:
    return (
        month,
        _salary_employee_key(employee_name),
        _normalize_decimal_text(gross),
        _normalize_decimal_text(net_pay),
    )


def _gestion_bank_comparison_signature(
    *,
    entry_date: date,
    amount: Decimal,
    description: str,
    reference: str | None,
) -> tuple[str, str, str, str]:
    return (
        entry_date.isoformat(),
        _normalize_decimal_text(amount),
        _normalize_text(description),
        _normalize_text(reference or ""),
    )


def _gestion_cash_comparison_signature(
    *,
    entry_date: date,
    movement_type: str,
    amount: Decimal,
    description: str,
    reference: str | None,
) -> tuple[str, str, str, str, str]:
    return (
        entry_date.isoformat(),
        _normalize_text(movement_type),
        _normalize_decimal_text(amount),
        _normalize_text(description),
        _normalize_text(reference or ""),
    )


def _build_extra_in_solde_summary(*parts: str | None) -> str:
    return " · ".join(part for part in parts if part)


def _make_gestion_invoice_extra_detail(number: str, invoice_date: date) -> dict[str, str]:
    number_value = number.strip()
    date_value = invoice_date.isoformat()
    return {
        "summary": _build_extra_in_solde_summary(number_value, date_value),
        "number": number_value,
        "date": date_value,
    }


def _make_gestion_payment_extra_detail(
    *,
    reference: str,
    payment_date: date,
    amount: Decimal,
    settlement_account: str,
    cheque_number: str | None,
    invoice_number: str | None,
    invoice_reference: str | None,
) -> dict[str, str]:
    reference_value = reference.strip()
    payment_date_value = payment_date.isoformat()
    amount_value = _normalize_decimal_text(amount)
    settlement_account_value = settlement_account.strip()
    detail = {
        "summary": _build_extra_in_solde_summary(
            reference_value,
            payment_date_value,
            amount_value,
            settlement_account_value,
        ),
        "reference": reference_value,
        "payment_date": payment_date_value,
        "amount": amount_value,
        "settlement_account": settlement_account_value,
    }
    if cheque_number:
        detail["cheque_number"] = cheque_number.strip()
    if invoice_number:
        detail["invoice_number"] = invoice_number.strip()
    if invoice_reference:
        detail["invoice_reference"] = invoice_reference.strip()
    return detail


def _make_gestion_salary_extra_detail(
    *,
    month: str,
    employee_name: str,
    gross: Decimal,
    net_pay: Decimal,
) -> dict[str, str]:
    month_value = month.strip()
    employee_name_value = employee_name.strip()
    gross_value = _normalize_decimal_text(gross)
    net_pay_value = _normalize_decimal_text(net_pay)
    return {
        "summary": _build_extra_in_solde_summary(
            month_value,
            employee_name_value,
            gross_value,
            net_pay_value,
        ),
        "month": month_value,
        "employee_name": employee_name_value,
        "gross": gross_value,
        "net_pay": net_pay_value,
    }


def _make_gestion_bank_extra_detail(
    *,
    entry_date: date,
    amount: Decimal,
    description: str,
    reference: str | None,
) -> dict[str, str]:
    entry_date_value = entry_date.isoformat()
    amount_value = _normalize_decimal_text(amount)
    description_value = description.strip()
    reference_value = reference.strip() if reference else ""
    detail = {
        "summary": _build_extra_in_solde_summary(
            entry_date_value,
            amount_value,
            description_value,
            reference_value,
        ),
        "entry_date": entry_date_value,
        "amount": amount_value,
        "description": description_value,
    }
    if reference_value:
        detail["reference"] = reference_value
    return detail


def _make_gestion_cash_extra_detail(
    *,
    entry_date: date,
    movement_type: str,
    amount: Decimal,
    description: str,
    reference: str | None,
) -> dict[str, str]:
    entry_date_value = entry_date.isoformat()
    movement_type_value = movement_type.strip()
    amount_value = _normalize_decimal_text(amount)
    description_value = description.strip()
    reference_value = reference.strip() if reference else ""
    detail = {
        "summary": _build_extra_in_solde_summary(
            entry_date_value,
            movement_type_value,
            amount_value,
            description_value,
            reference_value,
        ),
        "entry_date": entry_date_value,
        "movement_type": movement_type_value,
        "amount": amount_value,
        "description": description_value,
    }
    if reference_value:
        detail["reference"] = reference_value
    return detail


def _expand_date_bounds(
    start_date: date | None,
    end_date: date | None,
    candidate_date: date,
) -> tuple[date, date]:
    if start_date is None or candidate_date < start_date:
        start_date = candidate_date
    if end_date is None or candidate_date > end_date:
        end_date = candidate_date
    return start_date, end_date


def _is_within_date_bounds(
    candidate_date: date,
    start_date: date | None,
    end_date: date | None,
) -> bool:
    if start_date is not None and candidate_date < start_date:
        return False
    return end_date is None or candidate_date <= end_date


def _gestion_file_fiscal_year_bounds(file_name: str | None) -> tuple[date, date] | None:
    return _preview_file_fiscal_year_bounds(file_name)


def _preview_file_fiscal_year_bounds(file_name: str | None) -> tuple[date, date] | None:
    if not file_name:
        return None

    match = _GESTION_FILE_YEAR_RE.search(file_name)
    if match is None:
        match = _COMPTABILITE_FILE_YEAR_RE.search(file_name)
    if match is None:
        return None

    fiscal_year_start_month = get_settings().fiscal_year_start_month
    start_year = int(match.group(1))
    start_date = date(start_year, fiscal_year_start_month, 1)
    next_fiscal_year_start = date(start_year + 1, fiscal_year_start_month, 1)
    return start_date, next_fiscal_year_start - timedelta(days=1)


@dataclass(frozen=True, slots=True)
class _PreviewFilterWindow:
    start_date: date | None
    end_date: date | None


def _resolve_comparison_window(
    file_name: str | None,
    requested_start_date: date | None,
    requested_end_date: date | None,
) -> _PreviewFilterWindow | None:
    del file_name
    if requested_start_date is None and requested_end_date is None:
        return None
    return _PreviewFilterWindow(
        start_date=requested_start_date,
        end_date=requested_end_date,
    )


def _is_within_comparison_window(
    candidate_date: date,
    window: _PreviewFilterWindow | None,
) -> bool:
    if window is None:
        return True
    if window.start_date is not None and candidate_date < window.start_date:
        return False
    return window.end_date is None or candidate_date <= window.end_date


def _salary_month_to_date(month: str) -> date | None:
    match = _SALARY_MONTH_RE.match(month)
    if match is None:
        return None
    return date(int(match.group(1)), int(match.group(2)), 1)


def _is_salary_month_within_comparison_window(
    month: str,
    window: _PreviewFilterWindow | None,
) -> bool:
    if window is None:
        return True
    candidate_date = _salary_month_to_date(month)
    if candidate_date is None:
        return False
    return _is_within_comparison_window(candidate_date, window)


def _find_date_column_index(col_map: dict[str, int]) -> int | None:
    for column_name, column_index in col_map.items():
        if "date" in _normalize_text(column_name):
            return column_index
    return None


def _build_sheet_row_dates(
    ws: Any,
    parsed_sheet: ParsedSheet,
) -> dict[int, date]:
    date_column_index = _find_date_column_index(parsed_sheet.col_map)
    if date_column_index is None:
        return {}

    row_dates: dict[int, date] = {}
    for source_row_number, row in enumerate(ws.iter_rows(values_only=True), start=1):
        parsed_date = _parse_date(_get_row_value(row, date_column_index))
        if parsed_date is not None:
            row_dates[source_row_number] = parsed_date
    return row_dates


def _build_salary_row_months(ws: Any) -> dict[int, str]:
    row_months: dict[int, str] = {}
    current_month: str | None = None
    for source_row_number, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if all(cell is None for cell in row):
            continue
        first_cell = _parse_str(_get_row_value(row, 0))
        if not first_cell:
            continue
        if match := _SALARY_MONTH_RE.match(first_cell):
            current_month = f"{match.group(1)}-{match.group(2)}"
            continue
        if current_month is not None:
            row_months[source_row_number] = current_month
    return row_months


def _filter_date_issues_in_window(
    issues: list[_IssueT],
    row_dates: dict[int, date],
    window: _PreviewFilterWindow | None,
) -> list[_IssueT]:
    if window is None:
        return issues
    return [
        issue
        for issue in issues
        if (candidate_date := row_dates.get(issue.source_row_number)) is not None
        and _is_within_comparison_window(candidate_date, window)
    ]


def _filter_salary_issues_in_window(
    issues: list[RowValidationIssue],
    row_months: dict[int, str],
    window: _PreviewFilterWindow | None,
) -> list[RowValidationIssue]:
    if window is None:
        return issues
    return [
        issue
        for issue in issues
        if (month := row_months.get(issue.source_row_number)) is not None
        and _is_salary_month_within_comparison_window(month, window)
    ]


def _comparison_years_within_bounds(
    years: set[int],
    start_date: date | None,
    end_date: date | None,
) -> set[int]:
    if start_date is None or end_date is None:
        return years
    return set(range(start_date.year, end_date.year + 1))


async def _load_existing_client_invoice_comparison_signatures(
    db: AsyncSession,
    years: set[int],
    start_date: date | None = None,
    end_date: date | None = None,
) -> set[str]:
    if not years:
        return set()

    from sqlalchemy import select  # noqa: PLC0415

    from backend.models.invoice import Invoice, InvoiceType  # noqa: PLC0415

    result = await db.execute(
        select(Invoice.number, Invoice.date).where(Invoice.type == InvoiceType.CLIENT)
    )
    return {
        _normalize_text(number)
        for number, invoice_date in result.all()
        if number
        and invoice_date.year in years
        and _is_within_date_bounds(invoice_date, start_date, end_date)
    }


async def _load_existing_client_payment_comparison_signatures(
    db: AsyncSession,
    years: set[int],
    start_date: date | None = None,
    end_date: date | None = None,
) -> set[tuple[str, str, str, str, str]]:
    if not years:
        return set()

    from sqlalchemy import select  # noqa: PLC0415

    from backend.models.invoice import Invoice, InvoiceType  # noqa: PLC0415
    from backend.models.payment import Payment  # noqa: PLC0415

    result = await db.execute(
        select(
            Invoice.number,
            Invoice.reference,
            Payment.date,
            Payment.amount,
            Payment.method,
            Payment.cheque_number,
        )
        .join(Payment, Payment.invoice_id == Invoice.id)
        .where(Invoice.type == InvoiceType.CLIENT)
    )

    signatures: set[tuple[str, str, str, str, str]] = set()
    for number, reference, payment_date, amount, method, cheque_number in result.all():
        if payment_date.year not in years or not _is_within_date_bounds(
            payment_date,
            start_date,
            end_date,
        ):
            continue
        references = {number}
        if reference:
            references.add(reference)
        settlement_account = _client_settlement_account_from_method(str(method))
        for candidate_reference in references:
            if not candidate_reference:
                continue
            signatures.add(
                _gestion_payment_comparison_signature(
                    reference=candidate_reference,
                    payment_date=payment_date,
                    amount=amount,
                    settlement_account=settlement_account,
                    cheque_number=cheque_number,
                )
            )
    return signatures


async def _load_existing_client_invoice_comparison_items(
    db: AsyncSession,
    years: set[int],
    start_date: date | None = None,
    end_date: date | None = None,
) -> dict[str, dict[str, str]]:
    if not years:
        return {}

    from sqlalchemy import select  # noqa: PLC0415

    from backend.models.invoice import Invoice, InvoiceType  # noqa: PLC0415

    result = await db.execute(
        select(Invoice.number, Invoice.date).where(Invoice.type == InvoiceType.CLIENT)
    )
    details: dict[str, dict[str, str]] = {}
    for number, invoice_date in result.all():
        if (
            not number
            or invoice_date.year not in years
            or not _is_within_date_bounds(invoice_date, start_date, end_date)
        ):
            continue
        details[_normalize_text(number)] = _make_gestion_invoice_extra_detail(number, invoice_date)
    return details


async def _load_existing_client_payment_comparison_items(
    db: AsyncSession,
    years: set[int],
    start_date: date | None = None,
    end_date: date | None = None,
) -> dict[tuple[str, str, str, str, str], dict[str, str]]:
    if not years:
        return {}

    from sqlalchemy import select  # noqa: PLC0415

    from backend.models.invoice import Invoice, InvoiceType  # noqa: PLC0415
    from backend.models.payment import Payment  # noqa: PLC0415

    result = await db.execute(
        select(
            Invoice.number,
            Invoice.reference,
            Payment.date,
            Payment.amount,
            Payment.method,
            Payment.cheque_number,
        )
        .join(Payment, Payment.invoice_id == Invoice.id)
        .where(Invoice.type == InvoiceType.CLIENT)
    )

    details: dict[tuple[str, str, str, str, str], dict[str, str]] = {}
    for number, reference, payment_date, amount, method, cheque_number in result.all():
        if payment_date.year not in years or not _is_within_date_bounds(
            payment_date,
            start_date,
            end_date,
        ):
            continue
        candidate_references = [candidate for candidate in (number, reference) if candidate]
        seen_references: set[str] = set()
        settlement_account = _client_settlement_account_from_method(str(method))
        for candidate_reference in candidate_references:
            normalized_reference = _normalize_text(candidate_reference)
            if not normalized_reference or normalized_reference in seen_references:
                continue
            seen_references.add(normalized_reference)
            signature = _gestion_payment_comparison_signature(
                reference=candidate_reference,
                payment_date=payment_date,
                amount=amount,
                settlement_account=settlement_account,
                cheque_number=cheque_number,
            )
            details[signature] = _make_gestion_payment_extra_detail(
                reference=candidate_reference,
                payment_date=payment_date,
                amount=amount,
                settlement_account=settlement_account,
                cheque_number=cheque_number,
                invoice_number=number,
                invoice_reference=reference,
            )
    return details


async def _load_existing_salary_comparison_signatures(
    db: AsyncSession,
    months: set[str],
) -> set[tuple[str, str, str, str]]:
    if not months:
        return set()

    from sqlalchemy import select  # noqa: PLC0415

    from backend.models.contact import Contact  # noqa: PLC0415
    from backend.models.salary import Salary  # noqa: PLC0415

    result = await db.execute(
        select(
            Salary.month,
            Contact.nom,
            Contact.prenom,
            Salary.gross,
            Salary.net_pay,
        ).join(Contact)
    )
    return {
        _gestion_salary_comparison_signature(
            month=month,
            employee_name=_format_contact_display_name(nom, prenom) or nom,
            gross=Decimal(str(gross)),
            net_pay=Decimal(str(net_pay)),
        )
        for month, nom, prenom, gross, net_pay in result.all()
        if month in months
    }


async def _load_existing_salary_comparison_items(
    db: AsyncSession,
    months: set[str],
) -> dict[tuple[str, str, str, str], dict[str, str]]:
    if not months:
        return {}

    from sqlalchemy import select  # noqa: PLC0415

    from backend.models.contact import Contact  # noqa: PLC0415
    from backend.models.salary import Salary  # noqa: PLC0415

    result = await db.execute(
        select(
            Salary.month,
            Contact.nom,
            Contact.prenom,
            Salary.gross,
            Salary.net_pay,
        ).join(Contact)
    )
    details: dict[tuple[str, str, str, str], dict[str, str]] = {}
    for month, nom, prenom, gross, net_pay in result.all():
        if month not in months:
            continue
        employee_name = _format_contact_display_name(nom, prenom) or nom
        signature = _gestion_salary_comparison_signature(
            month=month,
            employee_name=employee_name,
            gross=Decimal(str(gross)),
            net_pay=Decimal(str(net_pay)),
        )
        details[signature] = _make_gestion_salary_extra_detail(
            month=month,
            employee_name=employee_name,
            gross=Decimal(str(gross)),
            net_pay=Decimal(str(net_pay)),
        )
    return details


async def _load_existing_bank_comparison_signatures(
    db: AsyncSession,
    years: set[int],
    start_date: date | None = None,
    end_date: date | None = None,
) -> set[tuple[str, str, str, str]]:
    if not years:
        return set()

    from sqlalchemy import select  # noqa: PLC0415

    from backend.models.bank import BankTransaction, BankTransactionSource  # noqa: PLC0415

    result = await db.execute(
        select(
            BankTransaction.date,
            BankTransaction.amount,
            BankTransaction.description,
            BankTransaction.reference,
            BankTransaction.source,
        )
    )
    return {
        _gestion_bank_comparison_signature(
            entry_date=entry_date,
            amount=amount,
            description=description,
            reference=reference,
        )
        for entry_date, amount, description, reference, source in result.all()
        if entry_date.year in years
        and _is_within_date_bounds(entry_date, start_date, end_date)
        and source != BankTransactionSource.SYSTEM_OPENING
    }


async def _load_existing_bank_comparison_items(
    db: AsyncSession,
    years: set[int],
    start_date: date | None = None,
    end_date: date | None = None,
) -> dict[tuple[str, str, str, str], dict[str, str]]:
    if not years:
        return {}

    from sqlalchemy import select  # noqa: PLC0415

    from backend.models.bank import BankTransaction, BankTransactionSource  # noqa: PLC0415

    result = await db.execute(
        select(
            BankTransaction.date,
            BankTransaction.amount,
            BankTransaction.description,
            BankTransaction.reference,
            BankTransaction.source,
        )
    )
    details: dict[tuple[str, str, str, str], dict[str, str]] = {}
    for entry_date, amount, description, reference, source in result.all():
        if (
            entry_date.year not in years
            or not _is_within_date_bounds(entry_date, start_date, end_date)
            or source == BankTransactionSource.SYSTEM_OPENING
        ):
            continue
        signature = _gestion_bank_comparison_signature(
            entry_date=entry_date,
            amount=amount,
            description=description,
            reference=reference,
        )
        details[signature] = _make_gestion_bank_extra_detail(
            entry_date=entry_date,
            amount=amount,
            description=description,
            reference=reference,
        )
    return details


async def _load_existing_cash_comparison_signatures(
    db: AsyncSession,
    years: set[int],
    start_date: date | None = None,
    end_date: date | None = None,
) -> set[tuple[str, str, str, str, str]]:
    if not years:
        return set()

    from sqlalchemy import select  # noqa: PLC0415

    from backend.models.cash import CashEntrySource, CashRegister  # noqa: PLC0415

    result = await db.execute(
        select(
            CashRegister.date,
            CashRegister.type,
            CashRegister.amount,
            CashRegister.description,
            CashRegister.reference,
            CashRegister.source,
        )
    )
    return {
        _gestion_cash_comparison_signature(
            entry_date=entry_date,
            movement_type=str(movement_type),
            amount=amount,
            description=description,
            reference=reference,
        )
        for entry_date, movement_type, amount, description, reference, source in result.all()
        if entry_date.year in years
        and _is_within_date_bounds(entry_date, start_date, end_date)
        and source != CashEntrySource.SYSTEM_OPENING
    }


async def _load_existing_cash_comparison_items(
    db: AsyncSession,
    years: set[int],
    start_date: date | None = None,
    end_date: date | None = None,
) -> dict[tuple[str, str, str, str, str], dict[str, str]]:
    if not years:
        return {}

    from sqlalchemy import select  # noqa: PLC0415

    from backend.models.cash import CashEntrySource, CashRegister  # noqa: PLC0415

    result = await db.execute(
        select(
            CashRegister.date,
            CashRegister.type,
            CashRegister.amount,
            CashRegister.description,
            CashRegister.reference,
            CashRegister.source,
        )
    )
    details: dict[tuple[str, str, str, str, str], dict[str, str]] = {}
    for entry_date, movement_type, amount, description, reference, source in result.all():
        if (
            entry_date.year not in years
            or not _is_within_date_bounds(entry_date, start_date, end_date)
            or source == CashEntrySource.SYSTEM_OPENING
        ):
            continue
        signature = _gestion_cash_comparison_signature(
            entry_date=entry_date,
            movement_type=str(movement_type),
            amount=amount,
            description=description,
            reference=reference,
        )
        details[signature] = _make_gestion_cash_extra_detail(
            entry_date=entry_date,
            movement_type=str(movement_type),
            amount=amount,
            description=description,
            reference=reference,
        )
    return details


def _make_preview_comparison_domain(
    *,
    kind: str,
    file_rows: int,
    already_in_solde: int,
    missing_in_solde: int,
    extra_in_solde: int,
    ignored_by_policy: int,
    blocked: int,
    extra_in_solde_details: list[dict[str, str]] | None = None,
) -> dict[str, Any]:
    domain = {
        "kind": kind,
        "file_rows": file_rows,
        "already_in_solde": already_in_solde,
        "missing_in_solde": missing_in_solde,
        "extra_in_solde": extra_in_solde,
        "ignored_by_policy": ignored_by_policy,
        "blocked": blocked,
    }
    if extra_in_solde_details:
        domain["extra_in_solde_details"] = extra_in_solde_details
    return domain


async def _build_gestion_preview_comparison_domains(
    db: AsyncSession,
    file_bytes: bytes,
    file_name: str | None,
    comparison_window: _PreviewFilterWindow | None,
) -> list[dict[str, Any]]:
    from io import BytesIO  # noqa: PLC0415

    import openpyxl  # noqa: PLC0415

    existing_contacts_by_preview_key = await _load_existing_contacts_by_preview_key(db)
    default_extra_bounds = _gestion_file_fiscal_year_bounds(file_name)
    default_extra_start_date = default_extra_bounds[0] if default_extra_bounds else None
    default_extra_end_date = default_extra_bounds[1] if default_extra_bounds else None
    workbook_candidates: list[PaymentMatchCandidate] = []

    wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        kind, _ = _classify_gestion_sheet(sheet_name)
        if kind != "invoices":
            continue
        parsed_sheet, invoice_rows, _, _ = _parse_invoice_sheet(ws)
        if parsed_sheet is None or parsed_sheet.missing_columns:
            continue
        row_dates = {
            invoice_row.source_row_number: invoice_row.invoice_date for invoice_row in invoice_rows
        }
        deduped_invoice_rows, _ = filter_duplicate_invoice_rows(
            invoice_rows,
            normalize_text=_normalize_text,
        )
        workbook_candidates.extend(
            _make_workbook_invoice_candidate(invoice_row) for invoice_row in deduped_invoice_rows
        )

    domains: list[dict[str, Any]] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        kind, _ = _classify_gestion_sheet(sheet_name)
        if kind is None or kind == "contacts":
            continue

        if kind == "invoices":
            parsed_sheet, invoice_rows, row_issues, ignored_issues = _parse_invoice_sheet(ws)
            if parsed_sheet is None or parsed_sheet.missing_columns:
                continue
            row_dates = {
                invoice_row.source_row_number: invoice_row.invoice_date
                for invoice_row in invoice_rows
            }
            deduped_invoice_rows, duplicate_ignored_issues = filter_duplicate_invoice_rows(
                invoice_rows,
                normalize_text=_normalize_text,
            )
            filtered_invoice_rows = (
                [
                    invoice_row
                    for invoice_row in deduped_invoice_rows
                    if _is_within_comparison_window(invoice_row.invoice_date, comparison_window)
                ]
                if comparison_window is not None
                else deduped_invoice_rows
            )
            filtered_ignored_issues = (
                _filter_date_issues_in_window(
                    ignored_issues + duplicate_ignored_issues,
                    row_dates,
                    comparison_window,
                )
                if comparison_window is not None
                else ignored_issues + duplicate_ignored_issues
            )
            filtered_blocked_issues = (
                _filter_date_issues_in_window(
                    row_issues,
                    _build_sheet_row_dates(ws, parsed_sheet),
                    comparison_window,
                )
                if comparison_window is not None
                else row_issues
            )

            invoice_years = {invoice_row.invoice_date.year for invoice_row in filtered_invoice_rows}
            invoice_start_date = (
                comparison_window.start_date
                if comparison_window is not None
                else default_extra_start_date
            )
            invoice_end_date = (
                comparison_window.end_date
                if comparison_window is not None
                else default_extra_end_date
            )
            invoice_comparison_years = _comparison_years_within_bounds(
                invoice_years,
                invoice_start_date,
                invoice_end_date,
            )
            existing_invoice_numbers = await _load_existing_client_invoice_comparison_signatures(
                db,
                invoice_comparison_years,
                start_date=invoice_start_date,
                end_date=invoice_end_date,
            )
            existing_invoice_items = await _load_existing_client_invoice_comparison_items(
                db,
                invoice_comparison_years,
                start_date=invoice_start_date,
                end_date=invoice_end_date,
            )
            existing_invoice_issues = find_existing_invoice_issues(
                filtered_invoice_rows,
                existing_invoice_numbers,
                normalize_text=_normalize_text,
            )
            ambiguous_invoice_issues = find_ambiguous_invoice_contact_issues(
                filtered_invoice_rows,
                existing_contacts_by_preview_key,
                normalize_text=_normalize_text,
            )
            workbook_invoice_numbers = {
                _normalize_text(invoice_row.invoice_number)
                for invoice_row in filtered_invoice_rows
                if invoice_row.invoice_number
            }
            extra_invoice_numbers = sorted(existing_invoice_numbers - workbook_invoice_numbers)
            domains.append(
                _make_preview_comparison_domain(
                    kind="invoices",
                    file_rows=len(filtered_invoice_rows)
                    + len(filtered_ignored_issues)
                    + len(filtered_blocked_issues),
                    already_in_solde=len(existing_invoice_issues),
                    missing_in_solde=max(
                        0,
                        len(filtered_invoice_rows)
                        - len(existing_invoice_issues)
                        - len(ambiguous_invoice_issues),
                    ),
                    extra_in_solde=len(extra_invoice_numbers),
                    ignored_by_policy=len(filtered_ignored_issues),
                    blocked=len(filtered_blocked_issues) + len(ambiguous_invoice_issues),
                    extra_in_solde_details=[
                        existing_invoice_items[number] for number in extra_invoice_numbers
                    ],
                )
            )
            continue

        if kind == "payments":
            parsed_sheet, payment_rows, row_issues = _parse_payment_sheet(ws)
            if parsed_sheet is None or parsed_sheet.missing_columns:
                continue
            filtered_payment_rows = (
                [
                    payment_row
                    for payment_row in payment_rows
                    if _is_within_comparison_window(payment_row.payment_date, comparison_window)
                ]
                if comparison_window is not None
                else payment_rows
            )
            filtered_blocked_issues = (
                _filter_date_issues_in_window(
                    row_issues,
                    _build_sheet_row_dates(ws, parsed_sheet),
                    comparison_window,
                )
                if comparison_window is not None
                else row_issues
            )
            payment_start_date = (
                comparison_window.start_date
                if comparison_window is not None
                else default_extra_start_date
            )
            payment_end_date = (
                comparison_window.end_date
                if comparison_window is not None
                else default_extra_end_date
            )
            payment_comparison_years = _comparison_years_within_bounds(
                {payment_row.payment_date.year for payment_row in filtered_payment_rows},
                payment_start_date,
                payment_end_date,
            )
            existing_payment_signatures = await _load_existing_client_payment_comparison_signatures(
                db,
                payment_comparison_years,
                start_date=payment_start_date,
                end_date=payment_end_date,
            )
            existing_payment_items = await _load_existing_client_payment_comparison_items(
                db,
                payment_comparison_years,
                start_date=payment_start_date,
                end_date=payment_end_date,
            )
            already_in_solde = 0
            blocked_validation = 0
            workbook_payment_signatures: set[tuple[str, str, str, str, str]] = set()
            for payment_row in filtered_payment_rows:
                resolution = await _resolve_payment_match(
                    db,
                    payment_row,
                    workbook_candidates,
                )
                blocking_issue = make_payment_resolution_issue(
                    source_row_number=payment_row.source_row_number,
                    status=resolution.status,
                    candidate=resolution.candidate,
                    message=resolution.message,
                    require_persistable_candidate=False,
                )
                if blocking_issue is not None:
                    blocked_validation += 1
                    continue
                if resolution.candidate is None or not resolution.candidate.invoice_number:
                    continue
                payment_signature = _gestion_payment_comparison_signature(
                    reference=resolution.candidate.invoice_number,
                    payment_date=payment_row.payment_date,
                    amount=payment_row.amount,
                    settlement_account=_client_settlement_account_from_method(payment_row.method),
                    cheque_number=payment_row.cheque_number,
                )
                workbook_payment_signatures.add(payment_signature)
                if payment_signature in existing_payment_signatures:
                    already_in_solde += 1
            extra_payment_signatures = sorted(
                existing_payment_signatures - workbook_payment_signatures
            )
            domains.append(
                _make_preview_comparison_domain(
                    kind="payments",
                    file_rows=len(filtered_payment_rows) + len(filtered_blocked_issues),
                    already_in_solde=already_in_solde,
                    missing_in_solde=max(
                        0,
                        len(filtered_payment_rows) - already_in_solde - blocked_validation,
                    ),
                    extra_in_solde=len(extra_payment_signatures),
                    ignored_by_policy=0,
                    blocked=len(filtered_blocked_issues) + blocked_validation,
                    extra_in_solde_details=[
                        existing_payment_items[signature] for signature in extra_payment_signatures
                    ],
                )
            )
            continue

        if kind == "salaries":
            parsed_sheet, salary_rows, row_issues = _parse_salary_sheet(ws)
            if parsed_sheet is None:
                continue
            filtered_salary_rows = (
                [
                    salary_row
                    for salary_row in salary_rows
                    if _is_salary_month_within_comparison_window(
                        salary_row.month, comparison_window
                    )
                ]
                if comparison_window is not None
                else salary_rows
            )
            filtered_blocked_issues = (
                _filter_salary_issues_in_window(
                    row_issues,
                    _build_salary_row_months(ws),
                    comparison_window,
                )
                if comparison_window is not None
                else row_issues
            )
            salary_months = {salary_row.month for salary_row in filtered_salary_rows}
            existing_salary_signatures = await _load_existing_salary_comparison_signatures(
                db,
                salary_months,
            )
            existing_salary_items = await _load_existing_salary_comparison_items(
                db,
                salary_months,
            )
            workbook_salary_signatures = {
                _gestion_salary_comparison_signature(
                    month=salary_row.month,
                    employee_name=salary_row.employee_name,
                    gross=salary_row.gross,
                    net_pay=salary_row.net_pay,
                )
                for salary_row in filtered_salary_rows
            }
            already_in_solde = sum(
                1
                for signature in workbook_salary_signatures
                if signature in existing_salary_signatures
            )
            extra_salary_signatures = sorted(
                existing_salary_signatures - workbook_salary_signatures
            )
            domains.append(
                _make_preview_comparison_domain(
                    kind="salaries",
                    file_rows=len(filtered_salary_rows) + len(filtered_blocked_issues),
                    already_in_solde=already_in_solde,
                    missing_in_solde=max(0, len(filtered_salary_rows) - already_in_solde),
                    extra_in_solde=len(extra_salary_signatures),
                    ignored_by_policy=0,
                    blocked=len(filtered_blocked_issues),
                    extra_in_solde_details=[
                        existing_salary_items[signature] for signature in extra_salary_signatures
                    ],
                )
            )
            continue

        if kind == "bank":
            parsed_sheet, bank_rows, row_issues, ignored_issues = _parse_bank_sheet(ws)
            if parsed_sheet is None or parsed_sheet.missing_columns:
                continue
            filtered_bank_rows = (
                [
                    bank_row
                    for bank_row in bank_rows
                    if _is_within_comparison_window(bank_row.entry_date, comparison_window)
                ]
                if comparison_window is not None
                else bank_rows
            )
            filtered_ignored_issues = (
                _filter_date_issues_in_window(
                    ignored_issues,
                    _build_sheet_row_dates(ws, parsed_sheet),
                    comparison_window,
                )
                if comparison_window is not None
                else ignored_issues
            )
            filtered_blocked_issues = (
                _filter_date_issues_in_window(
                    row_issues,
                    _build_sheet_row_dates(ws, parsed_sheet),
                    comparison_window,
                )
                if comparison_window is not None
                else row_issues
            )
            bank_start_date = (
                comparison_window.start_date if comparison_window is not None else None
            )
            bank_end_date = comparison_window.end_date if comparison_window is not None else None
            if comparison_window is None:
                for bank_row in filtered_bank_rows:
                    bank_start_date, bank_end_date = _expand_date_bounds(
                        bank_start_date,
                        bank_end_date,
                        bank_row.entry_date,
                    )
            bank_comparison_years = _comparison_years_within_bounds(
                {bank_row.entry_date.year for bank_row in filtered_bank_rows},
                bank_start_date,
                bank_end_date,
            )
            existing_bank_signatures = await _load_existing_bank_comparison_signatures(
                db,
                bank_comparison_years,
                start_date=bank_start_date,
                end_date=bank_end_date,
            )
            existing_bank_items = await _load_existing_bank_comparison_items(
                db,
                bank_comparison_years,
                start_date=bank_start_date,
                end_date=bank_end_date,
            )
            filtered_invoice_numbers = await _load_existing_client_invoice_comparison_signatures(
                db,
                _comparison_years_within_bounds(set(), bank_start_date, bank_end_date),
                start_date=bank_start_date,
                end_date=bank_end_date,
            )
            already_in_solde = 0
            blocked_supplier = 0
            workbook_bank_signatures: set[tuple[str, str, str, str]] = set()
            for bank_row in filtered_bank_rows:
                bank_signature = _gestion_bank_comparison_signature(
                    entry_date=bank_row.entry_date,
                    amount=bank_row.amount,
                    description=bank_row.description,
                    reference=bank_row.reference,
                )
                workbook_bank_signatures.add(bank_signature)
                candidate = _supplier_invoice_candidate_from_bank_row(bank_row)
                if bank_signature in existing_bank_signatures:
                    already_in_solde += 1
                    continue
                if candidate is None:
                    continue
                if _normalize_text(candidate.invoice_number) in filtered_invoice_numbers:
                    already_in_solde += 1
                    continue
                _, supplier_blocked_issue = resolve_invoice_contact_match(
                    _invoice_row_from_supplier_candidate(candidate),
                    existing_contacts_by_preview_key,
                    normalize_text=_normalize_text,
                )
                if supplier_blocked_issue is not None:
                    blocked_supplier += 1
            extra_bank_signatures = sorted(existing_bank_signatures - workbook_bank_signatures)
            domains.append(
                _make_preview_comparison_domain(
                    kind="bank",
                    file_rows=len(filtered_bank_rows)
                    + len(filtered_ignored_issues)
                    + len(filtered_blocked_issues),
                    already_in_solde=already_in_solde,
                    missing_in_solde=max(
                        0,
                        len(filtered_bank_rows) - already_in_solde - blocked_supplier,
                    ),
                    extra_in_solde=len(extra_bank_signatures),
                    ignored_by_policy=len(filtered_ignored_issues),
                    blocked=len(filtered_blocked_issues) + blocked_supplier,
                    extra_in_solde_details=[
                        existing_bank_items[signature] for signature in extra_bank_signatures
                    ],
                )
            )
            continue

        if kind == "cash":
            parsed_sheet, cash_rows, row_issues, ignored_issues = _parse_cash_sheet(ws)
            if parsed_sheet is None or parsed_sheet.missing_columns:
                continue
            filtered_cash_rows = (
                [
                    cash_row
                    for cash_row in cash_rows
                    if _is_within_comparison_window(cash_row.entry_date, comparison_window)
                ]
                if comparison_window is not None
                else cash_rows
            )
            filtered_ignored_issues = (
                _filter_date_issues_in_window(
                    ignored_issues,
                    _build_sheet_row_dates(ws, parsed_sheet),
                    comparison_window,
                )
                if comparison_window is not None
                else ignored_issues
            )
            filtered_blocked_issues = (
                _filter_date_issues_in_window(
                    row_issues,
                    _build_sheet_row_dates(ws, parsed_sheet),
                    comparison_window,
                )
                if comparison_window is not None
                else row_issues
            )
            cash_start_date = (
                comparison_window.start_date if comparison_window is not None else None
            )
            cash_end_date = comparison_window.end_date if comparison_window is not None else None
            if comparison_window is None:
                for cash_row in filtered_cash_rows:
                    cash_start_date, cash_end_date = _expand_date_bounds(
                        cash_start_date,
                        cash_end_date,
                        cash_row.entry_date,
                    )
            cash_comparison_years = _comparison_years_within_bounds(
                {cash_row.entry_date.year for cash_row in filtered_cash_rows},
                cash_start_date,
                cash_end_date,
            )
            existing_cash_signatures = await _load_existing_cash_comparison_signatures(
                db,
                cash_comparison_years,
                start_date=cash_start_date,
                end_date=cash_end_date,
            )
            existing_cash_items = await _load_existing_cash_comparison_items(
                db,
                cash_comparison_years,
                start_date=cash_start_date,
                end_date=cash_end_date,
            )
            filtered_invoice_numbers = await _load_existing_client_invoice_comparison_signatures(
                db,
                _comparison_years_within_bounds(set(), cash_start_date, cash_end_date),
                start_date=cash_start_date,
                end_date=cash_end_date,
            )
            already_in_solde = 0
            blocked_supplier = 0
            workbook_cash_signatures: set[tuple[str, str, str, str, str]] = set()
            for cash_row in filtered_cash_rows:
                cash_signature = _gestion_cash_comparison_signature(
                    entry_date=cash_row.entry_date,
                    movement_type=cash_row.movement_type,
                    amount=cash_row.amount,
                    description=cash_row.description,
                    reference=cash_row.reference,
                )
                workbook_cash_signatures.add(cash_signature)
                candidate = _supplier_invoice_candidate_from_cash_row(cash_row)
                if cash_signature in existing_cash_signatures:
                    already_in_solde += 1
                    continue
                if candidate is None:
                    continue
                if _normalize_text(candidate.invoice_number) in filtered_invoice_numbers:
                    already_in_solde += 1
                    continue
                _, supplier_blocked_issue = resolve_invoice_contact_match(
                    _invoice_row_from_supplier_candidate(candidate),
                    existing_contacts_by_preview_key,
                    normalize_text=_normalize_text,
                )
                if supplier_blocked_issue is not None:
                    blocked_supplier += 1
            extra_cash_signatures = sorted(existing_cash_signatures - workbook_cash_signatures)
            domains.append(
                _make_preview_comparison_domain(
                    kind="cash",
                    file_rows=len(filtered_cash_rows)
                    + len(filtered_ignored_issues)
                    + len(filtered_blocked_issues),
                    already_in_solde=already_in_solde,
                    missing_in_solde=max(
                        0,
                        len(filtered_cash_rows) - already_in_solde - blocked_supplier,
                    ),
                    extra_in_solde=len(extra_cash_signatures),
                    ignored_by_policy=len(filtered_ignored_issues),
                    blocked=len(filtered_blocked_issues) + blocked_supplier,
                    extra_in_solde_details=[
                        existing_cash_items[signature] for signature in extra_cash_signatures
                    ],
                )
            )

    return domains


async def _build_comptabilite_preview_comparison_domains(
    db: AsyncSession,
    file_bytes: bytes,
    file_name: str | None,
    comparison_window: _PreviewFilterWindow | None,
) -> list[dict[str, Any]]:
    from io import BytesIO  # noqa: PLC0415

    import openpyxl  # noqa: PLC0415

    existing_entry_signatures = await _load_existing_accounting_entry_signatures(db)
    existing_invoice_numbers = await _load_existing_invoice_numbers(db)
    existing_client_payment_signatures = await _load_existing_client_payment_reference_signatures(
        db
    )
    existing_supplier_payment_signatures = (
        await _load_existing_supplier_payment_reference_signatures(db)
    )
    existing_salary_group_signatures = await _load_existing_salary_group_signatures(db)
    existing_generated_group_signatures = (
        await _load_existing_generated_accounting_group_signatures(db)
    )

    wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    domains: list[dict[str, Any]] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        kind, _ = _classify_comptabilite_sheet(sheet_name)
        if kind != "entries":
            continue
        parsed_sheet, normalized_rows, row_issues, ignored_issues = _parse_entries_sheet(ws)
        if parsed_sheet is None or parsed_sheet.missing_columns:
            continue

        filtered_rows = [
            entry_row
            for entry_row in normalized_rows
            if _is_within_comparison_window(entry_row.entry_date, comparison_window)
        ]
        filtered_ignored_issues = _filter_date_issues_in_window(
            ignored_issues,
            _build_sheet_row_dates(ws, parsed_sheet),
            comparison_window,
        )
        filtered_blocked_issues = _filter_date_issues_in_window(
            row_issues,
            _build_sheet_row_dates(ws, parsed_sheet),
            comparison_window,
        )

        entry_groups = _build_entry_row_groups(filtered_rows)
        already_in_solde = 0
        index = 0
        while index < len(entry_groups):
            entry_group, next_index = _merge_existing_client_invoice_entry_groups(
                entry_groups,
                index,
                existing_invoice_numbers,
            )
            if _matching_existing_client_invoice_reference(entry_group, existing_invoice_numbers):
                already_in_solde += len(entry_group)
                index = next_index
                continue
            if _matching_existing_client_payment_reference(
                entry_group,
                existing_client_payment_signatures,
            ):
                already_in_solde += len(entry_group)
                index = next_index
                continue
            if _matching_existing_supplier_invoice_payment_reference(
                entry_group,
                existing_supplier_payment_signatures,
            ):
                already_in_solde += len(entry_group)
                index = next_index
                continue
            if _matching_existing_salary_entry_group(
                entry_group,
                existing_salary_group_signatures,
            ):
                already_in_solde += len(entry_group)
                index = next_index
                continue
            if (
                _normalized_entry_group_signature(entry_group)
                in existing_generated_group_signatures
            ):
                already_in_solde += len(entry_group)
                index = next_index
                continue

            for entry_row in entry_group:
                signature = _accounting_entry_signature(
                    entry_date=entry_row.entry_date,
                    account_number=entry_row.account_number,
                    label=entry_row.label,
                    debit=entry_row.debit,
                    credit=entry_row.credit,
                )
                if signature in existing_entry_signatures:
                    already_in_solde += 1
            index = next_index

        comparison_years = _comparison_years_within_bounds(
            {entry_row.entry_date.year for entry_row in filtered_rows},
            comparison_window.start_date if comparison_window else None,
            comparison_window.end_date if comparison_window else None,
        )
        existing_comparison_signatures = (
            await _load_existing_accounting_entry_comparison_signatures(
                db,
                comparison_years,
                start_date=comparison_window.start_date if comparison_window else None,
                end_date=comparison_window.end_date if comparison_window else None,
            )
        )
        workbook_entry_signatures = {
            _accounting_entry_signature(
                entry_date=entry_row.entry_date,
                account_number=entry_row.account_number,
                label=entry_row.label,
                debit=entry_row.debit,
                credit=entry_row.credit,
            )
            for entry_row in filtered_rows
        }
        domains.append(
            _make_preview_comparison_domain(
                kind="entries",
                file_rows=(
                    len(filtered_rows) + len(filtered_ignored_issues) + len(filtered_blocked_issues)
                ),
                already_in_solde=already_in_solde,
                missing_in_solde=max(0, len(filtered_rows) - already_in_solde),
                extra_in_solde=len(existing_comparison_signatures - workbook_entry_signatures),
                ignored_by_policy=len(filtered_ignored_issues),
                blocked=len(filtered_blocked_issues),
            )
        )

    return domains


async def _collect_gestion_extra_in_solde_by_kind(
    db: AsyncSession,
    file_bytes: bytes,
    preview: PreviewResult,
    file_name: str | None = None,
) -> dict[str, Any]:
    from io import BytesIO  # noqa: PLC0415

    import openpyxl  # noqa: PLC0415

    workbook_invoice_numbers: set[str] = set()
    workbook_payment_signatures: set[tuple[str, str, str, str, str]] = set()
    workbook_salary_signatures: set[tuple[str, str, str, str]] = set()
    workbook_bank_signatures: set[tuple[str, str, str, str]] = set()
    workbook_cash_signatures: set[tuple[str, str, str, str, str]] = set()
    invoice_years: set[int] = set()
    payment_years: set[int] = set()
    bank_years: set[int] = set()
    cash_years: set[int] = set()
    salary_months: set[str] = set()
    gestion_exercise_bounds = _gestion_file_fiscal_year_bounds(file_name)
    invoice_exercise_bounds = gestion_exercise_bounds
    invoice_start_date: date | None = (
        invoice_exercise_bounds[0] if invoice_exercise_bounds else None
    )
    invoice_end_date: date | None = invoice_exercise_bounds[1] if invoice_exercise_bounds else None
    payment_start_date: date | None = (
        gestion_exercise_bounds[0] if gestion_exercise_bounds else None
    )
    payment_end_date: date | None = gestion_exercise_bounds[1] if gestion_exercise_bounds else None
    bank_start_date: date | None = None
    bank_end_date: date | None = None
    cash_start_date: date | None = None
    cash_end_date: date | None = None
    workbook_candidates: list[PaymentMatchCandidate] = []
    payment_rows: list[NormalizedPaymentRow] = []

    wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        kind, _ = _classify_gestion_sheet(sheet_name)
        sheet_preview = _find_sheet_preview(preview, sheet_name)
        if kind is None or sheet_preview is None or sheet_preview["status"] != "recognized":
            continue

        if kind == "invoices":
            parsed_sheet, invoice_rows, _, _ = _parse_invoice_sheet(ws)
            if parsed_sheet is None or parsed_sheet.missing_columns:
                continue
            invoice_rows, _ = filter_duplicate_invoice_rows(
                invoice_rows,
                normalize_text=_normalize_text,
            )
            for invoice_row in invoice_rows:
                if invoice_row.invoice_number:
                    workbook_invoice_numbers.add(_normalize_text(invoice_row.invoice_number))
                invoice_years.add(invoice_row.invoice_date.year)
                if invoice_exercise_bounds is None:
                    invoice_start_date, invoice_end_date = _expand_date_bounds(
                        invoice_start_date,
                        invoice_end_date,
                        invoice_row.invoice_date,
                    )
                workbook_candidates.append(_make_workbook_invoice_candidate(invoice_row))

        elif kind == "payments":
            parsed_sheet, parsed_payment_rows, _ = _parse_payment_sheet(ws)
            if parsed_sheet is None or parsed_sheet.missing_columns:
                continue
            payment_rows.extend(parsed_payment_rows)
            for payment_row in parsed_payment_rows:
                payment_years.add(payment_row.payment_date.year)
                if gestion_exercise_bounds is None:
                    payment_start_date, payment_end_date = _expand_date_bounds(
                        payment_start_date,
                        payment_end_date,
                        payment_row.payment_date,
                    )

        elif kind == "salaries":
            parsed_sheet, salary_rows, _ = _parse_salary_sheet(ws)
            if parsed_sheet is None or parsed_sheet.missing_columns:
                continue
            for salary_row in salary_rows:
                salary_months.add(salary_row.month)
                workbook_salary_signatures.add(
                    _gestion_salary_comparison_signature(
                        month=salary_row.month,
                        employee_name=salary_row.employee_name,
                        gross=salary_row.gross,
                        net_pay=salary_row.net_pay,
                    )
                )

        elif kind == "bank":
            parsed_sheet, bank_rows, _, _ = _parse_bank_sheet(ws)
            if parsed_sheet is None or parsed_sheet.missing_columns:
                continue
            for bank_row in bank_rows:
                bank_years.add(bank_row.entry_date.year)
                bank_start_date, bank_end_date = _expand_date_bounds(
                    bank_start_date,
                    bank_end_date,
                    bank_row.entry_date,
                )
                workbook_bank_signatures.add(
                    _gestion_bank_comparison_signature(
                        entry_date=bank_row.entry_date,
                        amount=bank_row.amount,
                        description=bank_row.description,
                        reference=bank_row.reference,
                    )
                )

        elif kind == "cash":
            parsed_sheet, cash_rows, _, _ = _parse_cash_sheet(ws)
            if parsed_sheet is None or parsed_sheet.missing_columns:
                continue
            for cash_row in cash_rows:
                cash_years.add(cash_row.entry_date.year)
                cash_start_date, cash_end_date = _expand_date_bounds(
                    cash_start_date,
                    cash_end_date,
                    cash_row.entry_date,
                )
                workbook_cash_signatures.add(
                    _gestion_cash_comparison_signature(
                        entry_date=cash_row.entry_date,
                        movement_type=cash_row.movement_type,
                        amount=cash_row.amount,
                        description=cash_row.description,
                        reference=cash_row.reference,
                    )
                )

    for payment_row in payment_rows:
        resolution = await _resolve_payment_match(
            db,
            payment_row,
            workbook_candidates=workbook_candidates,
        )
        if resolution.status != "matched" or resolution.candidate is None:
            continue
        if not resolution.candidate.invoice_number:
            continue
        workbook_payment_signatures.add(
            _gestion_payment_comparison_signature(
                reference=resolution.candidate.invoice_number,
                payment_date=payment_row.payment_date,
                amount=payment_row.amount,
                settlement_account=_client_settlement_account_from_method(payment_row.method),
                cheque_number=payment_row.cheque_number,
            )
        )

    extra_in_solde_by_kind: dict[str, int] = {}
    extra_in_solde_details_by_kind: dict[str, list[dict[str, str]]] = {}
    if invoice_years:
        invoice_comparison_years = _comparison_years_within_bounds(
            invoice_years,
            invoice_start_date,
            invoice_end_date,
        )
        existing_invoice_numbers = await _load_existing_client_invoice_comparison_signatures(
            db,
            invoice_comparison_years,
            start_date=invoice_start_date,
            end_date=invoice_end_date,
        )
        existing_invoice_items = await _load_existing_client_invoice_comparison_items(
            db,
            invoice_comparison_years,
            start_date=invoice_start_date,
            end_date=invoice_end_date,
        )
        extra_invoice_numbers = sorted(existing_invoice_numbers - workbook_invoice_numbers)
        extra_in_solde_by_kind["invoices"] = len(extra_invoice_numbers)
        if extra_invoice_numbers:
            extra_in_solde_details_by_kind["invoices"] = [
                existing_invoice_items[number] for number in extra_invoice_numbers
            ]
    if payment_years:
        payment_comparison_years = _comparison_years_within_bounds(
            payment_years,
            payment_start_date,
            payment_end_date,
        )
        existing_payment_signatures = await _load_existing_client_payment_comparison_signatures(
            db,
            payment_comparison_years,
            start_date=payment_start_date,
            end_date=payment_end_date,
        )
        existing_payment_items = await _load_existing_client_payment_comparison_items(
            db,
            payment_comparison_years,
            start_date=payment_start_date,
            end_date=payment_end_date,
        )
        extra_payment_signatures = sorted(existing_payment_signatures - workbook_payment_signatures)
        extra_in_solde_by_kind["payments"] = len(extra_payment_signatures)
        if extra_payment_signatures:
            extra_in_solde_details_by_kind["payments"] = [
                existing_payment_items[signature] for signature in extra_payment_signatures
            ]
    if salary_months:
        existing_salary_signatures = await _load_existing_salary_comparison_signatures(
            db,
            salary_months,
        )
        existing_salary_items = await _load_existing_salary_comparison_items(
            db,
            salary_months,
        )
        extra_salary_signatures = sorted(existing_salary_signatures - workbook_salary_signatures)
        extra_in_solde_by_kind["salaries"] = len(extra_salary_signatures)
        if extra_salary_signatures:
            extra_in_solde_details_by_kind["salaries"] = [
                existing_salary_items[signature] for signature in extra_salary_signatures
            ]
    if bank_years:
        bank_comparison_years = _comparison_years_within_bounds(
            bank_years,
            bank_start_date,
            bank_end_date,
        )
        existing_bank_signatures = await _load_existing_bank_comparison_signatures(
            db,
            bank_comparison_years,
            start_date=bank_start_date,
            end_date=bank_end_date,
        )
        existing_bank_items = await _load_existing_bank_comparison_items(
            db,
            bank_comparison_years,
            start_date=bank_start_date,
            end_date=bank_end_date,
        )
        extra_bank_signatures = sorted(existing_bank_signatures - workbook_bank_signatures)
        extra_in_solde_by_kind["bank"] = len(extra_bank_signatures)
        if extra_bank_signatures:
            extra_in_solde_details_by_kind["bank"] = [
                existing_bank_items[signature] for signature in extra_bank_signatures
            ]
    if cash_years:
        cash_comparison_years = _comparison_years_within_bounds(
            cash_years,
            cash_start_date,
            cash_end_date,
        )
        existing_cash_signatures = await _load_existing_cash_comparison_signatures(
            db,
            cash_comparison_years,
            start_date=cash_start_date,
            end_date=cash_end_date,
        )
        existing_cash_items = await _load_existing_cash_comparison_items(
            db,
            cash_comparison_years,
            start_date=cash_start_date,
            end_date=cash_end_date,
        )
        extra_cash_signatures = sorted(existing_cash_signatures - workbook_cash_signatures)
        extra_in_solde_by_kind["cash"] = len(extra_cash_signatures)
        if extra_cash_signatures:
            extra_in_solde_details_by_kind["cash"] = [
                existing_cash_items[signature] for signature in extra_cash_signatures
            ]

    return {
        "counts": extra_in_solde_by_kind,
        "details": extra_in_solde_details_by_kind,
    }


async def _load_existing_accounting_entry_comparison_signatures(
    db: AsyncSession,
    years: set[int],
    start_date: date | None = None,
    end_date: date | None = None,
) -> set[str]:
    if not years:
        return set()

    from sqlalchemy import select  # noqa: PLC0415

    from backend.models.accounting_entry import AccountingEntry  # noqa: PLC0415

    result = await db.execute(
        select(
            AccountingEntry.date,
            AccountingEntry.account_number,
            AccountingEntry.label,
            AccountingEntry.debit,
            AccountingEntry.credit,
        )
    )
    return {
        _accounting_entry_signature(
            entry_date=entry_date,
            account_number=account_number,
            label=label,
            debit=debit,
            credit=credit,
        )
        for entry_date, account_number, label, debit, credit in result.all()
        if entry_date.year in years and _is_within_date_bounds(entry_date, start_date, end_date)
    }


async def _collect_comptabilite_extra_in_solde_by_kind(
    db: AsyncSession,
    file_bytes: bytes,
    preview: PreviewResult,
) -> dict[str, int]:
    from io import BytesIO  # noqa: PLC0415

    import openpyxl  # noqa: PLC0415

    workbook_entry_signatures: set[str] = set()
    workbook_years: set[int] = set()

    wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        kind, _ = _classify_comptabilite_sheet(sheet_name)
        sheet_preview = _find_sheet_preview(preview, sheet_name)
        if kind != "entries" or sheet_preview is None or sheet_preview["status"] != "recognized":
            continue

        parsed_sheet, normalized_rows, _, _ = _parse_entries_sheet(ws)
        if parsed_sheet is None or parsed_sheet.missing_columns:
            continue

        for entry_row in normalized_rows:
            workbook_years.add(entry_row.entry_date.year)
            workbook_entry_signatures.add(
                _accounting_entry_signature(
                    entry_date=entry_row.entry_date,
                    account_number=entry_row.account_number,
                    label=entry_row.label,
                    debit=entry_row.debit,
                    credit=entry_row.credit,
                )
            )

    if not workbook_years:
        return {}

    existing_entry_signatures = await _load_existing_accounting_entry_comparison_signatures(
        db,
        workbook_years,
    )
    return {"entries": len(existing_entry_signatures - workbook_entry_signatures)}


def _matching_existing_salary_entry_group(
    entry_rows: list[NormalizedEntryRow],
    existing_salary_group_signatures: dict[str, set[tuple[str, ...]]],
) -> str | None:
    month = _extract_salary_month(*(entry_row.label for entry_row in entry_rows))
    if month is None:
        return None
    normalized_labels = [_normalize_text(entry_row.label) for entry_row in entry_rows]
    if not any(_is_salary_related_label(label) for label in normalized_labels):
        return None
    month_signatures = existing_salary_group_signatures.get(month)
    if month_signatures is None:
        return None
    if _entry_group_amount_signature(entry_rows) in month_signatures:
        return month
    if _is_salary_accrual_like_entry_group(entry_rows):
        return month
    if _is_salary_payment_like_entry_group(entry_rows):
        return month
    return None


def _supplier_invoice_candidate_from_bank_row(
    row: NormalizedBankRow,
) -> _SupplierInvoiceCandidate | None:
    if row.amount >= 0:
        return None
    invoice_number = _extract_supplier_invoice_reference(row.description, row.reference)
    if not invoice_number:
        return None
    description = _strip_supplier_invoice_reference(row.description) or invoice_number
    contact_name = _infer_supplier_contact_name(description) or description
    label = "cs" if _is_supplier_subcontracting_description(description) else "general"
    return _SupplierInvoiceCandidate(
        source_row_number=row.source_row_number,
        invoice_date=row.entry_date,
        amount=abs(row.amount),
        invoice_number=invoice_number,
        contact_name=contact_name,
        description=description,
        label=label,
        payment_method="virement",
    )


def _supplier_invoice_candidate_from_cash_row(
    row: NormalizedCashRow,
) -> _SupplierInvoiceCandidate | None:
    if row.movement_type != "out":
        return None
    invoice_number = _extract_supplier_invoice_reference(row.reference, row.description)
    if not invoice_number:
        return None
    description = _strip_supplier_invoice_reference(row.description) or invoice_number
    contact_name = row.contact_name or _infer_supplier_contact_name(description) or description
    label = "cs" if _is_supplier_subcontracting_description(description) else "general"
    return _SupplierInvoiceCandidate(
        source_row_number=row.source_row_number,
        invoice_date=row.entry_date,
        amount=row.amount,
        invoice_number=invoice_number,
        contact_name=contact_name,
        description=description,
        label=label,
        payment_method="especes",
    )


def _invoice_row_from_supplier_candidate(
    candidate: _SupplierInvoiceCandidate,
) -> NormalizedInvoiceRow:
    return NormalizedInvoiceRow(
        source_row_number=candidate.source_row_number,
        invoice_date=candidate.invoice_date,
        amount=candidate.amount,
        contact_name=candidate.contact_name,
        invoice_number=candidate.invoice_number,
        label=candidate.label,
    )


def _build_entry_row_groups(
    normalized_rows: list[NormalizedEntryRow],
) -> list[list[NormalizedEntryRow]]:
    if not normalized_rows:
        return []

    if not any(entry_row.change_marker is not None for entry_row in normalized_rows):
        return [[entry_row] for entry_row in normalized_rows]

    marker_groups: list[list[NormalizedEntryRow]] = []
    current_group: list[NormalizedEntryRow] = []
    last_change_marker: str | None = None
    for entry_row in normalized_rows:
        if (
            current_group
            and entry_row.change_marker is not None
            and last_change_marker is not None
            and entry_row.change_marker != last_change_marker
        ):
            marker_groups.append(current_group)
            current_group = []

        current_group.append(entry_row)
        if entry_row.change_marker is not None:
            last_change_marker = entry_row.change_marker

    if current_group:
        marker_groups.append(current_group)

    groups: list[list[NormalizedEntryRow]] = []
    for marker_group in marker_groups:
        balance = Decimal("0")
        current_subgroup: list[NormalizedEntryRow] = []
        for index, entry_row in enumerate(marker_group):
            current_subgroup.append(entry_row)
            balance += entry_row.debit - entry_row.credit
            if balance == 0 and index < len(marker_group) - 1:
                groups.append(current_subgroup)
                current_subgroup = []
        if current_subgroup:
            groups.append(current_subgroup)

    return groups


def _normalized_entry_group_signature(
    entry_rows: list[NormalizedEntryRow],
) -> tuple[str, ...]:
    return _accounting_entry_group_signature(
        [
            _accounting_entry_line_signature(
                entry_date=entry_row.entry_date,
                account_number=entry_row.account_number,
                debit=entry_row.debit,
                credit=entry_row.credit,
            )
            for entry_row in entry_rows
        ]
    )


async def _queue_client_payment_from_bank_row(
    db: AsyncSession,
    *,
    bank_row: NormalizedBankRow,
    existing_payment_signatures: set[tuple[int, str, str, str, str]],
    created_payments: list[tuple[Any, Any]],
    affected_invoice_ids: set[int],
    result: ImportResult,
    sheet_name: str,
) -> None:
    from backend.models.invoice import InvoiceType  # noqa: PLC0415
    from backend.models.payment import Payment, PaymentMethod  # noqa: PLC0415

    if bank_row.amount <= 0:
        return

    invoice_reference = _single_client_invoice_reference(bank_row.reference, bank_row.description)
    if invoice_reference is None:
        return

    resolution = await _resolve_payment_match(
        db,
        _payment_row_from_bank_row(bank_row, invoice_reference=invoice_reference),
    )
    blocking_issue = make_payment_resolution_issue(
        source_row_number=bank_row.source_row_number,
        status=resolution.status,
        candidate=resolution.candidate,
        message=resolution.message,
        require_persistable_candidate=True,
    )
    if blocking_issue is not None:
        return

    candidate = resolution.candidate
    assert candidate is not None
    assert candidate.invoice_id is not None
    signature = _payment_signature(
        invoice_id=candidate.invoice_id,
        payment_date=bank_row.entry_date,
        amount=bank_row.amount,
        method=PaymentMethod.VIREMENT.value,
    )
    if signature in existing_payment_signatures:
        return

    payment = Payment(
        invoice_id=candidate.invoice_id,
        contact_id=candidate.contact_id,
        date=bank_row.entry_date,
        amount=bank_row.amount,
        method=PaymentMethod.VIREMENT,
        reference=invoice_reference,
        notes=bank_row.description,
        deposited=True,
        deposit_date=bank_row.entry_date,
    )
    db.add(payment)
    created_payments.append((payment, candidate.invoice_type or InvoiceType.CLIENT))
    affected_invoice_ids.add(candidate.invoice_id)
    existing_payment_signatures.add(signature)
    result.payments_created += 1
    result.add_imported_row(sheet_name, "payments")


async def _ensure_supplier_invoice_payment(
    db: AsyncSession,
    *,
    candidate: _SupplierInvoiceCandidate,
    existing_contacts_by_preview_key: dict[str, list[Any]],
    result: ImportResult,
    sheet_name: str,
    default_invoice_due_days: int | None,
) -> tuple[int, int, int | None, bool]:
    from sqlalchemy import select  # noqa: PLC0415

    from backend.models.contact import Contact, ContactType  # noqa: PLC0415
    from backend.models.invoice import (  # noqa: PLC0415
        Invoice,
        InvoiceLabel,
        InvoiceStatus,
        InvoiceType,
    )
    from backend.models.payment import Payment, PaymentMethod  # noqa: PLC0415

    invoice_result = await db.execute(
        select(Invoice).where(Invoice.number == candidate.invoice_number)
    )
    invoice = invoice_result.scalar_one_or_none()
    created_invoice = False
    if invoice is None:
        contact_key = _normalize_text(candidate.contact_name)
        matched_contact, contact_issue = resolve_invoice_contact_match(
            _invoice_row_from_supplier_candidate(candidate),
            existing_contacts_by_preview_key,
            normalize_text=_normalize_text,
        )
        if contact_issue is not None:
            raise ValueError(format_row_issue(contact_issue))
        if matched_contact is None:
            contact = Contact(nom=candidate.contact_name, type=ContactType.FOURNISSEUR)
            db.add(contact)
            await db.flush()
            existing_contacts_by_preview_key.setdefault(contact_key, []).append(contact)
            result.contacts_created += 1
            result.record_created_object(
                sheet_name=sheet_name,
                kind="contacts",
                object_type="contact",
                object_id=contact.id,
                reference=contact.nom,
                details={"created_from": "supplier_import"},
            )
        else:
            contact = matched_contact
            if contact.type == ContactType.CLIENT:
                contact.type = ContactType.LES_DEUX
        invoice = Invoice(
            number=candidate.invoice_number,
            type=InvoiceType.FOURNISSEUR,
            contact_id=contact.id,
            date=candidate.invoice_date,
            due_date=apply_default_due_date(
                candidate.invoice_date,
                None,
                default_invoice_due_days,
            ),
            total_amount=candidate.amount,
            paid_amount=Decimal("0"),
            status=InvoiceStatus.SENT,
            label=InvoiceLabel(candidate.label),
            description=candidate.description,
            reference=candidate.invoice_number,
        )
        db.add(invoice)
        await db.flush()
        created_invoice = True
        result.invoices_created += 1
        result.record_created_object(
            sheet_name=sheet_name,
            kind="invoices",
            object_type="invoice",
            object_id=invoice.id,
            reference=invoice.number,
            details={"contact_id": invoice.contact_id, "imported_from": "supplier_import"},
        )
    payment = Payment(
        invoice_id=invoice.id,
        contact_id=invoice.contact_id,
        amount=candidate.amount,
        date=candidate.invoice_date,
        method=PaymentMethod(candidate.payment_method),
        reference=candidate.invoice_number,
        notes=candidate.description,
    )
    db.add(payment)
    await db.flush()
    result.payments_created += 1
    result.record_created_object(
        sheet_name=sheet_name,
        kind="payments",
        object_type="payment",
        object_id=payment.id,
        reference=candidate.invoice_number,
        details={
            "invoice_id": payment.invoice_id,
            "contact_id": payment.contact_id,
            "amount": str(payment.amount),
        },
    )
    return invoice.id, payment.id, invoice.contact_id, created_invoice


# ---------------------------------------------------------------------------
# Main import functions
# ---------------------------------------------------------------------------


async def import_gestion_file(
    db: AsyncSession, file_bytes: bytes, file_name: str | None = None
) -> ImportResult:
    """Import contacts, invoices and payments from a 'Gestion YYYY.xlsx' file.

    Expected sheets:
    - 'Factures' or 'Clients' : contact, date, montant, type, statut
    - 'Paiements' : date, contact, montant, mode, N° chèque

    The parser blocks recognized sheets containing invalid rows or unresolved payments.
    """
    import openpyxl  # noqa: PLC0415

    result = ImportResult()

    file_hash = _compute_file_hash(file_bytes)
    preview = await preview_gestion_file(db, file_bytes, file_name)
    if preview.errors:
        result.absorb_preview(preview)
        await _record_import_log(
            db,
            import_type="gestion",
            status="blocked",
            file_hash=file_hash,
            file_name=file_name,
            result=result,
        )
        return result

    try:
        from io import BytesIO  # noqa: PLC0415

        wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        result.add_open_file_error(exc)
        await _record_import_log(
            db,
            import_type="gestion",
            status="failed",
            file_hash=file_hash,
            file_name=file_name,
            result=result,
        )
        return result

    sheets_by_kind: dict[str, list[Any]] = {kind: [] for kind in _GESTION_IMPORT_ORDER}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        kind, _ = _classify_gestion_sheet(sheet_name)
        if kind in sheets_by_kind:
            sheets_by_kind[kind].append(ws)

    for kind in _GESTION_IMPORT_ORDER:
        try:
            for ws in sheets_by_kind[kind]:
                if kind == "invoices":
                    await _import_invoices_sheet(db, ws, result)
                elif kind == "payments":
                    await _import_payments_sheet(db, ws, result)
                elif kind == "contacts":
                    await _import_contacts_sheet(db, ws, result)
                elif kind == "salaries":
                    await _import_salaries_sheet(db, ws, result)
                elif kind == "cash":
                    await _import_cash_sheet(db, ws, result)
                elif kind == "bank":
                    await _import_bank_sheet(db, ws, result)
        except Exception as exc:
            logger.error("Gestion import aborted during %s: %s", kind, exc, exc_info=True)
            await db.rollback()
            result.reset_persisted_counts()
            if not isinstance(exc, _ImportSheetFailure):
                result.add_import_error("gestion", exc)
            await _record_import_log(
                db,
                import_type="gestion",
                status="failed",
                file_hash=file_hash,
                file_name=file_name,
                result=result,
            )
            return result

    await _record_import_log(
        db,
        import_type="gestion",
        status="success",
        file_hash=file_hash,
        file_name=file_name,
        result=result,
    )
    return result


async def import_comptabilite_file(
    db: AsyncSession, file_bytes: bytes, file_name: str | None = None
) -> ImportResult:
    """Import accounting entries from a 'Comptabilité YYYY.xlsx' file.

    Expected sheet columns (flexible column detection):
    date | N° pièce | compte | libellé | débit | crédit
    """
    import openpyxl  # noqa: PLC0415

    result = ImportResult()

    file_hash = _compute_file_hash(file_bytes)
    preview = await preview_comptabilite_file(db, file_bytes)
    if preview.errors:
        result.absorb_preview(preview)
        await _record_import_log(
            db,
            import_type="comptabilite",
            status="blocked",
            file_hash=file_hash,
            file_name=file_name,
            result=result,
        )
        return result

    try:
        from io import BytesIO  # noqa: PLC0415

        wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        result.add_open_file_error(exc)
        await _record_import_log(
            db,
            import_type="comptabilite",
            status="failed",
            file_hash=file_hash,
            file_name=file_name,
            result=result,
        )
        return result

    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            kind, _ = _classify_comptabilite_sheet(sheet_name)
            if kind == "entries":
                await _import_entries_sheet(db, ws, result)
    except Exception as exc:
        logger.error("Comptabilite import aborted: %s", exc, exc_info=True)
        await db.rollback()
        result.reset_persisted_counts()
        if not isinstance(exc, _ImportSheetFailure):
            result.add_import_error("comptabilite", exc)
        await _record_import_log(
            db,
            import_type="comptabilite",
            status="failed",
            file_hash=file_hash,
            file_name=file_name,
            result=result,
        )
        return result

    await _record_import_log(
        db,
        import_type="comptabilite",
        status="success",
        file_hash=file_hash,
        file_name=file_name,
        result=result,
    )
    return result


def _is_cash_pending_deposit_forecast(
    row: tuple[Any, ...],
    *,
    tiers_idx: int | None,
    libelle_idx: int | None,
    raw_amount: Decimal | None,
) -> bool:
    """Recognize a cash-to-bank forecast row that should be ignored until dated."""
    return should_ignore_cash_pending_deposit_forecast(
        row,
        tiers_idx=tiers_idx,
        libelle_idx=libelle_idx,
        raw_amount=raw_amount,
        get_row_value=_get_row_value,
        parse_str=_parse_str,
        normalize_text=_normalize_text,
    )


def _parse_invoice_sheet(
    ws: Any,
) -> tuple[
    ParsedSheet | None,
    list[NormalizedInvoiceRow],
    list[RowValidationIssue],
    list[RowIgnoredIssue],
]:
    return _parse_invoice_sheet_impl(ws)


def _parse_payment_sheet(
    ws: Any,
) -> tuple[ParsedSheet | None, list[NormalizedPaymentRow], list[RowValidationIssue]]:
    return _parse_payment_sheet_impl(ws)


def _parse_salary_sheet(
    ws: Any,
) -> tuple[ParsedSheet | None, list[NormalizedSalaryRow], list[RowValidationIssue]]:
    return _parse_salary_sheet_impl(ws)


def _parse_contact_sheet(
    ws: Any,
) -> tuple[ParsedSheet | None, list[NormalizedContactRow], list[RowValidationIssue]]:
    return _parse_contact_sheet_impl(ws)


def _parse_cash_sheet(
    ws: Any,
) -> tuple[
    ParsedSheet | None,
    list[NormalizedCashRow],
    list[RowValidationIssue],
    list[RowIgnoredIssue],
]:
    return _parse_cash_sheet_impl(ws)


def _parse_bank_sheet(
    ws: Any,
) -> tuple[
    ParsedSheet | None,
    list[NormalizedBankRow],
    list[RowValidationIssue],
    list[RowIgnoredIssue],
]:
    return _parse_bank_sheet_impl(ws)


def _parse_entries_sheet(
    ws: Any,
) -> tuple[
    ParsedSheet | None,
    list[NormalizedEntryRow],
    list[RowValidationIssue],
    list[RowIgnoredIssue],
]:
    return _parse_entries_sheet_impl(ws)


async def _import_contacts_sheet(db: AsyncSession, ws: Any, result: ImportResult) -> None:
    """Import contacts from a sheet."""
    from backend.models.contact import Contact, ContactType  # noqa: PLC0415

    parsed_sheet, normalized_rows, _ = _parse_contact_sheet(ws)
    if parsed_sheet is None or parsed_sheet.missing_columns:
        result.skipped += 1
        return

    created_contacts: list[Contact] = []
    existing_contacts_by_preview_key = await _load_existing_contacts_by_preview_key(db)
    normalized_rows, ignored_issues = filter_duplicate_contact_rows(
        normalized_rows,
        normalize_text=_normalize_text,
    )
    for ignored_issue in ignored_issues:
        result.add_ignored_row(
            ws.title,
            "contacts",
            format_row_issue(ignored_issue),
        )

    for contact_row in normalized_rows:
        preview_key = _contact_preview_key(contact_row.nom, contact_row.prenom)
        if existing_contacts_by_preview_key.get(preview_key):
            ignored_issue = make_existing_contact_issue(contact_row.source_row_number)
            result.add_ignored_row(
                ws.title,
                "contacts",
                format_row_issue(ignored_issue),
            )
            continue

        contact = Contact(
            nom=contact_row.nom,
            prenom=contact_row.prenom,
            email=contact_row.email,
            type=ContactType.CLIENT,
        )
        db.add(contact)
        created_contacts.append(contact)
        existing_contacts_by_preview_key.setdefault(preview_key, []).append(contact)
        result.contacts_created += 1
        result.add_imported_row(ws.title, "contacts")

    try:
        await db.flush()
        for contact in created_contacts:
            result.record_created_object(
                sheet_name=ws.title,
                kind="contacts",
                object_type="contact",
                object_id=contact.id,
                reference=_format_contact_display_name(contact.nom, contact.prenom),
            )
        logger.debug("Contacts import done — created=%d", result.contacts_created)
    except Exception as exc:
        logger.error("Contacts flush failed: %s", exc, exc_info=True)
        result.add_import_error("contacts", exc)
        await db.rollback()
        raise _ImportSheetFailure from exc


async def _import_invoices_sheet(db: AsyncSession, ws: Any, result: ImportResult) -> None:
    """Import invoices from a sheet with flexible column detection."""
    from sqlalchemy import select  # noqa: PLC0415
    from sqlalchemy.orm import selectinload  # noqa: PLC0415

    from backend.models.contact import Contact, ContactType  # noqa: PLC0415
    from backend.models.invoice import (  # noqa: PLC0415
        Invoice,
        InvoiceLine,
        InvoiceStatus,
        InvoiceType,
        derive_client_invoice_label,
    )

    parsed_sheet, normalized_rows, _, ignored_issues = _parse_invoice_sheet(ws)
    if parsed_sheet is None or parsed_sheet.missing_columns:
        result.skipped += 1
        return

    logger.debug(
        "Importing invoices sheet — rows=%s missing_columns=%s",
        len(normalized_rows),
        parsed_sheet.missing_columns,
    )

    for ignored_issue in ignored_issues:
        result.add_ignored_row(
            ws.title,
            "invoices",
            format_row_issue(ignored_issue),
        )

    normalized_rows, ignored_issues = filter_duplicate_invoice_rows(
        normalized_rows,
        normalize_text=_normalize_text,
    )
    for ignored_issue in ignored_issues:
        result.add_ignored_row(
            ws.title,
            "invoices",
            format_row_issue(ignored_issue),
        )

    created_invoices: list[Invoice] = []
    created_contacts: list[Contact] = []
    existing_contacts_by_preview_key = await _load_existing_contacts_by_preview_key(db)
    generated_number_index = 0
    default_invoice_due_days = await settings_service.get_default_invoice_due_days(db)
    for invoice_row in normalized_rows:
        invoice_date = invoice_row.invoice_date
        total = invoice_row.amount

        # Find or create contact
        contact_id: int | None = None
        contact_key = _normalize_text(invoice_row.contact_name)
        matched_contact, contact_issue = resolve_invoice_contact_match(
            invoice_row,
            existing_contacts_by_preview_key,
            normalize_text=_normalize_text,
        )
        if contact_issue is not None:
            raise ValueError(format_row_issue(contact_issue))
        if matched_contact is not None:
            contact_id = matched_contact.id
        else:
            contact_nom, contact_prenom = _split_contact_full_name(invoice_row.contact_name)
            new_contact = Contact(
                nom=contact_nom,
                prenom=contact_prenom,
                type=ContactType.CLIENT,
            )
            db.add(new_contact)
            await db.flush()
            created_contacts.append(new_contact)
            contact_id = new_contact.id
            existing_contacts_by_preview_key.setdefault(contact_key, []).append(new_contact)
            result.contacts_created += 1
            logger.debug(
                "Row %d — created contact '%s' (id=%s)",
                invoice_row.source_row_number,
                invoice_row.contact_name,
                contact_id,
            )
        if contact_id is None:
            result.skipped += 1
            continue

        number_raw = invoice_row.invoice_number or ""
        if not number_raw:
            generated_number_index += 1
            number_raw = f"IMP-{invoice_date.year}-{generated_number_index:04d}"

        # Idempotency: skip if already in DB
        existing_inv = await db.execute(select(Invoice).where(Invoice.number == number_raw))
        if existing_inv.scalar_one_or_none() is not None:
            ignored_issue = make_existing_invoice_issue(invoice_row.source_row_number)
            logger.debug(
                "Row %d — invoice '%s' already exists, skipping",
                invoice_row.source_row_number,
                number_raw,
            )
            result.add_ignored_row(
                ws.title,
                "invoices",
                format_row_issue(ignored_issue),
            )
            continue

        invoice_lines = _build_client_invoice_lines_from_import_row(invoice_row)
        derived_label = derive_client_invoice_label(
            {line["line_type"] for line in invoice_lines if line["amount"] > 0}
        )

        invoice = Invoice(
            number=number_raw,
            type=InvoiceType.CLIENT,
            contact_id=contact_id,
            date=invoice_date,
            due_date=apply_default_due_date(invoice_date, None, default_invoice_due_days),
            total_amount=total,
            paid_amount=Decimal("0"),
            status=InvoiceStatus.SENT,
            label=derived_label,
            has_explicit_breakdown=len(invoice_lines) > 1,
        )
        db.add(invoice)
        await db.flush()
        db.add_all(
            [
                InvoiceLine(
                    invoice_id=invoice.id,
                    description=line["description"],
                    line_type=line["line_type"],
                    quantity=Decimal("1"),
                    unit_price=line["amount"],
                    amount=line["amount"],
                )
                for line in invoice_lines
            ]
        )
        created_invoices.append(invoice)
        logger.debug(
            "Row %d — invoice '%s' queued (contact_id=%d, amount=%s)",
            invoice_row.source_row_number,
            number_raw,
            contact_id,
            total,
        )
        result.invoices_created += 1
        result.add_imported_row(ws.title, "invoices")

    try:
        await db.flush()
        for contact in created_contacts:
            result.record_created_object(
                sheet_name=ws.title,
                kind="contacts",
                object_type="contact",
                object_id=contact.id,
                reference=_format_contact_display_name(contact.nom, contact.prenom),
                details={"created_from": "invoice_sheet"},
            )
        for inv_obj in created_invoices:
            result.record_created_object(
                sheet_name=ws.title,
                kind="invoices",
                object_type="invoice",
                object_id=inv_obj.id,
                reference=inv_obj.number,
                details={"contact_id": inv_obj.contact_id},
            )
        # Generate accounting entries for each new invoice (no-op if no rules configured)
        from backend.services.accounting_engine import (
            generate_entries_for_invoice,
        )  # noqa: PLC0415

        for inv_obj in created_invoices:
            try:
                refreshed_invoice = (
                    (
                        await db.execute(
                            select(Invoice)
                            .where(Invoice.id == inv_obj.id)
                            .options(selectinload(Invoice.lines))
                        )
                    )
                    .scalars()
                    .one()
                )
                entries = await generate_entries_for_invoice(db, refreshed_invoice)
                result.entries_created += len(entries)
            except Exception as e:
                logger.debug("Accounting entries skipped for invoice '%s': %s", inv_obj.number, e)
                result.add_warning(
                    ws.title,
                    "invoices",
                    f"Ecritures comptables non generees pour la facture {inv_obj.number} : {e}",
                )
        logger.debug(
            "Invoices import done — created=%d skipped=%d entries=%d",
            result.invoices_created,
            result.skipped,
            result.entries_created,
        )
    except Exception as exc:
        logger.error("Invoices flush failed: %s", exc, exc_info=True)
        result.add_import_error("factures", exc)
        await db.rollback()
        raise _ImportSheetFailure from exc


async def _import_payments_sheet(db: AsyncSession, ws: Any, result: ImportResult) -> None:
    """Import payments from a sheet."""
    from backend.models.invoice import InvoiceType  # noqa: PLC0415
    from backend.models.payment import Payment  # noqa: PLC0415

    parsed_sheet, normalized_rows, _ = _parse_payment_sheet(ws)
    if parsed_sheet is None or parsed_sheet.missing_columns:
        result.skipped += 1
        return

    logger.debug(
        "Importing payments sheet — rows=%s missing_columns=%s",
        len(normalized_rows),
        parsed_sheet.missing_columns,
    )

    created_payments: list[tuple[Payment, InvoiceType]] = []
    affected_invoice_ids: set[int] = set()
    for payment_row in normalized_rows:
        pay_date = payment_row.payment_date
        amount = payment_row.amount
        method = payment_row.method

        resolution = await _resolve_payment_match(db, payment_row)
        blocking_issue = make_payment_resolution_issue(
            source_row_number=payment_row.source_row_number,
            status=resolution.status,
            candidate=resolution.candidate,
            message=resolution.message,
            require_persistable_candidate=True,
        )
        if blocking_issue is not None:
            raise ValueError(format_row_issue(blocking_issue))

        candidate = resolution.candidate
        assert candidate is not None
        assert candidate.invoice_id is not None
        invoice_id = candidate.invoice_id
        contact_id = candidate.contact_id
        inv_type = candidate.invoice_type or InvoiceType.CLIENT

        payment = Payment(
            invoice_id=invoice_id,
            contact_id=contact_id,
            date=pay_date,
            amount=amount,
            method=method,
            cheque_number=payment_row.cheque_number,
            deposited=payment_row.deposited,
            deposit_date=payment_row.deposit_date,
        )
        db.add(payment)
        created_payments.append((payment, inv_type))
        affected_invoice_ids.add(invoice_id)
        logger.debug(
            "Row %d — payment %s queued (invoice_id=%d, amount=%s)",
            payment_row.source_row_number,
            pay_date,
            invoice_id,
            amount,
        )
        result.payments_created += 1
        result.add_imported_row(ws.title, "payments")

    try:
        await db.flush()
        for payment, _ in created_payments:
            result.record_created_object(
                sheet_name=ws.title,
                kind="payments",
                object_type="payment",
                object_id=payment.id,
                reference=str(payment.id) if payment.id is not None else None,
                details={
                    "invoice_id": payment.invoice_id,
                    "contact_id": payment.contact_id,
                    "amount": str(payment.amount),
                    "date": payment.date.isoformat(),
                },
            )
        # Refresh invoice statuses and generate accounting entries
        from backend.services.accounting_engine import (
            generate_entries_for_payment,
        )  # noqa: PLC0415
        from backend.services.payment import _refresh_invoice_status  # noqa: PLC0415

        for inv_id in affected_invoice_ids:
            await _refresh_invoice_status(db, inv_id)
        for p, p_inv_type in created_payments:
            try:
                entries = await generate_entries_for_payment(db, p, p_inv_type)
                result.entries_created += len(entries)
            except Exception as e:
                logger.debug("Accounting entries skipped for payment: %s", e)
                result.add_warning(
                    ws.title,
                    "payments",
                    f"Ecritures comptables non generees pour un paiement importe : {e}",
                )
        await db.flush()
        logger.debug(
            "Payments import done — created=%d skipped=%d entries=%d",
            result.payments_created,
            result.skipped,
            result.entries_created,
        )
    except Exception as exc:
        logger.error("Payments flush failed: %s", exc, exc_info=True)
        result.add_import_error("paiements", exc)
        await db.rollback()
        raise _ImportSheetFailure from exc


async def _import_salaries_sheet(db: AsyncSession, ws: Any, result: ImportResult) -> None:
    """Import salary records from the historical `Aide Salaires` worksheet."""
    from sqlalchemy import select  # noqa: PLC0415

    from backend.models.accounting_entry import AccountingEntry, EntrySourceType  # noqa: PLC0415
    from backend.models.contact import Contact, ContactType  # noqa: PLC0415
    from backend.models.salary import Salary  # noqa: PLC0415
    from backend.services.accounting_engine import _next_entry_number  # noqa: PLC0415
    from backend.services.fiscal_year_service import find_fiscal_year_id_for_date  # noqa: PLC0415

    parsed_sheet, normalized_rows, _ = _parse_salary_sheet(ws)
    if parsed_sheet is None or parsed_sheet.missing_columns:
        result.skipped += 1
        return

    existing_contacts_by_salary_key = await _load_existing_contacts_by_salary_key(db)
    existing_salary_keys = await _load_existing_salary_keys(db)
    created_contacts: list[Contact] = []
    created_salaries: list[Salary] = []
    touched_months: set[str] = set()

    for salary_row in normalized_rows:
        employee_key = _salary_employee_key(salary_row.employee_name)
        contact = existing_contacts_by_salary_key.get(employee_key)
        if contact is None:
            contact = Contact(nom=salary_row.employee_name, type=ContactType.FOURNISSEUR)
            db.add(contact)
            await db.flush()
            created_contacts.append(contact)
            existing_contacts_by_salary_key[employee_key] = contact
            result.contacts_created += 1

        salary_key = (salary_row.month, employee_key)
        if salary_key in existing_salary_keys:
            result.add_ignored_row(
                ws.title,
                "salaries",
                format_row_issue(
                    RowIgnoredIssue(
                        source_row_number=salary_row.source_row_number,
                        message=EXISTING_SALARY_MESSAGE,
                    )
                ),
            )
            continue

        salary = Salary(
            employee_id=contact.id,
            month=salary_row.month,
            hours=salary_row.hours,
            gross=salary_row.gross,
            employee_charges=salary_row.employee_charges,
            employer_charges=salary_row.employer_charges,
            tax=salary_row.tax,
            net_pay=salary_row.net_pay,
            notes="Imported from Gestion Excel",
        )
        db.add(salary)
        created_salaries.append(salary)
        existing_salary_keys.add(salary_key)
        touched_months.add(salary_row.month)
        result.salaries_created += 1
        result.add_imported_row(ws.title, "salaries")

    try:
        await db.flush()
        for contact in created_contacts:
            result.record_created_object(
                sheet_name=ws.title,
                kind="salaries",
                object_type="contact",
                object_id=contact.id,
                reference=contact.nom,
                details={"created_from": "salary_sheet"},
            )
        for salary in created_salaries:
            result.record_created_object(
                sheet_name=ws.title,
                kind="salaries",
                object_type="salary",
                object_id=salary.id,
                reference=salary.month,
                details={"employee_id": salary.employee_id, "net_pay": str(salary.net_pay)},
            )

        existing_salary_entry_group_keys = {
            group_key
            for (group_key,) in (
                await db.execute(
                    select(AccountingEntry.group_key).where(
                        AccountingEntry.source_type == EntrySourceType.SALARY,
                        AccountingEntry.group_key.is_not(None),
                    )
                )
            ).all()
            if group_key
        }

        for month in sorted(touched_months):
            month_group_prefix = f"salary-import:{month}:"
            if any(
                group_key.startswith(month_group_prefix)
                for group_key in existing_salary_entry_group_keys
            ):
                continue

            month_result = await db.execute(
                select(
                    Salary.hours,
                    Salary.gross,
                    Salary.employee_charges,
                    Salary.employer_charges,
                    Salary.tax,
                    Salary.net_pay,
                    Contact.nom,
                    Contact.prenom,
                )
                .join(Contact, Contact.id == Salary.employee_id)
                .where(Salary.month == month)
                .order_by(Contact.nom, Contact.prenom, Salary.employee_id)
            )
            month_salary_rows = month_result.all()
            if not month_salary_rows:
                continue

            month_rows = [
                NormalizedSalaryRow(
                    source_row_number=0,
                    month=month,
                    employee_name=_format_contact_display_name(nom, prenom) or nom,
                    hours=Decimal(str(hours)),
                    gross=Decimal(str(gross)),
                    employee_charges=Decimal(str(employee_charges)),
                    employer_charges=Decimal(str(employer_charges)),
                    tax=Decimal(str(tax)),
                    net_pay=Decimal(str(net_pay)),
                )
                for (
                    hours,
                    gross,
                    employee_charges,
                    employer_charges,
                    tax,
                    net_pay,
                    nom,
                    prenom,
                ) in month_salary_rows
            ]
            entry_date = _salary_entry_date(month)
            fiscal_year_id = await find_fiscal_year_id_for_date(db, entry_date)

            for group_kind, group_lines in zip(
                ("accrual", "payment"),
                _salary_month_group_lines(month, month_rows),
                strict=True,
            ):
                group_key = f"salary-import:{month}:{group_kind}"
                for account_number, label, debit, credit in group_lines:
                    if debit <= 0 and credit <= 0:
                        continue
                    entry = AccountingEntry(
                        entry_number=await _next_entry_number(db),
                        date=entry_date,
                        account_number=account_number,
                        label=label,
                        debit=debit,
                        credit=credit,
                        fiscal_year_id=fiscal_year_id,
                        source_type=EntrySourceType.SALARY,
                        source_id=None,
                        group_key=group_key,
                    )
                    db.add(entry)
                    result.entries_created += 1

            existing_salary_entry_group_keys.add(f"salary-import:{month}:accrual")
            existing_salary_entry_group_keys.add(f"salary-import:{month}:payment")

        await db.flush()
        logger.debug("Salaries import done — created=%d", result.salaries_created)
    except Exception as exc:
        logger.error("Salaries flush failed: %s", exc, exc_info=True)
        result.add_import_error("salaires", exc)
        await db.rollback()
        raise _ImportSheetFailure from exc


async def _import_cash_sheet(db: AsyncSession, ws: Any, result: ImportResult) -> None:
    """Import cash register movements from a 'Caisse' sheet.

    Expected columns (flexible): date | libellé/description | entrée | sortie
    or: date | description | montant | type (E/S or in/out)
    """
    from backend.models.cash import CashEntrySource, CashMovementType, CashRegister  # noqa: PLC0415
    from backend.models.invoice import Invoice, InvoiceType  # noqa: PLC0415
    from backend.models.payment import Payment  # noqa: PLC0415
    from backend.services.accounting_engine import (  # noqa: PLC0415
        generate_entries_for_invoice,
        generate_entries_for_payment,
    )
    from backend.services.cash_service import recompute_cash_balances  # noqa: PLC0415
    from backend.services.payment import _refresh_invoice_status  # noqa: PLC0415

    parsed_sheet, normalized_rows, _, ignored_issues = _parse_cash_sheet(ws)
    if parsed_sheet is None or parsed_sheet.missing_columns:
        logger.debug(
            "Cash sheet skipped — header not detected or missing columns=%s",
            parsed_sheet.missing_columns if parsed_sheet else None,
        )
        return

    for ignored_issue in ignored_issues:
        result.add_ignored_row(
            ws.title,
            "cash",
            format_row_issue(ignored_issue),
        )

    created_cash_entries: list[CashRegister] = []
    created_supplier_invoice_ids: set[int] = set()
    created_supplier_payment_ids: list[int] = []
    affected_invoice_ids: set[int] = set()
    existing_contacts_by_preview_key = await _load_existing_contacts_by_preview_key(db)
    default_invoice_due_days = await settings_service.get_default_invoice_due_days(db)
    for cash_row in normalized_rows:
        contact_id: int | None = None
        payment_id: int | None = None
        supplier_candidate = _supplier_invoice_candidate_from_cash_row(cash_row)
        if supplier_candidate is not None:
            (
                invoice_id,
                payment_id,
                contact_id,
                created_invoice,
            ) = await _ensure_supplier_invoice_payment(
                db,
                candidate=supplier_candidate,
                existing_contacts_by_preview_key=existing_contacts_by_preview_key,
                result=result,
                sheet_name=ws.title,
                default_invoice_due_days=default_invoice_due_days,
            )
            affected_invoice_ids.add(invoice_id)
            created_supplier_payment_ids.append(payment_id)
            if created_invoice:
                created_supplier_invoice_ids.add(invoice_id)
        entry = CashRegister(
            date=cash_row.entry_date,
            amount=cash_row.amount,
            type=CashMovementType(cash_row.movement_type),
            contact_id=contact_id,
            payment_id=payment_id,
            reference=cash_row.reference,
            description=cash_row.description,
            source=CashEntrySource.MANUAL,
            balance_after=Decimal("0"),
        )
        db.add(entry)
        created_cash_entries.append(entry)
        result.cash_created += 1
        result.add_imported_row(ws.title, "cash")

    try:
        await db.flush()
        if created_cash_entries:
            await recompute_cash_balances(db)
        for entry in created_cash_entries:
            result.record_created_object(
                sheet_name=ws.title,
                kind="cash",
                object_type="cash_register",
                object_id=entry.id,
                reference=entry.date.isoformat(),
                details={
                    "amount": str(entry.amount),
                    "type": str(entry.type),
                    "description": entry.description,
                },
            )
        for invoice_id in created_supplier_invoice_ids:
            invoice = await db.get(Invoice, invoice_id)
            if invoice is None:
                continue
            result.entries_created += len(await generate_entries_for_invoice(db, invoice))
        for invoice_id in affected_invoice_ids:
            await _refresh_invoice_status(db, invoice_id)
        for payment_id in created_supplier_payment_ids:
            supplier_payment = await db.get(Payment, payment_id)
            if supplier_payment is None:
                continue
            result.entries_created += len(
                await generate_entries_for_payment(db, supplier_payment, InvoiceType.FOURNISSEUR)
            )
        await db.flush()
        logger.debug("Cash import done — created=%d", result.cash_created)
    except Exception as exc:
        logger.error("Cash flush failed: %s", exc, exc_info=True)
        result.add_import_error("caisse", exc)
        await db.rollback()
        raise _ImportSheetFailure from exc


async def _import_bank_sheet(db: AsyncSession, ws: Any, result: ImportResult) -> None:
    """Import bank transactions from a 'Banque' or 'Relevé' sheet.

    Expected columns (flexible): date | libellé | débit | crédit | solde
    or: date | description | montant (positive=credit, negative=debit)
    """
    from backend.models.bank import (
        BankTransaction,
        BankTransactionSource,
    )  # noqa: PLC0415
    from backend.models.invoice import Invoice, InvoiceType  # noqa: PLC0415
    from backend.models.payment import Payment  # noqa: PLC0415
    from backend.services.accounting_engine import (  # noqa: PLC0415
        generate_entries_for_invoice,
        generate_entries_for_payment,
    )
    from backend.services.payment import _refresh_invoice_status  # noqa: PLC0415

    parsed_sheet, normalized_rows, _, ignored_issues = _parse_bank_sheet(ws)
    if parsed_sheet is None or parsed_sheet.missing_columns:
        logger.debug(
            "Bank sheet skipped — header not detected or missing columns=%s",
            parsed_sheet.missing_columns if parsed_sheet else None,
        )
        return

    for ignored_issue in ignored_issues:
        result.add_ignored_row(
            ws.title,
            "bank",
            format_row_issue(ignored_issue),
        )

    created_bank_entries: list[tuple[BankTransaction, NormalizedBankRow]] = []
    created_payments: list[tuple[Payment, InvoiceType]] = []
    created_supplier_invoice_ids: set[int] = set()
    created_supplier_payment_ids: list[int] = []
    affected_invoice_ids: set[int] = set()
    existing_contacts_by_preview_key = await _load_existing_contacts_by_preview_key(db)
    existing_payment_signatures = await _load_existing_payment_signatures(db)
    default_invoice_due_days = await settings_service.get_default_invoice_due_days(db)
    for bank_row in normalized_rows:
        supplier_candidate = _supplier_invoice_candidate_from_bank_row(bank_row)
        if supplier_candidate is not None:
            invoice_id, payment_id, _, created_invoice = await _ensure_supplier_invoice_payment(
                db,
                candidate=supplier_candidate,
                existing_contacts_by_preview_key=existing_contacts_by_preview_key,
                result=result,
                sheet_name=ws.title,
                default_invoice_due_days=default_invoice_due_days,
            )
            affected_invoice_ids.add(invoice_id)
            created_supplier_payment_ids.append(payment_id)
            if created_invoice:
                created_supplier_invoice_ids.add(invoice_id)
        else:
            await _queue_client_payment_from_bank_row(
                db,
                bank_row=bank_row,
                existing_payment_signatures=existing_payment_signatures,
                created_payments=created_payments,
                affected_invoice_ids=affected_invoice_ids,
                result=result,
                sheet_name=ws.title,
            )
        entry = BankTransaction(
            date=bank_row.entry_date,
            amount=bank_row.amount,
            reference=bank_row.reference,
            description=bank_row.description,
            balance_after=bank_row.balance_after,
            source=BankTransactionSource.IMPORT,
        )
        db.add(entry)
        created_bank_entries.append((entry, bank_row))
        result.bank_created += 1
        result.add_imported_row(ws.title, "bank")

    try:
        await db.flush()
        for payment, _ in created_payments:
            result.record_created_object(
                sheet_name=ws.title,
                kind="payments",
                object_type="payment",
                object_id=payment.id,
                reference=payment.reference,
                details={
                    "invoice_id": payment.invoice_id,
                    "contact_id": payment.contact_id,
                    "amount": str(payment.amount),
                    "date": payment.date.isoformat(),
                },
            )
        for entry, _ in created_bank_entries:
            result.record_created_object(
                sheet_name=ws.title,
                kind="bank",
                object_type="bank_transaction",
                object_id=entry.id,
                reference=entry.date.isoformat(),
                details={
                    "amount": str(entry.amount),
                    "description": entry.description,
                    "balance_after": str(entry.balance_after),
                },
            )
        for entry, bank_row in created_bank_entries:
            result.entries_created += len(
                await _generate_direct_bank_entries(
                    db,
                    bank_row=bank_row,
                    bank_entry=entry,
                )
            )
        for invoice_id in created_supplier_invoice_ids:
            invoice = await db.get(Invoice, invoice_id)
            if invoice is None:
                continue
            result.entries_created += len(await generate_entries_for_invoice(db, invoice))
        for invoice_id in affected_invoice_ids:
            await _refresh_invoice_status(db, invoice_id)
        for payment_id in created_supplier_payment_ids:
            supplier_payment = await db.get(Payment, payment_id)
            if supplier_payment is None:
                continue
            result.entries_created += len(
                await generate_entries_for_payment(db, supplier_payment, InvoiceType.FOURNISSEUR)
            )
        for payment, invoice_type in created_payments:
            result.entries_created += len(
                await generate_entries_for_payment(db, payment, invoice_type)
            )
        await db.flush()
        logger.debug("Bank import done — created=%d", result.bank_created)
    except Exception as exc:
        logger.error("Bank flush failed: %s", exc, exc_info=True)
        result.add_import_error("banque", exc)
        await db.rollback()
        raise _ImportSheetFailure from exc


async def _import_entries_sheet(db: AsyncSession, ws: Any, result: ImportResult) -> None:
    """Import accounting entries from a comptabilité sheet."""
    from sqlalchemy import func, select  # noqa: PLC0415

    from backend.models.accounting_entry import (
        AccountingEntry,
        EntrySourceType,
    )  # noqa: PLC0415
    from backend.services.fiscal_year_service import find_fiscal_year_id_for_date  # noqa: PLC0415

    parsed_sheet, normalized_rows, _, ignored_issues = _parse_entries_sheet(ws)
    if parsed_sheet is None or parsed_sheet.missing_columns:
        result.skipped += 1
        return

    existing_entry_signatures = await _load_existing_accounting_entry_signatures(db)
    existing_invoice_numbers = await _load_existing_invoice_numbers(db)
    existing_client_invoices_by_number = await _load_existing_client_invoices_by_number(db)
    existing_client_payment_signatures = await _load_existing_client_payment_reference_signatures(
        db
    )
    existing_supplier_payment_signatures = (
        await _load_existing_supplier_payment_reference_signatures(db)
    )
    existing_salary_group_signatures = await _load_existing_salary_group_signatures(db)
    existing_manual_line_signatures = await load_existing_manual_accounting_line_signatures(db)
    existing_generated_group_signatures = (
        await _load_existing_generated_accounting_group_signatures(db)
    )
    for ignored_issue in ignored_issues:
        result.add_ignored_row(
            ws.title,
            "entries",
            format_row_issue(ignored_issue),
        )

    # Pre-compute next entry number offset to avoid per-row DB queries
    count_result = await db.execute(select(func.count(AccountingEntry.id)))
    base_count = count_result.scalar_one_or_none() or 0
    entries_to_add: list[AccountingEntry] = []
    next_offset = 0
    entry_groups = _build_entry_row_groups(normalized_rows)
    index = 0
    while index < len(entry_groups):
        entry_group, next_index = _merge_existing_client_invoice_entry_groups(
            entry_groups,
            index,
            existing_invoice_numbers,
        )
        existing_client_invoice_reference = _matching_existing_client_invoice_reference(
            entry_group,
            existing_invoice_numbers,
        )
        if existing_client_invoice_reference is not None:
            clarified_invoice = await _clarify_existing_client_invoice_from_entries(
                db,
                entry_group,
                existing_client_invoices_by_number,
            )
            message = (
                _CLIENT_INVOICE_CLARIFIED_MESSAGE
                if clarified_invoice is not None
                else ENTRY_COVERED_BY_SOLDE_MESSAGE
            )
            for entry_row in entry_group:
                result.add_ignored_row(
                    ws.title,
                    "entries",
                    format_row_issue(
                        RowIgnoredIssue(
                            source_row_number=entry_row.source_row_number,
                            message=message,
                        )
                    ),
                )
            if clarified_invoice is not None:
                result.record_created_object(
                    sheet_name=ws.title,
                    kind="entries",
                    object_type="invoice",
                    object_id=clarified_invoice.id,
                    reference=clarified_invoice.number,
                    details={"action": "clarified_from_accounting_entries"},
                )
            index = next_index
            continue

        if _matching_existing_client_payment_reference(
            entry_group,
            existing_client_payment_signatures,
        ):
            for entry_row in entry_group:
                result.add_ignored_row(
                    ws.title,
                    "entries",
                    format_row_issue(
                        RowIgnoredIssue(
                            source_row_number=entry_row.source_row_number,
                            message=ENTRY_COVERED_BY_SOLDE_MESSAGE,
                        )
                    ),
                )
            index = next_index
            continue

        if _matching_existing_supplier_invoice_payment_reference(
            entry_group,
            existing_supplier_payment_signatures,
        ):
            for entry_row in entry_group:
                result.add_ignored_row(
                    ws.title,
                    "entries",
                    format_row_issue(
                        RowIgnoredIssue(
                            source_row_number=entry_row.source_row_number,
                            message=ENTRY_COVERED_BY_SOLDE_MESSAGE,
                        )
                    ),
                )
            index = next_index
            continue

        if _matching_existing_salary_entry_group(
            entry_group,
            existing_salary_group_signatures,
        ):
            for entry_row in entry_group:
                result.add_ignored_row(
                    ws.title,
                    "entries",
                    format_row_issue(
                        RowIgnoredIssue(
                            source_row_number=entry_row.source_row_number,
                            message=ENTRY_COVERED_BY_SOLDE_MESSAGE,
                        )
                    ),
                )
            index = next_index
            continue

        if _normalized_entry_group_signature(entry_group) in existing_generated_group_signatures:
            for entry_row in entry_group:
                result.add_ignored_row(
                    ws.title,
                    "entries",
                    format_row_issue(
                        RowIgnoredIssue(
                            source_row_number=entry_row.source_row_number,
                            message=ENTRY_COVERED_BY_SOLDE_MESSAGE,
                        )
                    ),
                )
            index = next_index
            continue

        group_key = f"import:{uuid4().hex}"
        for entry_row in entry_group:
            signature = _accounting_entry_signature(
                entry_date=entry_row.entry_date,
                account_number=entry_row.account_number,
                label=entry_row.label,
                debit=entry_row.debit,
                credit=entry_row.credit,
            )
            if signature in existing_entry_signatures:
                result.add_ignored_row(
                    ws.title,
                    "entries",
                    format_row_issue(
                        RowIgnoredIssue(
                            source_row_number=entry_row.source_row_number,
                            message=ENTRY_EXISTING_MESSAGE,
                        )
                    ),
                )
                continue

            line_signature = _accounting_entry_line_signature(
                entry_date=entry_row.entry_date,
                account_number=entry_row.account_number,
                debit=entry_row.debit,
                credit=entry_row.credit,
            )
            if line_signature in existing_manual_line_signatures:
                result.add_warning(
                    ws.title,
                    "entries",
                    format_row_issue(
                        RowWarningIssue(
                            source_row_number=entry_row.source_row_number,
                            message=ENTRY_NEAR_EXISTING_MANUAL_MESSAGE,
                        )
                    ),
                )

            next_offset += 1
            entry = AccountingEntry(
                entry_number=f"{base_count + next_offset:06d}",
                date=entry_row.entry_date,
                account_number=entry_row.account_number,
                label=entry_row.label,
                debit=entry_row.debit,
                credit=entry_row.credit,
                fiscal_year_id=await find_fiscal_year_id_for_date(db, entry_row.entry_date),
                source_type=EntrySourceType.MANUAL,
                group_key=group_key,
            )
            entries_to_add.append(entry)
            existing_entry_signatures.add(signature)
            result.add_imported_row(ws.title, "entries")

        index = next_index

    try:
        for entry in entries_to_add:
            db.add(entry)
        await db.flush()
        for entry in entries_to_add:
            result.record_created_object(
                sheet_name=ws.title,
                kind="entries",
                object_type="accounting_entry",
                object_id=entry.id,
                reference=entry.entry_number,
                details={
                    "account_number": entry.account_number,
                    "date": entry.date.isoformat(),
                },
            )
        result.entries_created += len(entries_to_add)
        logger.debug("Entries import done — created=%d", result.entries_created)
    except Exception as exc:
        logger.error("Entries flush failed: %s", exc, exc_info=True)
        result.add_import_error("écritures", exc)
        await db.rollback()
        raise _ImportSheetFailure from exc


async def _add_gestion_existing_rows_preview(
    db: AsyncSession, file_bytes: bytes, preview: PreviewResult
) -> None:
    from io import BytesIO  # noqa: PLC0415

    import openpyxl  # noqa: PLC0415

    existing_contacts_by_preview_key = await _load_existing_contacts_by_preview_key(db)
    existing_contact_keys = set(existing_contacts_by_preview_key)
    existing_invoice_numbers = await _load_existing_invoice_numbers(db)
    existing_salary_keys = await _load_existing_salary_keys(db)

    wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    workbook_candidates: list[PaymentMatchCandidate] = []
    payment_years: set[int] = set()
    bank_years: set[int] = set()
    cash_years: set[int] = set()
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        kind, _ = _classify_gestion_sheet(sheet_name)
        if kind == "invoices":
            parsed_sheet, invoice_rows, _, _ = _parse_invoice_sheet(ws)
            if parsed_sheet is None or parsed_sheet.missing_columns:
                continue
            invoice_rows, _ = filter_duplicate_invoice_rows(
                invoice_rows,
                normalize_text=_normalize_text,
            )
            workbook_candidates.extend(
                _make_workbook_invoice_candidate(invoice_row) for invoice_row in invoice_rows
            )
        elif kind == "payments":
            parsed_sheet, payment_rows, _ = _parse_payment_sheet(ws)
            if parsed_sheet is None or parsed_sheet.missing_columns:
                continue
            payment_years.update(payment_row.payment_date.year for payment_row in payment_rows)
        elif kind == "bank":
            parsed_sheet, bank_rows, _, _ = _parse_bank_sheet(ws)
            if parsed_sheet is None or parsed_sheet.missing_columns:
                continue
            bank_years.update(bank_row.entry_date.year for bank_row in bank_rows)
        elif kind == "cash":
            parsed_sheet, cash_rows, _, _ = _parse_cash_sheet(ws)
            if parsed_sheet is None or parsed_sheet.missing_columns:
                continue
            cash_years.update(cash_row.entry_date.year for cash_row in cash_rows)

    existing_payment_signatures = await _load_existing_client_payment_comparison_signatures(
        db,
        payment_years,
    )
    existing_bank_signatures = await _load_existing_bank_comparison_signatures(db, bank_years)
    existing_cash_signatures = await _load_existing_cash_comparison_signatures(db, cash_years)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        kind, _ = _classify_gestion_sheet(sheet_name)
        sheet_preview = _find_sheet_preview(preview, sheet_name)
        if kind is None or sheet_preview is None or sheet_preview["status"] != "recognized":
            continue

        if kind == "contacts":
            parsed_sheet, contact_rows, _ = _parse_contact_sheet(ws)
            if parsed_sheet is None or parsed_sheet.missing_columns:
                continue
            contact_rows, _ = filter_duplicate_contact_rows(
                contact_rows,
                normalize_text=_normalize_text,
            )
            ignored_issues = find_existing_contact_issues(
                contact_rows,
                existing_contact_keys,
                contact_preview_key=_contact_preview_key,
            )
            for ignored_issue in ignored_issues:
                _append_preview_ignored_issue(
                    preview,
                    sheet_preview,
                    ignored_issue,
                )
            if ignored_count := len(ignored_issues):
                sheet_preview["rows"] = max(0, sheet_preview["rows"] - ignored_count)

        elif kind == "invoices":
            if (invoice_parsed := _parse_invoice_sheet(ws))[0] is None or invoice_parsed[
                0
            ].missing_columns:
                continue
            _, invoice_rows, _, _ = invoice_parsed
            invoice_rows, _ = filter_duplicate_invoice_rows(
                invoice_rows,
                normalize_text=_normalize_text,
            )
            ignored_issues = find_existing_invoice_issues(
                invoice_rows,
                existing_invoice_numbers,
                normalize_text=_normalize_text,
            )
            blocked_issues = find_ambiguous_invoice_contact_issues(
                invoice_rows,
                existing_contacts_by_preview_key,
                normalize_text=_normalize_text,
            )
            for ignored_issue in ignored_issues:
                _append_preview_ignored_issue(
                    preview,
                    sheet_preview,
                    ignored_issue,
                )
            for blocked_issue in blocked_issues:
                _append_preview_blocked_issue(
                    preview,
                    sheet_preview,
                    blocked_issue,
                )
            ignored_count = len(ignored_issues)
            blocked_count = len(blocked_issues)
            if ignored_count:
                sheet_preview["rows"] = max(0, sheet_preview["rows"] - ignored_count)
                preview.estimated_invoices = max(0, preview.estimated_invoices - ignored_count)
            if blocked_count:
                sheet_preview["rows"] = max(0, sheet_preview["rows"] - blocked_count)
                preview.estimated_invoices = max(0, preview.estimated_invoices - blocked_count)

        elif kind == "payments":
            parsed_sheet, payment_rows, _ = _parse_payment_sheet(ws)
            if parsed_sheet is None or parsed_sheet.missing_columns:
                continue
            for payment_row in payment_rows:
                resolution = await _resolve_payment_match(
                    db,
                    payment_row,
                    workbook_candidates,
                )
                if resolution.status != "matched" or resolution.candidate is None:
                    continue
                if not resolution.candidate.invoice_number:
                    continue
                payment_signature = _gestion_payment_comparison_signature(
                    reference=resolution.candidate.invoice_number,
                    payment_date=payment_row.payment_date,
                    amount=payment_row.amount,
                    settlement_account=_client_settlement_account_from_method(payment_row.method),
                    cheque_number=payment_row.cheque_number,
                )
                if payment_signature not in existing_payment_signatures:
                    continue
                _append_preview_ignored_issue(
                    preview,
                    sheet_preview,
                    RowIgnoredIssue(
                        source_row_number=payment_row.source_row_number,
                        message=EXISTING_GESTION_ROW_MESSAGE,
                    ),
                )
                sheet_preview["rows"] = max(0, sheet_preview["rows"] - 1)
                preview.estimated_payments = max(0, preview.estimated_payments - 1)

        elif kind == "bank":
            parsed_sheet, bank_rows, _, _ = _parse_bank_sheet(ws)
            if parsed_sheet is None or parsed_sheet.missing_columns:
                continue
            for bank_row in bank_rows:
                bank_signature = _gestion_bank_comparison_signature(
                    entry_date=bank_row.entry_date,
                    amount=bank_row.amount,
                    description=bank_row.description,
                    reference=bank_row.reference,
                )
                candidate = _supplier_invoice_candidate_from_bank_row(bank_row)
                if bank_signature in existing_bank_signatures:
                    _append_preview_ignored_issue(
                        preview,
                        sheet_preview,
                        RowIgnoredIssue(
                            source_row_number=bank_row.source_row_number,
                            message=EXISTING_GESTION_ROW_MESSAGE,
                        ),
                    )
                    sheet_preview["rows"] = max(0, sheet_preview["rows"] - 1)
                    if candidate is not None:
                        preview.estimated_invoices = max(0, preview.estimated_invoices - 1)
                        preview.estimated_payments = max(0, preview.estimated_payments - 1)
                    elif _single_client_invoice_reference(
                        bank_row.reference,
                        bank_row.description,
                    ):
                        preview.estimated_payments = max(0, preview.estimated_payments - 1)
                    continue
                if candidate is None:
                    continue
                if _normalize_text(candidate.invoice_number) in existing_invoice_numbers:
                    _append_preview_ignored_issue(
                        preview,
                        sheet_preview,
                        make_existing_invoice_issue(candidate.source_row_number),
                    )
                    sheet_preview["rows"] = max(0, sheet_preview["rows"] - 1)
                    preview.estimated_invoices = max(0, preview.estimated_invoices - 1)
                    preview.estimated_payments = max(0, preview.estimated_payments - 1)
                    continue
                _, supplier_blocked_issue = resolve_invoice_contact_match(
                    _invoice_row_from_supplier_candidate(candidate),
                    existing_contacts_by_preview_key,
                    normalize_text=_normalize_text,
                )
                if supplier_blocked_issue is not None:
                    _append_preview_blocked_issue(
                        preview,
                        sheet_preview,
                        supplier_blocked_issue,
                    )
                    preview.estimated_invoices = max(0, preview.estimated_invoices - 1)
                    preview.estimated_payments = max(0, preview.estimated_payments - 1)

        elif kind == "cash":
            parsed_sheet, cash_rows, _, _ = _parse_cash_sheet(ws)
            if parsed_sheet is None or parsed_sheet.missing_columns:
                continue
            for cash_row in cash_rows:
                cash_signature = _gestion_cash_comparison_signature(
                    entry_date=cash_row.entry_date,
                    movement_type=cash_row.movement_type,
                    amount=cash_row.amount,
                    description=cash_row.description,
                    reference=cash_row.reference,
                )
                candidate = _supplier_invoice_candidate_from_cash_row(cash_row)
                if cash_signature in existing_cash_signatures:
                    _append_preview_ignored_issue(
                        preview,
                        sheet_preview,
                        RowIgnoredIssue(
                            source_row_number=cash_row.source_row_number,
                            message=EXISTING_GESTION_ROW_MESSAGE,
                        ),
                    )
                    sheet_preview["rows"] = max(0, sheet_preview["rows"] - 1)
                    if candidate is not None:
                        preview.estimated_invoices = max(0, preview.estimated_invoices - 1)
                        preview.estimated_payments = max(0, preview.estimated_payments - 1)
                    continue
                if candidate is None:
                    continue
                if _normalize_text(candidate.invoice_number) in existing_invoice_numbers:
                    _append_preview_ignored_issue(
                        preview,
                        sheet_preview,
                        make_existing_invoice_issue(candidate.source_row_number),
                    )
                    sheet_preview["rows"] = max(0, sheet_preview["rows"] - 1)
                    preview.estimated_invoices = max(0, preview.estimated_invoices - 1)
                    preview.estimated_payments = max(0, preview.estimated_payments - 1)
                    continue
                _, supplier_blocked_issue = resolve_invoice_contact_match(
                    _invoice_row_from_supplier_candidate(candidate),
                    existing_contacts_by_preview_key,
                    normalize_text=_normalize_text,
                )
                if supplier_blocked_issue is not None:
                    _append_preview_blocked_issue(
                        preview,
                        sheet_preview,
                        supplier_blocked_issue,
                    )
                    preview.estimated_invoices = max(0, preview.estimated_invoices - 1)
                    preview.estimated_payments = max(0, preview.estimated_payments - 1)

        elif kind == "salaries":
            parsed_sheet, salary_rows, _ = _parse_salary_sheet(ws)
            if parsed_sheet is None or parsed_sheet.missing_columns:
                continue
            salary_ignored_issues: list[RowIgnoredIssue] = []
            seen_salary_keys = set(existing_salary_keys)
            for salary_row in salary_rows:
                salary_key = (salary_row.month, _salary_employee_key(salary_row.employee_name))
                if salary_key in seen_salary_keys:
                    salary_ignored_issues.append(
                        RowIgnoredIssue(
                            source_row_number=salary_row.source_row_number,
                            message=EXISTING_SALARY_MESSAGE,
                        )
                    )
                    continue
                seen_salary_keys.add(salary_key)
            for ignored_issue in salary_ignored_issues:
                _append_preview_ignored_issue(
                    preview,
                    sheet_preview,
                    ignored_issue,
                )
            if ignored_count := len(salary_ignored_issues):
                sheet_preview["rows"] = max(0, sheet_preview["rows"] - ignored_count)
                preview.estimated_salaries = max(0, preview.estimated_salaries - ignored_count)

    preview._candidate_contacts.difference_update(existing_contact_keys)
    preview.estimated_contacts = len(preview._candidate_contacts)
    _recompute_preview_can_import(preview)


async def _add_comptabilite_existing_rows_preview(
    db: AsyncSession, file_bytes: bytes, preview: PreviewResult
) -> None:
    from io import BytesIO  # noqa: PLC0415

    import openpyxl  # noqa: PLC0415

    existing_entry_signatures = await _load_existing_accounting_entry_signatures(db)
    existing_invoice_numbers = await _load_existing_invoice_numbers(db)
    existing_client_payment_signatures = await _load_existing_client_payment_reference_signatures(
        db
    )
    existing_supplier_payment_signatures = (
        await _load_existing_supplier_payment_reference_signatures(db)
    )
    existing_client_invoices_by_number = await _load_existing_client_invoices_by_number(db)
    existing_salary_group_signatures = await _load_existing_salary_group_signatures(db)
    existing_manual_line_signatures = await load_existing_manual_accounting_line_signatures(db)
    existing_generated_group_signatures = (
        await _load_existing_generated_accounting_group_signatures(db)
    )
    if (
        not existing_entry_signatures
        and not existing_generated_group_signatures
        and not existing_invoice_numbers
    ):
        return

    wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        kind, _ = _classify_comptabilite_sheet(sheet_name)
        sheet_preview = _find_sheet_preview(preview, sheet_name)
        if kind != "entries" or sheet_preview is None or sheet_preview["status"] != "recognized":
            continue

        parsed_sheet, normalized_rows, _, _ = _parse_entries_sheet(ws)
        if parsed_sheet is None or parsed_sheet.missing_columns:
            continue

        entry_groups = _build_entry_row_groups(normalized_rows)
        index = 0
        while index < len(entry_groups):
            entry_group, next_index = _merge_existing_client_invoice_entry_groups(
                entry_groups,
                index,
                existing_invoice_numbers,
            )
            existing_client_invoice_reference = _matching_existing_client_invoice_reference(
                entry_group,
                existing_invoice_numbers,
            )
            if existing_client_invoice_reference is not None:
                clarified_invoice = await _find_clarifiable_existing_client_invoice(
                    db,
                    entry_group,
                    existing_client_invoices_by_number,
                )
                for entry_row in entry_group:
                    _append_preview_ignored_issue(
                        preview,
                        sheet_preview,
                        RowIgnoredIssue(
                            source_row_number=entry_row.source_row_number,
                            message=(
                                _CLIENT_INVOICE_CLARIFIED_MESSAGE
                                if clarified_invoice is not None
                                else ENTRY_COVERED_BY_SOLDE_MESSAGE
                            ),
                        ),
                    )
                    sheet_preview["rows"] = max(0, sheet_preview["rows"] - 1)
                    preview.estimated_entries = max(0, preview.estimated_entries - 1)
                index = next_index
                continue

            if _matching_existing_client_payment_reference(
                entry_group,
                existing_client_payment_signatures,
            ):
                for entry_row in entry_group:
                    _append_preview_ignored_issue(
                        preview,
                        sheet_preview,
                        RowIgnoredIssue(
                            source_row_number=entry_row.source_row_number,
                            message=ENTRY_COVERED_BY_SOLDE_MESSAGE,
                        ),
                    )
                    sheet_preview["rows"] = max(0, sheet_preview["rows"] - 1)
                    preview.estimated_entries = max(0, preview.estimated_entries - 1)
                index = next_index
                continue

            if _matching_existing_supplier_invoice_payment_reference(
                entry_group,
                existing_supplier_payment_signatures,
            ):
                for entry_row in entry_group:
                    _append_preview_ignored_issue(
                        preview,
                        sheet_preview,
                        RowIgnoredIssue(
                            source_row_number=entry_row.source_row_number,
                            message=ENTRY_COVERED_BY_SOLDE_MESSAGE,
                        ),
                    )
                    sheet_preview["rows"] = max(0, sheet_preview["rows"] - 1)
                    preview.estimated_entries = max(0, preview.estimated_entries - 1)
                index = next_index
                continue

            if _matching_existing_salary_entry_group(
                entry_group,
                existing_salary_group_signatures,
            ):
                for entry_row in entry_group:
                    _append_preview_ignored_issue(
                        preview,
                        sheet_preview,
                        RowIgnoredIssue(
                            source_row_number=entry_row.source_row_number,
                            message=ENTRY_COVERED_BY_SOLDE_MESSAGE,
                        ),
                    )
                    sheet_preview["rows"] = max(0, sheet_preview["rows"] - 1)
                    preview.estimated_entries = max(0, preview.estimated_entries - 1)
                index = next_index
                continue

            if (
                _normalized_entry_group_signature(entry_group)
                in existing_generated_group_signatures
            ):
                for entry_row in entry_group:
                    _append_preview_ignored_issue(
                        preview,
                        sheet_preview,
                        RowIgnoredIssue(
                            source_row_number=entry_row.source_row_number,
                            message=ENTRY_COVERED_BY_SOLDE_MESSAGE,
                        ),
                    )
                    sheet_preview["rows"] = max(0, sheet_preview["rows"] - 1)
                    preview.estimated_entries = max(0, preview.estimated_entries - 1)
                index = next_index
                continue

            for entry_row in entry_group:
                signature = _accounting_entry_signature(
                    entry_date=entry_row.entry_date,
                    account_number=entry_row.account_number,
                    label=entry_row.label,
                    debit=entry_row.debit,
                    credit=entry_row.credit,
                )
                if signature not in existing_entry_signatures:
                    line_signature = _accounting_entry_line_signature(
                        entry_date=entry_row.entry_date,
                        account_number=entry_row.account_number,
                        debit=entry_row.debit,
                        credit=entry_row.credit,
                    )
                    if line_signature in existing_manual_line_signatures:
                        _append_preview_warning_issue(
                            preview,
                            sheet_preview,
                            RowWarningIssue(
                                source_row_number=entry_row.source_row_number,
                                message=ENTRY_NEAR_EXISTING_MANUAL_MESSAGE,
                            ),
                        )
                    continue
                _append_preview_ignored_issue(
                    preview,
                    sheet_preview,
                    RowIgnoredIssue(
                        source_row_number=entry_row.source_row_number,
                        message=ENTRY_EXISTING_MESSAGE,
                    ),
                )
                sheet_preview["rows"] = max(0, sheet_preview["rows"] - 1)
                preview.estimated_entries = max(0, preview.estimated_entries - 1)
            index = next_index

    _recompute_preview_can_import(preview)


async def _add_gestion_payment_validation(
    db: AsyncSession, file_bytes: bytes, preview: PreviewResult
) -> None:
    from io import BytesIO  # noqa: PLC0415

    import openpyxl  # noqa: PLC0415

    workbook_candidates: list[PaymentMatchCandidate] = []
    payment_rows_by_sheet: dict[str, list[NormalizedPaymentRow]] = {}

    wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        kind, _ = _classify_gestion_sheet(sheet_name)
        if kind == "invoices":
            parsed_sheet, invoice_rows, _, _ = _parse_invoice_sheet(ws)
            if parsed_sheet is None or parsed_sheet.missing_columns:
                continue
            workbook_candidates.extend(
                _make_workbook_invoice_candidate(invoice_row) for invoice_row in invoice_rows
            )
        elif kind == "payments":
            parsed_sheet, payment_rows, _ = _parse_payment_sheet(ws)
            if parsed_sheet is None or parsed_sheet.missing_columns:
                continue
            payment_rows_by_sheet[sheet_name] = payment_rows

    for sheet_name, payment_rows in payment_rows_by_sheet.items():
        sheet_preview = _find_sheet_preview(preview, sheet_name)
        if sheet_preview is None:
            continue

        blocked_count = 0
        for payment_row in payment_rows:
            resolution = await _resolve_payment_match(
                db,
                payment_row,
                workbook_candidates,
            )
            blocking_issue = make_payment_resolution_issue(
                source_row_number=payment_row.source_row_number,
                status=resolution.status,
                candidate=resolution.candidate,
                message=resolution.message,
                require_persistable_candidate=False,
            )
            if blocking_issue is not None:
                _append_preview_blocked_issue(preview, sheet_preview, blocking_issue)
                blocked_count += 1

        if blocked_count:
            sheet_preview["rows"] = max(0, sheet_preview["rows"] - blocked_count)
            preview.estimated_payments = max(0, preview.estimated_payments - blocked_count)


def _collect_sample_rows(
    ws: Any, header_row: int, col_map: dict[str, int], *, limit: int = 3
) -> list[dict[str, str]]:
    samples: list[dict[str, str]] = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if all(cell is None for cell in row):
            continue
        samples.append(
            {key: _parse_str(row[idx] if idx < len(row) else None) for key, idx in col_map.items()}
        )
        if len(samples) >= limit:
            break
    return samples


def _preview_sheet_gestion(ws: Any, sheet_name: str, preview: PreviewResult) -> None:
    """Count parseable rows and report diagnostics for gestion files."""
    kind, reason = _classify_gestion_sheet(sheet_name)
    if kind is None:
        _append_reasoned_ignored_sheet_preview(
            preview,
            sheet_name=sheet_name,
            has_content=_sheet_has_content(ws),
            warning=preview_warning_for_gestion_reason(reason),
        )
        return

    header_info = detect_gestion_preview_header(
        ws,
        kind,
        detect_header_row=_detect_header_row,
    )

    if header_info is None:
        _append_unknown_structure_sheet_preview(
            preview,
            sheet_name=sheet_name,
            kind=kind,
            warning=GESTION_UNKNOWN_STRUCTURE_MESSAGE,
        )
        return

    header_row, col_map = header_info
    detected_columns = list(col_map.keys())
    count = 0
    missing_columns: list[str] = []
    ignored_rows = 0
    errors: list[str] = []
    warnings: list[str] = []

    if kind == "contacts":
        parsed_sheet, contact_rows, row_issues = _parse_contact_sheet(ws)
        if parsed_sheet is None:
            _append_unknown_structure_sheet_preview(
                preview,
                sheet_name=sheet_name,
                kind=kind,
                warning=GESTION_UNKNOWN_STRUCTURE_MESSAGE,
            )
            return
        header_row = parsed_sheet.header_row
        detected_columns = list(parsed_sheet.col_map.keys())
        missing_columns = parsed_sheet.missing_columns
        contact_rows, ignored_issues = filter_duplicate_contact_rows(
            contact_rows,
            normalize_text=_normalize_text,
        )
        ignored_rows = len(ignored_issues)
        count = len(contact_rows)
        _append_row_issues(errors, row_issues)
        _append_ignored_issues(warnings, ignored_issues)
        if not missing_columns:
            for contact_row in contact_rows:
                _register_preview_contact(
                    preview,
                    _format_contact_display_name(contact_row.nom, contact_row.prenom),
                )
    elif kind == "invoices":
        parsed_sheet, invoice_rows, row_issues, ignored_issues = _parse_invoice_sheet(ws)
        if parsed_sheet is None:
            _append_unknown_structure_sheet_preview(
                preview,
                sheet_name=sheet_name,
                kind=kind,
                warning=GESTION_UNKNOWN_STRUCTURE_MESSAGE,
            )
            return
        header_row = parsed_sheet.header_row
        detected_columns = list(parsed_sheet.col_map.keys())
        missing_columns = parsed_sheet.missing_columns
        invoice_rows, duplicate_ignored_issues = filter_duplicate_invoice_rows(
            invoice_rows,
            normalize_text=_normalize_text,
        )
        ignored_rows = len(ignored_issues) + len(duplicate_ignored_issues)
        count = len(invoice_rows)
        _append_row_issues(errors, row_issues)
        _append_ignored_issues(warnings, ignored_issues)
        _append_ignored_issues(warnings, duplicate_ignored_issues)
        if not missing_columns:
            for invoice_row in invoice_rows:
                _register_preview_contact(preview, invoice_row.contact_name)
            preview.estimated_invoices += count
    elif kind == "payments":
        parsed_sheet, payment_rows, row_issues = _parse_payment_sheet(ws)
        if parsed_sheet is None:
            _append_unknown_structure_sheet_preview(
                preview,
                sheet_name=sheet_name,
                kind=kind,
                warning=GESTION_UNKNOWN_STRUCTURE_MESSAGE,
            )
            return
        header_row = parsed_sheet.header_row
        detected_columns = list(parsed_sheet.col_map.keys())
        missing_columns = parsed_sheet.missing_columns
        count = len(payment_rows)
        _append_row_issues(errors, row_issues)
        if not missing_columns:
            for payment_row in payment_rows:
                if payment_row.contact_name:
                    _register_preview_contact(preview, payment_row.contact_name)
            preview.estimated_payments += count
    elif kind == "salaries":
        parsed_sheet, salary_rows, row_issues = _parse_salary_sheet(ws)
        if parsed_sheet is None:
            _append_unknown_structure_sheet_preview(
                preview,
                sheet_name=sheet_name,
                kind=kind,
                warning=GESTION_UNKNOWN_STRUCTURE_MESSAGE,
            )
            return
        header_row = parsed_sheet.header_row
        detected_columns = list(parsed_sheet.col_map.keys())
        missing_columns = parsed_sheet.missing_columns
        count = len(salary_rows)
        _append_row_issues(errors, row_issues)
        if not missing_columns:
            for salary_row in salary_rows:
                _register_preview_contact(preview, salary_row.employee_name)
            preview.estimated_salaries += count
    elif kind == "bank":
        parsed_sheet, bank_rows, row_issues, ignored_issues = _parse_bank_sheet(ws)
        if parsed_sheet is None:
            _append_unknown_structure_sheet_preview(
                preview,
                sheet_name=sheet_name,
                kind=kind,
                warning=GESTION_UNKNOWN_STRUCTURE_MESSAGE,
            )
            return
        header_row = parsed_sheet.header_row
        detected_columns = list(parsed_sheet.col_map.keys())
        missing_columns = parsed_sheet.missing_columns
        count = len(bank_rows)
        _append_row_issues(errors, row_issues)
        ignored_rows = len(ignored_issues)
        _append_ignored_issues(warnings, ignored_issues)
        if not missing_columns:
            for bank_row in bank_rows:
                candidate = _supplier_invoice_candidate_from_bank_row(bank_row)
                if candidate is not None:
                    preview.estimated_invoices += 1
                    preview.estimated_payments += 1
                    _register_preview_contact(preview, candidate.contact_name)
                    continue
                if _single_client_invoice_reference(bank_row.reference, bank_row.description):
                    preview.estimated_payments += 1
    elif kind == "cash":
        parsed_sheet, cash_rows, row_issues, ignored_issues = _parse_cash_sheet(ws)
        if parsed_sheet is None:
            _append_unknown_structure_sheet_preview(
                preview,
                sheet_name=sheet_name,
                kind=kind,
                warning=GESTION_UNKNOWN_STRUCTURE_MESSAGE,
            )
            return
        header_row = parsed_sheet.header_row
        detected_columns = list(parsed_sheet.col_map.keys())
        missing_columns = parsed_sheet.missing_columns
        count = len(cash_rows)
        _append_row_issues(errors, row_issues)
        ignored_rows = len(ignored_issues)
        _append_ignored_issues(warnings, ignored_issues)
        if not missing_columns:
            for cash_row in cash_rows:
                candidate = _supplier_invoice_candidate_from_cash_row(cash_row)
                if candidate is None:
                    continue
                preview.estimated_invoices += 1
                preview.estimated_payments += 1
                _register_preview_contact(preview, candidate.contact_name)

    assert parsed_sheet is not None
    col_map = parsed_sheet.col_map
    _append_finalized_sheet_preview(
        preview,
        sheet_name=sheet_name,
        kind=kind,
        header_row=header_row,
        rows=count,
        detected_columns=detected_columns,
        missing_columns=missing_columns,
        ignored_rows=ignored_rows,
        sample_rows=_collect_sample_rows(ws, header_row, col_map),
        warnings=warnings,
        errors=errors,
    )


def _preview_sheet_comptabilite(ws: Any, sheet_name: str, preview: PreviewResult) -> None:
    """Count parseable accounting entry rows and report diagnostics."""
    kind, reason = _classify_comptabilite_sheet(sheet_name)
    if kind is None:
        _append_reasoned_ignored_sheet_preview(
            preview,
            sheet_name=sheet_name,
            has_content=_sheet_has_content(ws),
            warning=preview_warning_for_comptabilite_reason(reason),
        )
        return

    parsed_sheet, normalized_rows, row_issues, ignored_issues = _parse_entries_sheet(ws)
    if parsed_sheet is None:
        _append_unknown_structure_sheet_preview(
            preview,
            sheet_name=sheet_name,
            kind=kind,
            warning=COMPTABILITE_UNKNOWN_STRUCTURE_MESSAGE,
        )
        return

    header_row = parsed_sheet.header_row
    detected_columns = list(parsed_sheet.col_map.keys())
    missing_columns = parsed_sheet.missing_columns
    count = len(normalized_rows)
    warnings: list[str] = []
    errors: list[str] = []
    _append_row_issues(errors, row_issues)
    _append_ignored_issues(warnings, ignored_issues)
    if not missing_columns:
        preview.estimated_entries += count

    _append_finalized_sheet_preview(
        preview,
        sheet_name=sheet_name,
        kind=kind,
        header_row=header_row,
        rows=count,
        detected_columns=detected_columns,
        missing_columns=missing_columns,
        ignored_rows=len(ignored_issues),
        sample_rows=_collect_sample_rows(ws, header_row, parsed_sheet.col_map),
        warnings=warnings,
        errors=errors,
    )


def _preview_gestion_file(file_bytes: bytes) -> PreviewResult:
    """Dry-run parse of a Gestion file using file-only validation."""
    from io import BytesIO  # noqa: PLC0415

    import openpyxl  # noqa: PLC0415

    preview = PreviewResult()
    try:
        wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        _append_preview_open_error(preview, exc)
        return preview

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        _preview_sheet_gestion(ws, sheet_name, preview)

    return preview


async def preview_gestion_file(
    db: AsyncSession,
    file_bytes: bytes,
    file_name: str | None = None,
    comparison_start_date: date | None = None,
    comparison_end_date: date | None = None,
) -> PreviewResult:
    """Dry-run parse of a Gestion file with shared business validation."""
    preview = _preview_gestion_file(file_bytes)
    preview.comparison_mode = "gestion-excel-to-solde"
    if preview.errors:
        preview.can_import = False
        return preview

    comparison_window = _resolve_comparison_window(
        file_name,
        comparison_start_date,
        comparison_end_date,
    )
    preview.comparison_context["domains"] = await _build_gestion_preview_comparison_domains(
        db,
        file_bytes,
        file_name,
        comparison_window,
    )

    await _add_gestion_existing_rows_preview(db, file_bytes, preview)

    import_log = await _find_successful_import_log(db, "gestion", _compute_file_hash(file_bytes))
    if import_log is not None:
        _mark_preview_as_already_imported(preview, "gestion", import_log)
        return preview

    await _add_gestion_payment_validation(db, file_bytes, preview)
    _finalize_preview_can_import(preview)
    return preview


async def preview_comptabilite_file(
    db: AsyncSession,
    file_bytes: bytes,
    file_name: str | None = None,
    comparison_start_date: date | None = None,
    comparison_end_date: date | None = None,
) -> PreviewResult:
    """Dry-run parse of a Comptabilité file — no DB writes."""
    from io import BytesIO  # noqa: PLC0415

    import openpyxl  # noqa: PLC0415

    preview = PreviewResult()
    try:
        wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        _append_preview_open_error(preview, exc)
        return preview

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        _preview_sheet_comptabilite(ws, sheet_name, preview)

    comparison_window = _resolve_comparison_window(
        file_name,
        comparison_start_date,
        comparison_end_date,
    )
    preview.comparison_context["domains"] = await _build_comptabilite_preview_comparison_domains(
        db,
        file_bytes,
        file_name,
        comparison_window,
    )

    await _add_comptabilite_coexistence_validation(db, preview)
    await _add_comptabilite_existing_rows_preview(db, file_bytes, preview)
    preview.comparison_mode = "global-convergence"
    preview.comparison_context[
        "extra_in_solde_by_kind"
    ] = await _collect_comptabilite_extra_in_solde_by_kind(db, file_bytes, preview)
    if preview.errors:
        preview.can_import = False
        return preview

    import_log = await _find_successful_import_log(
        db, "comptabilite", _compute_file_hash(file_bytes)
    )
    if import_log is not None:
        _mark_preview_as_already_imported(preview, "comptabilite", import_log)
        return preview

    _finalize_preview_can_import(preview)

    return preview
