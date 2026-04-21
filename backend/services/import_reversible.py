"""Reversible Excel import runs with prepared operations and generic effect replay."""

from __future__ import annotations

import json
import re
from dataclasses import asdict, dataclass
from datetime import date, datetime
from decimal import Decimal
from hashlib import sha256
from typing import Any, TypeVar, cast

import openpyxl
from sqlalchemy import desc, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.inspection import inspect as sa_inspect
from sqlalchemy.orm import selectinload
from sqlalchemy.sql.sqltypes import Boolean, Date, DateTime, Integer, Numeric

from backend.models.accounting_entry import AccountingEntry, EntrySourceType
from backend.models.accounting_rule import AccountingRule, TriggerType
from backend.models.bank import BankTransaction, BankTransactionSource
from backend.models.cash import CashEntrySource, CashMovementType, CashRegister
from backend.models.contact import Contact, ContactType
from backend.models.import_log import ImportLog
from backend.models.import_run import (
    ImportEffect,
    ImportEffectAction,
    ImportEffectStatus,
    ImportOperation,
    ImportOperationDecision,
    ImportOperationStatus,
    ImportRun,
    ImportRunStatus,
)
from backend.models.invoice import (
    Invoice,
    InvoiceLabel,
    InvoiceLine,
    InvoiceStatus,
    InvoiceType,
    derive_client_invoice_label,
)
from backend.models.payment import Payment, PaymentMethod
from backend.models.salary import Salary
from backend.services import excel_import as legacy_excel_import
from backend.services import settings as settings_service
from backend.services.accounting_engine import (
    generate_entries_for_invoice,
    generate_entries_for_payment,
)
from backend.services.bank_service import create_bank_transaction_record, recompute_bank_balances
from backend.services.cash_service import recompute_cash_balances
from backend.services.excel_import_classification import (
    classify_comptabilite_sheet,
    classify_gestion_sheet,
)
from backend.services.excel_import_parsing import (
    format_contact_display_name,
    normalize_text,
    split_contact_full_name,
)
from backend.services.excel_import_payment_matching import (
    PaymentMatchCandidate,
    resolve_payment_match_with_database,
)
from backend.services.excel_import_policy import (
    ENTRY_COVERED_BY_SOLDE_MESSAGE,
    ENTRY_EXISTING_MESSAGE,
    ENTRY_NEAR_EXISTING_MANUAL_MESSAGE,
    EXISTING_GESTION_ROW_MESSAGE,
    EXISTING_SALARY_MESSAGE,
    filter_duplicate_contact_rows,
    filter_duplicate_invoice_rows,
    format_row_issue,
    make_existing_contact_issue,
    make_existing_invoice_issue,
    make_payment_resolution_issue,
    resolve_invoice_contact_match,
)
from backend.services.excel_import_preview_helpers import contact_preview_key
from backend.services.excel_import_state import (
    accounting_entry_line_signature,
    accounting_entry_signature,
    compute_file_hash,
    load_existing_accounting_entry_signatures,
    load_existing_contacts_by_preview_key,
    load_existing_generated_accounting_group_signatures,
    load_existing_invoice_numbers,
    load_existing_manual_accounting_line_signatures,
)
from backend.services.excel_import_types import (
    NormalizedBankRow,
    NormalizedCashRow,
    NormalizedContactRow,
    NormalizedEntryRow,
    NormalizedInvoiceRow,
    NormalizedPaymentRow,
    NormalizedSalaryRow,
    RowIgnoredIssue,
    RowWarningIssue,
)
from backend.services.fiscal_year_service import find_fiscal_year_id_for_date
from backend.services.invoice import apply_default_due_date
from backend.services.payment import _refresh_invoice_status

_JsonValue = TypeVar("_JsonValue")


@dataclass(slots=True)
class _PreparedOperationSpec:
    position: int
    operation_type: str
    title: str
    description: str | None
    source_sheet: str | None
    source_row_numbers: list[int]
    decision: ImportOperationDecision
    diagnostics: list[str]
    payload: dict[str, Any]


@dataclass(slots=True)
class _SupplierOperationResult:
    contact: Contact | None
    contact_before: dict[str, Any] | None
    invoice: Invoice
    invoice_before: dict[str, Any] | None
    payment: Payment
    created_contact: bool
    created_invoice: bool


_IGNORED_FINGERPRINT_COLUMNS: dict[str, set[str]] = {
    "contact": {"created_at", "updated_at"},
    "invoice": {"created_at", "updated_at"},
    "invoice_line": set(),
    "payment": {"created_at"},
    "salary": {"created_at"},
    "cash_register": {"created_at", "balance_after"},
    "bank_transaction": {"created_at", "balance_after"},
    "accounting_entry": {"created_at"},
}

_DUPLICATE_WORKBOOK_PAYMENT_MESSAGE = (
    "paiement deja couvert par une autre ligne du fichier, ligne ignoree"
)


_ENTITY_MODELS: dict[str, type[Any]] = {
    "contact": Contact,
    "invoice": Invoice,
    "invoice_line": InvoiceLine,
    "payment": Payment,
    "salary": Salary,
    "cash_register": CashRegister,
    "bank_transaction": BankTransaction,
    "accounting_entry": AccountingEntry,
}


def _json_dumps(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, sort_keys=True)


def _jsonable(value: Any) -> Any:
    if isinstance(value, Decimal):
        return format(value, "f")
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if isinstance(value, tuple):
        return [_jsonable(item) for item in value]
    if isinstance(value, list):
        return [_jsonable(item) for item in value]
    if isinstance(value, dict):
        return {key: _jsonable(item) for key, item in value.items()}
    if hasattr(value, "value") and value.__class__.__module__.startswith("backend.models"):
        return value.value
    return value


def _canonical_decimal_string(value: Any) -> str:
    text = format(Decimal(str(value)), "f")
    if "." in text:
        text = text.rstrip("0").rstrip(".")
    if text in {"", "-0"}:
        return "0"
    return text


def _canonical_numeric_column_value(column: Any, value: Any) -> str:
    decimal_value = Decimal(str(value))
    scale = getattr(column.type, "scale", None)
    if scale is not None:
        decimal_value = decimal_value.quantize(Decimal("1").scaleb(-scale))
    return _canonical_decimal_string(decimal_value)


def _normalize_snapshot_for_fingerprint(
    entity_type: str,
    snapshot: dict[str, Any],
) -> dict[str, Any]:
    model_cls = _ENTITY_MODELS.get(entity_type)
    if model_cls is None:
        return snapshot

    normalized = dict(snapshot)
    for column in sa_inspect(model_cls).columns:
        if column.key not in normalized:
            continue
        value = normalized[column.key]
        if value is None:
            continue
        if isinstance(column.type, Numeric):
            normalized[column.key] = _canonical_numeric_column_value(column, value)
    return normalized


def _filtered_snapshot_for_fingerprint(
    entity_type: str,
    snapshot: dict[str, Any],
) -> dict[str, Any]:
    normalized_snapshot = _normalize_snapshot_for_fingerprint(entity_type, snapshot)
    return {
        key: value
        for key, value in normalized_snapshot.items()
        if key not in _IGNORED_FINGERPRINT_COLUMNS.get(entity_type, set())
    }


def _deserialize_column_value(column: Any, value: Any) -> Any:
    if value is None:
        return None
    if isinstance(column.type, Numeric):
        return Decimal(str(value))
    if isinstance(column.type, DateTime):
        return datetime.fromisoformat(str(value))
    if isinstance(column.type, Date):
        return date.fromisoformat(str(value))
    if isinstance(column.type, Boolean):
        return bool(value)
    if isinstance(column.type, Integer):
        return int(value)
    return value


def _entity_type_for_instance(instance: Any) -> str:
    for entity_type, model_cls in _ENTITY_MODELS.items():
        if isinstance(instance, model_cls):
            return entity_type
    raise ValueError(f"Unsupported entity instance: {type(instance)!r}")


def _serialize_instance(instance: Any) -> dict[str, Any]:
    payload: dict[str, Any] = {}
    for column in sa_inspect(type(instance)).columns:
        payload[column.key] = _jsonable(getattr(instance, column.key))
    return payload


def _snapshot_fingerprint(entity_type: str, snapshot: dict[str, Any] | None) -> str | None:
    if snapshot is None:
        return None
    filtered = _filtered_snapshot_for_fingerprint(entity_type, snapshot)
    return sha256(_json_dumps(filtered).encode("utf-8")).hexdigest()


def _restore_instance(model_cls: type[Any], snapshot: dict[str, Any]) -> Any:
    values: dict[str, Any] = {}
    for column in sa_inspect(model_cls).columns:
        if column.key not in snapshot:
            continue
        values[column.key] = _deserialize_column_value(column, snapshot[column.key])
    return model_cls(**values)


def _apply_snapshot(instance: Any, snapshot: dict[str, Any]) -> None:
    for column in sa_inspect(type(instance)).columns:
        if column.key == "id" or column.key not in snapshot:
            continue
        setattr(instance, column.key, _deserialize_column_value(column, snapshot[column.key]))


def _row_to_payload(row: Any) -> dict[str, Any]:
    return cast(dict[str, Any], _jsonable(asdict(row)))


def _worksheet_row_to_payload(worksheet: Any, source_row_number: int) -> dict[str, Any]:
    header_values = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    row_values = next(
        worksheet.iter_rows(min_row=source_row_number, max_row=source_row_number, values_only=True),
        (),
    )
    if not row_values:
        return {"source_row_number": source_row_number}

    payload: dict[str, Any] = {"source_row_number": source_row_number}
    for index, value in enumerate(row_values, start=1):
        header = header_values[index - 1] if index - 1 < len(header_values) else None
        key = str(header).strip() if header not in (None, "") else f"column_{index}"
        payload[key] = _jsonable(value)
    return payload


def _payload_to_contact_row(payload: dict[str, Any]) -> NormalizedContactRow:
    return NormalizedContactRow(
        source_row_number=int(payload["source_row_number"]),
        nom=str(payload["nom"]),
        prenom=payload.get("prenom"),
        email=payload.get("email"),
    )


def _payload_to_invoice_row(payload: dict[str, Any]) -> NormalizedInvoiceRow:
    return NormalizedInvoiceRow(
        source_row_number=int(payload["source_row_number"]),
        invoice_date=date.fromisoformat(str(payload["invoice_date"])),
        amount=Decimal(str(payload["amount"])),
        contact_name=str(payload["contact_name"]),
        invoice_number=payload.get("invoice_number"),
        label=str(payload["label"]),
        course_amount=(
            Decimal(str(payload["course_amount"]))
            if payload.get("course_amount") is not None
            else None
        ),
        adhesion_amount=(
            Decimal(str(payload["adhesion_amount"]))
            if payload.get("adhesion_amount") is not None
            else None
        ),
    )


def _payload_to_payment_row(payload: dict[str, Any]) -> NormalizedPaymentRow:
    return NormalizedPaymentRow(
        source_row_number=int(payload["source_row_number"]),
        payment_date=date.fromisoformat(str(payload["payment_date"])),
        amount=Decimal(str(payload["amount"])),
        method=str(payload["method"]),
        invoice_ref=str(payload["invoice_ref"]),
        contact_name=str(payload["contact_name"]),
        cheque_number=payload.get("cheque_number"),
        deposited=bool(payload["deposited"]),
        deposit_date=(
            date.fromisoformat(str(payload["deposit_date"]))
            if payload.get("deposit_date")
            else None
        ),
    )


def _payload_to_salary_row(payload: dict[str, Any]) -> NormalizedSalaryRow:
    return NormalizedSalaryRow(
        source_row_number=int(payload["source_row_number"]),
        month=str(payload["month"]),
        employee_name=str(payload["employee_name"]),
        hours=Decimal(str(payload["hours"])),
        gross=Decimal(str(payload["gross"])),
        employee_charges=Decimal(str(payload["employee_charges"])),
        employer_charges=Decimal(str(payload["employer_charges"])),
        tax=Decimal(str(payload["tax"])),
        net_pay=Decimal(str(payload["net_pay"])),
    )


def _payload_to_cash_row(payload: dict[str, Any]) -> NormalizedCashRow:
    return NormalizedCashRow(
        source_row_number=int(payload["source_row_number"]),
        entry_date=date.fromisoformat(str(payload["entry_date"])),
        amount=Decimal(str(payload["amount"])),
        movement_type=str(payload["movement_type"]),
        reference=payload.get("reference"),
        contact_name=payload.get("contact_name"),
        description=str(payload["description"]),
    )


def _payload_to_bank_row(payload: dict[str, Any]) -> NormalizedBankRow:
    return NormalizedBankRow(
        source_row_number=int(payload["source_row_number"]),
        entry_date=date.fromisoformat(str(payload["entry_date"])),
        amount=Decimal(str(payload["amount"])),
        reference=payload.get("reference"),
        description=str(payload["description"]),
        balance_after=Decimal(str(payload["balance_after"])),
    )


def _payload_to_entry_row(payload: dict[str, Any]) -> NormalizedEntryRow:
    return NormalizedEntryRow(
        source_row_number=int(payload["source_row_number"]),
        entry_date=date.fromisoformat(str(payload["entry_date"])),
        account_number=str(payload["account_number"]),
        label=str(payload["label"]),
        debit=Decimal(str(payload["debit"])),
        credit=Decimal(str(payload["credit"])),
        change_marker=payload.get("change_marker"),
    )


def _deserialize_json(value: str | None, default: _JsonValue) -> _JsonValue:
    if not value:
        return default
    try:
        return cast(_JsonValue, json.loads(value))
    except json.JSONDecodeError:
        return default


def _summary_base() -> dict[str, Any]:
    return {
        "contacts_created": 0,
        "invoices_created": 0,
        "payments_created": 0,
        "salaries_created": 0,
        "entries_created": 0,
        "cash_created": 0,
        "bank_created": 0,
        "skipped": 0,
        "ignored_rows": 0,
        "blocked_rows": 0,
        "warnings": [],
        "errors": [],
        "warning_details": [],
        "error_details": [],
        "sheets": [],
        "created_objects": [],
    }


def _increment_summary_count(summary: dict[str, Any], entity_type: str) -> None:
    if entity_type == "contact":
        summary["contacts_created"] += 1
    elif entity_type == "invoice":
        summary["invoices_created"] += 1
    elif entity_type == "payment":
        summary["payments_created"] += 1
    elif entity_type == "salary":
        summary["salaries_created"] += 1
    elif entity_type == "accounting_entry":
        summary["entries_created"] += 1
    elif entity_type == "cash_register":
        summary["cash_created"] += 1
    elif entity_type == "bank_transaction":
        summary["bank_created"] += 1


def _make_effect_reference(entity_type: str, snapshot: dict[str, Any] | None) -> str | None:
    if snapshot is None:
        return None
    if entity_type == "contact":
        nom = snapshot.get("nom")
        prenom = snapshot.get("prenom")
        return format_contact_display_name(nom or "", prenom) or nom
    if entity_type == "invoice":
        return snapshot.get("number")
    if entity_type == "invoice_line":
        return snapshot.get("description")
    if entity_type == "payment":
        return snapshot.get("reference") or str(snapshot.get("id"))
    if entity_type == "salary":
        return snapshot.get("month")
    if entity_type == "cash_register":
        return snapshot.get("reference") or snapshot.get("date")
    if entity_type == "bank_transaction":
        return snapshot.get("reference") or snapshot.get("date")
    if entity_type == "accounting_entry":
        return snapshot.get("entry_number")
    return str(snapshot.get("id")) if snapshot.get("id") is not None else None


def _make_effect_details(entity_type: str, snapshot: dict[str, Any] | None) -> dict[str, Any]:
    if snapshot is None:
        return {}
    if entity_type == "contact":
        return {
            "nom": snapshot.get("nom"),
            "prenom": snapshot.get("prenom"),
            "email": snapshot.get("email"),
            "type": snapshot.get("type"),
        }
    if entity_type == "invoice":
        return {
            "number": snapshot.get("number"),
            "contact_id": snapshot.get("contact_id"),
            "status": snapshot.get("status"),
            "date": snapshot.get("date"),
            "amount": snapshot.get("total_amount"),
            "paid_amount": snapshot.get("paid_amount"),
            "label": snapshot.get("label"),
            "reference": snapshot.get("reference"),
        }
    if entity_type == "invoice_line":
        return {
            "description": snapshot.get("description"),
            "amount": snapshot.get("amount"),
            "line_type": snapshot.get("line_type"),
        }
    if entity_type == "payment":
        return {
            "invoice_id": snapshot.get("invoice_id"),
            "contact_id": snapshot.get("contact_id"),
            "amount": snapshot.get("amount"),
            "date": snapshot.get("date"),
            "method": snapshot.get("method"),
            "reference": snapshot.get("reference"),
            "cheque_number": snapshot.get("cheque_number"),
            "deposited": snapshot.get("deposited"),
            "deposit_date": snapshot.get("deposit_date"),
        }
    if entity_type == "salary":
        return {
            "employee_id": snapshot.get("employee_id"),
            "month": snapshot.get("month"),
            "gross": snapshot.get("gross"),
            "employee_charges": snapshot.get("employee_charges"),
            "employer_charges": snapshot.get("employer_charges"),
            "tax": snapshot.get("tax"),
            "net_pay": snapshot.get("net_pay"),
        }
    if entity_type == "cash_register":
        return {
            "amount": snapshot.get("amount"),
            "type": snapshot.get("type"),
            "description": snapshot.get("description"),
            "date": snapshot.get("date"),
            "reference": snapshot.get("reference"),
        }
    if entity_type == "bank_transaction":
        return {
            "amount": snapshot.get("amount"),
            "description": snapshot.get("description"),
            "date": snapshot.get("date"),
            "reference": snapshot.get("reference"),
        }
    if entity_type == "accounting_entry":
        return {
            "entry_number": snapshot.get("entry_number"),
            "account_number": snapshot.get("account_number"),
            "date": snapshot.get("date"),
            "label": snapshot.get("label"),
            "debit": snapshot.get("debit"),
            "credit": snapshot.get("credit"),
        }
    return {}


def _operation_status_for_decision(decision: ImportOperationDecision) -> ImportOperationStatus:
    if decision == ImportOperationDecision.IGNORE:
        return ImportOperationStatus.IGNORED
    if decision == ImportOperationDecision.BLOCK:
        return ImportOperationStatus.BLOCKED
    return ImportOperationStatus.PREPARED


def _operation_row_count(operation: ImportOperation) -> int:
    return len(_deserialize_json(operation.source_row_numbers_json, []))


def _sheet_summary(summary: dict[str, Any], name: str, kind: str) -> dict[str, Any]:
    sheets = cast(list[dict[str, Any]], summary["sheets"])
    for sheet in sheets:
        if sheet["name"] == name and sheet["kind"] == kind:
            return sheet
    sheet = {
        "name": name,
        "kind": kind,
        "imported_rows": 0,
        "ignored_rows": 0,
        "blocked_rows": 0,
        "warnings": [],
        "errors": [],
    }
    sheets.append(sheet)
    return sheet


def _summary_kind_from_operation(operation: ImportOperation) -> str:
    if operation.operation_type == "contact_row_import":
        return "contacts"
    if operation.operation_type == "client_invoice_row_import":
        return "invoices"
    if operation.operation_type == "client_payment_row_import":
        return "payments"
    if operation.operation_type == "salary_month_import":
        return "salaries"
    if operation.operation_type == "cash_row_import":
        return "cash"
    if operation.operation_type == "bank_row_import":
        return "bank"
    return "entries"


def _build_run_summary(run: ImportRun) -> dict[str, Any]:
    summary = _summary_base()
    preview: dict[str, Any] = _deserialize_json(run.preview_json, {})
    summary["warnings"].extend(preview.get("warnings", []))
    summary["errors"].extend(preview.get("errors", []))
    summary["warning_details"] = preview.get("warning_details", [])
    summary["error_details"] = preview.get("error_details", [])

    for operation in run.operations:
        source_rows = _operation_row_count(operation)
        kind = _summary_kind_from_operation(operation)
        diagnostics: list[str] = _deserialize_json(operation.diagnostics_json, [])
        if operation.source_sheet:
            sheet = _sheet_summary(summary, operation.source_sheet, kind)
        else:
            sheet = None

        if operation.status == ImportOperationStatus.IGNORED:
            summary["skipped"] += source_rows
            summary["ignored_rows"] += source_rows
            summary["warnings"].extend(diagnostics)
            if sheet is not None:
                sheet["ignored_rows"] += source_rows
                sheet["warnings"].extend(diagnostics)
            continue

        if operation.status == ImportOperationStatus.BLOCKED:
            summary["blocked_rows"] += source_rows
            summary["errors"].extend(diagnostics)
            if sheet is not None:
                sheet["blocked_rows"] += source_rows
                sheet["errors"].extend(diagnostics)
            continue

        if operation.status == ImportOperationStatus.FAILED:
            failure_messages = [message for message in diagnostics if message]
            if operation.error_message:
                failure_messages.append(operation.error_message)
            formatted_messages = [
                f"{operation.title} — {message}" if operation.title else message
                for message in dict.fromkeys(failure_messages)
            ]
            summary["errors"].extend(formatted_messages)
            if sheet is not None:
                sheet["errors"].extend(formatted_messages)
            continue

        if operation.status != ImportOperationStatus.APPLIED:
            continue

        if sheet is not None:
            sheet["imported_rows"] += source_rows
        for effect in operation.effects:
            if effect.status != ImportEffectStatus.APPLIED:
                continue
            snapshot = _deserialize_json(effect.after_snapshot_json, None)
            if effect.action == ImportEffectAction.CREATE:
                _increment_summary_count(summary, effect.entity_type)
                summary["created_objects"].append(
                    {
                        "sheet_name": operation.source_sheet,
                        "kind": kind,
                        "object_type": effect.entity_type,
                        "object_id": effect.entity_id,
                        "reference": effect.entity_reference,
                        "details": _deserialize_json(effect.details_json, {}),
                    }
                )
            elif effect.action == ImportEffectAction.UPDATE and snapshot is not None:
                summary["created_objects"].append(
                    {
                        "sheet_name": operation.source_sheet,
                        "kind": kind,
                        "object_type": effect.entity_type,
                        "object_id": effect.entity_id,
                        "reference": effect.entity_reference,
                        "details": {
                            **_deserialize_json(effect.details_json, {}),
                            "action": "updated",
                        },
                    }
                )
    return summary


def _can_run_execute(run: ImportRun) -> bool:
    return run.status in {ImportRunStatus.PREPARED} and all(
        operation.decision != ImportOperationDecision.BLOCK for operation in run.operations
    )


def _run_can_undo(run: ImportRun) -> bool:
    return any(operation.status == ImportOperationStatus.APPLIED for operation in run.operations)


def _run_can_redo(run: ImportRun) -> bool:
    return any(operation.status == ImportOperationStatus.UNDONE for operation in run.operations)


def _serialize_effect(effect: ImportEffect) -> dict[str, Any]:
    return {
        "id": effect.id,
        "entity_type": effect.entity_type,
        "action": effect.action,
        "entity_id": effect.entity_id,
        "entity_reference": effect.entity_reference,
        "details": _deserialize_json(effect.details_json, {}),
        "status": effect.status,
    }


def _serialize_planned_effect(effect: ImportEffect) -> dict[str, Any]:
    return {
        "id": None,
        "entity_type": effect.entity_type,
        "action": effect.action,
        "entity_id": None,
        "entity_reference": effect.entity_reference,
        "details": _deserialize_json(effect.details_json, {}),
        "status": "planned",
    }


def _serialize_operation_source_data(payload: dict[str, Any]) -> list[dict[str, Any]]:
    source_data: list[dict[str, Any]] = []
    single_row = payload.get("row")
    if isinstance(single_row, dict):
        source_data.append(single_row)

    grouped_rows = payload.get("rows")
    if isinstance(grouped_rows, list):
        source_data.extend(row for row in grouped_rows if isinstance(row, dict))

    return source_data


def _serialize_operation(operation: ImportOperation) -> dict[str, Any]:
    payload: dict[str, Any] = _deserialize_json(operation.payload_json, {})
    return {
        "id": operation.id,
        "position": operation.position,
        "operation_type": operation.operation_type,
        "title": operation.title,
        "description": operation.description,
        "source_sheet": operation.source_sheet,
        "source_row_numbers": _deserialize_json(operation.source_row_numbers_json, []),
        "decision": operation.decision,
        "status": operation.status,
        "diagnostics": _deserialize_json(operation.diagnostics_json, []),
        "error_message": operation.error_message,
        "can_undo": operation.status == ImportOperationStatus.APPLIED,
        "can_redo": operation.status == ImportOperationStatus.UNDONE,
        "effects": [_serialize_effect(effect) for effect in operation.effects],
        "planned_effects": payload.get("planned_effects", []),
        "source_data": _serialize_operation_source_data(payload),
    }


def serialize_run(run: ImportRun) -> dict[str, Any]:
    summary = _deserialize_json(run.summary_json, None)
    if summary is None:
        summary = _build_run_summary(run)
    return {
        "id": run.id,
        "kind": "run",
        "import_type": run.import_type,
        "status": run.status,
        "file_name": run.file_name,
        "file_hash": run.file_hash,
        "comparison_start_date": (
            run.comparison_start_date.isoformat() if run.comparison_start_date else None
        ),
        "comparison_end_date": run.comparison_end_date.isoformat()
        if run.comparison_end_date
        else None,
        "created_at": run.created_at.isoformat() if run.created_at else None,
        "updated_at": run.updated_at.isoformat() if run.updated_at else None,
        "preview": _deserialize_json(run.preview_json, None),
        "summary": summary,
        "operations": [_serialize_operation(operation) for operation in run.operations],
        "can_execute": _can_run_execute(run),
        "can_undo": _run_can_undo(run),
        "can_redo": _run_can_redo(run),
    }


def _serialize_legacy_log(log: ImportLog) -> dict[str, Any]:
    summary = None
    if log.summary:
        try:
            summary = json.loads(log.summary)
        except json.JSONDecodeError:
            summary = None
    return {
        "id": log.id,
        "kind": "legacy_log",
        "import_type": log.import_type,
        "status": log.status,
        "file_name": log.file_name,
        "file_hash": log.file_hash,
        "comparison_start_date": None,
        "comparison_end_date": None,
        "created_at": log.created_at.isoformat() if log.created_at else None,
        "updated_at": log.created_at.isoformat() if log.created_at else None,
        "preview": None,
        "summary": summary,
        "operations": [],
        "can_execute": False,
        "can_undo": False,
        "can_redo": False,
    }


async def list_import_history(db: AsyncSession, *, limit: int = 20) -> list[dict[str, Any]]:
    runs_result = await db.execute(
        select(ImportRun)
        .execution_options(populate_existing=True)
        .options(selectinload(ImportRun.operations).selectinload(ImportOperation.effects))
        .order_by(desc(ImportRun.created_at))
        .limit(limit)
    )
    logs_result = await db.execute(
        select(ImportLog).order_by(desc(ImportLog.created_at)).limit(limit)
    )

    items = [serialize_run(run) for run in runs_result.scalars().all()] + [
        _serialize_legacy_log(log) for log in logs_result.scalars().all()
    ]
    items.sort(key=lambda item: item.get("created_at") or "", reverse=True)
    return items[:limit]


async def get_import_run(db: AsyncSession, run_id: int) -> ImportRun | None:
    result = await db.execute(
        select(ImportRun)
        .where(ImportRun.id == run_id)
        .execution_options(populate_existing=True)
        .options(selectinload(ImportRun.operations).selectinload(ImportOperation.effects))
    )
    return result.scalar_one_or_none()


async def _load_workbook(file_bytes: bytes) -> Any:
    from io import BytesIO

    return openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)


def _diagnostics_for_ignored_issue(issue: RowIgnoredIssue) -> list[str]:
    return [format_row_issue(issue)]


def _append_spec(
    specs: list[_PreparedOperationSpec],
    *,
    operation_type: str,
    title: str,
    description: str | None,
    source_sheet: str | None,
    source_row_numbers: list[int],
    decision: ImportOperationDecision,
    diagnostics: list[str],
    payload: dict[str, Any],
) -> None:
    specs.append(
        _PreparedOperationSpec(
            position=len(specs) + 1,
            operation_type=operation_type,
            title=title,
            description=description,
            source_sheet=source_sheet,
            source_row_numbers=source_row_numbers,
            decision=decision,
            diagnostics=diagnostics,
            payload=payload,
        )
    )


async def _prepare_gestion_specs(
    db: AsyncSession,
    file_bytes: bytes,
) -> list[_PreparedOperationSpec]:
    workbook = await _load_workbook(file_bytes)
    specs: list[_PreparedOperationSpec] = []
    existing_contacts_by_key = await load_existing_contacts_by_preview_key(db)
    existing_invoice_numbers = await load_existing_invoice_numbers(db)
    existing_salary_keys = await legacy_excel_import._load_existing_salary_keys(db)
    planned_invoice_candidates: list[PaymentMatchCandidate] = []
    deferred_payment_sheets: list[tuple[str, list[NormalizedPaymentRow]]] = []
    planned_payment_signatures: set[tuple[str, str, str, str, str]] = set()
    existing_payment_signatures = await legacy_excel_import._load_existing_payment_signatures(db)

    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        kind, _ = classify_gestion_sheet(sheet_name)
        if kind == "contacts":
            parsed_sheet, contact_rows, _ = legacy_excel_import._parse_contact_sheet(ws)
            if parsed_sheet is None or parsed_sheet.missing_columns:
                continue
            contact_rows, ignored_issues = filter_duplicate_contact_rows(
                contact_rows,
                normalize_text=normalize_text,
            )
            for issue in ignored_issues:
                _append_spec(
                    specs,
                    operation_type="ignored_by_policy",
                    title=f"Contact ligne {issue.source_row_number}",
                    description="Ligne de contact ignorée par la politique d'import.",
                    source_sheet=sheet_name,
                    source_row_numbers=[issue.source_row_number],
                    decision=ImportOperationDecision.IGNORE,
                    diagnostics=_diagnostics_for_ignored_issue(issue),
                    payload={
                        "reason": issue.message,
                        "row": _worksheet_row_to_payload(ws, issue.source_row_number),
                    },
                )
            for row in contact_rows:
                preview_key = contact_preview_key(row.nom, row.prenom)
                if existing_contacts_by_key.get(preview_key):
                    issue = make_existing_contact_issue(row.source_row_number)
                    _append_spec(
                        specs,
                        operation_type="ignored_by_policy",
                        title=format_contact_display_name(row.nom, row.prenom) or row.nom,
                        description="Contact déjà présent dans Solde.",
                        source_sheet=sheet_name,
                        source_row_numbers=[row.source_row_number],
                        decision=ImportOperationDecision.IGNORE,
                        diagnostics=[format_row_issue(issue)],
                        payload={"reason": issue.message, "row": _row_to_payload(row)},
                    )
                    continue
                _append_spec(
                    specs,
                    operation_type="contact_row_import",
                    title=format_contact_display_name(row.nom, row.prenom) or row.nom,
                    description="Créer un contact client à partir du fichier Gestion.",
                    source_sheet=sheet_name,
                    source_row_numbers=[row.source_row_number],
                    decision=ImportOperationDecision.APPLY,
                    diagnostics=[],
                    payload={"row": _row_to_payload(row)},
                )

        elif kind == "invoices":
            parsed_sheet, invoice_rows, _, ignored_issues = (
                legacy_excel_import._parse_invoice_sheet(ws)
            )
            if parsed_sheet is None or parsed_sheet.missing_columns:
                continue
            for issue in ignored_issues:
                _append_spec(
                    specs,
                    operation_type="ignored_by_policy",
                    title=f"Facture ligne {issue.source_row_number}",
                    description="Ligne de facture ignorée par la politique d'import.",
                    source_sheet=sheet_name,
                    source_row_numbers=[issue.source_row_number],
                    decision=ImportOperationDecision.IGNORE,
                    diagnostics=_diagnostics_for_ignored_issue(issue),
                    payload={
                        "reason": issue.message,
                        "row": _worksheet_row_to_payload(ws, issue.source_row_number),
                    },
                )
            invoice_rows, duplicate_issues = filter_duplicate_invoice_rows(
                invoice_rows,
                normalize_text=normalize_text,
            )
            for issue in duplicate_issues:
                _append_spec(
                    specs,
                    operation_type="ignored_by_policy",
                    title=f"Facture ligne {issue.source_row_number}",
                    description="Ligne de facture dupliquée dans le fichier source.",
                    source_sheet=sheet_name,
                    source_row_numbers=[issue.source_row_number],
                    decision=ImportOperationDecision.IGNORE,
                    diagnostics=_diagnostics_for_ignored_issue(issue),
                    payload={
                        "reason": issue.message,
                        "row": _worksheet_row_to_payload(ws, issue.source_row_number),
                    },
                )
            generated_number_index = 0
            for invoice_row in invoice_rows:
                matched_contact, contact_issue = resolve_invoice_contact_match(
                    invoice_row,
                    existing_contacts_by_key,
                    normalize_text=normalize_text,
                )
                if contact_issue is not None:
                    _append_spec(
                        specs,
                        operation_type="blocked_by_validation",
                        title=(
                            invoice_row.invoice_number
                            or f"Facture ligne {invoice_row.source_row_number}"
                        ),
                        description="Correspondance contact ambiguë ou invalide.",
                        source_sheet=sheet_name,
                        source_row_numbers=[invoice_row.source_row_number],
                        decision=ImportOperationDecision.BLOCK,
                        diagnostics=[format_row_issue(contact_issue)],
                        payload={"row": _row_to_payload(invoice_row)},
                    )
                    continue
                invoice_number = invoice_row.invoice_number
                if not invoice_number:
                    generated_number_index += 1
                    invoice_number = (
                        f"IMP-{invoice_row.invoice_date.year}-{generated_number_index:04d}"
                    )
                if normalize_text(invoice_number) in existing_invoice_numbers:
                    issue = make_existing_invoice_issue(invoice_row.source_row_number)
                    _append_spec(
                        specs,
                        operation_type="ignored_by_policy",
                        title=invoice_number,
                        description="Facture déjà présente dans Solde.",
                        source_sheet=sheet_name,
                        source_row_numbers=[invoice_row.source_row_number],
                        decision=ImportOperationDecision.IGNORE,
                        diagnostics=[format_row_issue(issue)],
                        payload={
                            "reason": issue.message,
                            "row": _row_to_payload(invoice_row),
                        },
                    )
                    continue
                _append_spec(
                    specs,
                    operation_type="client_invoice_row_import",
                    title=invoice_number,
                    description="Créer une facture client et ses effets dérivés.",
                    source_sheet=sheet_name,
                    source_row_numbers=[invoice_row.source_row_number],
                    decision=ImportOperationDecision.APPLY,
                    diagnostics=[],
                    payload={
                        "row": _row_to_payload(invoice_row),
                        "invoice_number": invoice_number,
                        "matched_contact_id": matched_contact.id
                        if matched_contact is not None
                        else None,
                    },
                )
                planned_invoice_candidates.append(
                    PaymentMatchCandidate(
                        candidate_key=f"invoice:{normalize_text(invoice_number)}",
                        invoice_id=None,
                        contact_id=matched_contact.id if matched_contact is not None else None,
                        invoice_type=InvoiceType.CLIENT,
                        invoice_number=invoice_number,
                        contact_name=invoice_row.contact_name,
                    )
                )

        elif kind == "payments":
            parsed_sheet, payment_rows, _ = legacy_excel_import._parse_payment_sheet(ws)
            if parsed_sheet is None or parsed_sheet.missing_columns:
                continue
            deferred_payment_sheets.append((sheet_name, payment_rows))

        elif kind == "salaries":
            parsed_sheet, salary_rows, _ = legacy_excel_import._parse_salary_sheet(ws)
            if parsed_sheet is None or parsed_sheet.missing_columns:
                continue
            rows_by_month: dict[str, list[NormalizedSalaryRow]] = {}
            for salary_row in salary_rows:
                salary_key = (
                    salary_row.month,
                    legacy_excel_import._salary_employee_key(salary_row.employee_name),
                )
                if salary_key in existing_salary_keys:
                    issue = RowIgnoredIssue(
                        source_row_number=salary_row.source_row_number,
                        message=EXISTING_SALARY_MESSAGE,
                    )
                    _append_spec(
                        specs,
                        operation_type="ignored_by_policy",
                        title=f"Salaire ligne {salary_row.source_row_number}",
                        description="Ligne de salaire déjà présente dans Solde.",
                        source_sheet=sheet_name,
                        source_row_numbers=[salary_row.source_row_number],
                        decision=ImportOperationDecision.IGNORE,
                        diagnostics=[format_row_issue(issue)],
                        payload={"reason": issue.message, "row": _row_to_payload(salary_row)},
                    )
                    continue
                existing_salary_keys.add(salary_key)
                rows_by_month.setdefault(salary_row.month, []).append(salary_row)
            for month, month_rows in sorted(rows_by_month.items()):
                _append_spec(
                    specs,
                    operation_type="salary_month_import",
                    title=month,
                    description="Créer les salaires du mois et leurs écritures groupées.",
                    source_sheet=sheet_name,
                    source_row_numbers=[row.source_row_number for row in month_rows],
                    decision=ImportOperationDecision.APPLY,
                    diagnostics=[],
                    payload={"rows": [_row_to_payload(row) for row in month_rows]},
                )

        elif kind == "cash":
            parsed_sheet, cash_rows, _, ignored_issues = legacy_excel_import._parse_cash_sheet(ws)
            if parsed_sheet is None or parsed_sheet.missing_columns:
                continue
            for issue in ignored_issues:
                _append_spec(
                    specs,
                    operation_type="ignored_by_policy",
                    title=f"Caisse ligne {issue.source_row_number}",
                    description="Ligne de caisse ignorée par la politique d'import.",
                    source_sheet=sheet_name,
                    source_row_numbers=[issue.source_row_number],
                    decision=ImportOperationDecision.IGNORE,
                    diagnostics=_diagnostics_for_ignored_issue(issue),
                    payload={
                        "reason": issue.message,
                        "row": _worksheet_row_to_payload(ws, issue.source_row_number),
                    },
                )
            for cash_row in cash_rows:
                _append_spec(
                    specs,
                    operation_type="cash_row_import",
                    title=cash_row.reference or cash_row.description,
                    description="Créer un mouvement de caisse et ses effets dérivés.",
                    source_sheet=sheet_name,
                    source_row_numbers=[cash_row.source_row_number],
                    decision=ImportOperationDecision.APPLY,
                    diagnostics=[],
                    payload={"row": _row_to_payload(cash_row)},
                )

        elif kind == "bank":
            parsed_sheet, bank_rows, _, ignored_issues = legacy_excel_import._parse_bank_sheet(ws)
            if parsed_sheet is None or parsed_sheet.missing_columns:
                continue
            for issue in ignored_issues:
                _append_spec(
                    specs,
                    operation_type="ignored_by_policy",
                    title=f"Banque ligne {issue.source_row_number}",
                    description="Ligne bancaire ignorée par la politique d'import.",
                    source_sheet=sheet_name,
                    source_row_numbers=[issue.source_row_number],
                    decision=ImportOperationDecision.IGNORE,
                    diagnostics=_diagnostics_for_ignored_issue(issue),
                    payload={
                        "reason": issue.message,
                        "row": _worksheet_row_to_payload(ws, issue.source_row_number),
                    },
                )
            for bank_row in bank_rows:
                if bank_row.amount > 0:
                    supplier_candidate = (
                        legacy_excel_import._supplier_invoice_candidate_from_bank_row(bank_row)
                    )
                    if supplier_candidate is None:
                        invoice_reference = legacy_excel_import._single_client_invoice_reference(
                            bank_row.reference,
                            bank_row.description,
                        )
                        if invoice_reference is not None:
                            resolution = await resolve_payment_match_with_database(
                                db,
                                legacy_excel_import._payment_row_from_bank_row(
                                    bank_row,
                                    invoice_reference=invoice_reference,
                                ),
                                planned_invoice_candidates,
                            )
                            if (
                                resolution.status == "matched"
                                and resolution.candidate is not None
                                and resolution.candidate.invoice_id is not None
                            ):
                                payment_signature = legacy_excel_import._payment_signature(
                                    invoice_id=resolution.candidate.invoice_id,
                                    payment_date=bank_row.entry_date,
                                    amount=bank_row.amount,
                                    method=PaymentMethod.VIREMENT.value,
                                )
                                if payment_signature in existing_payment_signatures:
                                    issue = RowIgnoredIssue(
                                        source_row_number=bank_row.source_row_number,
                                        message=EXISTING_GESTION_ROW_MESSAGE,
                                    )
                                    _append_spec(
                                        specs,
                                        operation_type="ignored_by_policy",
                                        title=bank_row.reference or bank_row.description,
                                        description="Ligne bancaire déjà présente dans Solde.",
                                        source_sheet=sheet_name,
                                        source_row_numbers=[bank_row.source_row_number],
                                        decision=ImportOperationDecision.IGNORE,
                                        diagnostics=[format_row_issue(issue)],
                                        payload={
                                            "reason": issue.message,
                                            "row": _row_to_payload(bank_row),
                                        },
                                    )
                                    continue
                            if resolution.status == "matched" and resolution.candidate is not None:
                                planned_payment_signatures.add(
                                    legacy_excel_import._gestion_payment_comparison_signature(
                                        reference=(
                                            resolution.candidate.invoice_number or invoice_reference
                                        ),
                                        payment_date=bank_row.entry_date,
                                        amount=bank_row.amount,
                                        settlement_account=(
                                            legacy_excel_import._client_settlement_account_from_method(
                                                PaymentMethod.VIREMENT.value
                                            )
                                        ),
                                    )
                                )
                _append_spec(
                    specs,
                    operation_type="bank_row_import",
                    title=bank_row.reference or bank_row.description,
                    description="Créer un mouvement bancaire et ses effets dérivés.",
                    source_sheet=sheet_name,
                    source_row_numbers=[bank_row.source_row_number],
                    decision=ImportOperationDecision.APPLY,
                    diagnostics=[],
                    payload={"row": _row_to_payload(bank_row)},
                )

    for sheet_name, rows in deferred_payment_sheets:
        for payment_row in rows:
            resolution = await resolve_payment_match_with_database(
                db,
                payment_row,
                planned_invoice_candidates,
            )
            blocking_issue = make_payment_resolution_issue(
                source_row_number=payment_row.source_row_number,
                status=resolution.status,
                candidate=resolution.candidate,
                message=resolution.message,
                require_persistable_candidate=False,
            )
            if blocking_issue is not None:
                _append_spec(
                    specs,
                    operation_type="blocked_by_validation",
                    title=f"Paiement ligne {payment_row.source_row_number}",
                    description="Paiement client non résolu pendant la préparation.",
                    source_sheet=sheet_name,
                    source_row_numbers=[payment_row.source_row_number],
                    decision=ImportOperationDecision.BLOCK,
                    diagnostics=[format_row_issue(blocking_issue)],
                    payload={"row": _row_to_payload(payment_row)},
                )
                continue
            assert resolution.candidate is not None
            workbook_payment_signature = legacy_excel_import._gestion_payment_comparison_signature(
                reference=resolution.candidate.invoice_number or payment_row.invoice_ref,
                payment_date=payment_row.payment_date,
                amount=payment_row.amount,
                settlement_account=legacy_excel_import._client_settlement_account_from_method(
                    payment_row.method
                ),
                cheque_number=payment_row.cheque_number,
            )
            if workbook_payment_signature in planned_payment_signatures:
                issue = RowIgnoredIssue(
                    source_row_number=payment_row.source_row_number,
                    message=_DUPLICATE_WORKBOOK_PAYMENT_MESSAGE,
                )
                _append_spec(
                    specs,
                    operation_type="ignored_by_policy",
                    title=f"Paiement ligne {payment_row.source_row_number}",
                    description="Paiement client déjà couvert par une autre ligne du fichier.",
                    source_sheet=sheet_name,
                    source_row_numbers=[payment_row.source_row_number],
                    decision=ImportOperationDecision.IGNORE,
                    diagnostics=[format_row_issue(issue)],
                    payload={"reason": issue.message, "row": _row_to_payload(payment_row)},
                )
                continue
            planned_payment_signatures.add(workbook_payment_signature)
            if resolution.candidate.invoice_id is not None:
                payment_signature = legacy_excel_import._payment_signature(
                    invoice_id=resolution.candidate.invoice_id,
                    payment_date=payment_row.payment_date,
                    amount=payment_row.amount,
                    method=payment_row.method,
                    cheque_number=payment_row.cheque_number,
                )
                if payment_signature in existing_payment_signatures:
                    issue = RowIgnoredIssue(
                        source_row_number=payment_row.source_row_number,
                        message=EXISTING_GESTION_ROW_MESSAGE,
                    )
                    _append_spec(
                        specs,
                        operation_type="ignored_by_policy",
                        title=f"Paiement ligne {payment_row.source_row_number}",
                        description="Paiement client déjà présent dans Solde.",
                        source_sheet=sheet_name,
                        source_row_numbers=[payment_row.source_row_number],
                        decision=ImportOperationDecision.IGNORE,
                        diagnostics=[format_row_issue(issue)],
                        payload={"reason": issue.message, "row": _row_to_payload(payment_row)},
                    )
                    continue
                existing_payment_signatures.add(payment_signature)
            _append_spec(
                specs,
                operation_type="client_payment_row_import",
                title=payment_row.invoice_ref,
                description="Créer un paiement client et mettre à jour la facture liée.",
                source_sheet=sheet_name,
                source_row_numbers=[payment_row.source_row_number],
                decision=ImportOperationDecision.APPLY,
                diagnostics=[],
                payload={"row": _row_to_payload(payment_row)},
            )

    return specs


async def _prepare_comptabilite_specs(
    db: AsyncSession,
    file_bytes: bytes,
) -> list[_PreparedOperationSpec]:
    workbook = await _load_workbook(file_bytes)
    specs: list[_PreparedOperationSpec] = []
    existing_entry_signatures = await load_existing_accounting_entry_signatures(db)
    existing_invoice_numbers = await load_existing_invoice_numbers(db)
    existing_client_invoices_by_number = (
        await legacy_excel_import._load_existing_client_invoices_by_number(db)
    )
    existing_client_payment_signatures = (
        await legacy_excel_import._load_existing_client_payment_reference_signatures(db)
    )
    existing_supplier_payment_signatures = (
        await legacy_excel_import._load_existing_supplier_payment_reference_signatures(db)
    )
    existing_salary_group_signatures = (
        await legacy_excel_import._load_existing_salary_group_signatures(db)
    )
    existing_manual_line_signatures = await load_existing_manual_accounting_line_signatures(db)
    existing_generated_group_signatures = await load_existing_generated_accounting_group_signatures(
        db
    )

    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        kind, _ = classify_comptabilite_sheet(sheet_name)
        if kind != "entries":
            continue
        parsed_sheet, rows, _, ignored_issues = legacy_excel_import._parse_entries_sheet(ws)
        if parsed_sheet is None or parsed_sheet.missing_columns:
            continue
        for issue in ignored_issues:
            _append_spec(
                specs,
                operation_type="ignored_by_policy",
                title=f"Écriture ligne {issue.source_row_number}",
                description="Ligne comptable ignorée par la politique d'import.",
                source_sheet=sheet_name,
                source_row_numbers=[issue.source_row_number],
                decision=ImportOperationDecision.IGNORE,
                diagnostics=_diagnostics_for_ignored_issue(issue),
                payload={
                    "reason": issue.message,
                    "row": _worksheet_row_to_payload(ws, issue.source_row_number),
                },
            )

        entry_groups = legacy_excel_import._build_entry_row_groups(rows)
        index = 0
        while index < len(entry_groups):
            entry_group, next_index = (
                legacy_excel_import._merge_existing_client_invoice_entry_groups(
                    entry_groups,
                    index,
                    existing_invoice_numbers,
                )
            )
            row_numbers = [row.source_row_number for row in entry_group]
            title = (
                legacy_excel_import._single_client_invoice_reference(
                    *(entry_row.label for entry_row in entry_group)
                )
                or f"Groupe comptable ligne {row_numbers[0]}"
            )

            existing_client_invoice_reference = (
                legacy_excel_import._matching_existing_client_invoice_reference(
                    entry_group,
                    existing_invoice_numbers,
                )
            )
            if existing_client_invoice_reference is not None:
                clarifiable_invoice = (
                    await legacy_excel_import._find_clarifiable_existing_client_invoice(
                        db,
                        entry_group,
                        existing_client_invoices_by_number,
                    )
                )
                if clarifiable_invoice is not None:
                    _append_spec(
                        specs,
                        operation_type="client_invoice_clarification_from_entries",
                        title=existing_client_invoice_reference,
                        description=(
                            "Clarifier une facture client existante depuis "
                            "les écritures comptables."
                        ),
                        source_sheet=sheet_name,
                        source_row_numbers=row_numbers,
                        decision=ImportOperationDecision.APPLY,
                        diagnostics=[],
                        payload={"rows": [_row_to_payload(row) for row in entry_group]},
                    )
                else:
                    _append_spec(
                        specs,
                        operation_type="ignored_by_policy",
                        title=title,
                        description="Groupe déjà couvert par une facture cliente existante.",
                        source_sheet=sheet_name,
                        source_row_numbers=row_numbers,
                        decision=ImportOperationDecision.IGNORE,
                        diagnostics=[ENTRY_COVERED_BY_SOLDE_MESSAGE],
                        payload={"rows": [_row_to_payload(row) for row in entry_group]},
                    )
                index = next_index
                continue

            if (
                legacy_excel_import._matching_existing_client_payment_reference(
                    entry_group,
                    existing_client_payment_signatures,
                )
                or legacy_excel_import._matching_existing_supplier_invoice_payment_reference(
                    entry_group,
                    existing_supplier_payment_signatures,
                )
                or legacy_excel_import._matching_existing_salary_entry_group(
                    entry_group,
                    existing_salary_group_signatures,
                )
                or legacy_excel_import._normalized_entry_group_signature(entry_group)
                in existing_generated_group_signatures
            ):
                _append_spec(
                    specs,
                    operation_type="ignored_by_policy",
                    title=title,
                    description="Groupe comptable déjà couvert dans Solde.",
                    source_sheet=sheet_name,
                    source_row_numbers=row_numbers,
                    decision=ImportOperationDecision.IGNORE,
                    diagnostics=[ENTRY_COVERED_BY_SOLDE_MESSAGE],
                    payload={"rows": [_row_to_payload(row) for row in entry_group]},
                )
                index = next_index
                continue

            creatable_rows: list[NormalizedEntryRow] = []
            diagnostics: list[str] = []
            for row in entry_group:
                signature = accounting_entry_signature(
                    entry_date=row.entry_date,
                    account_number=row.account_number,
                    label=row.label,
                    debit=row.debit,
                    credit=row.credit,
                )
                if signature in existing_entry_signatures:
                    diagnostics.append(
                        format_row_issue(
                            RowIgnoredIssue(
                                source_row_number=row.source_row_number,
                                message=ENTRY_EXISTING_MESSAGE,
                            )
                        )
                    )
                    continue
                line_signature = accounting_entry_line_signature(
                    entry_date=row.entry_date,
                    account_number=row.account_number,
                    debit=row.debit,
                    credit=row.credit,
                )
                if line_signature in existing_manual_line_signatures:
                    diagnostics.append(
                        format_row_issue(
                            RowWarningIssue(
                                source_row_number=row.source_row_number,
                                message=ENTRY_NEAR_EXISTING_MANUAL_MESSAGE,
                            )
                        )
                    )
                creatable_rows.append(row)

            if not creatable_rows:
                _append_spec(
                    specs,
                    operation_type="ignored_by_policy",
                    title=title,
                    description="Toutes les lignes du groupe existent déjà à l'identique.",
                    source_sheet=sheet_name,
                    source_row_numbers=row_numbers,
                    decision=ImportOperationDecision.IGNORE,
                    diagnostics=diagnostics or [ENTRY_EXISTING_MESSAGE],
                    payload={"rows": [_row_to_payload(row) for row in entry_group]},
                )
            else:
                _append_spec(
                    specs,
                    operation_type="accounting_entry_group_import",
                    title=title,
                    description="Créer un groupe d'écritures comptables manuelles.",
                    source_sheet=sheet_name,
                    source_row_numbers=row_numbers,
                    decision=ImportOperationDecision.APPLY,
                    diagnostics=diagnostics,
                    payload={
                        "rows": [_row_to_payload(row) for row in entry_group],
                        "expected_creatable_count": len(creatable_rows),
                    },
                )
            index = next_index

    return specs


async def prepare_import_run(
    db: AsyncSession,
    *,
    import_type: str,
    file_bytes: bytes,
    file_name: str | None,
    comparison_start_date: date | None = None,
    comparison_end_date: date | None = None,
) -> ImportRun:
    if import_type == "gestion":
        preview = await legacy_excel_import.preview_gestion_file(
            db,
            file_bytes,
            file_name,
            comparison_start_date=comparison_start_date,
            comparison_end_date=comparison_end_date,
        )
        specs = await _prepare_gestion_specs(db, file_bytes)
    else:
        preview = await legacy_excel_import.preview_comptabilite_file(
            db,
            file_bytes,
            file_name,
            comparison_start_date=comparison_start_date,
            comparison_end_date=comparison_end_date,
        )
        specs = await _prepare_comptabilite_specs(db, file_bytes)

    planned_effects_by_position = await _build_planned_effects_for_specs(
        db=db,
        import_type=import_type,
        file_name=file_name,
        comparison_start_date=comparison_start_date,
        comparison_end_date=comparison_end_date,
        specs=specs,
    )
    for spec in specs:
        spec.payload["planned_effects"] = planned_effects_by_position.get(spec.position, [])

    run = ImportRun(
        import_type=import_type,
        status=(
            ImportRunStatus.BLOCKED
            if (
                not preview.can_import
                or any(spec.decision == ImportOperationDecision.BLOCK for spec in specs)
            )
            else ImportRunStatus.PREPARED
        ),
        file_hash=compute_file_hash(file_bytes),
        file_name=file_name,
        comparison_start_date=comparison_start_date,
        comparison_end_date=comparison_end_date,
        preview_json=_json_dumps(preview.to_dict()),
    )
    db.add(run)
    await db.flush()

    for spec in specs:
        db.add(
            ImportOperation(
                run_id=run.id,
                position=spec.position,
                operation_type=spec.operation_type,
                title=spec.title,
                description=spec.description,
                source_sheet=spec.source_sheet,
                source_row_numbers_json=_json_dumps(spec.source_row_numbers),
                decision=spec.decision,
                status=_operation_status_for_decision(spec.decision),
                payload_json=_json_dumps(spec.payload),
                diagnostics_json=_json_dumps(spec.diagnostics),
            )
        )

    await db.flush()
    refreshed_run = await get_import_run(db, run.id)
    assert refreshed_run is not None
    refreshed_run.summary_json = _json_dumps(_build_run_summary(refreshed_run))
    await db.flush()
    refreshed_run = await get_import_run(db, run.id)
    assert refreshed_run is not None
    return refreshed_run


async def _record_effect(
    db: AsyncSession,
    *,
    operation: ImportOperation,
    position: int,
    entity_type: str,
    action: ImportEffectAction,
    before_snapshot: dict[str, Any] | None,
    after_snapshot: dict[str, Any] | None,
) -> ImportEffect:
    snapshot = after_snapshot if after_snapshot is not None else before_snapshot
    effect = ImportEffect(
        operation_id=operation.id,
        position=position,
        entity_type=entity_type,
        action=action,
        entity_id=(snapshot.get("id") if snapshot is not None else None),
        entity_reference=_make_effect_reference(entity_type, snapshot),
        details_json=_json_dumps(_make_effect_details(entity_type, snapshot)),
        before_snapshot_json=_json_dumps(before_snapshot) if before_snapshot is not None else None,
        after_snapshot_json=_json_dumps(after_snapshot) if after_snapshot is not None else None,
        before_fingerprint=_snapshot_fingerprint(entity_type, before_snapshot),
        after_fingerprint=_snapshot_fingerprint(entity_type, after_snapshot),
        status=ImportEffectStatus.APPLIED,
    )
    db.add(effect)
    return effect


def _planned_effect(
    *,
    entity_type: str,
    action: str,
    entity_reference: str | None,
    details: dict[str, Any],
) -> dict[str, Any]:
    return {
        "id": None,
        "entity_type": entity_type,
        "action": action,
        "entity_id": None,
        "entity_reference": entity_reference,
        "details": _jsonable(details),
        "status": "planned",
    }


def _render_planned_label(template: str, context: dict[str, Any]) -> str:
    def replace(match: re.Match[str]) -> str:
        key = match.group(1).strip()
        return str(context.get(key, match.group(0)))

    return re.sub(r"\{\{(\w+)\}\}", replace, template)


async def _load_active_rule(db: AsyncSession, trigger: TriggerType) -> AccountingRule | None:
    result = await db.execute(
        select(AccountingRule)
        .where(AccountingRule.trigger_type == trigger)
        .where(AccountingRule.is_active == True)  # noqa: E712
        .options(selectinload(AccountingRule.entries))
        .limit(1)
    )
    return result.scalar_one_or_none()


def _client_invoice_trigger_for_label(label: InvoiceLabel | None) -> TriggerType:
    label_map: dict[InvoiceLabel | None, TriggerType] = {
        InvoiceLabel.CS: TriggerType.INVOICE_CLIENT_CS,
        InvoiceLabel.ADHESION: TriggerType.INVOICE_CLIENT_A,
        InvoiceLabel.CS_ADHESION: TriggerType.INVOICE_CLIENT_CS_A,
        InvoiceLabel.GENERAL: TriggerType.INVOICE_CLIENT_GENERAL,
        None: TriggerType.INVOICE_CLIENT_GENERAL,
    }
    return label_map.get(label, TriggerType.INVOICE_CLIENT_GENERAL)


def _client_invoice_trigger_for_line_type(line_type: Any) -> TriggerType:
    line_value = getattr(line_type, "value", str(line_type))
    if line_value == "course":
        return TriggerType.INVOICE_CLIENT_CS
    if line_value == "adhesion":
        return TriggerType.INVOICE_CLIENT_A
    return TriggerType.INVOICE_CLIENT_GENERAL


async def _planned_accounting_effects_from_rule(
    db: AsyncSession,
    *,
    trigger: TriggerType,
    amount: Decimal,
    entry_date: date,
    context: dict[str, Any],
    side: str | None = None,
) -> list[dict[str, Any]]:
    if amount <= 0:
        return []
    rule = await _load_active_rule(db, trigger)
    if rule is None:
        return []

    planned_effects: list[dict[str, Any]] = []
    for rule_entry in rule.entries:
        side_value = getattr(rule_entry.side, "value", str(rule_entry.side))
        if side is not None and side_value != side:
            continue
        debit = amount if side_value == "debit" else Decimal("0")
        credit = amount if side_value == "credit" else Decimal("0")
        label = _render_planned_label(rule_entry.description_template, context)
        planned_effects.append(
            _planned_effect(
                entity_type="accounting_entry",
                action="create",
                entity_reference=rule_entry.account_number,
                details={
                    "account_number": rule_entry.account_number,
                    "date": entry_date,
                    "label": label,
                    "debit": debit,
                    "credit": credit,
                },
            )
        )
    return planned_effects


async def _planned_effects_for_contact_operation(payload: dict[str, Any]) -> list[dict[str, Any]]:
    row = _payload_to_contact_row(payload["row"])
    display_name = format_contact_display_name(row.nom, row.prenom) or row.nom
    return [
        _planned_effect(
            entity_type="contact",
            action="create",
            entity_reference=display_name,
            details={
                "nom": row.nom,
                "prenom": row.prenom,
                "email": row.email,
                "type": ContactType.CLIENT.value,
            },
        )
    ]


async def _planned_effects_for_invoice_operation(
    db: AsyncSession, payload: dict[str, Any]
) -> list[dict[str, Any]]:
    row = _payload_to_invoice_row(payload["row"])
    invoice_number = str(payload.get("invoice_number") or row.invoice_number or "")
    invoice_label = InvoiceLabel(str(row.label)) if row.label else None
    invoice_lines = legacy_excel_import._build_client_invoice_lines_from_import_row(row)
    planned_effects: list[dict[str, Any]] = []

    if payload.get("matched_contact_id") is None:
        planned_effects.append(
            _planned_effect(
                entity_type="contact",
                action="create",
                entity_reference=row.contact_name,
                details={"nom": row.contact_name, "type": ContactType.CLIENT.value},
            )
        )

    planned_effects.append(
        _planned_effect(
            entity_type="invoice",
            action="create",
            entity_reference=invoice_number,
            details={
                "number": invoice_number,
                "date": row.invoice_date,
                "amount": row.amount,
                "label": row.label,
                "reference": invoice_number,
                "status": InvoiceStatus.SENT.value,
            },
        )
    )

    grouped_amounts: dict[str, Decimal] = {}
    for line in invoice_lines:
        line_amount = Decimal(str(line["amount"]))
        line_type = getattr(line["line_type"], "value", str(line["line_type"]))
        grouped_amounts[line_type] = grouped_amounts.get(line_type, Decimal("0")) + line_amount
        planned_effects.append(
            _planned_effect(
                entity_type="invoice_line",
                action="create",
                entity_reference=str(line["description"]),
                details={
                    "description": line["description"],
                    "amount": line_amount,
                    "line_type": line_type,
                },
            )
        )

    accounting_context = {
        "number": invoice_number,
        "contact": row.contact_name,
        "label": invoice_number,
        "amount": format(row.amount, "f"),
        "date": row.invoice_date.isoformat(),
    }
    positive_line_types = {line_type for line_type, amount in grouped_amounts.items() if amount > 0}
    if positive_line_types:
        derived_label = derive_client_invoice_label(set())
        if "course" in positive_line_types and "adhesion" in positive_line_types:
            derived_label = InvoiceLabel.CS_ADHESION
        elif "course" in positive_line_types:
            derived_label = InvoiceLabel.CS
        elif "adhesion" in positive_line_types:
            derived_label = InvoiceLabel.ADHESION
        else:
            derived_label = invoice_label or InvoiceLabel.GENERAL

        planned_effects.extend(
            await _planned_accounting_effects_from_rule(
                db,
                trigger=_client_invoice_trigger_for_label(derived_label),
                amount=row.amount,
                entry_date=row.invoice_date,
                context=accounting_context,
                side="debit",
            )
        )
        for line_type in sorted(positive_line_types):
            planned_effects.extend(
                await _planned_accounting_effects_from_rule(
                    db,
                    trigger=_client_invoice_trigger_for_line_type(line_type),
                    amount=grouped_amounts[line_type],
                    entry_date=row.invoice_date,
                    context=accounting_context,
                    side="credit",
                )
            )

    if not any(effect["entity_type"] == "accounting_entry" for effect in planned_effects):
        planned_effects.extend(
            await _planned_accounting_effects_from_rule(
                db,
                trigger=_client_invoice_trigger_for_label(invoice_label),
                amount=row.amount,
                entry_date=row.invoice_date,
                context=accounting_context,
            )
        )

    return planned_effects


async def _planned_effects_for_payment_operation(
    db: AsyncSession, payload: dict[str, Any]
) -> list[dict[str, Any]]:
    row = _payload_to_payment_row(payload["row"])
    method = PaymentMethod(str(row.method))
    planned_effects: list[dict[str, Any]] = [
        _planned_effect(
            entity_type="payment",
            action="create",
            entity_reference=row.invoice_ref,
            details={
                "reference": row.invoice_ref,
                "date": row.payment_date,
                "amount": row.amount,
                "method": method.value,
                "cheque_number": row.cheque_number,
                "deposited": row.deposited,
                "deposit_date": row.deposit_date,
            },
        ),
        _planned_effect(
            entity_type="invoice",
            action="update",
            entity_reference=row.invoice_ref,
            details={"reference": row.invoice_ref, "status": "updated"},
        ),
    ]
    method_trigger_map = {
        PaymentMethod.ESPECES: TriggerType.PAYMENT_RECEIVED_ESPECES,
        PaymentMethod.CHEQUE: TriggerType.PAYMENT_RECEIVED_CHEQUE,
        PaymentMethod.VIREMENT: TriggerType.PAYMENT_RECEIVED_VIREMENT,
    }
    planned_effects.extend(
        await _planned_accounting_effects_from_rule(
            db,
            trigger=method_trigger_map.get(method, TriggerType.PAYMENT_RECEIVED_VIREMENT),
            amount=row.amount,
            entry_date=row.payment_date,
            context={
                "label": row.cheque_number and f"Chèque {row.cheque_number}" or row.invoice_ref,
                "amount": format(row.amount, "f"),
                "date": row.payment_date.isoformat(),
                "method": method.value,
            },
        )
    )
    return planned_effects


async def _planned_effects_for_accounting_group_operation(
    payload: dict[str, Any],
) -> list[dict[str, Any]]:
    rows = [_payload_to_entry_row(row_payload) for row_payload in payload.get("rows", [])]
    return [
        _planned_effect(
            entity_type="accounting_entry",
            action="create",
            entity_reference=row.account_number,
            details={
                "account_number": row.account_number,
                "date": row.entry_date,
                "label": row.label,
                "debit": row.debit,
                "credit": row.credit,
            },
        )
        for row in rows
    ]


async def _planned_effects_for_spec(
    db: AsyncSession, spec: _PreparedOperationSpec
) -> list[dict[str, Any]]:
    if spec.decision != ImportOperationDecision.APPLY:
        return []
    if spec.operation_type == "contact_row_import":
        return await _planned_effects_for_contact_operation(spec.payload)
    if spec.operation_type == "client_invoice_row_import":
        return await _planned_effects_for_invoice_operation(db, spec.payload)
    if spec.operation_type == "client_payment_row_import":
        return await _planned_effects_for_payment_operation(db, spec.payload)
    if spec.operation_type == "accounting_entry_group_import":
        return await _planned_effects_for_accounting_group_operation(spec.payload)
    return []


async def _build_planned_effects_for_specs(
    *,
    db: AsyncSession,
    import_type: str,
    file_name: str | None,
    comparison_start_date: date | None,
    comparison_end_date: date | None,
    specs: list[_PreparedOperationSpec],
) -> dict[int, list[dict[str, Any]]]:
    if not specs:
        return {}

    return {spec.position: await _planned_effects_for_spec(db, spec) for spec in specs}


async def _query_generated_invoice_entries(
    db: AsyncSession, invoice_id: int
) -> list[AccountingEntry]:
    result = await db.execute(
        select(AccountingEntry)
        .where(
            AccountingEntry.source_type == EntrySourceType.INVOICE,
            AccountingEntry.source_id == invoice_id,
        )
        .order_by(AccountingEntry.id.asc())
    )
    return list(result.scalars().all())


async def _query_generated_payment_entries(
    db: AsyncSession, payment_id: int
) -> list[AccountingEntry]:
    result = await db.execute(
        select(AccountingEntry)
        .where(
            AccountingEntry.source_type == EntrySourceType.PAYMENT,
            AccountingEntry.source_id == payment_id,
        )
        .order_by(AccountingEntry.id.asc())
    )
    return list(result.scalars().all())


async def _query_generated_bank_entries(
    db: AsyncSession, bank_entry_id: int
) -> list[AccountingEntry]:
    result = await db.execute(
        select(AccountingEntry)
        .where(
            AccountingEntry.source_type == EntrySourceType.GESTION,
            AccountingEntry.source_id == bank_entry_id,
        )
        .order_by(AccountingEntry.id.asc())
    )
    return list(result.scalars().all())


async def _ensure_supplier_invoice_payment(
    db: AsyncSession,
    *,
    candidate: Any,
    existing_contacts_by_preview_key: dict[str, list[Any]],
) -> _SupplierOperationResult:
    invoice_result = await db.execute(
        select(Invoice).where(Invoice.number == candidate.invoice_number)
    )
    invoice = invoice_result.scalar_one_or_none()
    created_invoice = False
    created_contact = False
    contact_before = None
    contact: Contact
    invoice_before = _serialize_instance(invoice) if invoice is not None else None
    if invoice is None:
        contact_key = normalize_text(candidate.contact_name)
        matched_contact, contact_issue = resolve_invoice_contact_match(
            legacy_excel_import._invoice_row_from_supplier_candidate(candidate),
            existing_contacts_by_preview_key,
            normalize_text=normalize_text,
        )
        if contact_issue is not None:
            raise ValueError(format_row_issue(contact_issue))
        if matched_contact is None:
            contact = Contact(nom=candidate.contact_name, type=ContactType.FOURNISSEUR)
            db.add(contact)
            await db.flush()
            existing_contacts_by_preview_key.setdefault(contact_key, []).append(contact)
            created_contact = True
        else:
            contact = cast(Contact, matched_contact)
            if contact.type == ContactType.CLIENT:
                contact_before = _serialize_instance(contact)
                contact.type = ContactType.LES_DEUX
        invoice = Invoice(
            number=candidate.invoice_number,
            type=InvoiceType.FOURNISSEUR,
            contact_id=contact.id,
            date=candidate.invoice_date,
            due_date=apply_default_due_date(
                candidate.invoice_date,
                None,
                await settings_service.get_default_invoice_due_days(db),
            ),
            total_amount=candidate.amount,
            paid_amount=Decimal("0"),
            status=InvoiceStatus.SENT,
            label=InvoiceLabel(candidate.label),
            description=candidate.description,
            reference=candidate.invoice_number,
        )
        db.add(invoice)
        await db.flush()
        created_invoice = True
    else:
        contact = cast(Contact, await db.get(Contact, invoice.contact_id))
        assert contact is not None
    payment = Payment(
        invoice_id=invoice.id,
        contact_id=invoice.contact_id,
        amount=candidate.amount,
        date=candidate.invoice_date,
        method=PaymentMethod(candidate.payment_method),
        reference=candidate.invoice_number,
        notes=candidate.description,
    )
    db.add(payment)
    await db.flush()
    return _SupplierOperationResult(
        contact=contact,
        contact_before=contact_before,
        invoice=invoice,
        invoice_before=invoice_before,
        payment=payment,
        created_contact=created_contact,
        created_invoice=created_invoice,
    )


async def _execute_contact_row_import(db: AsyncSession, operation: ImportOperation) -> None:
    payload: dict[str, Any] = _deserialize_json(operation.payload_json, {})
    row = _payload_to_contact_row(payload["row"])
    existing_contacts = await load_existing_contacts_by_preview_key(db)
    preview_key = contact_preview_key(row.nom, row.prenom)
    if existing_contacts.get(preview_key):
        raise ValueError("Le contact préparé existe déjà")
    contact = Contact(
        nom=row.nom,
        prenom=row.prenom,
        email=row.email,
        type=ContactType.CLIENT,
    )
    db.add(contact)
    await db.flush()
    await _record_effect(
        db,
        operation=operation,
        position=1,
        entity_type="contact",
        action=ImportEffectAction.CREATE,
        before_snapshot=None,
        after_snapshot=_serialize_instance(contact),
    )


async def _execute_client_invoice_row_import(db: AsyncSession, operation: ImportOperation) -> None:
    payload: dict[str, Any] = _deserialize_json(operation.payload_json, {})
    row = _payload_to_invoice_row(payload["row"])
    invoice_number = str(payload["invoice_number"])
    existing_invoice = await db.execute(select(Invoice).where(Invoice.number == invoice_number))
    if existing_invoice.scalar_one_or_none() is not None:
        raise ValueError("La facture préparée existe déjà")
    existing_contacts_by_key = await load_existing_contacts_by_preview_key(db)
    matched_contact, contact_issue = resolve_invoice_contact_match(
        row,
        existing_contacts_by_key,
        normalize_text=normalize_text,
    )
    if contact_issue is not None:
        raise ValueError(format_row_issue(contact_issue))
    created_contact: Contact | None = None
    if matched_contact is None:
        contact_nom, contact_prenom = split_contact_full_name(row.contact_name)
        matched_contact = Contact(
            nom=contact_nom,
            prenom=contact_prenom,
            type=ContactType.CLIENT,
        )
        db.add(matched_contact)
        await db.flush()
        created_contact = matched_contact
    invoice_lines = legacy_excel_import._build_client_invoice_lines_from_import_row(row)
    derived_label = derive_client_invoice_label(
        {line["line_type"] for line in invoice_lines if line["amount"] > 0}
    )
    invoice = Invoice(
        number=invoice_number,
        type=InvoiceType.CLIENT,
        contact_id=matched_contact.id,
        date=row.invoice_date,
        due_date=apply_default_due_date(
            row.invoice_date,
            None,
            await settings_service.get_default_invoice_due_days(db),
        ),
        total_amount=row.amount,
        paid_amount=Decimal("0"),
        status=InvoiceStatus.SENT,
        label=derived_label,
        has_explicit_breakdown=len(invoice_lines) > 1,
    )
    db.add(invoice)
    await db.flush()
    db.add_all(
        [
            InvoiceLine(
                invoice_id=invoice.id,
                description=line["description"],
                line_type=line["line_type"],
                quantity=Decimal("1"),
                unit_price=line["amount"],
                amount=line["amount"],
            )
            for line in invoice_lines
        ]
    )
    await db.flush()
    refreshed_invoice_result = await db.execute(
        select(Invoice).where(Invoice.id == invoice.id).options(selectinload(Invoice.lines))
    )
    invoice = refreshed_invoice_result.scalar_one()
    await generate_entries_for_invoice(db, invoice)
    await db.flush()

    effect_position = 1
    if created_contact is not None:
        await _record_effect(
            db,
            operation=operation,
            position=effect_position,
            entity_type="contact",
            action=ImportEffectAction.CREATE,
            before_snapshot=None,
            after_snapshot=_serialize_instance(created_contact),
        )
        effect_position += 1
    await _record_effect(
        db,
        operation=operation,
        position=effect_position,
        entity_type="invoice",
        action=ImportEffectAction.CREATE,
        before_snapshot=None,
        after_snapshot=_serialize_instance(invoice),
    )
    effect_position += 1
    for line in invoice.lines:
        await _record_effect(
            db,
            operation=operation,
            position=effect_position,
            entity_type="invoice_line",
            action=ImportEffectAction.CREATE,
            before_snapshot=None,
            after_snapshot=_serialize_instance(line),
        )
        effect_position += 1
    for entry in await _query_generated_invoice_entries(db, invoice.id):
        await _record_effect(
            db,
            operation=operation,
            position=effect_position,
            entity_type="accounting_entry",
            action=ImportEffectAction.CREATE,
            before_snapshot=None,
            after_snapshot=_serialize_instance(entry),
        )
        effect_position += 1


async def _execute_client_payment_row_import(db: AsyncSession, operation: ImportOperation) -> None:
    payload: dict[str, Any] = _deserialize_json(operation.payload_json, {})
    row = _payload_to_payment_row(payload["row"])
    resolution = await resolve_payment_match_with_database(db, row)
    blocking_issue = make_payment_resolution_issue(
        source_row_number=row.source_row_number,
        status=resolution.status,
        candidate=resolution.candidate,
        message=resolution.message,
        require_persistable_candidate=True,
    )
    if (
        blocking_issue is not None
        or resolution.candidate is None
        or resolution.candidate.invoice_id is None
    ):
        raise ValueError(
            format_row_issue(blocking_issue)
            if blocking_issue is not None
            else "Paiement introuvable"
        )

    payment_signature = legacy_excel_import._payment_signature(
        invoice_id=resolution.candidate.invoice_id,
        payment_date=row.payment_date,
        amount=row.amount,
        method=row.method,
        cheque_number=row.cheque_number,
    )
    existing_payment_signatures = await legacy_excel_import._load_existing_payment_signatures(db)
    if payment_signature in existing_payment_signatures:
        raise ValueError("Le paiement préparé existe déjà")

    invoice = await db.get(Invoice, resolution.candidate.invoice_id)
    if invoice is None:
        raise ValueError("Facture de paiement introuvable")
    invoice_before = _serialize_instance(invoice)

    payment = Payment(
        invoice_id=resolution.candidate.invoice_id,
        contact_id=resolution.candidate.contact_id,
        date=row.payment_date,
        amount=row.amount,
        method=row.method,
        cheque_number=row.cheque_number,
        deposited=row.deposited,
        deposit_date=row.deposit_date,
    )
    db.add(payment)
    await db.flush()
    await _refresh_invoice_status(db, invoice.id)
    await db.flush()
    await generate_entries_for_payment(
        db, payment, resolution.candidate.invoice_type or InvoiceType.CLIENT
    )
    await db.flush()
    refreshed_invoice = await db.get(Invoice, invoice.id)
    assert refreshed_invoice is not None
    await db.refresh(refreshed_invoice)
    effect_position = 1
    await _record_effect(
        db,
        operation=operation,
        position=effect_position,
        entity_type="payment",
        action=ImportEffectAction.CREATE,
        before_snapshot=None,
        after_snapshot=_serialize_instance(payment),
    )
    effect_position += 1
    for entry in await _query_generated_payment_entries(db, payment.id):
        await _record_effect(
            db,
            operation=operation,
            position=effect_position,
            entity_type="accounting_entry",
            action=ImportEffectAction.CREATE,
            before_snapshot=None,
            after_snapshot=_serialize_instance(entry),
        )
        effect_position += 1
    await _record_effect(
        db,
        operation=operation,
        position=effect_position,
        entity_type="invoice",
        action=ImportEffectAction.UPDATE,
        before_snapshot=invoice_before,
        after_snapshot=_serialize_instance(refreshed_invoice),
    )


async def _execute_salary_month_import(db: AsyncSession, operation: ImportOperation) -> None:
    payload: dict[str, Any] = _deserialize_json(operation.payload_json, {})
    rows = [_payload_to_salary_row(row_payload) for row_payload in payload.get("rows", [])]
    existing_contacts_by_key = await legacy_excel_import._load_existing_contacts_by_salary_key(db)
    existing_salary_keys = await legacy_excel_import._load_existing_salary_keys(db)
    created_contacts: list[Contact] = []
    created_salaries: list[Salary] = []
    touched_months: set[str] = set()

    for row in rows:
        employee_key = legacy_excel_import._salary_employee_key(row.employee_name)
        contact = existing_contacts_by_key.get(employee_key)
        if contact is None:
            contact = Contact(nom=row.employee_name, type=ContactType.FOURNISSEUR)
            db.add(contact)
            await db.flush()
            created_contacts.append(contact)
            existing_contacts_by_key[employee_key] = contact
        salary_key = (row.month, employee_key)
        if salary_key in existing_salary_keys:
            raise ValueError(f"Salaire déjà présent pour {row.employee_name} {row.month}")
        salary = Salary(
            employee_id=contact.id,
            month=row.month,
            hours=row.hours,
            gross=row.gross,
            employee_charges=row.employee_charges,
            employer_charges=row.employer_charges,
            tax=row.tax,
            net_pay=row.net_pay,
            notes="Imported from Gestion Excel",
        )
        db.add(salary)
        created_salaries.append(salary)
        existing_salary_keys.add(salary_key)
        touched_months.add(row.month)
    await db.flush()

    created_entries: list[AccountingEntry] = []
    existing_group_keys = {
        group_key
        for (group_key,) in (
            await db.execute(
                select(AccountingEntry.group_key).where(
                    AccountingEntry.source_type == EntrySourceType.SALARY,
                    AccountingEntry.group_key.is_not(None),
                )
            )
        ).all()
        if group_key
    }
    for month in sorted(touched_months):
        if any(
            group_key.startswith(f"salary-import:{month}:") for group_key in existing_group_keys
        ):
            raise ValueError(f"Écritures de salaire déjà présentes pour {month}")
        month_rows = [row for row in rows if row.month == month]
        entry_date = legacy_excel_import._salary_entry_date(month)
        fiscal_year_id = await find_fiscal_year_id_for_date(db, entry_date)
        for group_kind, group_lines in zip(
            ("accrual", "payment"),
            legacy_excel_import._salary_month_group_lines(month, month_rows),
            strict=True,
        ):
            group_key = f"salary-import:{month}:{group_kind}"
            for account_number, label, debit, credit in group_lines:
                if debit <= 0 and credit <= 0:
                    continue
                entry_suffix = "acr" if group_kind == "accrual" else "pay"
                entry = AccountingEntry(
                    entry_number=(f"SAL-{month}-{entry_suffix}-{len(created_entries) + 1:02d}"),
                    date=entry_date,
                    account_number=account_number,
                    label=label,
                    debit=debit,
                    credit=credit,
                    fiscal_year_id=fiscal_year_id,
                    source_type=EntrySourceType.SALARY,
                    source_id=None,
                    group_key=group_key,
                )
                db.add(entry)
                created_entries.append(entry)
            existing_group_keys.add(group_key)
    await db.flush()

    effect_position = 1
    for contact in created_contacts:
        await _record_effect(
            db,
            operation=operation,
            position=effect_position,
            entity_type="contact",
            action=ImportEffectAction.CREATE,
            before_snapshot=None,
            after_snapshot=_serialize_instance(contact),
        )
        effect_position += 1
    for salary in created_salaries:
        await _record_effect(
            db,
            operation=operation,
            position=effect_position,
            entity_type="salary",
            action=ImportEffectAction.CREATE,
            before_snapshot=None,
            after_snapshot=_serialize_instance(salary),
        )
        effect_position += 1
    for entry in created_entries:
        await _record_effect(
            db,
            operation=operation,
            position=effect_position,
            entity_type="accounting_entry",
            action=ImportEffectAction.CREATE,
            before_snapshot=None,
            after_snapshot=_serialize_instance(entry),
        )
        effect_position += 1


async def _execute_cash_row_import(db: AsyncSession, operation: ImportOperation) -> None:
    payload: dict[str, Any] = _deserialize_json(operation.payload_json, {})
    row = _payload_to_cash_row(payload["row"])
    created_entries: list[AccountingEntry] = []
    effect_position = 1
    existing_contacts_by_key = await load_existing_contacts_by_preview_key(db)
    supplier_candidate = legacy_excel_import._supplier_invoice_candidate_from_cash_row(row)
    supplier_result: _SupplierOperationResult | None = None
    refreshed_invoice: Invoice | None = None
    if supplier_candidate is not None:
        supplier_result = await _ensure_supplier_invoice_payment(
            db,
            candidate=supplier_candidate,
            existing_contacts_by_preview_key=existing_contacts_by_key,
        )
    cash_entry = CashRegister(
        date=row.entry_date,
        amount=row.amount,
        type=CashMovementType(row.movement_type),
        contact_id=(supplier_result.invoice.contact_id if supplier_result is not None else None),
        payment_id=(supplier_result.payment.id if supplier_result is not None else None),
        reference=row.reference,
        description=row.description,
        source=CashEntrySource.MANUAL,
        balance_after=Decimal("0"),
    )
    db.add(cash_entry)
    await db.flush()
    await recompute_cash_balances(db)
    await db.flush()

    if supplier_result is not None:
        if supplier_result.created_invoice:
            created_entries.extend(
                await _query_generated_invoice_entries(db, supplier_result.invoice.id)
            )
        await _refresh_invoice_status(db, supplier_result.invoice.id)
        await db.flush()
        refreshed_invoice = await db.get(Invoice, supplier_result.invoice.id)
        created_entries.extend(
            await _query_generated_payment_entries(db, supplier_result.payment.id)
        )
        if not created_entries:
            if supplier_result.created_invoice:
                created_entries.extend(
                    await generate_entries_for_invoice(db, supplier_result.invoice)
                )
            created_entries.extend(
                await generate_entries_for_payment(
                    db, supplier_result.payment, InvoiceType.FOURNISSEUR
                )
            )
            await db.flush()
            created_entries = []
            if supplier_result.created_invoice:
                created_entries.extend(
                    await _query_generated_invoice_entries(db, supplier_result.invoice.id)
                )
            created_entries.extend(
                await _query_generated_payment_entries(db, supplier_result.payment.id)
            )

    if refreshed_invoice is not None:
        await db.refresh(refreshed_invoice)

    if (
        supplier_result is not None
        and supplier_result.created_contact
        and supplier_result.contact is not None
    ):
        await _record_effect(
            db,
            operation=operation,
            position=effect_position,
            entity_type="contact",
            action=ImportEffectAction.CREATE,
            before_snapshot=None,
            after_snapshot=_serialize_instance(supplier_result.contact),
        )
        effect_position += 1
    elif (
        supplier_result is not None
        and supplier_result.contact_before is not None
        and supplier_result.contact is not None
    ):
        await _record_effect(
            db,
            operation=operation,
            position=effect_position,
            entity_type="contact",
            action=ImportEffectAction.UPDATE,
            before_snapshot=supplier_result.contact_before,
            after_snapshot=_serialize_instance(supplier_result.contact),
        )
        effect_position += 1
    if supplier_result is not None and supplier_result.created_invoice:
        await _record_effect(
            db,
            operation=operation,
            position=effect_position,
            entity_type="invoice",
            action=ImportEffectAction.CREATE,
            before_snapshot=None,
            after_snapshot=_serialize_instance(supplier_result.invoice),
        )
        effect_position += 1
    elif (
        supplier_result is not None
        and supplier_result.invoice_before is not None
        and refreshed_invoice is not None
    ):
        await _record_effect(
            db,
            operation=operation,
            position=effect_position,
            entity_type="invoice",
            action=ImportEffectAction.UPDATE,
            before_snapshot=supplier_result.invoice_before,
            after_snapshot=_serialize_instance(refreshed_invoice),
        )
        effect_position += 1
    if supplier_result is not None:
        await _record_effect(
            db,
            operation=operation,
            position=effect_position,
            entity_type="payment",
            action=ImportEffectAction.CREATE,
            before_snapshot=None,
            after_snapshot=_serialize_instance(supplier_result.payment),
        )
        effect_position += 1
    await _record_effect(
        db,
        operation=operation,
        position=effect_position,
        entity_type="cash_register",
        action=ImportEffectAction.CREATE,
        before_snapshot=None,
        after_snapshot=_serialize_instance(cash_entry),
    )
    effect_position += 1
    for entry in created_entries:
        await _record_effect(
            db,
            operation=operation,
            position=effect_position,
            entity_type="accounting_entry",
            action=ImportEffectAction.CREATE,
            before_snapshot=None,
            after_snapshot=_serialize_instance(entry),
        )
        effect_position += 1


async def _execute_bank_row_import(db: AsyncSession, operation: ImportOperation) -> None:
    payload: dict[str, Any] = _deserialize_json(operation.payload_json, {})
    row = _payload_to_bank_row(payload["row"])
    existing_contacts_by_key = await load_existing_contacts_by_preview_key(db)
    created_entries: list[AccountingEntry] = []
    effect_position = 1
    supplier_result: _SupplierOperationResult | None = None
    refreshed_invoice: Invoice | None = None
    created_payment: Payment | None = None
    client_invoice_before: dict[str, Any] | None = None

    supplier_candidate = legacy_excel_import._supplier_invoice_candidate_from_bank_row(row)
    if supplier_candidate is not None:
        supplier_result = await _ensure_supplier_invoice_payment(
            db,
            candidate=supplier_candidate,
            existing_contacts_by_preview_key=existing_contacts_by_key,
        )
    elif row.amount > 0:
        invoice_reference = legacy_excel_import._single_client_invoice_reference(
            row.reference, row.description
        )
        if invoice_reference is not None:
            resolution = await resolve_payment_match_with_database(
                db,
                legacy_excel_import._payment_row_from_bank_row(
                    row, invoice_reference=invoice_reference
                ),
            )
            blocking_issue = make_payment_resolution_issue(
                source_row_number=row.source_row_number,
                status=resolution.status,
                candidate=resolution.candidate,
                message=resolution.message,
                require_persistable_candidate=True,
            )
            if blocking_issue is not None:
                raise ValueError(format_row_issue(blocking_issue))
            assert resolution.candidate is not None and resolution.candidate.invoice_id is not None
            client_invoice = await db.get(Invoice, resolution.candidate.invoice_id)
            if client_invoice is None:
                raise ValueError("Facture client introuvable")
            client_invoice_before = _serialize_instance(client_invoice)
            payment_signature = legacy_excel_import._payment_signature(
                invoice_id=resolution.candidate.invoice_id,
                payment_date=row.entry_date,
                amount=row.amount,
                method=PaymentMethod.VIREMENT.value,
            )
            if payment_signature in await legacy_excel_import._load_existing_payment_signatures(db):
                raise ValueError("Le virement préparé existe déjà")
            created_payment = Payment(
                invoice_id=resolution.candidate.invoice_id,
                contact_id=resolution.candidate.contact_id,
                date=row.entry_date,
                amount=row.amount,
                method=PaymentMethod.VIREMENT,
                reference=invoice_reference,
                notes=row.description,
                deposited=True,
                deposit_date=row.entry_date,
            )
            db.add(created_payment)
            await db.flush()

    bank_entry = await create_bank_transaction_record(
        db,
        date=row.entry_date,
        amount=row.amount,
        reference=row.reference,
        description=row.description,
        source=BankTransactionSource.IMPORT,
    )
    await db.flush()

    if supplier_result is not None:
        if supplier_result.created_invoice:
            await generate_entries_for_invoice(db, supplier_result.invoice)
        await _refresh_invoice_status(db, supplier_result.invoice.id)
        await db.flush()
        refreshed_invoice = await db.get(Invoice, supplier_result.invoice.id)
        await generate_entries_for_payment(db, supplier_result.payment, InvoiceType.FOURNISSEUR)
        await db.flush()
        if supplier_result.created_invoice:
            created_entries.extend(
                await _query_generated_invoice_entries(db, supplier_result.invoice.id)
            )
        created_entries.extend(
            await _query_generated_payment_entries(db, supplier_result.payment.id)
        )
    elif created_payment is not None:
        await _refresh_invoice_status(db, created_payment.invoice_id)
        await db.flush()
        refreshed_invoice = await db.get(Invoice, created_payment.invoice_id)
        await generate_entries_for_payment(db, created_payment, InvoiceType.CLIENT)
        await db.flush()
        created_entries.extend(await _query_generated_payment_entries(db, created_payment.id))

    created_entries.extend(
        await legacy_excel_import._generate_direct_bank_entries(
            db, bank_row=row, bank_entry=bank_entry
        )
    )
    await db.flush()

    if refreshed_invoice is not None:
        await db.refresh(refreshed_invoice)

    created_entries = await _query_generated_bank_entries(db, bank_entry.id) + [
        entry
        for entry in created_entries
        if getattr(entry, "id", None) is not None and entry.source_id != bank_entry.id
    ]

    if (
        supplier_result is not None
        and supplier_result.created_contact
        and supplier_result.contact is not None
    ):
        await _record_effect(
            db,
            operation=operation,
            position=effect_position,
            entity_type="contact",
            action=ImportEffectAction.CREATE,
            before_snapshot=None,
            after_snapshot=_serialize_instance(supplier_result.contact),
        )
        effect_position += 1
    elif (
        supplier_result is not None
        and supplier_result.contact_before is not None
        and supplier_result.contact is not None
    ):
        await _record_effect(
            db,
            operation=operation,
            position=effect_position,
            entity_type="contact",
            action=ImportEffectAction.UPDATE,
            before_snapshot=supplier_result.contact_before,
            after_snapshot=_serialize_instance(supplier_result.contact),
        )
        effect_position += 1
    if supplier_result is not None and supplier_result.created_invoice:
        await _record_effect(
            db,
            operation=operation,
            position=effect_position,
            entity_type="invoice",
            action=ImportEffectAction.CREATE,
            before_snapshot=None,
            after_snapshot=_serialize_instance(supplier_result.invoice),
        )
        effect_position += 1
    elif (
        supplier_result is not None
        and supplier_result.invoice_before is not None
        and refreshed_invoice is not None
    ):
        await _record_effect(
            db,
            operation=operation,
            position=effect_position,
            entity_type="invoice",
            action=ImportEffectAction.UPDATE,
            before_snapshot=supplier_result.invoice_before,
            after_snapshot=_serialize_instance(refreshed_invoice),
        )
        effect_position += 1
    if supplier_result is not None:
        await _record_effect(
            db,
            operation=operation,
            position=effect_position,
            entity_type="payment",
            action=ImportEffectAction.CREATE,
            before_snapshot=None,
            after_snapshot=_serialize_instance(supplier_result.payment),
        )
        effect_position += 1
    elif created_payment is not None:
        await _record_effect(
            db,
            operation=operation,
            position=effect_position,
            entity_type="payment",
            action=ImportEffectAction.CREATE,
            before_snapshot=None,
            after_snapshot=_serialize_instance(created_payment),
        )
        effect_position += 1
        if client_invoice_before is not None and refreshed_invoice is not None:
            await _record_effect(
                db,
                operation=operation,
                position=effect_position,
                entity_type="invoice",
                action=ImportEffectAction.UPDATE,
                before_snapshot=client_invoice_before,
                after_snapshot=_serialize_instance(refreshed_invoice),
            )
            effect_position += 1
    await _record_effect(
        db,
        operation=operation,
        position=effect_position,
        entity_type="bank_transaction",
        action=ImportEffectAction.CREATE,
        before_snapshot=None,
        after_snapshot=_serialize_instance(bank_entry),
    )
    effect_position += 1
    seen_entry_ids: set[int] = set()
    for entry in created_entries:
        if entry.id in seen_entry_ids:
            continue
        seen_entry_ids.add(entry.id)
        await _record_effect(
            db,
            operation=operation,
            position=effect_position,
            entity_type="accounting_entry",
            action=ImportEffectAction.CREATE,
            before_snapshot=None,
            after_snapshot=_serialize_instance(entry),
        )
        effect_position += 1


async def _execute_accounting_entry_group_import(
    db: AsyncSession, operation: ImportOperation
) -> None:
    payload: dict[str, Any] = _deserialize_json(operation.payload_json, {})
    rows = [_payload_to_entry_row(row_payload) for row_payload in payload.get("rows", [])]
    expected_creatable_count = int(payload.get("expected_creatable_count", len(rows)))
    existing_entry_signatures = await load_existing_accounting_entry_signatures(db)
    existing_manual_line_signatures = await load_existing_manual_accounting_line_signatures(db)
    creatable_rows: list[NormalizedEntryRow] = []
    for row in rows:
        signature = accounting_entry_signature(
            entry_date=row.entry_date,
            account_number=row.account_number,
            label=row.label,
            debit=row.debit,
            credit=row.credit,
        )
        if signature in existing_entry_signatures:
            continue
        line_signature = accounting_entry_line_signature(
            entry_date=row.entry_date,
            account_number=row.account_number,
            debit=row.debit,
            credit=row.credit,
        )
        if line_signature in existing_manual_line_signatures:
            continue
        creatable_rows.append(row)
    if len(creatable_rows) != expected_creatable_count:
        raise ValueError("Le groupe préparé ne correspond plus à l'état courant de la base")

    entries_to_add: list[AccountingEntry] = []
    group_key = f"import-run:{operation.id}"
    for row in creatable_rows:
        entries_to_add.append(
            AccountingEntry(
                entry_number=f"RUN-{operation.id}-{row.source_row_number}",
                date=row.entry_date,
                account_number=row.account_number,
                label=row.label,
                debit=row.debit,
                credit=row.credit,
                fiscal_year_id=await find_fiscal_year_id_for_date(db, row.entry_date),
                source_type=EntrySourceType.MANUAL,
                group_key=group_key,
            )
        )
    db.add_all(entries_to_add)
    await db.flush()
    for position, entry in enumerate(entries_to_add, start=1):
        await _record_effect(
            db,
            operation=operation,
            position=position,
            entity_type="accounting_entry",
            action=ImportEffectAction.CREATE,
            before_snapshot=None,
            after_snapshot=_serialize_instance(entry),
        )


async def _execute_client_invoice_clarification(
    db: AsyncSession, operation: ImportOperation
) -> None:
    payload: dict[str, Any] = _deserialize_json(operation.payload_json, {})
    rows = [_payload_to_entry_row(row_payload) for row_payload in payload.get("rows", [])]
    invoice = await legacy_excel_import._find_clarifiable_existing_client_invoice(db, rows)
    if invoice is None:
        raise ValueError("La facture préparée ne peut plus être clarifiée")
    invoice_before = _serialize_instance(invoice)
    old_lines = [_serialize_instance(line) for line in invoice.lines]
    old_entries = [
        _serialize_instance(entry)
        for entry in await _query_generated_invoice_entries(db, invoice.id)
    ]
    clarified_invoice = await legacy_excel_import._clarify_existing_client_invoice_from_entries(
        db, rows
    )
    if clarified_invoice is None:
        raise ValueError("Clarification comptable impossible")
    refreshed_invoice_result = await db.execute(
        select(Invoice)
        .where(Invoice.id == clarified_invoice.id)
        .options(selectinload(Invoice.lines))
    )
    refreshed_invoice = refreshed_invoice_result.scalar_one()
    new_lines = [_serialize_instance(line) for line in refreshed_invoice.lines]
    new_entries = [
        _serialize_instance(entry)
        for entry in await _query_generated_invoice_entries(db, refreshed_invoice.id)
    ]

    effect_position = 1
    for snapshot in old_entries:
        await _record_effect(
            db,
            operation=operation,
            position=effect_position,
            entity_type="accounting_entry",
            action=ImportEffectAction.DELETE,
            before_snapshot=snapshot,
            after_snapshot=None,
        )
        effect_position += 1
    for snapshot in old_lines:
        await _record_effect(
            db,
            operation=operation,
            position=effect_position,
            entity_type="invoice_line",
            action=ImportEffectAction.DELETE,
            before_snapshot=snapshot,
            after_snapshot=None,
        )
        effect_position += 1
    await _record_effect(
        db,
        operation=operation,
        position=effect_position,
        entity_type="invoice",
        action=ImportEffectAction.UPDATE,
        before_snapshot=invoice_before,
        after_snapshot=_serialize_instance(refreshed_invoice),
    )
    effect_position += 1
    for snapshot in new_lines:
        await _record_effect(
            db,
            operation=operation,
            position=effect_position,
            entity_type="invoice_line",
            action=ImportEffectAction.CREATE,
            before_snapshot=None,
            after_snapshot=snapshot,
        )
        effect_position += 1
    for snapshot in new_entries:
        await _record_effect(
            db,
            operation=operation,
            position=effect_position,
            entity_type="accounting_entry",
            action=ImportEffectAction.CREATE,
            before_snapshot=None,
            after_snapshot=snapshot,
        )
        effect_position += 1


async def _execute_operation(db: AsyncSession, operation: ImportOperation) -> None:
    if operation.operation_type == "contact_row_import":
        await _execute_contact_row_import(db, operation)
        return
    if operation.operation_type == "client_invoice_row_import":
        await _execute_client_invoice_row_import(db, operation)
        return
    if operation.operation_type == "client_payment_row_import":
        await _execute_client_payment_row_import(db, operation)
        return
    if operation.operation_type == "salary_month_import":
        await _execute_salary_month_import(db, operation)
        return
    if operation.operation_type == "cash_row_import":
        await _execute_cash_row_import(db, operation)
        return
    if operation.operation_type == "bank_row_import":
        await _execute_bank_row_import(db, operation)
        return
    if operation.operation_type == "accounting_entry_group_import":
        await _execute_accounting_entry_group_import(db, operation)
        return
    if operation.operation_type == "client_invoice_clarification_from_entries":
        await _execute_client_invoice_clarification(db, operation)
        return
    raise ValueError(f"Unsupported operation type: {operation.operation_type}")


async def execute_import_run(db: AsyncSession, run_id: int) -> ImportRun:
    run = await get_import_run(db, run_id)
    if run is None:
        raise LookupError("Import run not found")
    if not _can_run_execute(run):
        raise ValueError("Import run is not executable")

    for operation in run.operations:
        if operation.decision != ImportOperationDecision.APPLY:
            continue
        if operation.status != ImportOperationStatus.PREPARED:
            continue
        try:
            async with db.begin_nested():
                await _execute_operation(db, operation)
                operation.status = ImportOperationStatus.APPLIED
                operation.error_message = None
        except Exception as exc:
            operation.status = ImportOperationStatus.FAILED
            operation.error_message = str(exc)
            run.status = ImportRunStatus.FAILED
            break
    else:
        run.status = ImportRunStatus.COMPLETED

    await db.flush()
    run = await get_import_run(db, run_id)
    assert run is not None
    run.summary_json = _json_dumps(_build_run_summary(run))
    await db.flush()
    run = await get_import_run(db, run_id)
    assert run is not None
    return run


async def _ensure_effect_matches_state(
    db: AsyncSession,
    *,
    effect: ImportEffect,
    expected: str,
) -> dict[str, Any] | None:
    snapshot = _deserialize_json(
        effect.after_snapshot_json if expected == "after" else effect.before_snapshot_json,
        None,
    )
    fingerprint = effect.after_fingerprint if expected == "after" else effect.before_fingerprint
    if snapshot is None:
        return None
    model_cls = _ENTITY_MODELS[effect.entity_type]
    entity = await db.get(model_cls, effect.entity_id)
    if entity is None:
        raise ValueError("Objet introuvable pour vérification stricte")
    await db.refresh(entity)
    current_snapshot = _serialize_instance(entity)
    current_fingerprint = _snapshot_fingerprint(effect.entity_type, current_snapshot)
    if current_fingerprint != fingerprint:
        expected_filtered = _filtered_snapshot_for_fingerprint(effect.entity_type, snapshot)
        current_filtered = _filtered_snapshot_for_fingerprint(effect.entity_type, current_snapshot)
        if current_filtered != expected_filtered:
            raise ValueError("L'état courant ne correspond plus à l'état attendu")
    return snapshot


async def _recreate_from_effect(db: AsyncSession, effect: ImportEffect) -> None:
    snapshot = _deserialize_json(effect.before_snapshot_json, None)
    if snapshot is None:
        raise ValueError("Aucun instantané avant pour restaurer l'objet")
    model_cls = _ENTITY_MODELS[effect.entity_type]
    if await db.get(model_cls, effect.entity_id) is not None:
        raise ValueError("L'objet à restaurer existe déjà")
    instance = _restore_instance(model_cls, snapshot)
    db.add(instance)
    await db.flush()


async def _recreate_after_snapshot(db: AsyncSession, effect: ImportEffect) -> None:
    snapshot = _deserialize_json(effect.after_snapshot_json, None)
    if snapshot is None:
        raise ValueError("Aucun instantané après pour recréer l'objet")
    model_cls = _ENTITY_MODELS[effect.entity_type]
    if await db.get(model_cls, effect.entity_id) is not None:
        raise ValueError("L'objet à recréer existe déjà")
    instance = _restore_instance(model_cls, snapshot)
    db.add(instance)
    await db.flush()


async def _delete_for_effect(db: AsyncSession, effect: ImportEffect, *, expected: str) -> None:
    await _ensure_effect_matches_state(db, effect=effect, expected=expected)
    model_cls = _ENTITY_MODELS[effect.entity_type]
    entity = await db.get(model_cls, effect.entity_id)
    if entity is None:
        raise ValueError("Objet introuvable pour suppression")
    await db.delete(entity)
    await db.flush()


async def _update_from_snapshot(
    db: AsyncSession, effect: ImportEffect, *, target: str, current: str
) -> None:
    await _ensure_effect_matches_state(db, effect=effect, expected=current)
    snapshot = _deserialize_json(
        effect.before_snapshot_json if target == "before" else effect.after_snapshot_json,
        None,
    )
    if snapshot is None:
        raise ValueError("Instantané de mise à jour manquant")
    model_cls = _ENTITY_MODELS[effect.entity_type]
    entity = await db.get(model_cls, effect.entity_id)
    if entity is None:
        raise ValueError("Objet introuvable pour mise à jour")
    _apply_snapshot(entity, snapshot)
    await db.flush()


async def _post_process_effects(db: AsyncSession, effects: list[ImportEffect]) -> None:
    invoice_ids: set[int] = set()
    recompute_cash = False
    recompute_bank = False
    for effect in effects:
        before_snapshot = _deserialize_json(effect.before_snapshot_json, None)
        after_snapshot = _deserialize_json(effect.after_snapshot_json, None)
        if effect.entity_type == "payment":
            if before_snapshot and before_snapshot.get("invoice_id") is not None:
                invoice_ids.add(int(before_snapshot["invoice_id"]))
            if after_snapshot and after_snapshot.get("invoice_id") is not None:
                invoice_ids.add(int(after_snapshot["invoice_id"]))
        elif effect.entity_type == "cash_register":
            recompute_cash = True
        elif effect.entity_type == "bank_transaction":
            recompute_bank = True
    for invoice_id in sorted(invoice_ids):
        await _refresh_invoice_status(db, invoice_id)
    if recompute_cash:
        await recompute_cash_balances(db)
    if recompute_bank:
        await recompute_bank_balances(db)
    await db.flush()


async def undo_import_operation(db: AsyncSession, operation_id: int) -> ImportOperation:
    result = await db.execute(
        select(ImportOperation)
        .where(ImportOperation.id == operation_id)
        .options(
            selectinload(ImportOperation.effects),
            selectinload(ImportOperation.run).selectinload(ImportRun.operations),
        )
    )
    operation = result.scalar_one_or_none()
    if operation is None:
        raise LookupError("Import operation not found")
    if operation.status != ImportOperationStatus.APPLIED:
        raise ValueError("Import operation cannot be undone")

    for effect in sorted(operation.effects, key=lambda item: item.position, reverse=True):
        if effect.action == ImportEffectAction.CREATE:
            await _delete_for_effect(db, effect, expected="after")
        elif effect.action == ImportEffectAction.UPDATE:
            await _update_from_snapshot(db, effect, target="before", current="after")
        elif effect.action == ImportEffectAction.DELETE:
            if await db.get(_ENTITY_MODELS[effect.entity_type], effect.entity_id) is not None:
                raise ValueError("Un objet supprimé par l'opération existe déjà")
            await _recreate_from_effect(db, effect)
        effect.status = ImportEffectStatus.UNDONE
    await _post_process_effects(db, operation.effects)
    operation.status = ImportOperationStatus.UNDONE
    operation.run.status = (
        ImportRunStatus.PARTIALLY_REVERTED
        if any(item.status == ImportOperationStatus.APPLIED for item in operation.run.operations)
        else ImportRunStatus.REVERTED
    )
    await db.flush()
    run = await get_import_run(db, operation.run_id)
    assert run is not None
    run.summary_json = _json_dumps(_build_run_summary(run))
    await db.flush()
    return operation


async def redo_import_operation(db: AsyncSession, operation_id: int) -> ImportOperation:
    result = await db.execute(
        select(ImportOperation)
        .where(ImportOperation.id == operation_id)
        .options(
            selectinload(ImportOperation.effects),
            selectinload(ImportOperation.run).selectinload(ImportRun.operations),
        )
    )
    operation = result.scalar_one_or_none()
    if operation is None:
        raise LookupError("Import operation not found")
    if operation.status != ImportOperationStatus.UNDONE:
        raise ValueError("Import operation cannot be redone")

    for effect in sorted(operation.effects, key=lambda item: item.position):
        if effect.action == ImportEffectAction.CREATE:
            await _recreate_after_snapshot(db, effect)
        elif effect.action == ImportEffectAction.UPDATE:
            await _update_from_snapshot(db, effect, target="after", current="before")
        elif effect.action == ImportEffectAction.DELETE:
            await _delete_for_effect(db, effect, expected="before")
        effect.status = ImportEffectStatus.APPLIED
    await _post_process_effects(db, operation.effects)
    operation.status = ImportOperationStatus.APPLIED
    operation.run.status = ImportRunStatus.COMPLETED
    await db.flush()
    run = await get_import_run(db, operation.run_id)
    assert run is not None
    run.summary_json = _json_dumps(_build_run_summary(run))
    await db.flush()
    return operation


async def undo_import_run(db: AsyncSession, run_id: int) -> ImportRun:
    run = await get_import_run(db, run_id)
    if run is None:
        raise LookupError("Import run not found")
    if not _run_can_undo(run):
        raise ValueError("Import run cannot be undone")
    for operation in sorted(run.operations, key=lambda item: item.position, reverse=True):
        if operation.status == ImportOperationStatus.APPLIED:
            await undo_import_operation(db, operation.id)
    run = await get_import_run(db, run_id)
    assert run is not None
    return run


async def redo_import_run(db: AsyncSession, run_id: int) -> ImportRun:
    run = await get_import_run(db, run_id)
    if run is None:
        raise LookupError("Import run not found")
    if not _run_can_redo(run):
        raise ValueError("Import run cannot be redone")
    for operation in sorted(run.operations, key=lambda item: item.position):
        if operation.status == ImportOperationStatus.UNDONE:
            await redo_import_operation(db, operation.id)
    run = await get_import_run(db, run_id)
    assert run is not None
    return run
