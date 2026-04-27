"""Integration tests for Excel import API."""

import io
import json
from datetime import date, datetime
from decimal import Decimal
from unittest.mock import patch

import pytest
from httpx import AsyncClient
from sqlalchemy import select

from backend.models.import_log import ImportLog, ImportLogStatus, ImportLogType
from backend.models.import_run import (
    ImportOperation,
    ImportOperationDecision,
    ImportOperationStatus,
    ImportRun,
    ImportRunStatus,
)

try:
    import openpyxl

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _build_import_log_summary(**overrides: object) -> dict[str, object]:
    return {
        "contacts_created": 0,
        "invoices_created": 0,
        "payments_created": 0,
        "salaries_created": 0,
        "entries_created": 0,
        "cash_created": 0,
        "bank_created": 0,
        "skipped": 0,
        "ignored_rows": 0,
        "blocked_rows": 0,
        "warnings": [],
        "errors": [],
        "warning_details": [],
        "error_details": [],
        "sheets": [],
        "created_objects": [],
        **overrides,
    }


def _make_simple_xlsx(headers: list[str], rows: list[list]) -> bytes:
    """Create a minimal in-memory xlsx with given headers and data rows."""
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.append(headers)
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_multi_sheet_xlsx(sheets: dict[str, tuple[list[str], list[list]]]) -> bytes:
    """Create an in-memory xlsx containing multiple sheets."""
    wb = openpyxl.Workbook()
    default_ws = wb.active
    assert default_ws is not None
    default_ws.title = next(iter(sheets))

    for index, (sheet_name, (headers, rows)) in enumerate(sheets.items()):
        ws = default_ws if index == 0 else wb.create_sheet(title=sheet_name)
        ws.append(headers)
        for row in rows:
            ws.append(row)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_multi_sheet_xlsx_rows(sheets: dict[str, list[list]]) -> bytes:
    """Create an in-memory xlsx containing multiple sheets with raw rows."""
    wb = openpyxl.Workbook()
    default_ws = wb.active
    assert default_ws is not None
    default_ws.title = next(iter(sheets))

    for index, (sheet_name, rows) in enumerate(sheets.items()):
        ws = default_ws if index == 0 else wb.create_sheet(title=sheet_name)
        for row in rows:
            ws.append(row)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.mark.asyncio
async def test_import_requires_auth(client: AsyncClient) -> None:
    """Import endpoints require authentication."""
    response = await client.post("/api/import/excel/gestion")
    assert response.status_code in (401, 422)


@pytest.mark.asyncio
async def test_import_history_requires_auth(client: AsyncClient) -> None:
    """Import history endpoint requires authentication."""
    response = await client.get("/api/import/history")
    assert response.status_code == 401


@pytest.mark.asyncio
async def test_import_endpoints_are_reserved_to_admin(
    client: AsyncClient, tresorier_auth_headers: dict
) -> None:
    """Import endpoints reject authenticated non-admin users."""
    response = await client.get("/api/import/history", headers=tresorier_auth_headers)
    assert response.status_code == 403


@pytest.mark.asyncio
async def test_import_history_lists_recent_logs_with_parsed_summary(
    client: AsyncClient,
    auth_headers: dict,
    db_session,
) -> None:
    """Import history should expose recent logs with parsed counters and traceability."""
    older_log = ImportLog(
        import_type=ImportLogType.GESTION,
        status=ImportLogStatus.SUCCESS,
        file_hash="a" * 64,
        file_name="Gestion 2024.xlsx",
        summary=json.dumps(
            _build_import_log_summary(
                contacts_created=1,
                warnings=["Factures — Ligne ignorée"],
                created_objects=[
                    {
                        "sheet_name": "Factures",
                        "kind": "invoices",
                        "object_type": "invoice",
                        "object_id": 12,
                        "reference": "FAC-2024-001",
                        "details": {"contact_name": "Alice Martin"},
                    }
                ],
            ),
            ensure_ascii=False,
        ),
        created_at=date(2026, 4, 18),
    )
    newer_log = ImportLog(
        import_type=ImportLogType.COMPTABILITE,
        status=ImportLogStatus.FAILED,
        file_hash="b" * 64,
        file_name="Comptabilite 2025.xlsx",
        summary=json.dumps(
            _build_import_log_summary(
                entries_created=4,
                blocked_rows=2,
                errors=["Journal — Compte manquant"],
                sheets=[
                    {
                        "name": "Journal",
                        "kind": "entries",
                        "imported_rows": 4,
                        "ignored_rows": 0,
                        "blocked_rows": 2,
                        "warnings": [],
                        "errors": ["Compte manquant"],
                    }
                ],
            ),
            ensure_ascii=False,
        ),
        created_at=date(2026, 4, 20),
    )
    db_session.add_all([older_log, newer_log])
    await db_session.commit()

    response = await client.get("/api/import/history", headers=auth_headers)

    assert response.status_code == 200
    data = response.json()
    assert [item["file_name"] for item in data] == [
        "Comptabilite 2025.xlsx",
        "Gestion 2024.xlsx",
    ]
    assert data[0]["status"] == "failed"
    assert data[0]["summary"]["entries_created"] == 4
    assert data[0]["summary"]["blocked_rows"] == 2
    assert data[0]["summary"]["errors"] == ["Journal — Compte manquant"]
    assert data[0]["summary"]["sheets"][0]["name"] == "Journal"
    assert data[1]["summary"]["contacts_created"] == 1
    assert data[1]["summary"]["created_objects"][0]["reference"] == "FAC-2024-001"


@pytest.mark.asyncio
async def test_import_history_tolerates_invalid_serialized_summary(
    client: AsyncClient,
    auth_headers: dict,
    db_session,
) -> None:
    """Import history should not fail when a legacy log contains invalid JSON."""
    db_session.add(
        ImportLog(
            import_type=ImportLogType.GESTION,
            status=ImportLogStatus.BLOCKED,
            file_hash="c" * 64,
            file_name="Gestion legacy.xlsx",
            summary="{invalid",
            created_at=date(2026, 4, 19),
        )
    )
    await db_session.commit()

    response = await client.get("/api/import/history", headers=auth_headers)

    assert response.status_code == 200
    data = response.json()
    assert data[0]["file_name"] == "Gestion legacy.xlsx"
    assert data[0]["summary"] is None


@pytest.mark.asyncio
async def test_get_import_run_exposes_source_data_for_blocked_operations(
    client: AsyncClient,
    auth_headers: dict,
    db_session,
) -> None:
    """Prepared runs should expose normalized Excel source data for blocked operations."""
    run = ImportRun(
        import_type="comptabilite",
        status=ImportRunStatus.BLOCKED,
        file_hash="d" * 64,
        file_name="Comptabilite 2025.xlsx",
        created_at=datetime(2026, 4, 20, 10, 0, 0),
        updated_at=datetime(2026, 4, 20, 10, 0, 0),
    )
    db_session.add(run)
    await db_session.flush()

    db_session.add(
        ImportOperation(
            run_id=run.id,
            position=1,
            operation_type="blocked_by_validation",
            title="Ligne 14",
            description="Compte 706200 absent du plan comptable",
            source_sheet="Journal",
            source_row_numbers_json=json.dumps([14], ensure_ascii=False),
            decision=ImportOperationDecision.BLOCK,
            status=ImportOperationStatus.BLOCKED,
            payload_json=json.dumps(
                {
                    "row": {
                        "source_row_number": 14,
                        "entry_date": "2025-09-15",
                        "account_number": "706200",
                        "label": "Cours septembre",
                        "debit": "0",
                        "credit": "85",
                    }
                },
                ensure_ascii=False,
            ),
            diagnostics_json=json.dumps(
                ["Journal — Ligne 14 : compte comptable introuvable"],
                ensure_ascii=False,
            ),
            error_message="Compte comptable introuvable",
        )
    )
    await db_session.commit()

    response = await client.get(f"/api/import/runs/{run.id}", headers=auth_headers)

    assert response.status_code == 200
    data = response.json()
    assert data["operations"][0]["source_data"] == [
        {
            "source_row_number": 14,
            "entry_date": "2025-09-15",
            "account_number": "706200",
            "label": "Cours septembre",
            "debit": "0",
            "credit": "85",
        }
    ]


@pytest.mark.asyncio
async def test_get_import_run_returns_localized_404(
    client: AsyncClient,
    auth_headers: dict,
) -> None:
    """Missing reversible runs should return a localized not-found message."""
    response = await client.get("/api/import/runs/999999", headers=auth_headers)

    assert response.status_code == 404
    assert response.json() == {"detail": "Import préparé introuvable (id : 999999)"}


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_import_gestion_empty_sheet(client: AsyncClient, auth_headers: dict) -> None:
    """Uploading an xlsx with empty valid sheet returns a result dict."""
    content = _make_simple_xlsx(["date", "montant", "nom"], [[]])
    response = await client.post(
        "/api/import/excel/gestion",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )
    assert response.status_code == 200
    data = response.json()
    assert "contacts_created" in data
    assert "invoices_created" in data
    assert "payments_created" in data
    assert "salaries_created" in data
    assert "entries_created" in data
    assert "skipped" in data
    assert "errors" in data


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_import_comptabilite_empty_sheet(client: AsyncClient, auth_headers: dict) -> None:
    """Uploading an empty comptabilité sheet returns zero counts."""
    content = _make_simple_xlsx(["date", "compte", "libellé", "débit", "crédit"], [])
    response = await client.post(
        "/api/import/excel/comptabilite",
        files={"file": ("Comptabilite 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )
    assert response.status_code == 200
    data = response.json()
    assert data["entries_created"] == 0


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_test_import_shortcuts_list_and_run_configured_file(
    client: AsyncClient,
    auth_headers: dict,
    tmp_path,
) -> None:
    """Configured test shortcuts should create and execute a reversible import run."""
    from backend.config import Settings, get_settings

    shortcut_file = tmp_path / "Gestion 2024.xlsx"
    shortcut_file.write_bytes(
        _make_multi_sheet_xlsx(
            {
                "Contacts": (
                    ["Nom", "Email"],
                    [
                        ["Christine LOPES", "christine@example.test"],
                        ["Thi BE NGUYEN", "thi@example.test"],
                    ],
                )
            }
        )
    )

    test_settings = Settings(
        jwt_secret_key=get_settings().jwt_secret_key,
        enable_test_import_shortcuts=True,
        test_import_gestion_2024_path=str(shortcut_file),
    )
    with patch("backend.routers.excel_import.get_settings", return_value=test_settings):
        shortcuts_response = await client.get(
            "/api/import/excel/test-shortcuts",
            headers=auth_headers,
        )

        assert shortcuts_response.status_code == 200
        shortcuts = {item["alias"]: item for item in shortcuts_response.json()}
        assert shortcuts["gestion-2024"]["available"] is True
        assert shortcuts["gestion-2024"]["file_name"] == "Gestion 2024.xlsx"
        assert shortcuts["gestion-2025"]["available"] is False

        import_response = await client.post(
            "/api/import/excel/test-shortcuts/gestion-2024",
            params={
                "comparison_start_date": "2024-08-01",
                "comparison_end_date": "2025-07-31",
            },
            headers=auth_headers,
        )

        assert import_response.status_code == 200
        data = import_response.json()
        assert data["kind"] == "run"
        assert data["status"] == "completed"
        assert data["comparison_start_date"] == "2024-08-01"
        assert data["comparison_end_date"] == "2025-07-31"
        assert data["summary"]["contacts_created"] == 2
        assert data["summary"]["entries_created"] == 0
        assert data["can_undo"] is True


@pytest.mark.asyncio
async def test_import_rejects_non_excel(client: AsyncClient, auth_headers: dict) -> None:
    """Uploading a non-Excel file returns 422."""
    response = await client.post(
        "/api/import/excel/gestion",
        files={"file": ("data.csv", b"nom,montant\nDupont,100", "text/csv")},
        headers=auth_headers,
    )
    assert response.status_code == 422


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_preview_gestion_rejects_invalid_comparison_date_range(
    client: AsyncClient,
    auth_headers: dict,
) -> None:
    """Preview should reject a comparison window where start_date is after end_date."""
    content = _make_multi_sheet_xlsx(
        {
            "Factures": (
                ["Date facture", "Réf facture", "Client", "Montant"],
                [["2025-08-01", "2025-0142", "Christine LOPES", 55]],
            ),
        }
    )

    response = await client.post(
        "/api/import/excel/gestion/preview",
        data={
            "comparison_start_date": "2025-08-02",
            "comparison_end_date": "2025-08-01",
        },
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert response.status_code == 422
    assert response.json() == {
        "detail": "La date de début doit être inférieure ou égale à la date de fin"
    }


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_preview_comptabilite_rejects_invalid_comparison_date_range(
    client: AsyncClient,
    auth_headers: dict,
) -> None:
    """Comptabilité preview should reject a comparison window where start_date is after end_date."""
    content = _make_multi_sheet_xlsx(
        {
            "Journal": (
                ["Date", "N° compte", "Libellé de l'écriture", "Débit", "Crédit"],
                [["2025-08-01", "401100", "Facture fournisseur", 10, None]],
            ),
        }
    )

    response = await client.post(
        "/api/import/excel/comptabilite/preview",
        data={
            "comparison_start_date": "2025-08-02",
            "comparison_end_date": "2025-08-01",
        },
        files={"file": ("Comptabilite 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert response.status_code == 422
    assert response.json() == {
        "detail": "La date de début doit être inférieure ou égale à la date de fin"
    }


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_preview_gestion_reports_auxiliary_sheets(
    client: AsyncClient, auth_headers: dict
) -> None:
    """Preview should distinguish recognized sheets from ignored helper sheets."""
    content = _make_multi_sheet_xlsx(
        {
            "Factures": (
                [
                    "Date facture",
                    "Réf facture",
                    "Client",
                    "Montant",
                ],
                [["2025-08-01", "2025-0142", "Christine LOPES", 55]],
            ),
            "Aide - Factures": (
                [
                    "Date",
                    "Numéro",
                    "Nom",
                    "Montant",
                    "N° compte",
                    "Débit",
                    "Crédit",
                ],
                [["2025-08-01", "AIDE-001", "Aide", 55, 411100, 55, None]],
            ),
            "TODO": (
                ["Date", "Factures", "Banque", "Paiements"],
                [["2025-08-31", "DONE", "DONE", "DONE"]],
            ),
        }
    )

    response = await client.post(
        "/api/import/excel/gestion/preview",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert response.status_code == 200
    data = response.json()
    assert data["can_import"] is True
    assert "warnings" in data

    sheets = {sheet["name"]: sheet for sheet in data["sheets"]}
    assert sheets["Factures"]["status"] == "recognized"
    assert sheets["Factures"]["kind"] == "invoices"
    assert sheets["Factures"]["rows"] == 1
    assert sheets["Aide - Factures"]["status"] == "ignored"
    assert sheets["TODO"]["status"] == "ignored"
    assert any("Aide - Factures" in warning for warning in data["warnings"])


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_preview_gestion_estimates_contacts_from_invoices_and_payments(
    client: AsyncClient, auth_headers: dict
) -> None:
    """Preview should estimate candidate contacts from invoice and payment sheets."""
    content = _make_multi_sheet_xlsx(
        {
            "Factures": (
                ["Date facture", "Réf facture", "Client", "Montant"],
                [
                    ["2025-08-01", "2025-0142", "Christine LOPES", 55],
                    ["2025-08-02", "2025-0143", "Thi BE NGUYEN", 154],
                ],
            ),
            "Paiements": (
                ["Réf facture", "Adhérent", "Montant", "Date paiement"],
                [
                    ["2025-0142", "Christine LOPES", 55, "2025-08-02"],
                    ["2025-0144", "Franck LEVY", 208, "2025-08-03"],
                ],
            ),
        }
    )

    response = await client.post(
        "/api/import/excel/gestion/preview",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert response.status_code == 200
    data = response.json()
    assert data["estimated_invoices"] == 2
    assert data["estimated_payments"] == 1
    assert data["estimated_contacts"] == 3


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_preview_gestion_comparison_summarizes_existing_missing_and_ignored_invoices(
    client: AsyncClient,
    auth_headers: dict,
    db_session,
) -> None:
    from backend.models.contact import Contact, ContactType
    from backend.models.invoice import Invoice, InvoiceStatus, InvoiceType

    contact = Contact(nom="LOPES", prenom="Christine", type=ContactType.CLIENT)
    db_session.add(contact)
    await db_session.flush()
    db_session.add(
        Invoice(
            number="2025-0142",
            type=InvoiceType.CLIENT,
            contact_id=contact.id,
            date=date(2025, 8, 1),
            total_amount=Decimal("55.00"),
            paid_amount=Decimal("0.00"),
            status=InvoiceStatus.DRAFT,
        )
    )
    await db_session.commit()

    content = _make_multi_sheet_xlsx(
        {
            "Factures": (
                ["Date facture", "Réf facture", "Client", "Montant"],
                [
                    ["2025-08-01", "2025-0142", "Christine LOPES", 55],
                    ["2025-08-02", "2025-0143", "Christine LOPES", 75],
                    ["2025-08-02", "2025-0143", "Christine LOPES", 75],
                ],
            ),
        }
    )

    response = await client.post(
        "/api/import/excel/gestion/preview",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert response.status_code == 200
    data = response.json()
    assert data["comparison"]["mode"] == "gestion-excel-to-solde"
    domains = {domain["kind"]: domain for domain in data["comparison"]["domains"]}
    assert domains["invoices"] == {
        "kind": "invoices",
        "file_rows": 3,
        "already_in_solde": 1,
        "missing_in_solde": 1,
        "extra_in_solde": 0,
        "ignored_by_policy": 1,
        "blocked": 0,
    }


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_prepare_run_keeps_source_data_for_ignored_operations(
    client: AsyncClient,
    auth_headers: dict,
    db_session,
) -> None:
    from backend.models.contact import Contact, ContactType
    from backend.models.invoice import Invoice, InvoiceStatus, InvoiceType

    contact = Contact(nom="LOPES", prenom="Christine", type=ContactType.CLIENT)
    db_session.add(contact)
    await db_session.flush()
    db_session.add(
        Invoice(
            number="2025-0142",
            type=InvoiceType.CLIENT,
            contact_id=contact.id,
            date=date(2025, 8, 1),
            total_amount=Decimal("55.00"),
            paid_amount=Decimal("0.00"),
            status=InvoiceStatus.DRAFT,
        )
    )
    await db_session.commit()

    content = _make_multi_sheet_xlsx(
        {
            "Factures": (
                ["Date facture", "Réf facture", "Client", "Montant"],
                [
                    ["2025-08-01", "2025-0142", "Christine LOPES", 55],
                    ["2025-08-02", "2025-0143", "Christine LOPES", 75],
                    ["2025-08-02", "2025-0143", "Christine LOPES", 75],
                ],
            ),
        }
    )

    response = await client.post(
        "/api/import/runs/prepare/gestion",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert response.status_code == 200
    data = response.json()
    ignored_operations = [
        operation for operation in data["operations"] if operation["decision"] == "ignore"
    ]
    assert ignored_operations
    assert all(operation["source_data"] for operation in ignored_operations)
    assert any(
        source_row.get("invoice_number") == "2025-0142"
        for operation in ignored_operations
        for source_row in operation["source_data"]
    )
    assert any(
        source_row.get("Réf facture") == "2025-0143"
        for operation in ignored_operations
        for source_row in operation["source_data"]
    )


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_reversible_prepare_accepts_payment_matched_against_planned_invoice(
    client: AsyncClient,
    auth_headers: dict,
) -> None:
    """Prepared runs should accept payments matched to invoices planned in the same workbook."""
    content = _make_multi_sheet_xlsx(
        {
            "Paiements": (
                ["Réf facture", "Adhérent", "Montant", "Date paiement"],
                [["2025-0142", "Christine LOPES", 55, "2025-08-02"]],
            ),
            "Factures": (
                ["Date facture", "Réf facture", "Client", "Montant"],
                [["2025-08-01", "2025-0142", "Christine LOPES", 55]],
            ),
        }
    )

    prepare_response = await client.post(
        "/api/import/runs/prepare/gestion",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert prepare_response.status_code == 200
    prepared_run = prepare_response.json()
    assert prepared_run["status"] == "prepared"
    assert prepared_run["can_execute"] is True
    assert [operation["operation_type"] for operation in prepared_run["operations"]] == [
        "client_invoice_row_import",
        "client_payment_row_import",
    ]
    assert all(operation["decision"] == "apply" for operation in prepared_run["operations"])

    execute_response = await client.post(
        f"/api/import/runs/{prepared_run['id']}/execute",
        headers=auth_headers,
    )

    assert execute_response.status_code == 200
    executed_run = execute_response.json()
    assert executed_run["status"] == "completed"
    assert executed_run["summary"]["contacts_created"] == 1
    assert executed_run["summary"]["invoices_created"] == 1
    assert executed_run["summary"]["payments_created"] == 1
    assert all(operation["status"] == "applied" for operation in executed_run["operations"])


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_reversible_execute_applies_default_due_date_to_imported_invoice(
    client: AsyncClient,
    auth_headers: dict,
    db_session,
) -> None:
    from sqlalchemy import select

    from backend.models.invoice import Invoice

    settings_response = await client.put(
        "/api/settings/",
        json={"default_invoice_due_days": 15},
        headers=auth_headers,
    )
    assert settings_response.status_code == 200

    content = _make_multi_sheet_xlsx(
        {
            "Factures": (
                ["Date facture", "Réf facture", "Client", "Montant"],
                [["2025-08-01", "2025-0143", "Christine LOPES", 55]],
            ),
        }
    )

    prepare_response = await client.post(
        "/api/import/runs/prepare/gestion",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )
    assert prepare_response.status_code == 200

    execute_response = await client.post(
        f"/api/import/runs/{prepare_response.json()['id']}/execute",
        headers=auth_headers,
    )
    assert execute_response.status_code == 200

    invoice = (
        (await db_session.execute(select(Invoice).where(Invoice.number == "2025-0143")))
        .scalars()
        .one()
    )
    assert invoice.due_date == date(2025, 8, 16)


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_reversible_prepare_accepts_distinct_cheques_same_invoice_amount_and_date(
    client: AsyncClient,
    auth_headers: dict,
    db_session,
) -> None:
    """Prepared runs should keep distinct cheque rows when only the cheque number differs."""
    from backend.models.payment import Payment

    content = _make_multi_sheet_xlsx(
        {
            "Factures": (
                ["Date facture", "Réf facture", "Client", "Montant"],
                [["2025-05-02", "2025-0072", "Christine LOPES", 60]],
            ),
            "Paiements": (
                [
                    "Date",
                    "Montant",
                    "Référence facture",
                    "Client",
                    "Mode",
                    "Numéro chèque",
                    "Déposé",
                    "Date dépôt",
                ],
                [
                    [
                        "2025-05-02",
                        30,
                        "2025-0072",
                        "Christine LOPES",
                        "Chèque",
                        "2025.05.02.05",
                        False,
                        None,
                    ],
                    [
                        "2025-05-02",
                        30,
                        "2025-0072",
                        "Christine LOPES",
                        "Chèque",
                        "2025.05.02.06",
                        False,
                        None,
                    ],
                ],
            ),
        }
    )

    prepare_response = await client.post(
        "/api/import/runs/prepare/gestion",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert prepare_response.status_code == 200
    prepared_run = prepare_response.json()
    assert prepared_run["status"] == "prepared"
    assert prepared_run["can_execute"] is True
    assert [operation["operation_type"] for operation in prepared_run["operations"]] == [
        "client_invoice_row_import",
        "client_payment_row_import",
        "client_payment_row_import",
    ]

    execute_response = await client.post(
        f"/api/import/runs/{prepared_run['id']}/execute",
        headers=auth_headers,
    )

    assert execute_response.status_code == 200
    executed_run = execute_response.json()
    assert executed_run["status"] == "completed"
    assert executed_run["summary"]["payments_created"] == 2

    payments = list(
        (await db_session.execute(select(Payment).order_by(Payment.cheque_number))).scalars()
    )
    assert [payment.cheque_number for payment in payments] == [
        "2025.05.02.05",
        "2025.05.02.06",
    ]


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_preview_gestion_comparison_counts_blocked_unmatched_payments(
    client: AsyncClient,
    auth_headers: dict,
) -> None:
    content = _make_multi_sheet_xlsx(
        {
            "Paiements": (
                ["Réf facture", "Adhérent", "Montant", "Date paiement"],
                [["2025-0999", "Christine LOPES", 55, "2025-08-02"]],
            ),
        }
    )

    response = await client.post(
        "/api/import/excel/gestion/preview",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert response.status_code == 200
    data = response.json()
    domains = {domain["kind"]: domain for domain in data["comparison"]["domains"]}
    assert domains["payments"] == {
        "kind": "payments",
        "file_rows": 1,
        "already_in_solde": 0,
        "missing_in_solde": 0,
        "extra_in_solde": 0,
        "ignored_by_policy": 0,
        "blocked": 1,
    }
    assert data["can_import"] is False


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_preview_gestion_comparison_counts_extra_invoices_in_solde(
    client: AsyncClient,
    auth_headers: dict,
    db_session,
) -> None:
    from backend.models.contact import Contact, ContactType
    from backend.models.invoice import Invoice, InvoiceStatus, InvoiceType

    contact = Contact(nom="LOPES", prenom="Christine", type=ContactType.CLIENT)
    db_session.add(contact)
    await db_session.flush()
    db_session.add_all(
        [
            Invoice(
                number="2025-0142",
                type=InvoiceType.CLIENT,
                contact_id=contact.id,
                date=date(2025, 8, 1),
                total_amount=Decimal("55.00"),
                paid_amount=Decimal("0.00"),
                status=InvoiceStatus.DRAFT,
            ),
            Invoice(
                number="2025-0199",
                type=InvoiceType.CLIENT,
                contact_id=contact.id,
                date=date(2025, 8, 2),
                total_amount=Decimal("80.00"),
                paid_amount=Decimal("0.00"),
                status=InvoiceStatus.DRAFT,
            ),
        ]
    )
    await db_session.commit()

    content = _make_multi_sheet_xlsx(
        {
            "Factures": (
                ["Date facture", "Réf facture", "Client", "Montant"],
                [
                    ["2025-08-01", "2025-0142", "Christine LOPES", 55],
                    ["2025-08-02", "2025-0143", "Christine LOPES", 75],
                ],
            ),
        }
    )

    response = await client.post(
        "/api/import/excel/gestion/preview",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert response.status_code == 200
    data = response.json()
    domains = {domain["kind"]: domain for domain in data["comparison"]["domains"]}
    assert domains["invoices"] == {
        "kind": "invoices",
        "file_rows": 2,
        "already_in_solde": 1,
        "missing_in_solde": 1,
        "extra_in_solde": 1,
        "extra_in_solde_details": [
            {
                "summary": "2025-0199 · 2025-08-02",
                "number": "2025-0199",
                "date": "2025-08-02",
            }
        ],
        "ignored_by_policy": 0,
        "blocked": 0,
    }
    assert data["comparison"]["totals"]["extra_in_solde"] == 1


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_preview_gestion_comparison_filters_only_convergence_section_by_date_window(
    client: AsyncClient,
    auth_headers: dict,
    db_session,
) -> None:
    from backend.models.contact import Contact, ContactType
    from backend.models.invoice import Invoice, InvoiceStatus, InvoiceType

    contact = Contact(nom="LOPES", prenom="Christine", type=ContactType.CLIENT)
    db_session.add(contact)
    await db_session.flush()
    db_session.add_all(
        [
            Invoice(
                number="2025-0142",
                type=InvoiceType.CLIENT,
                contact_id=contact.id,
                date=date(2025, 8, 1),
                total_amount=Decimal("55.00"),
                paid_amount=Decimal("0.00"),
                status=InvoiceStatus.DRAFT,
            ),
            Invoice(
                number="2025-0199",
                type=InvoiceType.CLIENT,
                contact_id=contact.id,
                date=date(2025, 8, 2),
                total_amount=Decimal("80.00"),
                paid_amount=Decimal("0.00"),
                status=InvoiceStatus.DRAFT,
            ),
        ]
    )
    await db_session.commit()

    content = _make_multi_sheet_xlsx(
        {
            "Factures": (
                ["Date facture", "Réf facture", "Client", "Montant"],
                [
                    ["2025-08-01", "2025-0142", "Christine LOPES", 55],
                    ["2025-08-02", "2025-0143", "Christine LOPES", 75],
                ],
            ),
        }
    )

    response = await client.post(
        "/api/import/excel/gestion/preview",
        data={
            "comparison_start_date": "2025-08-01",
            "comparison_end_date": "2025-08-01",
        },
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert response.status_code == 200
    data = response.json()
    domains = {domain["kind"]: domain for domain in data["comparison"]["domains"]}
    assert domains["invoices"] == {
        "kind": "invoices",
        "file_rows": 1,
        "already_in_solde": 1,
        "missing_in_solde": 0,
        "extra_in_solde": 0,
        "ignored_by_policy": 0,
        "blocked": 0,
    }
    sheets = {sheet["name"]: sheet for sheet in data["sheets"]}
    assert sheets["Factures"]["rows"] == 1


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_preview_gestion_invoice_extra_in_solde_uses_file_exercise_not_old_open_invoice_dates(
    client: AsyncClient,
    auth_headers: dict,
    db_session,
) -> None:
    from backend.models.contact import Contact, ContactType
    from backend.models.invoice import Invoice, InvoiceStatus, InvoiceType

    contact = Contact(nom="LOPES", prenom="Christine", type=ContactType.CLIENT)
    db_session.add(contact)
    await db_session.flush()
    db_session.add(
        Invoice(
            number="2024-0199",
            type=InvoiceType.CLIENT,
            contact_id=contact.id,
            date=date(2024, 10, 12),
            total_amount=Decimal("80.00"),
            paid_amount=Decimal("0.00"),
            status=InvoiceStatus.DRAFT,
        )
    )
    await db_session.commit()

    content = _make_multi_sheet_xlsx(
        {
            "Factures": (
                ["Date facture", "Réf facture", "Client", "Montant"],
                [
                    ["2025-08-01", "2025-0142", "Christine LOPES", 55],
                    ["2022-11-27", "2022-0007", "Christine LOPES", 45],
                ],
            ),
        }
    )

    response = await client.post(
        "/api/import/excel/gestion/preview",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert response.status_code == 200
    data = response.json()
    domains = {domain["kind"]: domain for domain in data["comparison"]["domains"]}
    assert domains["invoices"] == {
        "kind": "invoices",
        "file_rows": 2,
        "already_in_solde": 0,
        "missing_in_solde": 2,
        "extra_in_solde": 0,
        "ignored_by_policy": 0,
        "blocked": 0,
    }


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_preview_gestion_invoice_extra_in_solde_counts_next_year_in_exercise(
    client: AsyncClient,
    auth_headers: dict,
    db_session,
) -> None:
    from backend.models.contact import Contact, ContactType
    from backend.models.invoice import Invoice, InvoiceStatus, InvoiceType

    contact = Contact(nom="LOPES", prenom="Christine", type=ContactType.CLIENT)
    db_session.add(contact)
    await db_session.flush()
    db_session.add(
        Invoice(
            number="2026-0004",
            type=InvoiceType.CLIENT,
            contact_id=contact.id,
            date=date(2026, 3, 5),
            total_amount=Decimal("90.00"),
            paid_amount=Decimal("0.00"),
            status=InvoiceStatus.DRAFT,
        )
    )
    await db_session.commit()

    content = _make_multi_sheet_xlsx(
        {
            "Factures": (
                ["Date facture", "Réf facture", "Client", "Montant"],
                [["2025-08-01", "2025-0142", "Christine LOPES", 55]],
            ),
        }
    )

    response = await client.post(
        "/api/import/excel/gestion/preview",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert response.status_code == 200
    data = response.json()
    domains = {domain["kind"]: domain for domain in data["comparison"]["domains"]}
    assert domains["invoices"] == {
        "kind": "invoices",
        "file_rows": 1,
        "already_in_solde": 0,
        "missing_in_solde": 1,
        "extra_in_solde": 1,
        "extra_in_solde_details": [
            {
                "summary": "2026-0004 · 2026-03-05",
                "number": "2026-0004",
                "date": "2026-03-05",
            }
        ],
        "ignored_by_policy": 0,
        "blocked": 0,
    }


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_preview_gestion_payment_extra_in_solde_uses_file_exercise_window(
    client: AsyncClient,
    auth_headers: dict,
    db_session,
) -> None:
    from backend.models.contact import Contact, ContactType
    from backend.models.invoice import Invoice, InvoiceStatus, InvoiceType
    from backend.models.payment import Payment, PaymentMethod

    contact = Contact(nom="LOPES", prenom="Christine", type=ContactType.CLIENT)
    db_session.add(contact)
    await db_session.flush()

    extra_invoice = Invoice(
        number="2025-0199",
        type=InvoiceType.CLIENT,
        contact_id=contact.id,
        date=date(2025, 11, 2),
        total_amount=Decimal("80.00"),
        paid_amount=Decimal("80.00"),
        status=InvoiceStatus.PAID,
    )
    db_session.add(extra_invoice)
    await db_session.flush()
    db_session.add(
        Payment(
            invoice_id=extra_invoice.id,
            contact_id=contact.id,
            date=date(2026, 5, 10),
            amount=Decimal("80.00"),
            method=PaymentMethod.VIREMENT,
            reference="2025-0199",
        )
    )
    await db_session.commit()

    content = _make_multi_sheet_xlsx(
        {
            "Factures": (
                ["Date facture", "Réf facture", "Client", "Montant"],
                [["2025-08-01", "2025-0142", "Christine LOPES", 55]],
            ),
            "Paiements": (
                ["Réf facture", "Adhérent", "Montant", "Date paiement"],
                [["2025-0142", "Christine LOPES", 55, "2025-08-02"]],
            ),
        }
    )

    response = await client.post(
        "/api/import/excel/gestion/preview",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert response.status_code == 200
    data = response.json()
    domains = {domain["kind"]: domain for domain in data["comparison"]["domains"]}
    assert domains["payments"] == {
        "kind": "payments",
        "file_rows": 1,
        "already_in_solde": 0,
        "missing_in_solde": 1,
        "extra_in_solde": 1,
        "extra_in_solde_details": [
            {
                "summary": "2025-0199 · 2026-05-10 · 80 · 512100",
                "reference": "2025-0199",
                "payment_date": "2026-05-10",
                "amount": "80",
                "settlement_account": "512100",
                "invoice_number": "2025-0199",
            }
        ],
        "ignored_by_policy": 0,
        "blocked": 0,
    }


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_preview_gestion_comparison_counts_existing_payments_bank_and_cash_rows(
    client: AsyncClient,
    auth_headers: dict,
) -> None:
    content = _make_multi_sheet_xlsx(
        {
            "Factures": (
                ["Date facture", "Réf facture", "Client", "Montant"],
                [["2025-08-01", "2025-0142", "Christine LOPES", 55]],
            ),
            "Paiements": (
                ["Réf facture", "Adhérent", "Montant", "Date paiement"],
                [["2025-0142", "Christine LOPES", 55, "2025-08-02"]],
            ),
            "Banque": (
                ["Date", "Libellé", "Débit", "Crédit", "Solde"],
                [["2025-08-03", "Paiement CB cantine", 18, None, 482]],
            ),
            "Caisse": (
                ["Date", "Libellé", "Entrée", "Sortie"],
                [["2025-08-04", "Recette adhesion", 55, None]],
            ),
        }
    )

    import_response = await client.post(
        "/api/import/excel/gestion",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert import_response.status_code == 200

    preview_response = await client.post(
        "/api/import/excel/gestion/preview",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert preview_response.status_code == 200
    data = preview_response.json()
    domains = {domain["kind"]: domain for domain in data["comparison"]["domains"]}
    assert domains["invoices"] == {
        "kind": "invoices",
        "file_rows": 1,
        "already_in_solde": 1,
        "missing_in_solde": 0,
        "extra_in_solde": 0,
        "ignored_by_policy": 0,
        "blocked": 0,
    }
    assert domains["payments"] == {
        "kind": "payments",
        "file_rows": 1,
        "already_in_solde": 1,
        "missing_in_solde": 0,
        "extra_in_solde": 0,
        "ignored_by_policy": 0,
        "blocked": 0,
    }
    assert domains["bank"] == {
        "kind": "bank",
        "file_rows": 1,
        "already_in_solde": 1,
        "missing_in_solde": 0,
        "extra_in_solde": 0,
        "ignored_by_policy": 0,
        "blocked": 0,
    }
    assert domains["cash"] == {
        "kind": "cash",
        "file_rows": 1,
        "already_in_solde": 1,
        "missing_in_solde": 0,
        "extra_in_solde": 0,
        "ignored_by_policy": 0,
        "blocked": 0,
    }
    assert data["comparison"]["totals"]["missing_in_solde"] == 0


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_preview_gestion_comparison_distinguishes_cheques_with_same_amount(
    client: AsyncClient,
    auth_headers: dict,
    db_session,
) -> None:
    """Preview comparison should keep cheque rows distinct when cheque numbers differ."""
    from backend.models.contact import Contact, ContactType
    from backend.models.invoice import Invoice, InvoiceStatus, InvoiceType
    from backend.models.payment import Payment, PaymentMethod

    contact = Contact(nom="Christine", prenom="LOPES", type=ContactType.CLIENT)
    db_session.add(contact)
    await db_session.flush()

    invoice = Invoice(
        number="2025-0072",
        reference="2025-0072",
        type=InvoiceType.CLIENT,
        contact_id=contact.id,
        date=date(2025, 5, 2),
        due_date=date(2025, 5, 2),
        total_amount=Decimal("60"),
        paid_amount=Decimal("30"),
        status=InvoiceStatus.PARTIAL,
        description="Cours mai",
    )
    db_session.add(invoice)
    await db_session.flush()

    db_session.add(
        Payment(
            invoice_id=invoice.id,
            contact_id=contact.id,
            date=date(2025, 5, 2),
            amount=Decimal("30"),
            method=PaymentMethod.CHEQUE,
            cheque_number="2025.05.02.05",
            reference="2025-0072",
        )
    )
    await db_session.commit()

    content = _make_multi_sheet_xlsx(
        {
            "Paiements": (
                [
                    "Date",
                    "Montant",
                    "Référence facture",
                    "Client",
                    "Mode",
                    "Numéro chèque",
                    "Déposé",
                    "Date dépôt",
                ],
                [
                    [
                        "2025-05-02",
                        30,
                        "2025-0072",
                        "Christine LOPES",
                        "Chèque",
                        "2025.05.02.05",
                        False,
                        None,
                    ],
                    [
                        "2025-05-02",
                        30,
                        "2025-0072",
                        "Christine LOPES",
                        "Chèque",
                        "2025.05.02.06",
                        False,
                        None,
                    ],
                ],
            )
        }
    )

    preview_response = await client.post(
        "/api/import/excel/gestion/preview",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
        data={
            "comparison_start_date": "2025-05-01",
            "comparison_end_date": "2025-05-31",
        },
    )

    assert preview_response.status_code == 200
    preview_data = preview_response.json()
    assert preview_data["estimated_payments"] == 1
    domains = {domain["kind"]: domain for domain in preview_data["comparison"]["domains"]}
    assert domains["payments"] == {
        "kind": "payments",
        "file_rows": 2,
        "already_in_solde": 1,
        "missing_in_solde": 1,
        "extra_in_solde": 0,
        "ignored_by_policy": 0,
        "blocked": 0,
    }


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_preview_gestion_comparison_ignores_extra_bank_rows_outside_workbook_date_range(
    client: AsyncClient,
    auth_headers: dict,
    db_session,
) -> None:
    from backend.models.bank import BankTransaction

    db_session.add(
        BankTransaction(
            date=date(2024, 1, 15),
            amount=Decimal("18.00"),
            description="Paiement CB ancien",
            reference=None,
            balance_after=Decimal("100.00"),
        )
    )
    await db_session.commit()

    content = _make_multi_sheet_xlsx(
        {
            "Banque": (
                ["Date", "Libellé", "Débit", "Crédit", "Solde"],
                [
                    ["2024-09-03", "Paiement CB rentree", 18, None, 482],
                    ["2025-01-12", "Virement recu", None, 55, 537],
                ],
            ),
        }
    )

    response = await client.post(
        "/api/import/excel/gestion/preview",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert response.status_code == 200
    data = response.json()
    domains = {domain["kind"]: domain for domain in data["comparison"]["domains"]}
    assert domains["bank"] == {
        "kind": "bank",
        "file_rows": 2,
        "already_in_solde": 0,
        "missing_in_solde": 2,
        "extra_in_solde": 0,
        "ignored_by_policy": 0,
        "blocked": 0,
    }


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_preview_gestion_bank_extra_in_solde_counts_intermediate_year(
    client: AsyncClient,
    auth_headers: dict,
    db_session,
) -> None:
    from backend.models.bank import BankTransaction

    db_session.add(
        BankTransaction(
            date=date(2025, 5, 15),
            amount=Decimal("42.00"),
            description="Virement intermediaire",
            reference="INT-2025",
            balance_after=Decimal("210.00"),
        )
    )
    await db_session.commit()

    content = _make_multi_sheet_xlsx(
        {
            "Banque": (
                ["Date", "Libellé", "Débit", "Crédit", "Solde"],
                [
                    ["2024-09-03", "Paiement CB rentree", 18, None, 482],
                    ["2026-01-12", "Virement recu", None, 55, 537],
                ],
            ),
        }
    )

    response = await client.post(
        "/api/import/excel/gestion/preview",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert response.status_code == 200
    data = response.json()
    domains = {domain["kind"]: domain for domain in data["comparison"]["domains"]}
    assert domains["bank"] == {
        "kind": "bank",
        "file_rows": 2,
        "already_in_solde": 0,
        "missing_in_solde": 2,
        "extra_in_solde": 1,
        "extra_in_solde_details": [
            {
                "summary": "2025-05-15 · 42 · Virement intermediaire · INT-2025",
                "entry_date": "2025-05-15",
                "amount": "42",
                "description": "Virement intermediaire",
                "reference": "INT-2025",
            }
        ],
        "ignored_by_policy": 0,
        "blocked": 0,
    }


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_preview_gestion_cash_extra_in_solde_counts_intermediate_year(
    client: AsyncClient,
    auth_headers: dict,
    db_session,
) -> None:
    from backend.models.cash import CashMovementType, CashRegister

    db_session.add(
        CashRegister(
            date=date(2025, 5, 15),
            type=CashMovementType.IN,
            amount=Decimal("42.00"),
            description="Don intermediaire",
            reference="CAISSE-2025",
            balance_after=Decimal("210.00"),
        )
    )
    await db_session.commit()

    content = _make_multi_sheet_xlsx(
        {
            "Caisse": (
                ["Date", "Libellé", "Entrée", "Sortie"],
                [
                    ["2024-09-03", "Recette rentree", 18, None],
                    ["2026-01-12", "Achat fournitures", None, 55],
                ],
            ),
        }
    )

    response = await client.post(
        "/api/import/excel/gestion/preview",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert response.status_code == 200
    data = response.json()
    domains = {domain["kind"]: domain for domain in data["comparison"]["domains"]}
    assert domains["cash"] == {
        "kind": "cash",
        "file_rows": 2,
        "already_in_solde": 0,
        "missing_in_solde": 2,
        "extra_in_solde": 1,
        "extra_in_solde_details": [
            {
                "summary": "2025-05-15 · in · 42 · Don intermediaire · CAISSE-2025",
                "entry_date": "2025-05-15",
                "movement_type": "in",
                "amount": "42",
                "description": "Don intermediaire",
                "reference": "CAISSE-2025",
            }
        ],
        "ignored_by_policy": 0,
        "blocked": 0,
    }


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_preview_and_import_gestion_accept_contacts_sheet_without_prenom(
    client: AsyncClient, auth_headers: dict, db_session
) -> None:
    """Preview and import should accept a contacts sheet when only nom is required."""
    from backend.models.contact import Contact

    content = _make_multi_sheet_xlsx(
        {
            "Contacts": (
                ["Nom", "Email"],
                [
                    ["Christine LOPES", "christine@example.test"],
                    ["Thi BE NGUYEN", "thi@example.test"],
                ],
            )
        }
    )

    preview_response = await client.post(
        "/api/import/excel/gestion/preview",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert preview_response.status_code == 200
    preview_data = preview_response.json()
    sheets = {sheet["name"]: sheet for sheet in preview_data["sheets"]}
    assert sheets["Contacts"]["status"] == "recognized"
    assert sheets["Contacts"]["kind"] == "contacts"
    assert sheets["Contacts"]["rows"] == 2
    assert preview_data["estimated_contacts"] == 2
    assert preview_data["can_import"] is True

    import_response = await client.post(
        "/api/import/excel/gestion",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert import_response.status_code == 200
    import_data = import_response.json()
    assert import_data["contacts_created"] == 2

    contacts = list((await db_session.execute(select(Contact).order_by(Contact.nom))).scalars())
    assert [(contact.nom, contact.prenom) for contact in contacts] == [
        ("BE NGUYEN", "Thi"),
        ("LOPES", "Christine"),
    ]


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_reversible_import_run_prepare_execute_undo_redo_cycle(
    client: AsyncClient, auth_headers: dict, db_session
) -> None:
    """A prepared reversible run should execute, undo, and redo a simple contacts import."""
    from backend.models.contact import Contact

    content = _make_multi_sheet_xlsx(
        {
            "Contacts": (
                ["Nom", "Email"],
                [
                    ["Christine LOPES", "christine@example.test"],
                    ["Thi BE NGUYEN", "thi@example.test"],
                ],
            )
        }
    )

    prepare_response = await client.post(
        "/api/import/runs/prepare/gestion",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert prepare_response.status_code == 200
    prepared_run = prepare_response.json()
    run_id = prepared_run["id"]
    assert prepared_run["status"] == "prepared"
    assert prepared_run["can_execute"] is True
    assert prepared_run["preview"]["estimated_contacts"] == 2
    assert [operation["operation_type"] for operation in prepared_run["operations"]] == [
        "contact_row_import",
        "contact_row_import",
    ]

    execute_response = await client.post(
        f"/api/import/runs/{run_id}/execute",
        headers=auth_headers,
    )

    assert execute_response.status_code == 200
    executed_run = execute_response.json()
    assert executed_run["status"] == "completed"
    assert executed_run["summary"]["contacts_created"] == 2
    assert executed_run["can_undo"] is True
    assert all(operation["status"] == "applied" for operation in executed_run["operations"])

    contacts = list((await db_session.execute(select(Contact).order_by(Contact.nom))).scalars())
    assert [(contact.nom, contact.prenom) for contact in contacts] == [
        ("BE NGUYEN", "Thi"),
        ("LOPES", "Christine"),
    ]

    history_response = await client.get("/api/import/history", headers=auth_headers)

    assert history_response.status_code == 200
    history = history_response.json()
    assert history[0]["kind"] == "run"
    assert history[0]["id"] == run_id
    assert history[0]["status"] == "completed"
    assert history[0]["summary"]["contacts_created"] == 2

    undo_response = await client.post(
        f"/api/import/runs/{run_id}/undo",
        headers=auth_headers,
    )

    assert undo_response.status_code == 200
    undone_run = undo_response.json()
    assert undone_run["status"] == "reverted"
    assert undone_run["can_redo"] is True
    assert all(operation["status"] == "undone" for operation in undone_run["operations"])

    contacts_after_undo = list(
        (await db_session.execute(select(Contact).order_by(Contact.nom))).scalars()
    )
    assert contacts_after_undo == []

    redo_response = await client.post(
        f"/api/import/runs/{run_id}/redo",
        headers=auth_headers,
    )

    assert redo_response.status_code == 200
    redone_run = redo_response.json()
    assert redone_run["status"] == "completed"
    assert redone_run["can_undo"] is True
    assert all(operation["status"] == "applied" for operation in redone_run["operations"])


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_reversible_comptabilite_run_undo_tolerates_decimal_scale_differences(
    client: AsyncClient,
    auth_headers: dict,
    db_session,
) -> None:
    """Undo should accept semantically identical Numeric values.

    This remains true even if SQLite round-trips them with a different scale.
    """
    from backend.models.accounting_account import AccountingAccount, AccountType
    from backend.models.accounting_entry import AccountingEntry

    db_session.add(
        AccountingAccount(number="401100", label="Fournisseurs", type=AccountType.PASSIF)
    )
    await db_session.commit()

    content = _make_multi_sheet_xlsx(
        {
            "Journal": (
                ["Date", "N° compte", "Libellé de l'écriture", "Débit", "Crédit"],
                [["2025-08-01", "401100", "Facture fournisseur", None, 261.5]],
            ),
        }
    )

    prepare_response = await client.post(
        "/api/import/runs/prepare/comptabilite",
        files={"file": ("Comptabilite 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert prepare_response.status_code == 200
    run_id = prepare_response.json()["id"]

    execute_response = await client.post(
        f"/api/import/runs/{run_id}/execute",
        headers=auth_headers,
    )

    assert execute_response.status_code == 200
    executed_run = execute_response.json()
    assert executed_run["status"] == "completed"
    assert executed_run["can_undo"] is True
    assert executed_run["summary"]["entries_created"] == 1

    entries_before_undo = list((await db_session.execute(select(AccountingEntry))).scalars())
    assert len(entries_before_undo) == 1

    undo_response = await client.post(
        f"/api/import/runs/{run_id}/undo",
        headers=auth_headers,
    )

    assert undo_response.status_code == 200
    undone_run = undo_response.json()
    assert undone_run["status"] == "reverted"
    remaining_entries = list((await db_session.execute(select(AccountingEntry))).scalars())
    assert not remaining_entries


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_reversible_bank_operation_failure_rolls_back_payment_side_effects(
    client: AsyncClient,
    auth_headers: dict,
    db_session,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """A failed bank operation must not leave a paid invoice or created payment behind."""
    from backend.models.bank import BankTransaction
    from backend.models.invoice import Invoice, InvoiceStatus
    from backend.models.payment import Payment
    from backend.services import import_reversible

    content = _make_multi_sheet_xlsx(
        {
            "Factures": (
                ["Date facture", "Réf facture", "Client", "Montant"],
                [["2025-08-01", "2025-0142", "Christine LOPES", 55]],
            ),
            "Banque": (
                ["Date", "Montant", "Référence", "Libellé", "Solde"],
                [["2025-08-02", 55, "2025-0142", "Paiement facture client", 537]],
            ),
        }
    )

    async def _crash_generate_entries_for_payment(*args, **kwargs):
        raise RuntimeError("forced payment generation failure")

    monkeypatch.setattr(
        import_reversible,
        "generate_entries_for_payment",
        _crash_generate_entries_for_payment,
    )

    prepare_response = await client.post(
        "/api/import/runs/prepare/gestion",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert prepare_response.status_code == 200
    run_id = prepare_response.json()["id"]

    execute_response = await client.post(
        f"/api/import/runs/{run_id}/execute",
        headers=auth_headers,
    )

    assert execute_response.status_code == 200
    executed_run = execute_response.json()
    assert executed_run["status"] == "failed"
    assert executed_run["can_undo"] is True
    assert [operation["status"] for operation in executed_run["operations"]] == [
        "applied",
        "failed",
    ]
    assert executed_run["operations"][1]["error_message"] == "forced payment generation failure"
    assert executed_run["operations"][1]["effects"] == []
    assert any(
        "forced payment generation failure" in error for error in executed_run["summary"]["errors"]
    )

    invoices = list((await db_session.execute(select(Invoice).order_by(Invoice.id))).scalars())
    assert len(invoices) == 1
    assert invoices[0].number == "2025-0142"
    assert invoices[0].paid_amount == Decimal("0.00")
    assert invoices[0].status == InvoiceStatus.SENT

    payments = list((await db_session.execute(select(Payment).order_by(Payment.id))).scalars())
    bank_entries = list(
        (await db_session.execute(select(BankTransaction).order_by(BankTransaction.id))).scalars()
    )
    assert payments == []
    assert bank_entries == []

    undo_response = await client.post(
        f"/api/import/runs/{run_id}/undo",
        headers=auth_headers,
    )

    assert undo_response.status_code == 200
    undone_run = undo_response.json()
    assert undone_run["status"] == "reverted"
    remaining_invoices = list((await db_session.execute(select(Invoice))).scalars())
    remaining_payments = list((await db_session.execute(select(Payment))).scalars())
    assert remaining_invoices == []
    assert remaining_payments == []


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_reversible_bank_virement_run_executes_with_accounting_rules(
    client: AsyncClient,
    auth_headers: dict,
    db_session,
) -> None:
    """A reversible bank client-payment run should complete.

    This covers the case where payment accounting rules are active.
    """
    from backend.models.accounting_entry import AccountingEntry
    from backend.models.invoice import Invoice, InvoiceStatus
    from backend.models.payment import Payment, PaymentMethod
    from backend.services.accounting_engine import seed_default_rules

    await seed_default_rules(db_session)

    content = _make_multi_sheet_xlsx(
        {
            "Factures": (
                ["Date facture", "Réf facture", "Client", "Montant"],
                [["2025-08-01", "2025-0142", "Christine LOPES", 55]],
            ),
            "Banque": (
                ["Date", "Montant", "Référence", "Libellé", "Solde"],
                [["2025-08-02", 55, "2025-0142", "Paiement facture client", 537]],
            ),
        }
    )

    prepare_response = await client.post(
        "/api/import/runs/prepare/gestion",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert prepare_response.status_code == 200
    run_id = prepare_response.json()["id"]

    execute_response = await client.post(
        f"/api/import/runs/{run_id}/execute",
        headers=auth_headers,
    )

    assert execute_response.status_code == 200
    executed_run = execute_response.json()
    assert executed_run["status"] == "completed"
    assert all(operation["status"] == "applied" for operation in executed_run["operations"])
    assert executed_run["summary"]["payments_created"] == 1
    assert executed_run["summary"]["entries_created"] >= 2

    invoice = (await db_session.execute(select(Invoice))).scalar_one()
    payment = (await db_session.execute(select(Payment))).scalar_one()
    entries = list((await db_session.execute(select(AccountingEntry))).scalars())

    assert invoice.number == "2025-0142"
    assert invoice.paid_amount == Decimal("55.00")
    assert invoice.status == InvoiceStatus.PAID
    assert payment.method == PaymentMethod.VIREMENT
    assert len(entries) >= 2

    undo_response = await client.post(
        f"/api/import/runs/{run_id}/undo",
        headers=auth_headers,
    )

    assert undo_response.status_code == 200
    undone_run = undo_response.json()
    assert undone_run["status"] == "reverted"
    assert list((await db_session.execute(select(Invoice))).scalars()) == []
    assert list((await db_session.execute(select(Payment))).scalars()) == []


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_reversible_gestion_salary_run_undo_tolerates_float_rounding_noise(
    client: AsyncClient,
    auth_headers: dict,
    db_session,
) -> None:
    """Undo should succeed when salary imports contain Excel float rounding noise."""
    from backend.models.accounting_entry import AccountingEntry, EntrySourceType
    from backend.models.salary import Salary
    from backend.services.accounting_engine import seed_default_rules

    await seed_default_rules(db_session)

    content = _make_multi_sheet_xlsx_rows(
        {
            "Aide Salaires": [
                ["2026.03"],
                ["NOM", "Heures", "Brut", "Salariales", "Patronales", "Impôts", "Net"],
                ["LAY", 12, 1605, 355.36, 370.86, 0, "1249.6399999999999"],
            ]
        }
    )

    prepare_response = await client.post(
        "/api/import/runs/prepare/gestion",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert prepare_response.status_code == 200
    run_id = prepare_response.json()["id"]

    execute_response = await client.post(
        f"/api/import/runs/{run_id}/execute",
        headers=auth_headers,
    )

    assert execute_response.status_code == 200
    executed_run = execute_response.json()
    assert executed_run["status"] == "completed"
    assert executed_run["summary"]["salaries_created"] == 1
    assert executed_run["can_undo"] is True

    salary_entries = list(
        (
            await db_session.execute(
                select(AccountingEntry)
                .where(AccountingEntry.source_type == EntrySourceType.SALARY)
                .order_by(AccountingEntry.id)
            )
        ).scalars()
    )
    assert salary_entries
    assert any(entry.credit == Decimal("1249.64") for entry in salary_entries)

    undo_response = await client.post(
        f"/api/import/runs/{run_id}/undo",
        headers=auth_headers,
    )

    assert undo_response.status_code == 200
    undone_run = undo_response.json()
    assert undone_run["status"] == "reverted"
    assert list((await db_session.execute(select(Salary))).scalars()) == []
    assert (
        list(
            (
                await db_session.execute(
                    select(AccountingEntry).where(
                        AccountingEntry.source_type == EntrySourceType.SALARY
                    )
                )
            ).scalars()
        )
        == []
    )


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_reversible_prepare_and_execute_ignore_duplicate_salary_rows_in_same_file(
    client: AsyncClient,
    auth_headers: dict,
) -> None:
    """Reversible preview/prepare should ignore duplicate salary employee-month rows.

    This applies when the same workbook already contains the employee-month pair.
    """
    content = _make_multi_sheet_xlsx_rows(
        {
            "Aide Salaires": [
                ["2025.03"],
                ["NOM", "Heures", "Brut", "Salariales", "Patronales", "Impôts", "Net"],
                ["LAY", 104, 1605, 355.36, 350.47, 0, 1249.64],
                ["WOLFF P.", 10, 249.99, 54.33, 110.12, 1.62, 194.04],
                ["LAY", 104, 1605, 355.36, 350.47, 0, 1249.64],
                ["WOLFF P.", 2, 50, 11.16, 20.21, 0.32, 38.52],
            ]
        }
    )

    preview_response = await client.post(
        "/api/import/excel/gestion/preview",
        files={"file": ("Gestion 2024.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert preview_response.status_code == 200
    preview_data = preview_response.json()
    assert preview_data["estimated_salaries"] == 2
    assert preview_data["warnings"]
    assert any("ligne ignoree" in warning.lower() for warning in preview_data["warnings"])

    prepare_response = await client.post(
        "/api/import/runs/prepare/gestion",
        files={"file": ("Gestion 2024.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert prepare_response.status_code == 200
    prepared_run = prepare_response.json()
    assert prepared_run["can_execute"] is True
    salary_operations = [
        operation
        for operation in prepared_run["operations"]
        if operation["operation_type"] == "salary_month_import"
    ]
    ignored_operations = [
        operation
        for operation in prepared_run["operations"]
        if operation["operation_type"] == "ignored_by_policy"
    ]
    assert len(salary_operations) == 1
    assert salary_operations[0]["title"] == "2025-03"
    assert salary_operations[0]["source_row_numbers"] == [3, 4]
    assert len(ignored_operations) == 2
    assert all("Salaire ligne" in operation["title"] for operation in ignored_operations)

    execute_response = await client.post(
        f"/api/import/runs/{prepared_run['id']}/execute",
        headers=auth_headers,
    )

    assert execute_response.status_code == 200
    executed_run = execute_response.json()
    assert executed_run["status"] == "completed"
    assert executed_run["summary"]["salaries_created"] == 2
    assert executed_run["summary"]["ignored_rows"] == 2


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_reversible_prepare_ignores_existing_client_payment_rows(
    client: AsyncClient,
    auth_headers: dict,
    db_session,
) -> None:
    """Reversible gestion prepare should ignore a client payment row already present in Solde."""
    from backend.models.contact import Contact, ContactType
    from backend.models.invoice import Invoice, InvoiceStatus, InvoiceType
    from backend.models.payment import Payment, PaymentMethod

    contact = Contact(nom="Valérie", prenom="POIROT", type=ContactType.CLIENT)
    db_session.add(contact)
    await db_session.flush()

    invoice = Invoice(
        number="2024-0170",
        reference="2024-0170",
        type=InvoiceType.CLIENT,
        contact_id=contact.id,
        date=date(2024, 8, 7),
        due_date=date(2024, 8, 7),
        total_amount=Decimal("78"),
        paid_amount=Decimal("78"),
        status=InvoiceStatus.PAID,
        description="Cours août",
    )
    db_session.add(invoice)
    await db_session.flush()

    db_session.add(
        Payment(
            invoice_id=invoice.id,
            contact_id=contact.id,
            date=date(2024, 8, 7),
            amount=Decimal("78"),
            method=PaymentMethod.VIREMENT,
            reference="2024-0170",
            deposited=True,
            deposit_date=date(2024, 8, 7),
        )
    )
    await db_session.commit()

    content = _make_multi_sheet_xlsx(
        {
            "Paiements": (
                [
                    "Date",
                    "Montant",
                    "Référence facture",
                    "Client",
                    "Mode",
                    "Numéro chèque",
                    "Déposé",
                    "Date dépôt",
                ],
                [
                    [
                        "2024-08-07",
                        78,
                        "2024-0170",
                        "Valérie POIROT",
                        "virement",
                        None,
                        True,
                        "2024-08-07",
                    ]
                ],
            )
        }
    )

    preview_response = await client.post(
        "/api/import/excel/gestion/preview",
        files={"file": ("Gestion 2024.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert preview_response.status_code == 200
    preview_data = preview_response.json()
    assert preview_data["estimated_payments"] == 0
    assert any("ligne ignoree" in warning.lower() for warning in preview_data["warnings"])

    prepare_response = await client.post(
        "/api/import/runs/prepare/gestion",
        files={"file": ("Gestion 2024.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
        data={
            "comparison_start_date": "2024-08-01",
            "comparison_end_date": "2025-07-31",
        },
    )

    assert prepare_response.status_code == 200
    prepared_run = prepare_response.json()
    assert prepared_run["can_execute"] is False
    payment_operations = [
        operation
        for operation in prepared_run["operations"]
        if operation["operation_type"] == "client_payment_row_import"
    ]
    ignored_operations = [
        operation
        for operation in prepared_run["operations"]
        if operation["operation_type"] == "ignored_by_policy"
    ]
    assert not payment_operations
    assert len(ignored_operations) == 1
    assert ignored_operations[0]["title"] == "Paiement ligne 2"
    assert "ligne ignoree" in ignored_operations[0]["diagnostics"][0].lower()


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_reversible_prepare_ignores_existing_bank_client_transfer_rows(
    client: AsyncClient,
    auth_headers: dict,
    db_session,
) -> None:
    """Reversible gestion prepare should ignore a bank transfer row already present in Solde."""
    from backend.models.bank import BankTransaction, BankTransactionSource
    from backend.models.contact import Contact, ContactType
    from backend.models.invoice import Invoice, InvoiceStatus, InvoiceType
    from backend.models.payment import Payment, PaymentMethod

    contact = Contact(nom="RIBEIRO", prenom="", type=ContactType.CLIENT)
    db_session.add(contact)
    await db_session.flush()

    invoice = Invoice(
        number="2025-0141",
        reference="2025-0141",
        type=InvoiceType.CLIENT,
        contact_id=contact.id,
        date=date(2025, 7, 21),
        due_date=date(2025, 7, 21),
        total_amount=Decimal("170"),
        paid_amount=Decimal("170"),
        status=InvoiceStatus.PAID,
        description="Cours été",
    )
    db_session.add(invoice)
    await db_session.flush()

    payment = Payment(
        invoice_id=invoice.id,
        contact_id=contact.id,
        date=date(2025, 8, 1),
        amount=Decimal("170"),
        method=PaymentMethod.VIREMENT,
        reference="2025-0141",
        notes="Paiement facture client",
        deposited=True,
        deposit_date=date(2025, 8, 1),
    )
    db_session.add(payment)
    await db_session.flush()

    db_session.add(
        BankTransaction(
            date=date(2025, 8, 1),
            amount=Decimal("170"),
            reference="2025-0141",
            description="Paiement facture client",
            source=BankTransactionSource.IMPORT,
            payment_id=payment.id,
        )
    )
    await db_session.commit()

    content = _make_multi_sheet_xlsx(
        {
            "Banque": (
                ["Date", "Montant", "Référence", "Libellé", "Solde"],
                [["2025-08-01", 170, "2025-0141", "Paiement facture client", 2844.77]],
            ),
        }
    )

    prepare_response = await client.post(
        "/api/import/runs/prepare/gestion",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert prepare_response.status_code == 200
    prepared_run = prepare_response.json()
    bank_operations = [
        operation
        for operation in prepared_run["operations"]
        if operation["operation_type"] == "bank_row_import"
    ]
    ignored_operations = [
        operation
        for operation in prepared_run["operations"]
        if operation["operation_type"] == "ignored_by_policy"
    ]
    assert not bank_operations
    assert len(ignored_operations) == 1
    assert ignored_operations[0]["title"] == "2025-0141"
    assert "ligne ignoree" in ignored_operations[0]["diagnostics"][0].lower()


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_reversible_prepare_ignores_payment_row_already_covered_by_bank_row(
    client: AsyncClient,
    auth_headers: dict,
) -> None:
    """A bank client transfer should suppress the duplicate payment-sheet row.

    This applies when the duplicate appears in the same workbook.
    """
    content = _make_multi_sheet_xlsx(
        {
            "Factures": (
                ["Date facture", "Réf facture", "Client", "Montant"],
                [["2025-08-01", "2025-0142", "Christine LOPES", 55]],
            ),
            "Banque": (
                ["Date", "Montant", "Référence", "Libellé", "Solde"],
                [["2025-08-02", 55, "2025-0142", "Paiement facture client", 1000]],
            ),
            "Paiements": (
                ["Réf facture", "Adhérent", "Montant", "Date paiement"],
                [["2025-0142", "Christine LOPES", 55, "2025-08-02"]],
            ),
        }
    )

    prepare_response = await client.post(
        "/api/import/runs/prepare/gestion",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert prepare_response.status_code == 200
    prepared_run = prepare_response.json()
    assert prepared_run["status"] == "prepared"
    assert [operation["operation_type"] for operation in prepared_run["operations"]] == [
        "client_invoice_row_import",
        "bank_row_import",
        "ignored_by_policy",
    ]
    ignored_operation = prepared_run["operations"][2]
    assert ignored_operation["title"] == "Paiement ligne 2"
    assert "paiement deja couvert" in ignored_operation["diagnostics"][0].lower()

    execute_response = await client.post(
        f"/api/import/runs/{prepared_run['id']}/execute",
        headers=auth_headers,
    )

    assert execute_response.status_code == 200
    executed_run = execute_response.json()
    assert executed_run["status"] == "completed"
    assert executed_run["summary"]["payments_created"] == 1
    assert executed_run["summary"]["ignored_rows"] == 1


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_reversible_import_prepare_exposes_planned_effects_for_invoice_operation(
    client: AsyncClient,
    auth_headers: dict,
    db_session,
) -> None:
    """Prepared runs should expose detailed planned effects before execution."""
    from backend.services.accounting_engine import seed_default_rules

    await seed_default_rules(db_session)

    content = _make_multi_sheet_xlsx(
        {
            "Factures": (
                ["Date facture", "Réf facture", "Client", "Montant"],
                [["2025-08-01", "2025-0142", "Christine LOPES", 55]],
            ),
        }
    )

    prepare_response = await client.post(
        "/api/import/runs/prepare/gestion",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert prepare_response.status_code == 200
    prepared_run = prepare_response.json()
    invoice_operation = next(
        operation
        for operation in prepared_run["operations"]
        if operation["operation_type"] == "client_invoice_row_import"
    )

    invoice_effect_types = {
        effect["entity_type"] for effect in invoice_operation["planned_effects"]
    }
    assert {"contact", "invoice", "invoice_line", "accounting_entry"} <= invoice_effect_types

    accounting_effect = next(
        effect
        for effect in invoice_operation["planned_effects"]
        if effect["entity_type"] == "accounting_entry"
    )
    assert accounting_effect["status"] == "planned"
    assert accounting_effect["details"]["account_number"]
    assert accounting_effect["details"]["date"] == "2025-08-01"


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_reversible_import_operation_undo_redo_targets_single_prepared_effect(
    client: AsyncClient, auth_headers: dict, db_session
) -> None:
    """Per-operation undo/redo should only affect the selected imported object."""
    from backend.models.contact import Contact

    content = _make_multi_sheet_xlsx(
        {
            "Contacts": (
                ["Nom", "Email"],
                [
                    ["Christine LOPES", "christine@example.test"],
                    ["Thi BE NGUYEN", "thi@example.test"],
                ],
            )
        }
    )

    prepare_response = await client.post(
        "/api/import/runs/prepare/gestion",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )
    run_id = prepare_response.json()["id"]

    execute_response = await client.post(
        f"/api/import/runs/{run_id}/execute",
        headers=auth_headers,
    )

    assert execute_response.status_code == 200
    executed_run = execute_response.json()
    operation_id = executed_run["operations"][0]["id"]

    undo_operation_response = await client.post(
        f"/api/import/operations/{operation_id}/undo",
        headers=auth_headers,
    )

    assert undo_operation_response.status_code == 200
    partially_undone_run = undo_operation_response.json()
    assert partially_undone_run["status"] == "partially_reverted"
    assert partially_undone_run["operations"][0]["status"] == "undone"
    assert partially_undone_run["operations"][1]["status"] == "applied"

    contacts_after_operation_undo = list(
        (await db_session.execute(select(Contact).order_by(Contact.nom))).scalars()
    )
    assert [(contact.nom, contact.prenom) for contact in contacts_after_operation_undo] == [
        ("BE NGUYEN", "Thi"),
    ]

    redo_operation_response = await client.post(
        f"/api/import/operations/{operation_id}/redo",
        headers=auth_headers,
    )

    assert redo_operation_response.status_code == 200
    redone_run = redo_operation_response.json()
    assert redone_run["status"] == "completed"
    assert redone_run["operations"][0]["status"] == "applied"
    assert redone_run["operations"][1]["status"] == "applied"

    contacts_after_operation_redo = list(
        (await db_session.execute(select(Contact).order_by(Contact.nom))).scalars()
    )
    assert [(contact.nom, contact.prenom) for contact in contacts_after_operation_redo] == [
        ("BE NGUYEN", "Thi"),
        ("LOPES", "Christine"),
    ]


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_reversible_import_run_undo_rejects_manually_modified_object(
    client: AsyncClient, auth_headers: dict, db_session
) -> None:
    """Undo must fail once an imported object diverged from the expected strict state."""
    from backend.models.contact import Contact

    content = _make_multi_sheet_xlsx(
        {
            "Contacts": (
                ["Nom", "Email"],
                [["Christine LOPES", "christine@example.test"]],
            )
        }
    )

    prepare_response = await client.post(
        "/api/import/runs/prepare/gestion",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )
    run_id = prepare_response.json()["id"]

    execute_response = await client.post(
        f"/api/import/runs/{run_id}/execute",
        headers=auth_headers,
    )

    assert execute_response.status_code == 200

    imported_contact = (
        await db_session.execute(select(Contact).where(Contact.nom == "LOPES"))
    ).scalar_one()

    update_response = await client.put(
        f"/api/contacts/{imported_contact.id}",
        json={"notes": "Retouche manuelle"},
        headers=auth_headers,
    )

    assert update_response.status_code == 200
    assert update_response.json()["notes"] == "Retouche manuelle"

    undo_response = await client.post(
        f"/api/import/runs/{run_id}/undo",
        headers=auth_headers,
    )

    assert undo_response.status_code == 409
    assert undo_response.json()["detail"] == "L'état courant ne correspond plus à l'état attendu"


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_preview_and_import_gestion_accept_payment_matched_by_contact(
    client: AsyncClient, auth_headers: dict
) -> None:
    """Preview and import should both accept payments matched through the contact."""
    content = _make_multi_sheet_xlsx(
        {
            "Factures": (
                ["Date facture", "Réf facture", "Client", "Montant"],
                [["2025-08-01", "2025-0142", "Christine LOPES", 55]],
            ),
            "Paiements": (
                ["Adhérent", "Montant", "Date paiement", "Mode de règlement"],
                [["Christine LOPES", 55, "2025-08-02", "Chèque"]],
            ),
        }
    )

    preview_response = await client.post(
        "/api/import/excel/gestion/preview",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert preview_response.status_code == 200
    preview_data = preview_response.json()
    assert preview_data["estimated_invoices"] == 1
    assert preview_data["estimated_payments"] == 1
    assert preview_data["can_import"] is True

    import_response = await client.post(
        "/api/import/excel/gestion",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert import_response.status_code == 200
    import_data = import_response.json()
    assert import_data["invoices_created"] == 1
    assert import_data["payments_created"] == 1


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_import_gestion_applies_default_due_date_to_imported_invoice(
    client: AsyncClient,
    auth_headers: dict,
    db_session,
) -> None:
    from sqlalchemy import select

    from backend.models.invoice import Invoice

    settings_response = await client.put(
        "/api/settings/",
        json={"default_invoice_due_days": 30},
        headers=auth_headers,
    )
    assert settings_response.status_code == 200

    content = _make_multi_sheet_xlsx(
        {
            "Factures": (
                ["Date facture", "Réf facture", "Client", "Montant"],
                [["2025-08-01", "2025-0144", "Christine LOPES", 55]],
            ),
        }
    )

    import_response = await client.post(
        "/api/import/excel/gestion",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert import_response.status_code == 200
    invoice = (
        (await db_session.execute(select(Invoice).where(Invoice.number == "2025-0144")))
        .scalars()
        .one()
    )
    assert invoice.due_date == date(2025, 8, 31)


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_import_gestion_splits_cs_a_invoice_when_component_columns_exist(
    client: AsyncClient,
    auth_headers: dict,
    db_session,
) -> None:
    from sqlalchemy.orm import selectinload

    from backend.models.accounting_entry import AccountingEntry, EntrySourceType
    from backend.models.invoice import Invoice, InvoiceLabel, InvoiceLineType
    from backend.services.accounting_engine import seed_default_rules

    await seed_default_rules(db_session)

    content = _make_multi_sheet_xlsx(
        {
            "Factures": (
                [
                    "Date facture",
                    "Réf facture",
                    "Client",
                    "Montant",
                    "Montant cours",
                    "Montant adhésion",
                    "Type",
                ],
                [["2024-08-26", "2024-0186", "Alexandre ELEZI", 160, 130, 30, "cs+a"]],
            ),
        }
    )

    import_response = await client.post(
        "/api/import/excel/gestion",
        files={"file": ("Gestion 2024.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert import_response.status_code == 200
    import_data = import_response.json()
    assert import_data["invoices_created"] == 1
    assert import_data["entries_created"] == 3

    invoice = (
        (
            await db_session.execute(
                select(Invoice)
                .where(Invoice.number == "2024-0186")
                .options(selectinload(Invoice.lines))
            )
        )
        .scalars()
        .one()
    )
    assert [(line.description, line.line_type, line.amount) for line in invoice.lines] == [
        ("Cours de soutien", InvoiceLineType.COURSE, Decimal("130.00")),
        ("Adhesion annuelle", InvoiceLineType.ADHESION, Decimal("30.00")),
    ]
    assert invoice.has_explicit_breakdown is True
    assert invoice.label == InvoiceLabel.CS_ADHESION

    entries = (
        (
            await db_session.execute(
                select(AccountingEntry)
                .where(AccountingEntry.source_type == EntrySourceType.INVOICE)
                .where(AccountingEntry.source_id == invoice.id)
                .order_by(AccountingEntry.entry_number.asc())
            )
        )
        .scalars()
        .all()
    )
    assert len(entries) == 3
    assert any(
        entry.account_number == "411100" and entry.debit == Decimal("160.00") for entry in entries
    )
    assert any(
        entry.account_number == "706110" and entry.credit == Decimal("130.00") for entry in entries
    )
    assert any(
        entry.account_number == "756000" and entry.credit == Decimal("30.00") for entry in entries
    )


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_import_gestion_maps_mixed_client_invoice_to_other_line_when_no_split_exists(
    client: AsyncClient,
    auth_headers: dict,
    db_session,
) -> None:
    from sqlalchemy.orm import selectinload

    from backend.models.accounting_entry import AccountingEntry, EntrySourceType
    from backend.models.invoice import Invoice, InvoiceLabel, InvoiceLineType
    from backend.services.accounting_engine import seed_default_rules

    await seed_default_rules(db_session)

    content = _make_multi_sheet_xlsx(
        {
            "Factures": (
                ["Date facture", "Réf facture", "Client", "Montant", "Type"],
                [["2024-08-30", "2024-0192", "Aleksander ELEZI", 100, "cs+a"]],
            ),
        }
    )

    import_response = await client.post(
        "/api/import/excel/gestion",
        files={"file": ("Gestion 2024.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert import_response.status_code == 200
    import_data = import_response.json()
    assert import_data["invoices_created"] == 1
    assert import_data["entries_created"] == 2

    invoice = (
        (
            await db_session.execute(
                select(Invoice)
                .where(Invoice.number == "2024-0192")
                .options(selectinload(Invoice.lines))
            )
        )
        .scalars()
        .one()
    )
    assert invoice.label == InvoiceLabel.GENERAL
    assert invoice.has_explicit_breakdown is False
    assert [(line.line_type, line.amount) for line in invoice.lines] == [
        (InvoiceLineType.OTHER, Decimal("100.00"))
    ]

    entries = (
        (
            await db_session.execute(
                select(AccountingEntry)
                .where(AccountingEntry.source_type == EntrySourceType.INVOICE)
                .where(AccountingEntry.source_id == invoice.id)
                .order_by(AccountingEntry.entry_number.asc())
            )
        )
        .scalars()
        .all()
    )
    assert len(entries) == 2
    assert any(
        entry.account_number == "758000" and entry.credit == Decimal("100.00") for entry in entries
    )


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_preview_gestion_blocks_inconsistent_cs_a_component_columns(
    client: AsyncClient,
    auth_headers: dict,
) -> None:
    content = _make_multi_sheet_xlsx(
        {
            "Factures": (
                [
                    "Date facture",
                    "Réf facture",
                    "Client",
                    "Montant",
                    "Montant cours",
                    "Montant adhésion",
                    "Type",
                ],
                [["2024-08-26", "2024-0186", "Alexandre ELEZI", 160, 120, 30, "cs+a"]],
            ),
        }
    )

    preview_response = await client.post(
        "/api/import/excel/gestion/preview",
        files={"file": ("Gestion 2024.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert preview_response.status_code == 200
    preview_data = preview_response.json()
    sheets = {sheet["name"]: sheet for sheet in preview_data["sheets"]}
    assert preview_data["can_import"] is False
    assert preview_data["estimated_invoices"] == 0
    assert sheets["Factures"]["rows"] == 0
    assert any(
        "ventilation cours/adhesion incoherente" in error.lower()
        for error in sheets["Factures"]["errors"]
    )


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_preview_and_import_gestion_create_supplier_invoice_and_payment_from_bank_row(
    client: AsyncClient,
    auth_headers: dict,
    db_session,
) -> None:
    """Banque rows with an FF reference should create a supplier invoice and its payment."""
    from backend.models.bank import BankTransaction
    from backend.models.contact import Contact, ContactType
    from backend.models.invoice import Invoice, InvoiceType
    from backend.models.payment import Payment, PaymentMethod

    content = _make_multi_sheet_xlsx(
        {
            "Banque": (
                ["Date", "Montant", "Référence", "Libellé", "Solde"],
                [
                    [
                        "2025-11-30",
                        -15,
                        "VIR INST FF-2025113019.02.03 VG53343MNDFS3S01",
                        "COJFA - 2025.10 - FF-2025113019.02.01",
                        1298.13,
                    ]
                ],
            )
        }
    )

    preview_response = await client.post(
        "/api/import/excel/gestion/preview",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert preview_response.status_code == 200
    preview_data = preview_response.json()
    sheets = {sheet["name"]: sheet for sheet in preview_data["sheets"]}
    assert preview_data["can_import"] is True
    assert preview_data["estimated_invoices"] == 1
    assert preview_data["estimated_payments"] == 1
    assert preview_data["estimated_contacts"] == 1
    assert sheets["Banque"]["rows"] == 1

    import_response = await client.post(
        "/api/import/excel/gestion",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert import_response.status_code == 200
    import_data = import_response.json()
    assert import_data["bank_created"] == 1
    assert import_data["invoices_created"] == 1
    assert import_data["payments_created"] == 1

    contact = (await db_session.execute(select(Contact))).scalar_one()
    assert contact.nom == "COJFA"
    assert contact.prenom is None
    assert contact.type == ContactType.FOURNISSEUR

    invoice = (await db_session.execute(select(Invoice))).scalar_one()
    assert invoice.number == "FF-2025113019.02.01"
    assert invoice.type == InvoiceType.FOURNISSEUR
    assert invoice.reference == "FF-2025113019.02.01"
    assert invoice.description == "COJFA - 2025.10"
    assert invoice.total_amount == 15
    assert invoice.paid_amount == 15

    payment = (await db_session.execute(select(Payment))).scalar_one()
    assert payment.invoice_id == invoice.id
    assert payment.contact_id == contact.id
    assert payment.method == PaymentMethod.VIREMENT
    assert payment.reference == "FF-2025113019.02.01"

    bank_entry = (await db_session.execute(select(BankTransaction))).scalar_one()
    assert bank_entry.reference == "VIR INST FF-2025113019.02.03 VG53343MNDFS3S01"
    assert bank_entry.description == "COJFA - 2025.10 - FF-2025113019.02.01"


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_preview_and_import_gestion_create_client_payment_from_bank_virement_row(
    client: AsyncClient,
    auth_headers: dict,
    db_session,
) -> None:
    """A positive bank transfer referencing one client invoice should create a payment."""
    from backend.models.bank import BankTransaction
    from backend.models.invoice import Invoice
    from backend.models.payment import Payment, PaymentMethod

    content = _make_multi_sheet_xlsx(
        {
            "Factures": (
                ["Date facture", "Réf facture", "Client", "Montant"],
                [["2025-08-01", "2025-0142", "Christine LOPES", 55]],
            ),
            "Banque": (
                ["Date", "Montant", "Référence", "Libellé", "Solde"],
                [["2025-08-02", 55, "2025-0142", "Paiement facture client", 537]],
            ),
        }
    )

    preview_response = await client.post(
        "/api/import/excel/gestion/preview",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert preview_response.status_code == 200
    preview_data = preview_response.json()
    assert preview_data["estimated_invoices"] == 1
    assert preview_data["estimated_payments"] == 1
    assert preview_data["can_import"] is True

    import_response = await client.post(
        "/api/import/excel/gestion",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert import_response.status_code == 200
    import_data = import_response.json()
    assert import_data["invoices_created"] == 1
    assert import_data["payments_created"] == 1
    assert import_data["bank_created"] == 1

    invoice = (await db_session.execute(select(Invoice))).scalar_one()
    payment = (await db_session.execute(select(Payment))).scalar_one()
    bank_entry = (await db_session.execute(select(BankTransaction))).scalar_one()

    assert payment.invoice_id == invoice.id
    assert payment.contact_id == invoice.contact_id
    assert payment.method == PaymentMethod.VIREMENT
    assert payment.reference == "2025-0142"
    assert payment.deposited is True
    assert payment.deposit_date == date(2025, 8, 2)
    assert bank_entry.reference == "2025-0142"
    assert bank_entry.description == "Paiement facture client"


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_preview_and_import_gestion_create_supplier_invoice_and_payment_from_cash_row(
    client: AsyncClient,
    auth_headers: dict,
    db_session,
) -> None:
    """Caisse rows with an FF reference should create a supplier invoice and its payment."""
    from backend.models.cash import CashRegister
    from backend.models.contact import Contact, ContactType
    from backend.models.invoice import Invoice, InvoiceType
    from backend.models.payment import Payment, PaymentMethod

    content = _make_multi_sheet_xlsx(
        {
            "Caisse": (
                ["Date", "Montant", "Tiers", "Commentaire", "Solde"],
                [["2025-08-04", -20, "Papeterie ABC", "FF-2025080412.00.00", 100]],
            )
        }
    )

    preview_response = await client.post(
        "/api/import/excel/gestion/preview",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert preview_response.status_code == 200
    preview_data = preview_response.json()
    sheets = {sheet["name"]: sheet for sheet in preview_data["sheets"]}
    assert preview_data["can_import"] is True
    assert preview_data["estimated_invoices"] == 1
    assert preview_data["estimated_payments"] == 1
    assert preview_data["estimated_contacts"] == 1
    assert sheets["Caisse"]["rows"] == 1

    import_response = await client.post(
        "/api/import/excel/gestion",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert import_response.status_code == 200
    import_data = import_response.json()
    assert import_data["cash_created"] == 1
    assert import_data["invoices_created"] == 1
    assert import_data["payments_created"] == 1

    contact = (await db_session.execute(select(Contact))).scalar_one()
    assert contact.nom == "Papeterie ABC"
    assert contact.prenom is None
    assert contact.type == ContactType.FOURNISSEUR

    invoice = (await db_session.execute(select(Invoice))).scalar_one()
    assert invoice.number == "FF-2025080412.00.00"
    assert invoice.type == InvoiceType.FOURNISSEUR
    assert invoice.reference == "FF-2025080412.00.00"
    assert invoice.description == "Papeterie ABC"
    assert invoice.total_amount == 20
    assert invoice.paid_amount == 20

    payment = (await db_session.execute(select(Payment))).scalar_one()
    assert payment.invoice_id == invoice.id
    assert payment.contact_id == contact.id
    assert payment.method == PaymentMethod.ESPECES
    assert payment.reference == "FF-2025080412.00.00"

    cash_entry = (await db_session.execute(select(CashRegister))).scalar_one()
    assert cash_entry.reference == "FF-2025080412.00.00"
    assert cash_entry.contact_id == contact.id
    assert cash_entry.payment_id == payment.id
    assert cash_entry.balance_after == Decimal("-20")


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_preview_and_import_gestion_generate_internal_transfer_entries_from_bank_rows(
    client: AsyncClient,
    auth_headers: dict,
    db_session,
) -> None:
    """Gestion import should generate internal transfer accounting entries for savings moves."""
    from backend.models.accounting_entry import AccountingEntry, EntrySourceType
    from backend.services.accounting_engine import seed_default_rules

    await seed_default_rules(db_session)
    await db_session.commit()

    content = _make_multi_sheet_xlsx(
        {
            "Banque": (
                ["Date", "Montant", "Référence", "Libellé", "Solde"],
                [
                    ["2025-09-06", 3000, "VIR DE LES ETUDES", "Virement interne", 5000],
                    ["2026-02-03", -2000, "VIR VERS LIVRET", "Virement interne", 3000],
                ],
            ),
        }
    )

    preview_response = await client.post(
        "/api/import/excel/gestion/preview",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert preview_response.status_code == 200
    assert preview_response.json()["can_import"] is True

    import_response = await client.post(
        "/api/import/excel/gestion",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert import_response.status_code == 200
    import_data = import_response.json()
    assert import_data["bank_created"] == 2
    assert import_data["entries_created"] == 4

    gestion_entries = list(
        (
            await db_session.execute(
                select(AccountingEntry)
                .where(AccountingEntry.source_type == EntrySourceType.GESTION)
                .order_by(AccountingEntry.date, AccountingEntry.entry_number)
            )
        ).scalars()
    )

    assert {
        (entry.account_number, Decimal(str(entry.debit)), Decimal(str(entry.credit)))
        for entry in gestion_entries
    } == {
        ("512100", Decimal("3000"), Decimal("0")),
        ("512102", Decimal("0"), Decimal("3000")),
        ("512102", Decimal("2000"), Decimal("0")),
        ("512100", Decimal("0"), Decimal("2000")),
    }


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_preview_and_import_gestion_generate_direct_bank_entries_and_deposit(
    client: AsyncClient,
    auth_headers: dict,
    db_session,
) -> None:
    """Gestion import should generate direct bank/deposit accounting coverage for known cases."""
    from backend.models.accounting_entry import AccountingEntry, EntrySourceType
    from backend.models.bank import BankTransaction
    from backend.models.invoice import Invoice
    from backend.models.payment import Payment
    from backend.services.accounting_engine import seed_default_rules

    await seed_default_rules(db_session)
    await db_session.commit()

    content = _make_multi_sheet_xlsx(
        {
            "Factures": (
                ["Date facture", "Réf facture", "Client", "Montant"],
                [["2025-08-01", "2025-0142", "Christine LOPES", 33]],
            ),
            "Paiements": (
                [
                    "Réf facture",
                    "Réf paiement",
                    "Adhérent",
                    "Montant",
                    "Date paiement",
                    "Numéro du chèque",
                    "Encaissé",
                    "Date encaissement",
                ],
                [
                    [
                        "2025-0142",
                        "Chèque",
                        "Christine LOPES",
                        33,
                        "2025-07-09",
                        "2025.08.02.01",
                        "OUI",
                        "2025-08-02",
                    ]
                ],
            ),
            "Banque": (
                ["Date", "Montant", "Référence", "Libellé", "Solde"],
                [
                    ["2025-08-02", 33, "REF05001A05", "Remise chèques", 599.76],
                    ["2025-08-11", -14.15, "SGT25050010010202", "Frais bancaires", 1069.61],
                    ["2025-08-20", -291.30, "PRLV SEPA MAIF", "Assurance MAIF - 2025", 778.31],
                    [
                        "2025-09-15",
                        -1474,
                        "TT404172509101922125314073410000511",
                        "Charges sociales (URSSAF) - 2025.07",
                        3265.82,
                    ],
                    [
                        "2025-02-18",
                        800,
                        "160570300002002025-02-141849 ACTE 131304",
                        "Subvention de la mairie (animations hiver 2025)",
                        3745.73,
                    ],
                    ["2025-12-31", 1563.47, "CM LIVRET", "INTERETS 2024 LIVRET BLEU", 5309.20],
                ],
            ),
        }
    )

    preview_response = await client.post(
        "/api/import/excel/gestion/preview",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert preview_response.status_code == 200
    assert preview_response.json()["can_import"] is True

    import_response = await client.post(
        "/api/import/excel/gestion",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert import_response.status_code == 200
    import_data = import_response.json()
    assert import_data["invoices_created"] == 1
    assert import_data["payments_created"] == 1
    assert import_data["bank_created"] == 6
    assert import_data["entries_created"] >= 12

    invoice = (await db_session.execute(select(Invoice))).scalar_one()
    payment = (await db_session.execute(select(Payment))).scalar_one()
    bank_rows = list(
        (
            await db_session.execute(
                select(BankTransaction).order_by(BankTransaction.date, BankTransaction.id)
            )
        ).scalars()
    )
    entries = list(
        (
            await db_session.execute(
                select(AccountingEntry).order_by(AccountingEntry.date, AccountingEntry.entry_number)
            )
        ).scalars()
    )

    assert invoice.number == "2025-0142"
    assert payment.reference is None
    assert len(bank_rows) == 6

    gestion_entries = [entry for entry in entries if entry.source_type == EntrySourceType.GESTION]

    assert {
        (entry.account_number, Decimal(str(entry.debit)), Decimal(str(entry.credit)))
        for entry in gestion_entries
    } >= {
        ("512100", Decimal("33"), Decimal("0")),
        ("511200", Decimal("0"), Decimal("33")),
        ("627000", Decimal("14.15"), Decimal("0")),
        ("512100", Decimal("0"), Decimal("14.15")),
        ("616100", Decimal("291.30"), Decimal("0")),
        ("512100", Decimal("0"), Decimal("291.30")),
        ("431100", Decimal("1474"), Decimal("0")),
        ("512100", Decimal("0"), Decimal("1474")),
        ("512100", Decimal("800"), Decimal("0")),
        ("740000", Decimal("0"), Decimal("800")),
        ("512102", Decimal("1563.47"), Decimal("0")),
        ("768100", Decimal("0"), Decimal("1563.47")),
    }


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_preview_and_import_gestion_import_salary_sheet(
    client: AsyncClient,
    auth_headers: dict,
    db_session,
) -> None:
    """Gestion import should create salary records and grouped salary accounting entries."""
    from backend.models.accounting_entry import AccountingEntry, EntrySourceType
    from backend.models.salary import Salary

    content = _make_multi_sheet_xlsx_rows(
        {
            "Aide Salaires": [
                ["2025.07"],
                ["NOM", "Heures", "Brut", "Salariales", "Patronales", "Impôts", "Net"],
                ["LAY", 12, 1200, 150, 250, 10, 1040],
                ["WOLFF P.", 8, 800, 100, 180, 0, 700],
                ["Total", None, None, None, None, None, None],
            ]
        }
    )

    preview_response = await client.post(
        "/api/import/excel/gestion/preview",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert preview_response.status_code == 200
    preview_data = preview_response.json()
    sheets = {sheet["name"]: sheet for sheet in preview_data["sheets"]}
    assert preview_data["estimated_salaries"] == 2
    assert sheets["Aide Salaires"]["kind"] == "salaries"
    assert sheets["Aide Salaires"]["rows"] == 2

    import_response = await client.post(
        "/api/import/excel/gestion",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert import_response.status_code == 200
    import_data = import_response.json()
    assert import_data["contacts_created"] == 2
    assert import_data["salaries_created"] == 2
    assert import_data["entries_created"] == 9

    salaries = list(
        (await db_session.execute(select(Salary).order_by(Salary.month, Salary.id))).scalars()
    )
    assert len(salaries) == 2
    assert {salary.month for salary in salaries} == {"2025-07"}

    salary_entries = list(
        (
            await db_session.execute(
                select(AccountingEntry)
                .where(AccountingEntry.source_type == EntrySourceType.SALARY)
                .order_by(AccountingEntry.entry_number)
            )
        ).scalars()
    )
    assert len(salary_entries) == 9
    assert {entry.group_key for entry in salary_entries} == {
        "salary-import:2025-07:accrual",
        "salary-import:2025-07:payment",
    }
    tax_entry = next((entry for entry in salary_entries if entry.account_number == "443000"), None)
    assert tax_entry is not None
    assert tax_entry.label == "Impôt sur le revenu 2025.07"
    assert tax_entry.debit == 0
    assert tax_entry.credit == 10


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_import_gestion_attaches_july_salary_entries_to_previous_fiscal_year(
    client: AsyncClient,
    auth_headers: dict,
    db_session,
) -> None:
    """July salary entries should stay inside the fiscal year ending on July 31."""
    from backend.models.accounting_entry import AccountingEntry, EntrySourceType
    from backend.models.fiscal_year import FiscalYear, FiscalYearStatus

    fiscal_year_2024 = FiscalYear(
        name="2024",
        start_date=date(2024, 8, 1),
        end_date=date(2025, 7, 31),
        status=FiscalYearStatus.OPEN,
    )
    fiscal_year_2025 = FiscalYear(
        name="2025",
        start_date=date(2025, 8, 1),
        end_date=date(2026, 7, 31),
        status=FiscalYearStatus.OPEN,
    )
    db_session.add_all([fiscal_year_2024, fiscal_year_2025])
    await db_session.commit()

    content = _make_multi_sheet_xlsx_rows(
        {
            "Aide Salaires": [
                ["2025.07"],
                ["NOM", "Heures", "Brut", "Salariales", "Patronales", "Impôts", "Net"],
                ["LAY", 12, 1200, 150, 250, 10, 1040],
            ]
        }
    )

    import_response = await client.post(
        "/api/import/excel/gestion",
        files={"file": ("Gestion 2024.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert import_response.status_code == 200

    salary_entries = list(
        (
            await db_session.execute(
                select(AccountingEntry)
                .where(AccountingEntry.source_type == EntrySourceType.SALARY)
                .order_by(AccountingEntry.entry_number)
            )
        ).scalars()
    )

    assert salary_entries
    assert {entry.date for entry in salary_entries} == {date(2025, 7, 1)}
    assert {entry.fiscal_year_id for entry in salary_entries} == {fiscal_year_2024.id}


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_import_gestion_salary_detailed_format_stores_cdd_fields(
    client: AsyncClient,
    auth_headers: dict,
    db_session,
) -> None:
    """Detailed Aide Salaires format should populate CDD-specific fields on CDD rows."""
    from backend.models.salary import Salary

    # Reproduce the "detailed" header format: month row col 1 == "Heures" triggers detailed mode
    content = _make_multi_sheet_xlsx_rows(
        {
            "Aide Salaires": [
                # Detailed header: month + column headers
                [
                    "2025.09",
                    "Heures",
                    "Brut déclaré",
                    "Congés",
                    "Précarité",
                    "Brut total",
                    "Brut",
                    "Salariales",
                    "Patronales",
                    "Impôts",
                    "Net",
                ],
                # CDI row: conges=0, precarite=0 → CDD fields should be None
                ["LAY", 104, 1605, 0, 0, 1605, 1605, 355, 350, 0, 1250],
                # CDD row: conges>0 → CDD fields should be stored
                ["WOLFF", 8, 200, 20, 22, 242, 242, 51, 97, 0, 191],
                ["Total charges fixes", 500],
            ]
        }
    )

    import_response = await client.post(
        "/api/import/excel/gestion",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert import_response.status_code == 200
    data = import_response.json()
    assert data["contacts_created"] == 2
    assert data["salaries_created"] == 2

    from sqlalchemy.orm import selectinload

    salary_rows = list(
        (
            await db_session.execute(
                select(Salary).options(selectinload(Salary.employee)).order_by(Salary.id)
            )
        ).scalars()
    )
    assert len(salary_rows) == 2

    lay_salary = next(s for s in salary_rows if "LAY" in s.employee.nom.upper())
    wolff_salary = next(s for s in salary_rows if "WOLFF" in s.employee.nom.upper())

    # CDI row: CDD fields must be None
    assert lay_salary.brut_declared is None
    assert lay_salary.conges_payes is None
    assert lay_salary.precarite is None

    # CDD row: CDD fields must be populated
    from decimal import Decimal

    assert wolff_salary.brut_declared == Decimal("200")
    assert wolff_salary.conges_payes == Decimal("20")
    assert wolff_salary.precarite == Decimal("22")
    assert wolff_salary.gross == Decimal("242")


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_import_comptabilite_ignores_salary_groups_already_generated_from_gestion(
    client: AsyncClient,
    auth_headers: dict,
    db_session,
) -> None:
    """Salary journal groups already recreated from Gestion should be ignored on compta import."""
    from backend.models.accounting_entry import AccountingEntry, EntrySourceType

    gestion_content = _make_multi_sheet_xlsx_rows(
        {
            "Aide Salaires": [
                ["2025.07"],
                ["NOM", "Heures", "Brut", "Salariales", "Patronales", "Impôts", "Net"],
                ["LAY", 12, 1200, 150, 250, 10, 1040],
                ["WOLFF", 8, 800, 100, 180, 0, 700],
            ]
        }
    )
    gestion_response = await client.post(
        "/api/import/excel/gestion",
        files={"file": ("Gestion 2025.xlsx", gestion_content, _XLSX_MIME)},
        headers=auth_headers,
    )
    assert gestion_response.status_code == 200
    assert gestion_response.json()["salaries_created"] == 2

    compta_content = _make_multi_sheet_xlsx(
        {
            "Journal": (
                ["Date", "N° compte", "Libellé de l'écriture", "Débit", "Crédit", "ChangeNum"],
                [
                    ["2025-07-31", "645100", "Charges patronales 2025.07", 430, None, "A"],
                    ["2025-07-31", "641000", "Salaires bruts 2025.07", 2000, None, "A"],
                    ["2025-07-31", "421000", "Salaires nets 2025.07", None, 1740, "A"],
                    ["2025-07-31", "431100", "Charges patronales 2025.07", None, 430, "A"],
                    ["2025-07-31", "431100", "Charges salariales 2025.07", None, 250, "A"],
                    ["2025-07-31", "443000", "Impôt sur le revenu 2025.07", None, 10, "A"],
                    ["2025-07-31", "421000", "Paiement salaires 2025.07", 1740, None, "B"],
                    ["2025-07-31", "512100", "Salaire LAY 2025.07", None, 1040, "B"],
                    ["2025-07-31", "512100", "Salaire WOLFF 2025.07", None, 700, "B"],
                ],
            )
        }
    )

    preview_response = await client.post(
        "/api/import/excel/comptabilite/preview",
        files={"file": ("Comptabilite 2025.xlsx", compta_content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert preview_response.status_code == 200
    preview_data = preview_response.json()
    assert preview_data["estimated_entries"] == 0

    import_response = await client.post(
        "/api/import/excel/comptabilite",
        files={"file": ("Comptabilite 2025.xlsx", compta_content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert import_response.status_code == 200
    import_data = import_response.json()
    assert import_data["entries_created"] == 0
    assert import_data["ignored_rows"] == 9

    manual_entries = list(
        (
            await db_session.execute(
                select(AccountingEntry).where(AccountingEntry.source_type == EntrySourceType.MANUAL)
            )
        ).scalars()
    )
    assert not manual_entries


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_import_comptabilite_ignores_salary_groups_with_float_rounding_noise(
    client: AsyncClient,
    auth_headers: dict,
) -> None:
    """Salary groups should still be ignored when Excel numbers carry float noise."""
    gestion_content = _make_multi_sheet_xlsx_rows(
        {
            "Aide Salaires": [
                ["2024.08"],
                ["NOM", "Heures", "Brut", "Salariales", "Patronales", "Impôts", "Net"],
                ["LAY", 12, 1605, 355.36, 370.86, 0, 1249.64],
            ]
        }
    )
    gestion_response = await client.post(
        "/api/import/excel/gestion",
        files={"file": ("Gestion 2024.xlsx", gestion_content, _XLSX_MIME)},
        headers=auth_headers,
    )
    assert gestion_response.status_code == 200

    compta_content = _make_multi_sheet_xlsx(
        {
            "Journal": (
                ["Date", "N° compte", "Libellé de l'écriture", "Débit", "Crédit", "ChangeNum"],
                [
                    [
                        "2024-09-07",
                        "645100",
                        "Charges patronales 2024.08",
                        370.86000000000005,
                        None,
                        "A",
                    ],
                    ["2024-09-07", "641000", "Salaires bruts 2024.08", 1605, None, "A"],
                    [
                        "2024-09-07",
                        "421000",
                        "Salaires nets 2024.08",
                        None,
                        1249.6399999999999,
                        "A",
                    ],
                    [
                        "2024-09-07",
                        "431100",
                        "Charges patronales 2024.08",
                        None,
                        370.86000000000005,
                        "A",
                    ],
                    ["2024-09-07", "431100", "Charges salariales 2024.08", None, 355.36, "A"],
                    [
                        "2024-09-07",
                        "421000",
                        "Paiement salaires 2024.08",
                        1249.6399999999999,
                        None,
                        "B",
                    ],
                    ["2024-09-07", "512100", "Salaire LAY 2024.08", None, 1249.6399999999999, "B"],
                ],
            )
        }
    )

    import_response = await client.post(
        "/api/import/excel/comptabilite",
        files={"file": ("Comptabilite 2024.xlsx", compta_content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert import_response.status_code == 200
    import_data = import_response.json()
    assert import_data["entries_created"] == 0
    assert import_data["ignored_rows"] == 7


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_import_comptabilite_ignores_standalone_salary_tax_line_from_gestion_month(
    client: AsyncClient,
    auth_headers: dict,
    db_session,
) -> None:
    """A standalone IR/PAS line should be ignored when Gestion already recreated that month."""
    from backend.models.accounting_entry import AccountingEntry, EntrySourceType

    gestion_content = _make_multi_sheet_xlsx_rows(
        {
            "Aide Salaires": [
                ["2024.11"],
                ["NOM", "Heures", "Brut", "Salariales", "Patronales", "Impôts", "Net"],
                ["LAY", 12, 1804.99, 398.85, 439.33, 1.3, 1404.84],
            ]
        }
    )
    gestion_response = await client.post(
        "/api/import/excel/gestion",
        files={"file": ("Gestion 2024.xlsx", gestion_content, _XLSX_MIME)},
        headers=auth_headers,
    )
    assert gestion_response.status_code == 200

    compta_content = _make_multi_sheet_xlsx(
        {
            "Journal": (
                ["Date", "N° compte", "Libellé de l'écriture", "Débit", "Crédit", "ChangeNum"],
                [["2024-12-01", "431100", "Impôt sur le revenu 2024.11", None, 1.3, "C"]],
            )
        }
    )

    preview_response = await client.post(
        "/api/import/excel/comptabilite/preview",
        files={"file": ("Comptabilite 2024.xlsx", compta_content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert preview_response.status_code == 200
    assert preview_response.json()["estimated_entries"] == 0

    import_response = await client.post(
        "/api/import/excel/comptabilite",
        files={"file": ("Comptabilite 2024.xlsx", compta_content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert import_response.status_code == 200
    import_data = import_response.json()
    assert import_data["entries_created"] == 0
    assert import_data["ignored_rows"] == 1

    manual_entries = list(
        (
            await db_session.execute(
                select(AccountingEntry).where(AccountingEntry.source_type == EntrySourceType.MANUAL)
            )
        ).scalars()
    )
    assert not manual_entries


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_import_comptabilite_ignores_salary_groups_with_variants(
    client: AsyncClient,
    auth_headers: dict,
    db_session,
) -> None:
    """Salary-like accrual and payment groups should be ignored.

    This still applies when the imported journal uses variant splits.
    """
    from backend.models.accounting_entry import AccountingEntry, EntrySourceType

    gestion_content = _make_multi_sheet_xlsx_rows(
        {
            "Aide Salaires": [
                ["2024.12"],
                ["NOM", "Heures", "Brut", "Salariales", "Patronales", "Impôts", "Net"],
                ["LAY", 12, 1200, 250, 250, 0.65, 949.35],
                ["WOLFF", 8, 504.99, 127.5, 133.84, 0, 377.49],
            ]
        }
    )
    gestion_response = await client.post(
        "/api/import/excel/gestion",
        files={"file": ("Gestion 2024.xlsx", gestion_content, _XLSX_MIME)},
        headers=auth_headers,
    )
    assert gestion_response.status_code == 200
    assert gestion_response.json()["salaries_created"] == 2

    compta_content = _make_multi_sheet_xlsx(
        {
            "Journal": (
                ["Date", "N° compte", "Libellé de l'écriture", "Débit", "Crédit", "ChangeNum"],
                [
                    ["2024-12-31", "645100", "Charges patronales 2024.12", 391.97, None, "A"],
                    ["2024-12-31", "641000", "Salaires bruts 2024.12", 1704.99, None, "A"],
                    ["2024-12-31", "421000", "Salaires nets 2024.12", None, 1326.69, "A"],
                    ["2024-12-31", "431100", "Charges patronales 2024.12", None, 391.97, "A"],
                    ["2024-12-31", "431100", "Charges salariales 2024.12", None, 377.66, "A"],
                    ["2024-12-31", "431100", "Impôt sur le revenu 2024.12", None, 0.64, "A"],
                    ["2024-12-31", "421000", "Paiement salaires 2024.12", 1326.69, None, "B"],
                    ["2024-12-31", "512100", "Salaire LAY 2024.12", None, 1249.64, "B"],
                    ["2024-12-31", "512100", "Salaire WOLFF 2024.12", None, 77.05, "B"],
                ],
            )
        }
    )

    preview_response = await client.post(
        "/api/import/excel/comptabilite/preview",
        files={"file": ("Comptabilite 2024.xlsx", compta_content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert preview_response.status_code == 200
    assert preview_response.json()["estimated_entries"] == 0

    import_response = await client.post(
        "/api/import/excel/comptabilite",
        files={"file": ("Comptabilite 2024.xlsx", compta_content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert import_response.status_code == 200
    import_data = import_response.json()
    assert import_data["entries_created"] == 0
    assert import_data["ignored_rows"] == 9

    manual_entries = list(
        (
            await db_session.execute(
                select(AccountingEntry).where(AccountingEntry.source_type == EntrySourceType.MANUAL)
            )
        ).scalars()
    )
    assert not manual_entries


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_import_comptabilite_ignores_partial_salary_accrual(
    client: AsyncClient,
    auth_headers: dict,
    db_session,
) -> None:
    """A partial salary accrual group should be ignored.

    The month is already present from Gestion import.
    """
    from backend.models.accounting_entry import AccountingEntry, EntrySourceType

    gestion_content = _make_multi_sheet_xlsx_rows(
        {
            "Aide Salaires": [
                ["2025.01"],
                ["NOM", "Heures", "Brut", "Salariales", "Patronales", "Impôts", "Net"],
                ["LAY", 12, 1704.99, 377.5, 383.84, 0.65, 1326.84],
            ]
        }
    )
    gestion_response = await client.post(
        "/api/import/excel/gestion",
        files={"file": ("Gestion 2025.xlsx", gestion_content, _XLSX_MIME)},
        headers=auth_headers,
    )
    assert gestion_response.status_code == 200
    assert gestion_response.json()["salaries_created"] == 1

    compta_content = _make_multi_sheet_xlsx(
        {
            "Journal": (
                ["Date", "N° compte", "Libellé de l'écriture", "Débit", "Crédit", "ChangeNum"],
                [
                    ["2025-02-01", "645100", "Charges patronales 2025.01", 384.64, None, "A"],
                    ["2025-02-01", "431100", "Charges patronales 2025.01", None, 384.64, "A"],
                    ["2025-02-01", "431100", "Impôt sur le revenu 2025.01", None, 0.65, "A"],
                ],
            )
        }
    )

    preview_response = await client.post(
        "/api/import/excel/comptabilite/preview",
        files={"file": ("Comptabilite 2025.xlsx", compta_content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert preview_response.status_code == 200
    assert preview_response.json()["estimated_entries"] == 0

    import_response = await client.post(
        "/api/import/excel/comptabilite",
        files={"file": ("Comptabilite 2025.xlsx", compta_content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert import_response.status_code == 200
    import_data = import_response.json()
    assert import_data["entries_created"] == 0
    assert import_data["ignored_rows"] == 3

    manual_entries = list(
        (
            await db_session.execute(
                select(AccountingEntry).where(AccountingEntry.source_type == EntrySourceType.MANUAL)
            )
        ).scalars()
    )
    assert not manual_entries


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_preview_and_import_gestion_generate_especes_deposit_from_cash_and_bank_rows(
    client: AsyncClient,
    auth_headers: dict,
    db_session,
) -> None:
    """A bank remittance should create a deposit slip for matching cash payments."""
    from backend.models.accounting_entry import AccountingEntry, EntrySourceType
    from backend.services.accounting_engine import seed_default_rules

    await seed_default_rules(db_session)
    await db_session.commit()
    content = _make_multi_sheet_xlsx(
        {
            "Factures": (
                ["Date facture", "Réf facture", "Client", "Montant"],
                [["2025-08-01", "2025-0142", "Christine LOPES", 20]],
            ),
            "Paiements": (
                [
                    "Réf facture",
                    "Réf paiement",
                    "Adhérent",
                    "Montant",
                    "Date paiement",
                    "Numéro du chèque",
                    "Encaissé",
                    "Date encaissement",
                ],
                [
                    [
                        "2025-0142",
                        "Espèces",
                        "Christine LOPES",
                        20,
                        "2025-08-01",
                        None,
                        "OUI",
                        "2025-08-02",
                    ]
                ],
            ),
            "Caisse": (
                ["Date", "Montant", "Tiers", "Commentaire", "Solde"],
                [["2025-08-02", -20, "CREDIT MUTUEL", "Remise espèces", 0]],
            ),
            "Banque": (
                ["Date", "Montant", "Référence", "Libellé", "Solde"],
                [["2025-08-02", 20, "REF-ESP-01", "Remise espèces", 520]],
            ),
        }
    )

    import_response = await client.post(
        "/api/import/excel/gestion",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert import_response.status_code == 200
    gestion_entries = list(
        (
            await db_session.execute(
                select(AccountingEntry).where(
                    AccountingEntry.source_type == EntrySourceType.GESTION
                )
            )
        ).scalars()
    )

    assert {
        (entry.account_number, Decimal(str(entry.debit)), Decimal(str(entry.credit)))
        for entry in gestion_entries
    } >= {
        ("512100", Decimal("20"), Decimal("0")),
        ("531000", Decimal("0"), Decimal("20")),
    }


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_preview_and_import_gestion_ignore_duplicate_invoice_rows_with_report(
    client: AsyncClient, auth_headers: dict
) -> None:
    """Preview and import should expose duplicate invoice rows as ignored, not blocked."""
    content = _make_multi_sheet_xlsx(
        {
            "Factures": (
                ["Date facture", "Réf facture", "Client", "Montant"],
                [
                    ["2025-08-01", "2025-0142", "Christine LOPES", 55],
                    ["2025-08-02", "2025-0142", "Christine LOPES", 55],
                ],
            ),
        }
    )

    preview_response = await client.post(
        "/api/import/excel/gestion/preview",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert preview_response.status_code == 200
    preview_data = preview_response.json()
    sheets = {sheet["name"]: sheet for sheet in preview_data["sheets"]}
    assert preview_data["can_import"] is True
    assert preview_data["estimated_invoices"] == 1
    assert sheets["Factures"]["rows"] == 1
    assert sheets["Factures"]["ignored_rows"] == 1
    assert sheets["Factures"]["blocked_rows"] == 0
    assert any("ignoree" in warning.lower() for warning in sheets["Factures"]["warnings"])

    import_response = await client.post(
        "/api/import/excel/gestion",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert import_response.status_code == 200
    import_data = import_response.json()
    assert import_data["invoices_created"] == 1
    assert import_data["ignored_rows"] == 1
    assert import_data["blocked_rows"] == 0
    assert any("ignoree" in warning.lower() for warning in import_data["warnings"])
    factures_sheet = next(sheet for sheet in import_data["sheets"] if sheet["name"] == "Factures")
    assert factures_sheet["imported_rows"] == 1
    assert factures_sheet["ignored_rows"] == 1
    assert factures_sheet["blocked_rows"] == 0


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_preview_and_import_gestion_ignore_invoice_total_rows(
    client: AsyncClient, auth_headers: dict
) -> None:
    """Preview and import should ignore invoice total rows instead of blocking the file."""
    content = _make_multi_sheet_xlsx(
        {
            "Factures": (
                ["Date facture", "Réf facture", "Client", "Montant"],
                [
                    ["2025-08-01", "2025-0142", "Christine LOPES", 55],
                    ["Total", None, None, 55],
                ],
            ),
        }
    )

    preview_response = await client.post(
        "/api/import/excel/gestion/preview",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert preview_response.status_code == 200
    preview_data = preview_response.json()
    sheets = {sheet["name"]: sheet for sheet in preview_data["sheets"]}
    assert preview_data["can_import"] is True
    assert preview_data["estimated_invoices"] == 1
    assert sheets["Factures"]["rows"] == 1
    assert sheets["Factures"]["ignored_rows"] == 1
    assert sheets["Factures"]["blocked_rows"] == 0
    assert any("total" in warning.lower() for warning in sheets["Factures"]["warnings"])

    import_response = await client.post(
        "/api/import/excel/gestion",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert import_response.status_code == 200
    import_data = import_response.json()
    assert import_data["invoices_created"] == 1
    assert import_data["ignored_rows"] == 1
    assert import_data["blocked_rows"] == 0
    assert any("total" in warning.lower() for warning in import_data["warnings"])


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_preview_and_import_gestion_ignore_existing_contact_rows_with_report(
    client: AsyncClient,
    auth_headers: dict,
    db_session,
) -> None:
    """Preview should already expose DB-existing contacts as ignored rows."""
    from backend.models.contact import Contact, ContactType

    db_session.add(Contact(nom="Christine LOPES", type=ContactType.CLIENT))
    await db_session.commit()

    content = _make_multi_sheet_xlsx(
        {
            "Contacts": (
                ["Nom", "Email"],
                [["Christine LOPES", "christine@example.test"]],
            )
        }
    )

    preview_response = await client.post(
        "/api/import/excel/gestion/preview",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert preview_response.status_code == 200
    preview_data = preview_response.json()
    sheets = {sheet["name"]: sheet for sheet in preview_data["sheets"]}
    assert preview_data["estimated_contacts"] == 0
    assert preview_data["can_import"] is False
    assert sheets["Contacts"]["rows"] == 0
    assert sheets["Contacts"]["ignored_rows"] == 1
    assert any("deja existant" in warning.lower() for warning in sheets["Contacts"]["warnings"])

    import_response = await client.post(
        "/api/import/excel/gestion",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert import_response.status_code == 200
    import_data = import_response.json()
    assert import_data["contacts_created"] == 0
    assert import_data["ignored_rows"] == 1
    assert any("deja existant" in warning.lower() for warning in import_data["warnings"])


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_preview_and_import_gestion_ignore_existing_invoice_rows_with_report(
    client: AsyncClient,
    auth_headers: dict,
    db_session,
) -> None:
    """Preview should already expose DB-existing invoices as ignored rows."""
    from datetime import date
    from decimal import Decimal

    from backend.models.contact import Contact, ContactType
    from backend.models.invoice import Invoice, InvoiceLabel, InvoiceStatus, InvoiceType

    contact = Contact(nom="Christine LOPES", type=ContactType.CLIENT)
    db_session.add(contact)
    await db_session.flush()
    db_session.add(
        Invoice(
            number="2025-0142",
            type=InvoiceType.CLIENT,
            contact_id=contact.id,
            date=date(2025, 8, 1),
            total_amount=Decimal("55"),
            paid_amount=Decimal("0"),
            status=InvoiceStatus.SENT,
            label=InvoiceLabel.CS,
        )
    )
    await db_session.commit()

    content = _make_multi_sheet_xlsx(
        {
            "Factures": (
                ["Date facture", "Réf facture", "Client", "Montant"],
                [["2025-08-01", "2025-0142", "Christine LOPES", 55]],
            ),
        }
    )

    preview_response = await client.post(
        "/api/import/excel/gestion/preview",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert preview_response.status_code == 200
    preview_data = preview_response.json()
    sheets = {sheet["name"]: sheet for sheet in preview_data["sheets"]}
    assert preview_data["estimated_invoices"] == 0
    assert preview_data["estimated_contacts"] == 0
    assert preview_data["can_import"] is False
    assert sheets["Factures"]["rows"] == 0
    assert sheets["Factures"]["ignored_rows"] == 1
    assert any("deja existante" in warning.lower() for warning in sheets["Factures"]["warnings"])

    import_response = await client.post(
        "/api/import/excel/gestion",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert import_response.status_code == 200
    import_data = import_response.json()
    assert import_data["invoices_created"] == 0
    assert import_data["contacts_created"] == 0
    assert import_data["ignored_rows"] == 1
    assert any("deja existante" in warning.lower() for warning in import_data["warnings"])


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_import_gestion_blocks_reimport_of_same_file_and_logs_attempts(
    client: AsyncClient,
    auth_headers: dict,
    db_session,
) -> None:
    """A successful Gestion import should block exact reimport of the same file."""
    content = _make_multi_sheet_xlsx(
        {
            "Factures": (
                ["Date facture", "Réf facture", "Client", "Montant"],
                [["2025-08-01", "2025-0142", "Christine LOPES", 55]],
            ),
        }
    )

    first_response = await client.post(
        "/api/import/excel/gestion",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert first_response.status_code == 200
    first_data = first_response.json()
    assert first_data["invoices_created"] == 1
    assert first_data["errors"] == []

    preview_response = await client.post(
        "/api/import/excel/gestion/preview",
        files={"file": ("Gestion copie.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert preview_response.status_code == 200
    preview_data = preview_response.json()
    assert preview_data["can_import"] is False
    assert any("deja importe" in error.lower() for error in preview_data["errors"])

    second_response = await client.post(
        "/api/import/excel/gestion",
        files={"file": ("Gestion copie.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert second_response.status_code == 200
    second_data = second_response.json()
    assert second_data["invoices_created"] == 0
    assert any("deja importe" in error.lower() for error in second_data["errors"])

    logs = (
        (
            await db_session.execute(
                select(ImportLog).order_by(ImportLog.created_at.asc(), ImportLog.id.asc())
            )
        )
        .scalars()
        .all()
    )

    assert len(logs) == 2
    assert logs[0].import_type == "gestion"
    assert logs[0].status == "success"
    assert logs[0].file_name == "Gestion 2025.xlsx"
    assert '"invoices_created": 1' in logs[0].summary
    first_summary = json.loads(logs[0].summary or "{}")
    assert any(
        created_object.get("object_type") == "invoice"
        and created_object.get("reference") == "2025-0142"
        for created_object in first_summary.get("created_objects", [])
    )
    assert logs[1].import_type == "gestion"
    assert logs[1].status == "blocked"
    assert logs[1].file_name == "Gestion copie.xlsx"
    assert logs[0].file_hash == logs[1].file_hash


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_import_gestion_blocks_rows_with_missing_required_invoice_data(
    client: AsyncClient, auth_headers: dict
) -> None:
    """Preview and import should block a recognized invoice sheet with invalid rows."""
    content = _make_multi_sheet_xlsx(
        {
            "Factures": (
                ["Date facture", "Réf facture", "Client", "Montant"],
                [
                    ["2025-08-01", "2025-0142", "Christine LOPES", 55],
                    ["2025-08-02", "2025-0143", None, 42],
                ],
            ),
        }
    )

    preview_response = await client.post(
        "/api/import/excel/gestion/preview",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert preview_response.status_code == 200
    preview_data = preview_response.json()
    assert preview_data["can_import"] is False
    assert any("Factures" in error and "Ligne 3" in error for error in preview_data["errors"])
    assert preview_data["error_details"] == [
        {
            "category": "invoice-missing-contact",
            "severity": "error",
            "sheet_name": "Factures",
            "kind": "invoices",
            "row_number": 3,
            "message": "client manquant",
            "display_message": "Factures — Ligne 3 : client manquant",
        }
    ]
    factures_preview_sheet = next(
        sheet for sheet in preview_data["sheets"] if sheet["name"] == "Factures"
    )
    assert factures_preview_sheet["error_details"][0]["row_number"] == 3
    assert factures_preview_sheet["error_details"][0]["message"] == "client manquant"

    import_response = await client.post(
        "/api/import/excel/gestion",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert import_response.status_code == 200
    import_data = import_response.json()
    assert import_data["invoices_created"] == 0
    assert import_data["contacts_created"] == 0
    assert any("Factures" in error and "Ligne 3" in error for error in import_data["errors"])
    assert import_data["error_details"] == [
        {
            "category": "invoice-missing-contact",
            "severity": "error",
            "sheet_name": "Factures",
            "kind": "invoices",
            "row_number": 3,
            "message": "client manquant",
            "display_message": "Factures — Ligne 3 : client manquant",
        }
    ]
    factures_import_sheet = next(
        sheet for sheet in import_data["sheets"] if sheet["name"] == "Factures"
    )
    assert factures_import_sheet["error_details"][0]["row_number"] == 3


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_import_gestion_blocks_invoices_with_missing_or_invalid_date(
    client: AsyncClient, auth_headers: dict
) -> None:
    content = _make_multi_sheet_xlsx(
        {
            "Factures": (
                ["Date facture", "Réf facture", "Client", "Montant"],
                [
                    [None, "2025-0142", "Christine LOPES", 55],
                    ["invalid-date", "2025-0143", "Thi BE NGUYEN", 42],
                ],
            ),
        }
    )

    preview_response = await client.post(
        "/api/import/excel/gestion/preview",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert preview_response.status_code == 200
    preview_data = preview_response.json()
    assert preview_data["can_import"] is False
    assert preview_data["error_details"] == [
        {
            "category": "invoice-invalid-date",
            "severity": "error",
            "sheet_name": "Factures",
            "kind": "invoices",
            "row_number": 2,
            "message": "date manquante ou invalide",
            "display_message": "Factures — Ligne 2 : date manquante ou invalide",
        },
        {
            "category": "invoice-invalid-date",
            "severity": "error",
            "sheet_name": "Factures",
            "kind": "invoices",
            "row_number": 3,
            "message": "date manquante ou invalide",
            "display_message": "Factures — Ligne 3 : date manquante ou invalide",
        },
    ]

    import_response = await client.post(
        "/api/import/excel/gestion",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert import_response.status_code == 200
    import_data = import_response.json()
    assert import_data["invoices_created"] == 0
    assert import_data["contacts_created"] == 0
    assert import_data["error_details"] == preview_data["error_details"]


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_preview_and_import_gestion_block_invoice_with_ambiguous_existing_contact(
    client: AsyncClient,
    auth_headers: dict,
    db_session,
) -> None:
    """Preview and import should block an invoice when several existing contacts match exactly."""
    from backend.models.contact import Contact, ContactType

    db_session.add_all(
        [
            Contact(nom="Dupont", prenom="Alice", type=ContactType.CLIENT),
            Contact(nom="Alice Dupont", prenom=None, type=ContactType.CLIENT),
        ]
    )
    await db_session.commit()

    content = _make_multi_sheet_xlsx(
        {
            "Factures": (
                ["Date facture", "Réf facture", "Client", "Montant"],
                [["2025-08-01", "2025-0142", "Alice Dupont", 55]],
            ),
        }
    )

    preview_response = await client.post(
        "/api/import/excel/gestion/preview",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert preview_response.status_code == 200
    preview_data = preview_response.json()
    assert preview_data["can_import"] is False
    assert preview_data["estimated_invoices"] == 0
    assert any(
        "Factures" in error and "ambigu" in error.lower() for error in preview_data["errors"]
    )
    assert preview_data["error_details"] == [
        {
            "category": "invoice-ambiguous-contact",
            "severity": "error",
            "sheet_name": "Factures",
            "kind": "invoices",
            "row_number": 2,
            "message": "client ambigu : plusieurs contacts correspondent",
            "display_message": (
                "Factures — Ligne 2 : client ambigu : plusieurs contacts correspondent"
            ),
        }
    ]

    import_response = await client.post(
        "/api/import/excel/gestion",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert import_response.status_code == 200
    import_data = import_response.json()
    assert import_data["invoices_created"] == 0
    assert import_data["contacts_created"] == 0
    assert any("Factures" in error and "ambigu" in error.lower() for error in import_data["errors"])
    assert import_data["error_details"] == [
        {
            "category": "invoice-ambiguous-contact",
            "severity": "error",
            "sheet_name": "Factures",
            "kind": "invoices",
            "row_number": 2,
            "message": "client ambigu : plusieurs contacts correspondent",
            "display_message": (
                "Factures — Ligne 2 : client ambigu : plusieurs contacts correspondent"
            ),
        }
    ]


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_preview_and_import_gestion_block_payment_without_match(
    client: AsyncClient, auth_headers: dict
) -> None:
    """Preview and import should block payments that cannot be matched to any invoice."""
    content = _make_multi_sheet_xlsx(
        {
            "Paiements": (
                ["Adhérent", "Montant", "Date paiement", "Mode de règlement"],
                [["Contact inconnu", 55, "2025-08-02", "Chèque"]],
            ),
        }
    )

    preview_response = await client.post(
        "/api/import/excel/gestion/preview",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert preview_response.status_code == 200
    preview_data = preview_response.json()
    assert preview_data["can_import"] is False
    assert any(
        "Paiements" in error and "rapprocher" in error.lower() for error in preview_data["errors"]
    )

    import_response = await client.post(
        "/api/import/excel/gestion",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert import_response.status_code == 200
    import_data = import_response.json()
    assert import_data["payments_created"] == 0
    assert any(
        "Paiements" in error and "rapprocher" in error.lower() for error in import_data["errors"]
    )


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_preview_and_import_gestion_accept_payment_matched_against_existing_invoice(
    client: AsyncClient,
    auth_headers: dict,
    db_session,
) -> None:
    """Preview and import should accept a payment-only file when DB data already resolves it."""
    from datetime import date
    from decimal import Decimal

    from backend.models.contact import Contact, ContactType
    from backend.models.invoice import Invoice, InvoiceLabel, InvoiceStatus, InvoiceType

    contact = Contact(nom="Christine LOPES", type=ContactType.CLIENT)
    db_session.add(contact)
    await db_session.flush()

    invoice = Invoice(
        number="2025-0142",
        type=InvoiceType.CLIENT,
        contact_id=contact.id,
        date=date(2025, 8, 1),
        total_amount=Decimal("55"),
        paid_amount=Decimal("0"),
        status=InvoiceStatus.SENT,
        label=InvoiceLabel.CS,
    )
    db_session.add(invoice)
    await db_session.commit()

    content = _make_multi_sheet_xlsx(
        {
            "Paiements": (
                ["Adhérent", "Montant", "Date paiement", "Mode de règlement"],
                [["Christine LOPES", 55, "2025-08-02", "Chèque"]],
            ),
        }
    )

    preview_response = await client.post(
        "/api/import/excel/gestion/preview",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert preview_response.status_code == 200
    preview_data = preview_response.json()
    assert preview_data["can_import"] is True
    assert preview_data["estimated_payments"] == 1

    import_response = await client.post(
        "/api/import/excel/gestion",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert import_response.status_code == 200
    import_data = import_response.json()
    assert import_data["payments_created"] == 1


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_preview_and_import_gestion_block_payment_with_ambiguous_invoice_reference(
    client: AsyncClient,
    auth_headers: dict,
    db_session,
) -> None:
    """Preview and import should block a payment when an invoice reference is ambiguous."""
    from datetime import date
    from decimal import Decimal

    from backend.models.contact import Contact, ContactType
    from backend.models.invoice import Invoice, InvoiceLabel, InvoiceStatus, InvoiceType

    first_contact = Contact(nom="Christine LOPES", type=ContactType.CLIENT)
    second_contact = Contact(nom="Thi BE NGUYEN", type=ContactType.CLIENT)
    db_session.add_all([first_contact, second_contact])
    await db_session.flush()

    db_session.add_all(
        [
            Invoice(
                number="2025-0142",
                type=InvoiceType.CLIENT,
                contact_id=first_contact.id,
                date=date(2025, 8, 1),
                total_amount=Decimal("55"),
                paid_amount=Decimal("0"),
                status=InvoiceStatus.SENT,
                label=InvoiceLabel.CS,
            ),
            Invoice(
                number="2025-0143",
                type=InvoiceType.CLIENT,
                contact_id=second_contact.id,
                date=date(2025, 8, 2),
                total_amount=Decimal("42"),
                paid_amount=Decimal("0"),
                status=InvoiceStatus.SENT,
                label=InvoiceLabel.CS,
            ),
        ]
    )
    await db_session.commit()

    content = _make_multi_sheet_xlsx(
        {
            "Paiements": (
                ["Réf facture", "Montant", "Date paiement", "Mode de règlement"],
                [["2025-014", 55, "2025-08-02", "Chèque"]],
            ),
        }
    )

    preview_response = await client.post(
        "/api/import/excel/gestion/preview",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert preview_response.status_code == 200
    preview_data = preview_response.json()
    assert preview_data["can_import"] is False
    assert any(
        "Paiements" in error and "ambigu" in error.lower() for error in preview_data["errors"]
    )

    import_response = await client.post(
        "/api/import/excel/gestion",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert import_response.status_code == 200
    import_data = import_response.json()
    assert import_data["payments_created"] == 0
    assert any(
        "Paiements" in error and "ambigu" in error.lower() for error in import_data["errors"]
    )


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_preview_and_import_gestion_block_payment_with_ambiguous_contact_match(
    client: AsyncClient, auth_headers: dict
) -> None:
    """Preview and import should block a contact-based payment when several invoices fit."""
    content = _make_multi_sheet_xlsx(
        {
            "Factures": (
                ["Date facture", "Réf facture", "Client", "Montant"],
                [
                    ["2025-08-01", "2025-0142", "Christine LOPES", 55],
                    ["2025-08-15", "2025-0149", "Christine LOPES", 42],
                ],
            ),
            "Paiements": (
                ["Adhérent", "Montant", "Date paiement", "Mode de règlement"],
                [["Christine LOPES", 55, "2025-08-20", "Chèque"]],
            ),
        }
    )

    preview_response = await client.post(
        "/api/import/excel/gestion/preview",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert preview_response.status_code == 200
    preview_data = preview_response.json()
    assert preview_data["can_import"] is False
    assert any(
        "Paiements" in error and "ambigu" in error.lower() for error in preview_data["errors"]
    )

    import_response = await client.post(
        "/api/import/excel/gestion",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert import_response.status_code == 200
    import_data = import_response.json()
    assert import_data["payments_created"] == 0
    assert any(
        "Paiements" in error and "ambigu" in error.lower() for error in import_data["errors"]
    )


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_import_gestion_blocks_partial_import_when_a_recognized_sheet_is_invalid(
    client: AsyncClient, auth_headers: dict
) -> None:
    """Import should refuse the whole file when a recognized sheet is unsupported."""
    content = _make_multi_sheet_xlsx(
        {
            "Factures": (
                ["Date facture", "Réf facture", "Client", "Montant"],
                [["2025-08-01", "2025-0142", "Christine LOPES", 55]],
            ),
            "Paiements": (
                ["Date paiement", "Montant", "Mode de règlement"],
                [["2025-08-02", 55, "Chèque"]],
            ),
        }
    )

    preview_response = await client.post(
        "/api/import/excel/gestion/preview",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert preview_response.status_code == 200
    preview_data = preview_response.json()
    assert preview_data["can_import"] is False
    assert any("Paiements" in error for error in preview_data["errors"])
    assert preview_data["error_details"] == [
        {
            "category": "payment-missing-columns",
            "severity": "error",
            "sheet_name": "Paiements",
            "kind": "payments",
            "row_number": None,
            "message": "Colonnes requises manquantes: référence facture ou contact",
            "display_message": (
                "Paiements — Colonnes requises manquantes: référence facture ou contact"
            ),
        }
    ]

    import_response = await client.post(
        "/api/import/excel/gestion",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert import_response.status_code == 200
    import_data = import_response.json()
    assert import_data["invoices_created"] == 0
    assert import_data["payments_created"] == 0
    assert import_data["contacts_created"] == 0
    assert any("Paiements" in error for error in import_data["errors"])
    assert import_data["error_details"] == [
        {
            "category": "payment-missing-columns",
            "severity": "error",
            "sheet_name": "Paiements",
            "kind": "payments",
            "row_number": None,
            "message": "Colonnes requises manquantes: référence facture ou contact",
            "display_message": (
                "Paiements — Colonnes requises manquantes: référence facture ou contact"
            ),
        }
    ]


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_import_gestion_rolls_back_all_changes_when_a_late_sheet_crashes(
    client: AsyncClient,
    auth_headers: dict,
    db_session,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """Import should rollback previously flushed rows if a later sheet crashes."""
    from backend.models.contact import Contact
    from backend.models.invoice import Invoice
    from backend.services import excel_import

    content = _make_multi_sheet_xlsx(
        {
            "Factures": (
                ["Date facture", "Réf facture", "Client", "Montant"],
                [["2025-08-01", "2025-0142", "Christine LOPES", 55]],
            ),
            "Paiements": (
                ["Adhérent", "Montant", "Date paiement", "Mode de règlement"],
                [["Christine LOPES", 55, "2025-08-02", "Chèque"]],
            ),
        }
    )

    async def _crash_import_payments_sheet(db, ws, result):
        raise RuntimeError("forced late failure")

    monkeypatch.setattr(
        excel_import,
        "_import_payments_sheet",
        _crash_import_payments_sheet,
    )

    response = await client.post(
        "/api/import/excel/gestion",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert response.status_code == 200
    data = response.json()
    assert data["contacts_created"] == 0
    assert data["invoices_created"] == 0
    assert data["payments_created"] == 0
    assert any("forced late failure" in error for error in data["errors"])

    contacts = (await db_session.execute(select(Contact))).scalars().all()
    invoices = (await db_session.execute(select(Invoice))).scalars().all()
    assert contacts == []
    assert invoices == []


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_import_gestion_fails_when_a_sheet_flush_error_is_caught_locally(
    client: AsyncClient,
    auth_headers: dict,
    db_session,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """A sheet-local flush error must still fail the whole import."""
    from backend.models.contact import Contact
    from backend.models.import_log import ImportLog

    content = _make_multi_sheet_xlsx(
        {
            "Contacts": (
                ["Nom", "Prénom", "Email"],
                [["Dupont", "Alice", "alice@example.test"]],
            ),
        }
    )

    session_type = type(db_session)
    original_flush = session_type.flush
    flush_calls = 0

    async def _fail_first_flush(self, *args, **kwargs):
        nonlocal flush_calls
        flush_calls += 1
        if flush_calls == 1:
            raise RuntimeError("forced contacts flush failure")
        return await original_flush(self, *args, **kwargs)

    monkeypatch.setattr(session_type, "flush", _fail_first_flush)

    response = await client.post(
        "/api/import/excel/gestion",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert response.status_code == 200
    data = response.json()
    assert data["contacts_created"] == 0
    assert data["invoices_created"] == 0
    assert data["payments_created"] == 0
    assert any("forced contacts flush failure" in error for error in data["errors"])

    contacts = (await db_session.execute(select(Contact))).scalars().all()
    assert contacts == []

    logs = (await db_session.execute(select(ImportLog))).scalars().all()
    assert len(logs) == 1
    assert logs[0].status == "failed"


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_import_gestion_processes_sheets_in_business_order(
    client: AsyncClient, auth_headers: dict
) -> None:
    """Import should not depend on workbook sheet ordering for invoice-linked payments."""
    content = _make_multi_sheet_xlsx(
        {
            "Paiements": (
                ["Adhérent", "Montant", "Date paiement", "Mode de règlement"],
                [["Christine LOPES", 55, "2025-08-02", "Chèque"]],
            ),
            "Factures": (
                ["Date facture", "Réf facture", "Client", "Montant"],
                [["2025-08-01", "2025-0142", "Christine LOPES", 55]],
            ),
        }
    )

    response = await client.post(
        "/api/import/excel/gestion",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert response.status_code == 200
    data = response.json()
    assert data["invoices_created"] == 1
    assert data["payments_created"] == 1


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_preview_and_import_gestion_accept_cash_sheet_with_entry_and_exit_columns(
    client: AsyncClient, auth_headers: dict
) -> None:
    """Preview and import should both process cash rows from entrée/sortie columns."""
    content = _make_multi_sheet_xlsx(
        {
            "Caisse": (
                ["Date", "Libellé", "Entrée", "Sortie"],
                [
                    ["2025-08-01", "Recette adhésion", 55, None],
                    ["2025-08-02", "Achat fournitures", None, 12],
                ],
            )
        }
    )

    preview_response = await client.post(
        "/api/import/excel/gestion/preview",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert preview_response.status_code == 200
    preview_data = preview_response.json()
    sheets = {sheet["name"]: sheet for sheet in preview_data["sheets"]}
    assert sheets["Caisse"]["status"] == "recognized"
    assert sheets["Caisse"]["kind"] == "cash"
    assert sheets["Caisse"]["rows"] == 2
    assert preview_data["can_import"] is True

    import_response = await client.post(
        "/api/import/excel/gestion",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert import_response.status_code == 200
    import_data = import_response.json()
    assert import_data["cash_created"] == 2


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_preview_and_import_gestion_accept_signed_cash_amounts_and_ignore_opening_balance(
    client: AsyncClient, auth_headers: dict
) -> None:
    """Preview and import should accept signed cash amounts and ignore opening balance rows."""
    content = _make_multi_sheet_xlsx(
        {
            "Caisse": (
                ["Date", "Montant", "Commentaire"],
                [
                    ["2025-08-01", None, "Solde positif en 2025"],
                    ["2025-08-02", 55, "Recette adhésion"],
                    ["2025-08-03", -12, "Achat fournitures"],
                ],
            )
        }
    )

    preview_response = await client.post(
        "/api/import/excel/gestion/preview",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert preview_response.status_code == 200
    preview_data = preview_response.json()
    sheets = {sheet["name"]: sheet for sheet in preview_data["sheets"]}
    assert preview_data["can_import"] is True
    assert sheets["Caisse"]["rows"] == 2
    assert sheets["Caisse"]["ignored_rows"] == 1
    assert sheets["Caisse"]["blocked_rows"] == 0
    assert any("solde" in warning.lower() for warning in sheets["Caisse"]["warnings"])

    import_response = await client.post(
        "/api/import/excel/gestion",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert import_response.status_code == 200
    import_data = import_response.json()
    assert import_data["cash_created"] == 2
    assert import_data["ignored_rows"] == 1
    assert import_data["blocked_rows"] == 0
    assert any("solde" in warning.lower() for warning in import_data["warnings"])


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_preview_and_import_gestion_ignore_cash_pending_deposit_without_date(
    client: AsyncClient, auth_headers: dict
) -> None:
    """Pending bank deposits without a date should be ignored as cash forecasts."""
    content = _make_multi_sheet_xlsx(
        {
            "Caisse": (
                ["Date", "Montant", "Tiers", "Commentaire"],
                [[None, -710, "CREDIT MUTUEL", "Remise espèces"]],
            )
        }
    )

    preview_response = await client.post(
        "/api/import/excel/gestion/preview",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert preview_response.status_code == 200
    preview_data = preview_response.json()
    sheets = {sheet["name"]: sheet for sheet in preview_data["sheets"]}
    assert preview_data["can_import"] is False
    assert preview_data["errors"] == []
    assert sheets["Caisse"]["rows"] == 0
    assert sheets["Caisse"]["ignored_rows"] == 1
    assert sheets["Caisse"]["blocked_rows"] == 0
    assert any(
        "prevision" in warning.lower() or "remise" in warning.lower()
        for warning in sheets["Caisse"]["warnings"]
    )

    import_response = await client.post(
        "/api/import/excel/gestion",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert import_response.status_code == 200
    import_data = import_response.json()
    assert import_data["cash_created"] == 0
    assert import_data["errors"] == []
    assert import_data["ignored_rows"] == 1
    assert any(
        "prevision" in warning.lower() or "remise" in warning.lower()
        for warning in import_data["warnings"]
    )


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_preview_and_import_gestion_keep_blocking_other_cash_rows_without_date(
    client: AsyncClient, auth_headers: dict
) -> None:
    """Cash rows without a date must still block when they are not pending deposits."""
    content = _make_multi_sheet_xlsx(
        {
            "Caisse": (
                ["Date", "Montant", "Commentaire"],
                [[None, -25, "Achat fournitures"]],
            )
        }
    )

    preview_response = await client.post(
        "/api/import/excel/gestion/preview",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert preview_response.status_code == 200
    preview_data = preview_response.json()
    assert preview_data["can_import"] is False
    assert any("Caisse" in error and "date" in error.lower() for error in preview_data["errors"])

    import_response = await client.post(
        "/api/import/excel/gestion",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert import_response.status_code == 200
    import_data = import_response.json()
    assert import_data["cash_created"] == 0
    assert any("Caisse" in error and "date" in error.lower() for error in import_data["errors"])


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_preview_and_import_gestion_block_cash_row_with_isolated_year(
    client: AsyncClient, auth_headers: dict
) -> None:
    """Cash preview and import should block an isolated row.

    The row year is inconsistent with neighboring rows.
    """
    content = _make_multi_sheet_xlsx(
        {
            "Caisse": (
                ["Date", "Montant", "Commentaire"],
                [
                    ["2024-12-31", 30, "Cotisation"],
                    ["2024-12-31", -10, "Achat fournitures"],
                    ["2025-12-31", 40, "Don en caisse"],
                    ["2025-01-02", 15, "Cotisation"],
                ],
            )
        }
    )

    preview_response = await client.post(
        "/api/import/excel/gestion/preview",
        files={"file": ("Gestion 2024.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert preview_response.status_code == 200
    preview_data = preview_response.json()
    assert preview_data["can_import"] is False
    assert any("Caisse" in error and "voisines" in error for error in preview_data["errors"])
    assert preview_data["error_details"] == [
        {
            "category": "cash-suspicious-date",
            "severity": "error",
            "sheet_name": "Caisse",
            "kind": "cash",
            "row_number": 4,
            "message": "date incoherente avec les lignes voisines",
            "display_message": "Caisse — Ligne 4 : date incoherente avec les lignes voisines",
        }
    ]

    import_response = await client.post(
        "/api/import/excel/gestion",
        files={"file": ("Gestion 2024.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert import_response.status_code == 200
    import_data = import_response.json()
    assert import_data["cash_created"] == 0
    assert import_data["error_details"] == preview_data["error_details"]


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_preview_and_import_gestion_block_cash_rows_with_clustered_dates(
    client: AsyncClient, auth_headers: dict
) -> None:
    """Cash preview and import should block a clustered date anomaly.

    The invalid dates are surrounded by nearby valid rows.
    """
    content = _make_multi_sheet_xlsx(
        {
            "Caisse": (
                ["Date", "Montant", "Commentaire"],
                [
                    ["2026-01-02", 30, "Cotisation"],
                    ["2026-12-16", -49.14, "THIRIET ; FF-2026010218.10.41"],
                    ["2026-12-11", -300, "FNAC ; FF-2026010218.11.51"],
                    ["2026-12-17", -8.14, "CARREFOUR ; FF-2026010218.12.58"],
                    ["2026-01-05", -25, "Achat fournitures"],
                ],
            )
        }
    )

    preview_response = await client.post(
        "/api/import/excel/gestion/preview",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert preview_response.status_code == 200
    preview_data = preview_response.json()
    assert preview_data["can_import"] is False
    assert any("Caisse" in error and "voisines" in error for error in preview_data["errors"])
    assert preview_data["error_details"] == [
        {
            "category": "cash-suspicious-date",
            "severity": "error",
            "sheet_name": "Caisse",
            "kind": "cash",
            "row_number": 3,
            "message": "date incoherente avec les lignes voisines",
            "display_message": "Caisse — Ligne 3 : date incoherente avec les lignes voisines",
        },
        {
            "category": "cash-suspicious-date",
            "severity": "error",
            "sheet_name": "Caisse",
            "kind": "cash",
            "row_number": 4,
            "message": "date incoherente avec les lignes voisines",
            "display_message": "Caisse — Ligne 4 : date incoherente avec les lignes voisines",
        },
        {
            "category": "cash-suspicious-date",
            "severity": "error",
            "sheet_name": "Caisse",
            "kind": "cash",
            "row_number": 5,
            "message": "date incoherente avec les lignes voisines",
            "display_message": "Caisse — Ligne 5 : date incoherente avec les lignes voisines",
        },
    ]

    import_response = await client.post(
        "/api/import/excel/gestion",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert import_response.status_code == 200
    import_data = import_response.json()
    assert import_data["cash_created"] == 0
    assert import_data["error_details"] == preview_data["error_details"]


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_preview_and_import_gestion_accept_bank_sheet_with_debit_and_credit_columns(
    client: AsyncClient, auth_headers: dict
) -> None:
    """Preview and import should both process bank rows from débit/crédit columns."""
    content = _make_multi_sheet_xlsx(
        {
            "Banque": (
                ["Date", "Libellé", "Débit", "Crédit", "Solde"],
                [
                    ["2025-08-01", "Paiement CB", 18, None, 482],
                    ["2025-08-02", "Virement reçu", None, 55, 537],
                ],
            )
        }
    )

    preview_response = await client.post(
        "/api/import/excel/gestion/preview",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert preview_response.status_code == 200
    preview_data = preview_response.json()
    sheets = {sheet["name"]: sheet for sheet in preview_data["sheets"]}
    assert sheets["Banque"]["status"] == "recognized"
    assert sheets["Banque"]["kind"] == "bank"
    assert sheets["Banque"]["rows"] == 2
    assert preview_data["can_import"] is True

    import_response = await client.post(
        "/api/import/excel/gestion",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert import_response.status_code == 200
    import_data = import_response.json()
    assert import_data["bank_created"] == 2


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_preview_and_import_gestion_ignore_bank_opening_descriptive_rows(
    client: AsyncClient, auth_headers: dict
) -> None:
    """Bank preview and import should ignore descriptive opening balance rows."""
    content = _make_multi_sheet_xlsx(
        {
            "Banque": (
                ["Date", "Montant", "Libellé", "Solde"],
                [
                    ["2025-08-01", 566.76, "Solde", 566.76],
                    [None, None, "Assurance MAIF", 566.76],
                    ["2025-08-02", -18, "Paiement CB", 548.76],
                ],
            )
        }
    )

    preview_response = await client.post(
        "/api/import/excel/gestion/preview",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert preview_response.status_code == 200
    preview_data = preview_response.json()
    sheets = {sheet["name"]: sheet for sheet in preview_data["sheets"]}
    assert preview_data["can_import"] is True
    assert sheets["Banque"]["rows"] == 2
    assert sheets["Banque"]["ignored_rows"] == 1
    assert sheets["Banque"]["blocked_rows"] == 0
    assert any(
        "descriptive" in warning.lower() or "solde" in warning.lower()
        for warning in sheets["Banque"]["warnings"]
    )

    import_response = await client.post(
        "/api/import/excel/gestion",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert import_response.status_code == 200
    import_data = import_response.json()
    assert import_data["bank_created"] == 2
    assert import_data["ignored_rows"] == 1
    assert import_data["blocked_rows"] == 0


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_preview_and_import_gestion_ignore_dated_bank_opening_balance_rows(
    client: AsyncClient, auth_headers: dict
) -> None:
    """Bank preview and import should ignore dated opening balance rows without movement."""
    content = _make_multi_sheet_xlsx(
        {
            "Banque": (
                ["Date", "Montant", "Libellé", "Solde"],
                [
                    ["2024-08-01", None, "Solde initial", 566.76],
                    [None, None, "Assurance MAIF", 566.76],
                    ["2024-08-02", -18, "Paiement CB", 548.76],
                ],
            )
        }
    )

    preview_response = await client.post(
        "/api/import/excel/gestion/preview",
        files={"file": ("Gestion 2024.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert preview_response.status_code == 200
    preview_data = preview_response.json()
    sheets = {sheet["name"]: sheet for sheet in preview_data["sheets"]}
    assert preview_data["can_import"] is True
    assert sheets["Banque"]["rows"] == 1
    assert sheets["Banque"]["ignored_rows"] == 2
    assert sheets["Banque"]["blocked_rows"] == 0
    assert any(
        "descriptive" in warning.lower() or "solde" in warning.lower()
        for warning in sheets["Banque"]["warnings"]
    )

    import_response = await client.post(
        "/api/import/excel/gestion",
        files={"file": ("Gestion 2024.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert import_response.status_code == 200
    import_data = import_response.json()
    assert import_data["bank_created"] == 1
    assert import_data["ignored_rows"] == 2
    assert import_data["blocked_rows"] == 0


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_preview_and_import_gestion_block_bank_row_with_isolated_year(
    client: AsyncClient, auth_headers: dict
) -> None:
    """Bank preview and import should block an isolated row.

    The row year is inconsistent with neighboring rows.
    """
    content = _make_multi_sheet_xlsx(
        {
            "Banque": (
                ["Date", "Montant", "Référence", "Libellé", "Solde"],
                [
                    ["2024-12-31", 91, "2024-0281", "Paiement facture client", 5676.34],
                    ["2024-12-31", -660, "VIR FF-2024123113.28.00", "Sous traitance", 5016.34],
                    [
                        "2024-12-31",
                        -1249.64,
                        "VIR SALAIRE LAY 2024.12",
                        "Salaire LAY 2024.12",
                        3766.70,
                    ],
                    ["2025-12-31", 753, "REM CHQ REF05001A05", "Remise chèques", 4442.65],
                    ["2025-01-02", 26, "2024-0279", "Paiement facture client", 4468.65],
                ],
            )
        }
    )

    preview_response = await client.post(
        "/api/import/excel/gestion/preview",
        files={"file": ("Gestion 2024.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert preview_response.status_code == 200
    preview_data = preview_response.json()
    assert preview_data["can_import"] is False
    assert any("Banque" in error and "voisines" in error for error in preview_data["errors"])
    assert preview_data["error_details"] == [
        {
            "category": "bank-suspicious-date",
            "severity": "error",
            "sheet_name": "Banque",
            "kind": "bank",
            "row_number": 5,
            "message": "date incoherente avec les lignes voisines",
            "display_message": "Banque — Ligne 5 : date incoherente avec les lignes voisines",
        }
    ]

    bank_preview_sheet = next(
        sheet for sheet in preview_data["sheets"] if sheet["name"] == "Banque"
    )
    assert bank_preview_sheet["blocked_rows"] == 1
    assert bank_preview_sheet["error_details"][0]["category"] == "bank-suspicious-date"
    assert bank_preview_sheet["error_details"][0]["row_number"] == 5
    assert bank_preview_sheet["error_details"][0]["message"] == (
        "date incoherente avec les lignes voisines"
    )

    import_response = await client.post(
        "/api/import/excel/gestion",
        files={"file": ("Gestion 2024.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert import_response.status_code == 200
    import_data = import_response.json()
    assert import_data["bank_created"] == 0
    assert import_data["error_details"] == preview_data["error_details"]


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_preview_and_import_gestion_block_payment_row_with_isolated_year(
    client: AsyncClient, auth_headers: dict
) -> None:
    """Payment preview and import should block an isolated row.

    The row year is inconsistent with neighboring rows.
    """
    content = _make_multi_sheet_xlsx(
        {
            "Paiements": (
                ["Réf facture", "Montant", "Date paiement", "Mode de règlement"],
                [
                    ["2024-0281", 91, "2024-12-31", "Chèque"],
                    ["2024-0282", 52, "2024-12-31", "Chèque"],
                    ["2024-0283", 44, "2025-12-31", "Chèque"],
                    ["2025-0001", 26, "2025-01-02", "Chèque"],
                ],
            )
        }
    )

    preview_response = await client.post(
        "/api/import/excel/gestion/preview",
        files={"file": ("Gestion 2024.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert preview_response.status_code == 200
    preview_data = preview_response.json()
    assert preview_data["can_import"] is False
    assert any("Paiements" in error and "voisines" in error for error in preview_data["errors"])
    assert preview_data["error_details"] == [
        {
            "category": "payment-suspicious-date",
            "severity": "error",
            "sheet_name": "Paiements",
            "kind": "payments",
            "row_number": 4,
            "message": "date incoherente avec les lignes voisines",
            "display_message": "Paiements — Ligne 4 : date incoherente avec les lignes voisines",
        }
    ]

    import_response = await client.post(
        "/api/import/excel/gestion",
        files={"file": ("Gestion 2024.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert import_response.status_code == 200
    import_data = import_response.json()
    assert import_data["payments_created"] == 0
    assert import_data["error_details"] == preview_data["error_details"]


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_import_gestion_ignores_auxiliary_invoice_sheets(
    client: AsyncClient, auth_headers: dict
) -> None:
    """Import should ignore helper invoice sheets even if they look parseable."""
    content = _make_multi_sheet_xlsx(
        {
            "Factures": (
                [
                    "Date facture",
                    "Réf facture",
                    "Client",
                    "Montant",
                ],
                [["2025-08-01", "2025-0142", "Christine LOPES", 55]],
            ),
            "Aide - Factures": (
                ["Date", "Numéro", "Nom", "Montant"],
                [["2025-08-02", "2025-9999", "Feuille aide", 99]],
            ),
        }
    )

    response = await client.post(
        "/api/import/excel/gestion",
        files={"file": ("Gestion 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert response.status_code == 200
    data = response.json()
    assert data["invoices_created"] == 1
    assert data["contacts_created"] == 1


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_preview_comptabilite_ignores_report_sheets(
    client: AsyncClient, auth_headers: dict
) -> None:
    """Preview should focus on journal sheets and ignore reporting sheets."""
    content = _make_multi_sheet_xlsx(
        {
            "Journal": (
                ["Date", "N° compte", "Libellé de l'écriture", "Débit", "Crédit"],
                [["2025-08-01", "411100", "Facture 2025-0142", 55, None]],
            ),
            "Grand Livre": (
                ["N° compte", "Intitulé", "Date", "Somme de Débit", "Somme de Crédit"],
                [["411100", "Clients", "2025-08-01", 55, 0]],
            ),
        }
    )

    response = await client.post(
        "/api/import/excel/comptabilite/preview",
        files={"file": ("Comptabilite 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert response.status_code == 200
    data = response.json()
    sheets = {sheet["name"]: sheet for sheet in data["sheets"]}
    assert sheets["Journal"]["status"] == "recognized"
    assert sheets["Journal"]["kind"] == "entries"
    assert sheets["Grand Livre"]["status"] == "ignored"
    assert any("Grand Livre" in warning for warning in data["warnings"])


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_preview_comptabilite_ignores_journal_saisie(
    client: AsyncClient, auth_headers: dict
) -> None:
    """Preview should not treat the helper input sheet as an importable journal."""
    content = _make_multi_sheet_xlsx(
        {
            "Journal": (
                ["Date", "N° compte", "Libellé de l'écriture", "Débit", "Crédit"],
                [["2025-08-01", "411100", "Facture 2025-0142", 55, None]],
            ),
            "Journal (saisie)": (
                ["379", "11152.22", "11152.22", "0", "1"],
                [["380", "999", "999", "0", "1"]],
            ),
        }
    )

    response = await client.post(
        "/api/import/excel/comptabilite/preview",
        files={"file": ("Comptabilite 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert response.status_code == 200
    data = response.json()
    sheets = {sheet["name"]: sheet for sheet in data["sheets"]}
    assert sheets["Journal"]["status"] == "recognized"
    assert sheets["Journal (saisie)"]["status"] == "ignored"
    assert any("Journal (saisie)" in warning for warning in data["warnings"])


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_preview_comptabilite_ignores_existing_generated_group(
    client: AsyncClient,
    auth_headers: dict,
    db_session,
) -> None:
    from backend.models.accounting_entry import AccountingEntry, EntrySourceType

    db_session.add_all(
        [
            AccountingEntry(
                entry_number="000001",
                date=date(2025, 8, 1),
                account_number="658000",
                label="Frais divers",
                debit=Decimal("10.00"),
                credit=Decimal("0.00"),
                source_type=EntrySourceType.GESTION,
                source_id=1,
                group_key="gestion:1",
            ),
            AccountingEntry(
                entry_number="000002",
                date=date(2025, 8, 1),
                account_number="512000",
                label="Frais divers",
                debit=Decimal("0.00"),
                credit=Decimal("10.00"),
                source_type=EntrySourceType.GESTION,
                source_id=1,
                group_key="gestion:1",
            ),
        ]
    )
    await db_session.commit()

    content = _make_multi_sheet_xlsx(
        {
            "Journal": (
                ["Date", "N° compte", "Libellé de l'écriture", "Débit", "Crédit", "ChangeNum"],
                [
                    ["2025-08-01", "658000", "Frais divers", 10, None, 1],
                    ["2025-08-01", "512000", "Frais divers", None, 10, 1],
                ],
            ),
        }
    )

    response = await client.post(
        "/api/import/excel/comptabilite/preview",
        files={"file": ("Comptabilite 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert response.status_code == 200
    data = response.json()
    assert data["estimated_entries"] == 0
    assert data["can_import"] is False
    journal_sheet = next(sheet for sheet in data["sheets"] if sheet["name"] == "Journal")
    assert journal_sheet["ignored_rows"] == 2


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_preview_comptabilite_comparison_counts_extra_entries_in_solde(
    client: AsyncClient,
    auth_headers: dict,
    db_session,
) -> None:
    from backend.models.accounting_entry import AccountingEntry, EntrySourceType

    db_session.add_all(
        [
            AccountingEntry(
                entry_number="000001",
                date=date(2025, 8, 1),
                account_number="401100",
                label="Facture fournisseur",
                debit=Decimal("10.00"),
                credit=Decimal("0.00"),
                source_type=EntrySourceType.MANUAL,
            ),
            AccountingEntry(
                entry_number="000002",
                date=date(2025, 8, 3),
                account_number="512000",
                label="Virement en trop",
                debit=Decimal("0.00"),
                credit=Decimal("20.00"),
                source_type=EntrySourceType.MANUAL,
            ),
        ]
    )
    await db_session.commit()

    content = _make_multi_sheet_xlsx(
        {
            "Journal": (
                ["Date", "N° compte", "Libellé de l'écriture", "Débit", "Crédit"],
                [
                    ["2025-08-01", "401100", "Facture fournisseur", 10, None],
                    ["2025-08-02", "512000", "Virement nouveau", None, 20],
                ],
            ),
        }
    )

    response = await client.post(
        "/api/import/excel/comptabilite/preview",
        files={"file": ("Comptabilite 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert response.status_code == 200
    data = response.json()
    assert data["comparison"]["mode"] == "global-convergence"
    domains = {domain["kind"]: domain for domain in data["comparison"]["domains"]}
    assert domains["entries"] == {
        "kind": "entries",
        "file_rows": 2,
        "already_in_solde": 1,
        "missing_in_solde": 1,
        "extra_in_solde": 1,
        "ignored_by_policy": 0,
        "blocked": 0,
    }


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_preview_comptabilite_comparison_filters_only_convergence_section_by_date_window(
    client: AsyncClient,
    auth_headers: dict,
    db_session,
) -> None:
    from backend.models.accounting_entry import AccountingEntry, EntrySourceType

    db_session.add_all(
        [
            AccountingEntry(
                entry_number="000001",
                date=date(2025, 8, 1),
                account_number="401100",
                label="Facture fournisseur",
                debit=Decimal("10.00"),
                credit=Decimal("0.00"),
                source_type=EntrySourceType.MANUAL,
            ),
            AccountingEntry(
                entry_number="000002",
                date=date(2025, 8, 3),
                account_number="512000",
                label="Virement en trop",
                debit=Decimal("0.00"),
                credit=Decimal("20.00"),
                source_type=EntrySourceType.MANUAL,
            ),
        ]
    )
    await db_session.commit()

    content = _make_multi_sheet_xlsx(
        {
            "Journal": (
                ["Date", "N° compte", "Libellé de l'écriture", "Débit", "Crédit"],
                [
                    ["2025-08-01", "401100", "Facture fournisseur", 10, None],
                    ["2025-08-02", "512000", "Virement nouveau", None, 20],
                ],
            ),
        }
    )

    response = await client.post(
        "/api/import/excel/comptabilite/preview",
        data={
            "comparison_start_date": "2025-08-01",
            "comparison_end_date": "2025-08-01",
        },
        files={"file": ("Comptabilite 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert response.status_code == 200
    data = response.json()
    domains = {domain["kind"]: domain for domain in data["comparison"]["domains"]}
    assert domains["entries"] == {
        "kind": "entries",
        "file_rows": 1,
        "already_in_solde": 1,
        "missing_in_solde": 0,
        "extra_in_solde": 0,
        "ignored_by_policy": 0,
        "blocked": 0,
    }
    journal_sheet = next(sheet for sheet in data["sheets"] if sheet["name"] == "Journal")
    assert journal_sheet["rows"] == 1


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_preview_and_import_comptabilite_block_row_with_missing_account(
    client: AsyncClient, auth_headers: dict
) -> None:
    """Preview and import should block a journal row with no account number."""
    content = _make_multi_sheet_xlsx(
        {
            "Journal": (
                ["Date", "N° compte", "Libellé de l'écriture", "Débit", "Crédit"],
                [
                    ["2025-08-01", "411100", "Facture 2025-0142", 55, None],
                    ["2025-08-02", None, "Ligne invalide", 12, None],
                ],
            ),
        }
    )

    preview_response = await client.post(
        "/api/import/excel/comptabilite/preview",
        files={"file": ("Comptabilite 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert preview_response.status_code == 200
    preview_data = preview_response.json()
    assert preview_data["can_import"] is False
    assert any("Journal" in error and "Ligne 3" in error for error in preview_data["errors"])

    import_response = await client.post(
        "/api/import/excel/comptabilite",
        files={"file": ("Comptabilite 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert import_response.status_code == 200
    import_data = import_response.json()
    assert import_data["entries_created"] == 0
    assert any("Journal" in error and "Ligne 3" in error for error in import_data["errors"])


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_preview_and_import_comptabilite_block_row_with_invalid_amounts(
    client: AsyncClient, auth_headers: dict
) -> None:
    """Preview and import should block a journal row when debit/credit are not usable."""
    content = _make_multi_sheet_xlsx(
        {
            "Journal": (
                ["Date", "N° compte", "Libellé de l'écriture", "Débit", "Crédit"],
                [
                    ["2025-08-01", "411100", "Facture 2025-0142", 55, None],
                    ["2025-08-02", "512000", "Ligne invalide", "abc", None],
                ],
            ),
        }
    )

    preview_response = await client.post(
        "/api/import/excel/comptabilite/preview",
        files={"file": ("Comptabilite 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert preview_response.status_code == 200
    preview_data = preview_response.json()
    assert preview_data["can_import"] is False
    assert any(
        "Journal" in error and "montant" in error.lower() for error in preview_data["errors"]
    )

    import_response = await client.post(
        "/api/import/excel/comptabilite",
        files={"file": ("Comptabilite 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert import_response.status_code == 200
    import_data = import_response.json()
    assert import_data["entries_created"] == 0
    assert any("Journal" in error and "montant" in error.lower() for error in import_data["errors"])


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_preview_and_import_comptabilite_ignore_zero_amount_rows(
    client: AsyncClient, auth_headers: dict
) -> None:
    """Zero-amount journal rows should be ignored instead of blocking the whole file."""
    content = _make_multi_sheet_xlsx(
        {
            "Journal": (
                ["Date", "N° compte", "Libellé de l'écriture", "Débit", "Crédit"],
                [
                    ["2025-08-01", "411100", "Facture 2025-0142", 55, None],
                    ["2025-08-02", "431100", "Impôt sur le revenu 2025.08", 0, 0],
                    ["2025-08-02", "512000", "Banque", None, 55],
                ],
            ),
        }
    )

    preview_response = await client.post(
        "/api/import/excel/comptabilite/preview",
        files={"file": ("Comptabilite 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert preview_response.status_code == 200
    preview_data = preview_response.json()
    sheets = {sheet["name"]: sheet for sheet in preview_data["sheets"]}
    assert preview_data["can_import"] is True
    assert preview_data["estimated_entries"] == 2
    assert sheets["Journal"]["rows"] == 2
    assert sheets["Journal"]["ignored_rows"] == 1
    assert sheets["Journal"]["blocked_rows"] == 0
    assert any(
        "nuls" in warning.lower() or "zero" in warning.lower()
        for warning in sheets["Journal"]["warnings"]
    )

    import_response = await client.post(
        "/api/import/excel/comptabilite",
        files={"file": ("Comptabilite 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert import_response.status_code == 200
    import_data = import_response.json()
    assert import_data["entries_created"] == 2
    assert import_data["ignored_rows"] == 1
    assert import_data["blocked_rows"] == 0


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_preview_and_import_comptabilite_block_journal_rows_with_clustered_dates(
    client: AsyncClient, auth_headers: dict
) -> None:
    """Comptabilite journal preview/import should block clustered dates.

    The affected run is inconsistent with surrounding rows.
    """
    content = _make_multi_sheet_xlsx(
        {
            "Journal": (
                ["Date", "N° compte", "Libellé de l'écriture", "Débit", "Crédit"],
                [
                    ["2025-12-31", "512100", "2025-0255 Marie-Claire RIBEIRO", 78, None],
                    ["2025-12-31", "411100", "2025-0255 Marie-Claire RIBEIRO", None, 78],
                    ["2026-12-03", "511200", "2025-0222 Christine LOPES", 33, None],
                    ["2026-12-03", "411100", "2025-0222 Christine LOPES", None, 33],
                    ["2026-12-02", "511200", "2025-0229 Zeina NOHRA", 78, None],
                    ["2026-12-02", "411100", "2025-0229 Zeina NOHRA", None, 78],
                    ["2026-12-05", "511200", "2025-0236 Franck LEVY", 208, None],
                    ["2026-12-05", "411100", "2025-0236 Franck LEVY", None, 208],
                    ["2026-01-02", "531000", "2025-0226 Moulay HACHIMI", 104, None],
                    ["2026-01-02", "411100", "2025-0226 Moulay HACHIMI", None, 104],
                ],
            ),
        }
    )

    preview_response = await client.post(
        "/api/import/excel/comptabilite/preview",
        files={"file": ("Comptabilite 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert preview_response.status_code == 200
    preview_data = preview_response.json()
    sheets = {sheet["name"]: sheet for sheet in preview_data["sheets"]}
    assert preview_data["can_import"] is False
    assert any("Journal" in error and "voisines" in error for error in preview_data["errors"])
    assert preview_data["error_details"] == [
        {
            "category": "entry-suspicious-date",
            "severity": "error",
            "sheet_name": "Journal",
            "kind": "entries",
            "row_number": 4,
            "message": "date incoherente avec les lignes voisines",
            "display_message": "Journal — Ligne 4 : date incoherente avec les lignes voisines",
        },
        {
            "category": "entry-suspicious-date",
            "severity": "error",
            "sheet_name": "Journal",
            "kind": "entries",
            "row_number": 5,
            "message": "date incoherente avec les lignes voisines",
            "display_message": "Journal — Ligne 5 : date incoherente avec les lignes voisines",
        },
        {
            "category": "entry-suspicious-date",
            "severity": "error",
            "sheet_name": "Journal",
            "kind": "entries",
            "row_number": 6,
            "message": "date incoherente avec les lignes voisines",
            "display_message": "Journal — Ligne 6 : date incoherente avec les lignes voisines",
        },
        {
            "category": "entry-suspicious-date",
            "severity": "error",
            "sheet_name": "Journal",
            "kind": "entries",
            "row_number": 7,
            "message": "date incoherente avec les lignes voisines",
            "display_message": "Journal — Ligne 7 : date incoherente avec les lignes voisines",
        },
        {
            "category": "entry-suspicious-date",
            "severity": "error",
            "sheet_name": "Journal",
            "kind": "entries",
            "row_number": 8,
            "message": "date incoherente avec les lignes voisines",
            "display_message": "Journal — Ligne 8 : date incoherente avec les lignes voisines",
        },
        {
            "category": "entry-suspicious-date",
            "severity": "error",
            "sheet_name": "Journal",
            "kind": "entries",
            "row_number": 9,
            "message": "date incoherente avec les lignes voisines",
            "display_message": "Journal — Ligne 9 : date incoherente avec les lignes voisines",
        },
    ]
    assert sheets["Journal"]["blocked_rows"] == 6

    import_response = await client.post(
        "/api/import/excel/comptabilite",
        files={"file": ("Comptabilite 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert import_response.status_code == 200
    import_data = import_response.json()
    assert import_data["entries_created"] == 0
    assert import_data["error_details"] == preview_data["error_details"]


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_import_comptabilite_groups_entries_using_change_num(
    client: AsyncClient,
    auth_headers: dict,
    db_session,
) -> None:
    from backend.models.accounting_entry import AccountingEntry, EntrySourceType

    content = _make_multi_sheet_xlsx(
        {
            "Journal": (
                ["Date", "N° compte", "Libellé de l'écriture", "Débit", "Crédit", "ChangeNum"],
                [
                    ["2025-08-01", "411100", "Facture 2025-0142", 55, None, 0],
                    ["2025-08-01", "706110", "Facture 2025-0142", None, 55, 0],
                    ["2025-08-02", "512000", "Reglement", 55, None, 1],
                    ["2025-08-02", "411100", "Reglement", None, 55, 1],
                ],
            ),
        }
    )

    response = await client.post(
        "/api/import/excel/comptabilite",
        files={"file": ("Comptabilite 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert response.status_code == 200
    result = await db_session.execute(
        select(AccountingEntry)
        .where(AccountingEntry.source_type == EntrySourceType.MANUAL)
        .order_by(AccountingEntry.id.asc())
    )
    entries = result.scalars().all()
    assert len(entries) == 4
    assert entries[0].group_key == entries[1].group_key
    assert entries[2].group_key == entries[3].group_key
    assert entries[0].group_key != entries[2].group_key


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_import_comptabilite_splits_balanced_subgroups_inside_same_change_num(
    client: AsyncClient,
    auth_headers: dict,
    db_session,
) -> None:
    from backend.models.accounting_entry import AccountingEntry, EntrySourceType

    content = _make_multi_sheet_xlsx(
        {
            "Journal": (
                ["Date", "N° compte", "Libellé de l'écriture", "Débit", "Crédit", "ChangeNum"],
                [
                    [
                        "2025-10-02",
                        "611100",
                        "Sous traitance des cours - LEXIO SAS - 2025.08 ; FF-2025100212.09.02",
                        120,
                        None,
                        1,
                    ],
                    [
                        "2025-10-02",
                        "512100",
                        "Sous traitance des cours - LEXIO SAS - 2025.08 ; FF-2025100212.09.02",
                        None,
                        120,
                        1,
                    ],
                    ["2025-10-02", "645100", "Charges patronales 2025.09", 333.63, None, 1],
                    ["2025-10-02", "641000", "Salaires bruts 2025.09", 1605, None, 1],
                    ["2025-10-02", "421000", "Salaires nets 2025.09", None, 1249.64, 1],
                    ["2025-10-02", "431100", "Charges patronales 2025.09", None, 333.63, 1],
                    ["2025-10-02", "431100", "Charges salariales 2025.09", None, 355.36, 1],
                ],
            ),
        }
    )

    response = await client.post(
        "/api/import/excel/comptabilite",
        files={"file": ("Comptabilite 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert response.status_code == 200
    result = await db_session.execute(
        select(AccountingEntry)
        .where(AccountingEntry.source_type == EntrySourceType.MANUAL)
        .order_by(AccountingEntry.id.asc())
    )
    entries = result.scalars().all()
    assert len(entries) == 7
    assert entries[0].group_key == entries[1].group_key
    assert entries[2].group_key == entries[3].group_key
    assert entries[3].group_key == entries[4].group_key
    assert entries[4].group_key == entries[5].group_key
    assert entries[5].group_key == entries[6].group_key
    assert entries[0].group_key != entries[2].group_key


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_import_comptabilite_ignores_generated_group_even_when_label_differs(
    client: AsyncClient,
    auth_headers: dict,
    db_session,
) -> None:
    from backend.models.accounting_entry import AccountingEntry, EntrySourceType

    db_session.add_all(
        [
            AccountingEntry(
                entry_number="000001",
                date=date(2025, 7, 21),
                account_number="411100",
                label="Fact. 2025-0141 32",
                debit=Decimal("170.00"),
                credit=Decimal("0.00"),
                source_type=EntrySourceType.INVOICE,
                source_id=257,
                group_key="invoice:257",
            ),
            AccountingEntry(
                entry_number="000002",
                date=date(2025, 7, 21),
                account_number="706110",
                label="Fact. 2025-0141 32",
                debit=Decimal("0.00"),
                credit=Decimal("170.00"),
                source_type=EntrySourceType.INVOICE,
                source_id=257,
                group_key="invoice:257",
            ),
        ]
    )
    await db_session.commit()

    content = _make_multi_sheet_xlsx(
        {
            "Journal": (
                ["Date", "N° compte", "Libellé de l'écriture", "Débit", "Crédit", "ChangeNum"],
                [
                    ["2025-07-21", "411100", "2025-0141 Marie-Claire RIBEIRO", 170, None, 0],
                    ["2025-07-21", "706110", "2025-0141 Marie-Claire RIBEIRO", None, 170, 0],
                ],
            ),
        }
    )

    preview_response = await client.post(
        "/api/import/excel/comptabilite/preview",
        files={"file": ("Comptabilite 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert preview_response.status_code == 200
    preview_data = preview_response.json()
    assert preview_data["estimated_entries"] == 0
    assert preview_data["sheets"][0]["ignored_rows"] == 2

    import_response = await client.post(
        "/api/import/excel/comptabilite",
        files={"file": ("Comptabilite 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert import_response.status_code == 200
    import_data = import_response.json()
    assert import_data["entries_created"] == 0
    assert import_data["ignored_rows"] == 2
    result = await db_session.execute(
        select(AccountingEntry).where(AccountingEntry.source_type == EntrySourceType.MANUAL)
    )
    assert result.scalars().all() == []


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_import_comptabilite_clarifies_existing_client_invoice_other_line_when_split_is_safe(
    client: AsyncClient,
    auth_headers: dict,
    db_session,
) -> None:
    """Client invoice journal rows can clarify an existing imported mixed invoice."""
    from sqlalchemy.orm import selectinload

    from backend.models.accounting_entry import AccountingEntry, EntrySourceType
    from backend.models.contact import Contact, ContactType
    from backend.models.invoice import (
        Invoice,
        InvoiceLabel,
        InvoiceLine,
        InvoiceLineType,
        InvoiceStatus,
        InvoiceType,
    )
    from backend.services.accounting_engine import seed_default_rules

    await seed_default_rules(db_session)
    await db_session.commit()

    contact = Contact(nom="ELEZI", prenom="Aleksander", type=ContactType.CLIENT)
    db_session.add(contact)
    await db_session.flush()

    invoice = Invoice(
        number="2024-0192",
        type=InvoiceType.CLIENT,
        contact_id=contact.id,
        date=date(2024, 8, 30),
        total_amount=Decimal("100.00"),
        paid_amount=Decimal("100.00"),
        status=InvoiceStatus.PAID,
        label=InvoiceLabel.GENERAL,
        has_explicit_breakdown=False,
    )
    db_session.add(invoice)
    await db_session.flush()

    db_session.add(
        InvoiceLine(
            invoice_id=invoice.id,
            description="Autres prestations",
            line_type=InvoiceLineType.OTHER,
            quantity=Decimal("1.00"),
            unit_price=Decimal("100.00"),
            amount=Decimal("100.00"),
        )
    )

    db_session.add_all(
        [
            AccountingEntry(
                entry_number="000001",
                date=date(2024, 8, 30),
                account_number="411100",
                label="Fact. 2024-0192 15",
                debit=Decimal("100.00"),
                credit=Decimal("0.00"),
                source_type=EntrySourceType.INVOICE,
                source_id=invoice.id,
                group_key=f"invoice:{invoice.id}",
            ),
            AccountingEntry(
                entry_number="000002",
                date=date(2024, 8, 30),
                account_number="706110",
                label="Fact. 2024-0192 15",
                debit=Decimal("0.00"),
                credit=Decimal("100.00"),
                source_type=EntrySourceType.INVOICE,
                source_id=invoice.id,
                group_key=f"invoice:{invoice.id}",
            ),
        ]
    )
    await db_session.commit()

    content = _make_multi_sheet_xlsx(
        {
            "Journal": (
                ["Date", "N° compte", "Libellé de l'écriture", "Débit", "Crédit", "ChangeNum"],
                [
                    ["2024-08-30", "411100", "2024-0192 Aleksander ELEZI", 50, None, 15],
                    ["2024-08-30", "706110", "2024-0192 Aleksander ELEZI", None, 50, 15],
                    ["2024-08-30", "411100", "2024-0192 Aleksander ELEZI", 50, None, 16],
                    ["2024-08-30", "756000", "2024-0192 Aleksander ELEZI", None, 50, 16],
                ],
            ),
        }
    )

    preview_response = await client.post(
        "/api/import/excel/comptabilite/preview",
        files={"file": ("Comptabilite 2024.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert preview_response.status_code == 200
    preview_data = preview_response.json()
    sheets = {sheet["name"]: sheet for sheet in preview_data["sheets"]}
    assert preview_data["estimated_entries"] == 0
    assert sheets["Journal"]["ignored_rows"] == 4
    assert sheets["Journal"]["blocked_rows"] == 0

    import_response = await client.post(
        "/api/import/excel/comptabilite",
        files={"file": ("Comptabilite 2024.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert import_response.status_code == 200
    import_data = import_response.json()
    assert import_data["entries_created"] == 0
    assert import_data["ignored_rows"] == 4
    result = await db_session.execute(
        select(AccountingEntry).where(AccountingEntry.source_type == EntrySourceType.MANUAL)
    )
    assert result.scalars().all() == []

    refreshed_invoice = (
        (
            await db_session.execute(
                select(Invoice).where(Invoice.id == invoice.id).options(selectinload(Invoice.lines))
            )
        )
        .scalars()
        .one()
    )
    assert refreshed_invoice.label == InvoiceLabel.CS_ADHESION
    assert refreshed_invoice.has_explicit_breakdown is True
    assert {(line.line_type, line.amount) for line in refreshed_invoice.lines} == {
        (InvoiceLineType.COURSE, Decimal("50.00")),
        (InvoiceLineType.ADHESION, Decimal("50.00")),
    }

    invoice_entries = (
        (
            await db_session.execute(
                select(AccountingEntry)
                .where(AccountingEntry.source_type == EntrySourceType.INVOICE)
                .where(AccountingEntry.source_id == invoice.id)
                .order_by(AccountingEntry.entry_number.asc())
            )
        )
        .scalars()
        .all()
    )
    assert len(invoice_entries) == 3
    assert any(
        entry.account_number == "706110" and entry.credit == Decimal("50.00")
        for entry in invoice_entries
    )
    assert any(
        entry.account_number == "756000" and entry.credit == Decimal("50.00")
        for entry in invoice_entries
    )


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_import_comptabilite_ignores_client_payment_group_when_date_differs(
    client: AsyncClient,
    auth_headers: dict,
    db_session,
) -> None:
    """Client payment rows should be ignored when an equivalent payment already exists."""
    from backend.models.accounting_entry import AccountingEntry, EntrySourceType
    from backend.models.contact import Contact, ContactType
    from backend.models.invoice import Invoice, InvoiceLabel, InvoiceStatus, InvoiceType
    from backend.models.payment import Payment, PaymentMethod

    contact = Contact(nom="NOHRA", prenom="Zeina", type=ContactType.CLIENT)
    db_session.add(contact)
    await db_session.flush()

    invoice = Invoice(
        number="2025-0229",
        type=InvoiceType.CLIENT,
        contact_id=contact.id,
        date=date(2025, 11, 29),
        total_amount=Decimal("78.00"),
        paid_amount=Decimal("78.00"),
        status=InvoiceStatus.PAID,
        label=InvoiceLabel.CS,
        description="2025-0229 Zeina NOHRA",
    )
    db_session.add(invoice)
    await db_session.flush()

    payment = Payment(
        invoice_id=invoice.id,
        contact_id=contact.id,
        date=date(2025, 12, 2),
        amount=Decimal("78.00"),
        method=PaymentMethod.CHEQUE,
        reference="2025-0229",
        notes="Cheque 2026.01.02.02",
        deposited=True,
        deposit_date=date(2025, 12, 3),
    )
    db_session.add(payment)
    await db_session.flush()

    db_session.add_all(
        [
            AccountingEntry(
                entry_number="000001",
                date=date(2025, 11, 29),
                account_number="411100",
                label="Fact. 2025-0229 Zeina NOHRA",
                debit=Decimal("78.00"),
                credit=Decimal("0.00"),
                source_type=EntrySourceType.INVOICE,
                source_id=invoice.id,
                group_key=f"invoice:{invoice.id}",
            ),
            AccountingEntry(
                entry_number="000002",
                date=date(2025, 11, 29),
                account_number="706000",
                label="Fact. 2025-0229 Zeina NOHRA",
                debit=Decimal("0.00"),
                credit=Decimal("78.00"),
                source_type=EntrySourceType.INVOICE,
                source_id=invoice.id,
                group_key=f"invoice:{invoice.id}",
            ),
            AccountingEntry(
                entry_number="000003",
                date=date(2025, 12, 2),
                account_number="511200",
                label="Règt. Chèque 2026.01.02.02",
                debit=Decimal("78.00"),
                credit=Decimal("0.00"),
                source_type=EntrySourceType.PAYMENT,
                source_id=payment.id,
                group_key=f"payment:{payment.id}",
            ),
            AccountingEntry(
                entry_number="000004",
                date=date(2025, 12, 2),
                account_number="411100",
                label="Règt. Chèque 2026.01.02.02",
                debit=Decimal("0.00"),
                credit=Decimal("78.00"),
                source_type=EntrySourceType.PAYMENT,
                source_id=payment.id,
                group_key=f"payment:{payment.id}",
            ),
        ]
    )
    await db_session.commit()

    content = _make_multi_sheet_xlsx(
        {
            "Journal": (
                ["Date", "N° compte", "Libellé de l'écriture", "Débit", "Crédit", "ChangeNum"],
                [
                    ["2025-12-03", "511200", "2025-0229 Zeina NOHRA", 78, None, 18],
                    ["2025-12-03", "411100", "2025-0229 Zeina NOHRA", None, 78, 18],
                ],
            ),
        }
    )

    preview_response = await client.post(
        "/api/import/excel/comptabilite/preview",
        files={"file": ("Comptabilite 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert preview_response.status_code == 200
    preview_data = preview_response.json()
    assert preview_data["estimated_entries"] == 0

    import_response = await client.post(
        "/api/import/excel/comptabilite",
        files={"file": ("Comptabilite 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert import_response.status_code == 200
    import_data = import_response.json()
    assert import_data["entries_created"] == 0
    assert import_data["ignored_rows"] == 2

    manual_entries = list(
        (
            await db_session.execute(
                select(AccountingEntry).where(AccountingEntry.source_type == EntrySourceType.MANUAL)
            )
        ).scalars()
    )
    assert not manual_entries


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_import_comptabilite_ignores_existing_supplier_invoice_and_payment_group_by_reference(
    client: AsyncClient,
    auth_headers: dict,
    db_session,
) -> None:
    """Supplier direct-settlement rows should be ignored when invoice and payment already exist."""
    from backend.models.accounting_entry import AccountingEntry, EntrySourceType
    from backend.models.contact import Contact, ContactType
    from backend.models.invoice import Invoice, InvoiceLabel, InvoiceStatus, InvoiceType
    from backend.models.payment import Payment, PaymentMethod

    contact = Contact(nom="Théo DAUPHY", type=ContactType.FOURNISSEUR)
    db_session.add(contact)
    await db_session.flush()

    invoice = Invoice(
        number="FF-2024090718.25.00",
        reference="FF-2024090718.25.00",
        type=InvoiceType.FOURNISSEUR,
        contact_id=contact.id,
        date=date(2024, 9, 7),
        total_amount=Decimal("300.00"),
        paid_amount=Decimal("300.00"),
        status=InvoiceStatus.PAID,
        label=InvoiceLabel.CS,
        description="Sous traitance des cours - Théo DAUPHY - 2024.08",
    )
    db_session.add(invoice)
    await db_session.flush()

    payment = Payment(
        invoice_id=invoice.id,
        contact_id=contact.id,
        date=date(2024, 9, 7),
        amount=Decimal("300.00"),
        method=PaymentMethod.VIREMENT,
        reference="FF-2024090718.25.00",
        notes="Sous traitance des cours - Théo DAUPHY - 2024.08",
    )
    db_session.add(payment)
    await db_session.flush()

    db_session.add_all(
        [
            AccountingEntry(
                entry_number="000001",
                date=date(2024, 9, 7),
                account_number="611100",
                label="Fact. FF-2024090718.25.00 Théo DAUPHY",
                debit=Decimal("300.00"),
                credit=Decimal("0.00"),
                source_type=EntrySourceType.INVOICE,
                source_id=invoice.id,
                group_key=f"invoice:{invoice.id}",
            ),
            AccountingEntry(
                entry_number="000002",
                date=date(2024, 9, 7),
                account_number="401000",
                label="Fact. FF-2024090718.25.00 Théo DAUPHY",
                debit=Decimal("0.00"),
                credit=Decimal("300.00"),
                source_type=EntrySourceType.INVOICE,
                source_id=invoice.id,
                group_key=f"invoice:{invoice.id}",
            ),
            AccountingEntry(
                entry_number="000003",
                date=date(2024, 9, 7),
                account_number="401000",
                label="Règt. FF-2024090718.25.00",
                debit=Decimal("300.00"),
                credit=Decimal("0.00"),
                source_type=EntrySourceType.PAYMENT,
                source_id=payment.id,
                group_key=f"payment:{payment.id}",
            ),
            AccountingEntry(
                entry_number="000004",
                date=date(2024, 9, 7),
                account_number="512100",
                label="Règt. FF-2024090718.25.00",
                debit=Decimal("0.00"),
                credit=Decimal("300.00"),
                source_type=EntrySourceType.PAYMENT,
                source_id=payment.id,
                group_key=f"payment:{payment.id}",
            ),
        ]
    )
    await db_session.commit()

    content = _make_multi_sheet_xlsx(
        {
            "Journal": (
                ["Date", "N° compte", "Libellé de l'écriture", "Débit", "Crédit", "ChangeNum"],
                [
                    [
                        "2024-09-07",
                        "611100",
                        "Sous traitance des cours - Théo DAUPHY - 2024.08 ; FF-2024090718.25.00",
                        300,
                        None,
                        15,
                    ],
                    [
                        "2024-09-07",
                        "512100",
                        "Sous traitance des cours - Théo DAUPHY - 2024.08 ; FF-2024090718.25.00",
                        None,
                        300,
                        15,
                    ],
                ],
            ),
        }
    )

    preview_response = await client.post(
        "/api/import/excel/comptabilite/preview",
        files={"file": ("Comptabilite 2024.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert preview_response.status_code == 200
    preview_data = preview_response.json()
    assert preview_data["estimated_entries"] == 0

    import_response = await client.post(
        "/api/import/excel/comptabilite",
        files={"file": ("Comptabilite 2024.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert import_response.status_code == 200
    import_data = import_response.json()
    assert import_data["entries_created"] == 0
    manual_entries = list(
        (
            await db_session.execute(
                select(AccountingEntry).where(AccountingEntry.source_type == EntrySourceType.MANUAL)
            )
        ).scalars()
    )
    assert not manual_entries


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_import_comptabilite_ignores_supplier_and_salary_subgroups_same_change_num(
    client: AsyncClient,
    auth_headers: dict,
    db_session,
) -> None:
    from backend.models.accounting_entry import AccountingEntry, EntrySourceType
    from backend.models.contact import Contact, ContactType
    from backend.models.invoice import Invoice, InvoiceLabel, InvoiceStatus, InvoiceType
    from backend.models.payment import Payment, PaymentMethod
    from backend.models.salary import Salary

    supplier = Contact(nom="LEXIO SAS", type=ContactType.FOURNISSEUR)
    employee = Contact(nom="Martin", prenom="Alice", type=ContactType.FOURNISSEUR)
    db_session.add_all([supplier, employee])
    await db_session.flush()

    invoice = Invoice(
        number="FF-2025100212.09.02",
        reference="FF-2025100212.09.02",
        type=InvoiceType.FOURNISSEUR,
        contact_id=supplier.id,
        date=date(2025, 10, 2),
        total_amount=Decimal("120.00"),
        paid_amount=Decimal("120.00"),
        status=InvoiceStatus.PAID,
        label=InvoiceLabel.CS,
        description="Sous traitance des cours - LEXIO SAS - 2025.08",
    )
    db_session.add(invoice)
    await db_session.flush()

    payment = Payment(
        invoice_id=invoice.id,
        contact_id=supplier.id,
        date=date(2025, 10, 2),
        amount=Decimal("120.00"),
        method=PaymentMethod.VIREMENT,
        reference="FF-2025100212.09.02",
        notes="Sous traitance des cours - LEXIO SAS - 2025.08",
    )
    salary = Salary(
        employee_id=employee.id,
        month="2025-09",
        hours=Decimal("35.00"),
        gross=Decimal("1605.00"),
        employee_charges=Decimal("355.36"),
        employer_charges=Decimal("333.63"),
        tax=Decimal("0.00"),
        net_pay=Decimal("1249.64"),
    )
    db_session.add_all([payment, salary])
    await db_session.flush()

    db_session.add_all(
        [
            AccountingEntry(
                entry_number="000001",
                date=date(2025, 10, 2),
                account_number="611100",
                label="Fact. FF-2025100212.09.02 LEXIO SAS",
                debit=Decimal("120.00"),
                credit=Decimal("0.00"),
                source_type=EntrySourceType.INVOICE,
                source_id=invoice.id,
                group_key=f"invoice:{invoice.id}",
            ),
            AccountingEntry(
                entry_number="000002",
                date=date(2025, 10, 2),
                account_number="401000",
                label="Fact. FF-2025100212.09.02 LEXIO SAS",
                debit=Decimal("0.00"),
                credit=Decimal("120.00"),
                source_type=EntrySourceType.INVOICE,
                source_id=invoice.id,
                group_key=f"invoice:{invoice.id}",
            ),
            AccountingEntry(
                entry_number="000003",
                date=date(2025, 10, 2),
                account_number="401000",
                label="Règt. FF-2025100212.09.02",
                debit=Decimal("120.00"),
                credit=Decimal("0.00"),
                source_type=EntrySourceType.PAYMENT,
                source_id=payment.id,
                group_key=f"payment:{payment.id}",
            ),
            AccountingEntry(
                entry_number="000004",
                date=date(2025, 10, 2),
                account_number="512100",
                label="Règt. FF-2025100212.09.02",
                debit=Decimal("0.00"),
                credit=Decimal("120.00"),
                source_type=EntrySourceType.PAYMENT,
                source_id=payment.id,
                group_key=f"payment:{payment.id}",
            ),
        ]
    )
    await db_session.commit()

    content = _make_multi_sheet_xlsx(
        {
            "Journal": (
                ["Date", "N° compte", "Libellé de l'écriture", "Débit", "Crédit", "ChangeNum"],
                [
                    [
                        "2025-10-02",
                        "611100",
                        "Sous traitance des cours - LEXIO SAS - 2025.08 ; FF-2025100212.09.02",
                        120,
                        None,
                        1,
                    ],
                    [
                        "2025-10-02",
                        "512100",
                        "Sous traitance des cours - LEXIO SAS - 2025.08 ; FF-2025100212.09.02",
                        None,
                        120,
                        1,
                    ],
                    ["2025-10-02", "645100", "Charges patronales 2025.09", 333.63, None, 1],
                    ["2025-10-02", "641000", "Salaires bruts 2025.09", 1605, None, 1],
                    ["2025-10-02", "421000", "Salaires nets 2025.09", None, 1249.64, 1],
                    ["2025-10-02", "431100", "Charges patronales 2025.09", None, 333.63, 1],
                    ["2025-10-02", "431100", "Charges salariales 2025.09", None, 355.36, 1],
                ],
            ),
        }
    )

    preview_response = await client.post(
        "/api/import/excel/comptabilite/preview",
        files={"file": ("Comptabilite 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert preview_response.status_code == 200
    preview_data = preview_response.json()
    assert preview_data["estimated_entries"] == 0
    assert preview_data["sheets"][0]["ignored_rows"] == 7

    import_response = await client.post(
        "/api/import/excel/comptabilite",
        files={"file": ("Comptabilite 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert import_response.status_code == 200
    import_data = import_response.json()
    assert import_data["entries_created"] == 0
    assert import_data["ignored_rows"] == 7
    manual_entries = list(
        (
            await db_session.execute(
                select(AccountingEntry).where(AccountingEntry.source_type == EntrySourceType.MANUAL)
            )
        ).scalars()
    )
    assert not manual_entries


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_import_comptabilite_blocks_reimport_of_same_file(
    client: AsyncClient,
    auth_headers: dict,
) -> None:
    """A successful Comptabilité import should block exact reimport of the same file."""
    content = _make_multi_sheet_xlsx(
        {
            "Journal": (
                ["Date", "N° compte", "Libellé de l'écriture", "Débit", "Crédit"],
                [["2025-08-01", "411100", "Facture 2025-0142", 55, None]],
            ),
        }
    )

    first_response = await client.post(
        "/api/import/excel/comptabilite",
        files={"file": ("Comptabilite 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert first_response.status_code == 200
    first_data = first_response.json()
    assert first_data["entries_created"] == 1
    assert first_data["errors"] == []

    preview_response = await client.post(
        "/api/import/excel/comptabilite/preview",
        files={"file": ("Comptabilite copie.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert preview_response.status_code == 200
    preview_data = preview_response.json()
    assert preview_data["can_import"] is False
    assert any("deja importe" in error.lower() for error in preview_data["errors"])

    second_response = await client.post(
        "/api/import/excel/comptabilite",
        files={"file": ("Comptabilite copie.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert second_response.status_code == 200
    second_data = second_response.json()
    assert second_data["entries_created"] == 0
    assert any("deja importe" in error.lower() for error in second_data["errors"])


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_preview_and_import_comptabilite_ignores_rows_already_covered_by_solde(
    client: AsyncClient,
    auth_headers: dict,
    db_session,
) -> None:
    """Comptabilité import should ignore rows already covered by Solde and keep new rows."""
    from datetime import date
    from decimal import Decimal

    from backend.models.accounting_entry import AccountingEntry, EntrySourceType
    from backend.models.fiscal_year import FiscalYear, FiscalYearStatus

    fiscal_year = FiscalYear(
        name="2025",
        start_date=date(2025, 1, 1),
        end_date=date(2025, 12, 31),
        status=FiscalYearStatus.OPEN,
    )
    db_session.add(fiscal_year)

    db_session.add(
        AccountingEntry(
            entry_number="000001",
            date=date(2025, 8, 1),
            account_number="411100",
            label="Facture 2025-0142",
            debit=Decimal("55"),
            credit=Decimal("0"),
            source_type=EntrySourceType.INVOICE,
            source_id=1,
        )
    )
    await db_session.commit()

    content = _make_multi_sheet_xlsx(
        {
            "Journal": (
                ["Date", "N° compte", "Libellé de l'écriture", "Débit", "Crédit"],
                [
                    ["2025-08-01", "411100", "Facture 2025-0142", 55, None],
                    ["2025-08-02", "706200", "Nouvelle ecriture", None, 55],
                ],
            ),
        }
    )

    preview_response = await client.post(
        "/api/import/excel/comptabilite/preview",
        files={"file": ("Comptabilite 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert preview_response.status_code == 200
    preview_data = preview_response.json()
    assert preview_data["can_import"] is True
    assert any("auto-generees" in warning.lower() for warning in preview_data["warnings"])
    assert preview_data["estimated_entries"] == 1
    assert preview_data["sheets"][0]["ignored_rows"] == 1
    assert any(
        "deja couverte par solde" in warning.lower()
        for warning in preview_data["sheets"][0]["warnings"]
    )
    assert any(
        detail["category"] == "entry-covered-by-solde"
        for detail in preview_data["sheets"][0]["warning_details"]
    )

    import_response = await client.post(
        "/api/import/excel/comptabilite",
        files={"file": ("Comptabilite 2025.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert import_response.status_code == 200
    import_data = import_response.json()
    assert import_data["entries_created"] == 1
    assert import_data["ignored_rows"] == 1
    assert import_data["errors"] == []


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_preview_and_import_comptabilite_ignores_exact_manual_duplicates_with_exact_category(
    client: AsyncClient,
    auth_headers: dict,
    db_session,
) -> None:
    """Exact duplicates against existing manual entries stay categorized as entry-existing."""
    from datetime import date
    from decimal import Decimal

    from sqlalchemy import select

    from backend.models.accounting_entry import AccountingEntry, EntrySourceType

    db_session.add(
        AccountingEntry(
            entry_number="000001",
            date=date(2025, 8, 1),
            account_number="411100",
            label="Ecriture manuelle existante",
            debit=Decimal("55"),
            credit=Decimal("0"),
            source_type=EntrySourceType.MANUAL,
            source_id=None,
        )
    )
    await db_session.commit()

    content = _make_multi_sheet_xlsx(
        {
            "Journal": (
                ["Date", "N° compte", "Libellé de l'écriture", "Débit", "Crédit"],
                [["2025-08-01", "411100", "Ecriture manuelle existante", 55, None]],
            ),
        }
    )

    preview_response = await client.post(
        "/api/import/excel/comptabilite/preview",
        files={"file": ("Comptabilite exact-duplicate.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert preview_response.status_code == 200
    preview_data = preview_response.json()
    assert preview_data["estimated_entries"] == 0
    assert preview_data["sheets"][0]["ignored_rows"] == 1
    assert any(
        "deja existante" in warning.lower() for warning in preview_data["sheets"][0]["warnings"]
    )
    assert any(
        detail["category"] == "entry-existing"
        for detail in preview_data["sheets"][0]["warning_details"]
    )
    assert all(
        detail["category"] != "entry-near-manual"
        for detail in preview_data["sheets"][0]["warning_details"]
    )

    import_response = await client.post(
        "/api/import/excel/comptabilite",
        files={"file": ("Comptabilite exact-duplicate.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert import_response.status_code == 200
    import_data = import_response.json()
    assert import_data["entries_created"] == 0
    assert import_data["ignored_rows"] == 1
    assert import_data["errors"] == []

    manual_entries = (
        (
            await db_session.execute(
                select(AccountingEntry).where(AccountingEntry.source_type == EntrySourceType.MANUAL)
            )
        )
        .scalars()
        .all()
    )
    assert len(manual_entries) == 1


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_preview_and_import_comptabilite_warns_for_near_manual_entries_without_blocking(
    client: AsyncClient,
    auth_headers: dict,
    db_session,
) -> None:
    """Near matches with manual entries should warn but still import."""
    from datetime import date
    from decimal import Decimal

    from sqlalchemy import select

    from backend.models.accounting_entry import AccountingEntry, EntrySourceType

    db_session.add(
        AccountingEntry(
            entry_number="000001",
            date=date(2025, 8, 1),
            account_number="411100",
            label="Libelle manuel",
            debit=Decimal("55"),
            credit=Decimal("0"),
            source_type=EntrySourceType.MANUAL,
            source_id=None,
        )
    )
    await db_session.commit()

    content = _make_multi_sheet_xlsx(
        {
            "Journal": (
                ["Date", "N° compte", "Libellé de l'écriture", "Débit", "Crédit"],
                [["2025-08-01", "411100", "Libelle importe", 55, None]],
            ),
        }
    )

    preview_response = await client.post(
        "/api/import/excel/comptabilite/preview",
        files={"file": ("Comptabilite near-manual.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert preview_response.status_code == 200
    preview_data = preview_response.json()
    assert preview_data["can_import"] is True
    assert preview_data["estimated_entries"] == 1
    assert preview_data["sheets"][0]["ignored_rows"] == 0
    assert any(
        "proche d'une ecriture manuelle existante" in warning.lower()
        for warning in preview_data["sheets"][0]["warnings"]
    )
    assert any(
        detail["category"] == "entry-near-manual"
        for detail in preview_data["sheets"][0]["warning_details"]
    )

    import_response = await client.post(
        "/api/import/excel/comptabilite",
        files={"file": ("Comptabilite near-manual.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert import_response.status_code == 200
    import_data = import_response.json()
    assert import_data["entries_created"] == 1
    assert import_data["ignored_rows"] == 0
    assert import_data["errors"] == []
    assert any(
        "proche d'une ecriture manuelle existante" in warning.lower()
        for warning in import_data["warnings"]
    )
    assert any(
        detail["category"] == "entry-near-manual" for detail in import_data["warning_details"]
    )

    manual_entries = (
        (
            await db_session.execute(
                select(AccountingEntry).where(AccountingEntry.source_type == EntrySourceType.MANUAL)
            )
        )
        .scalars()
        .all()
    )
    assert len(manual_entries) == 2


@pytest.mark.asyncio
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_preview_and_import_comptabilite_allow_existing_manual_entries(
    client: AsyncClient,
    auth_headers: dict,
    db_session,
) -> None:
    """Manual accounting entries do not block a comptabilité import."""
    from datetime import date
    from decimal import Decimal

    from backend.models.accounting_entry import AccountingEntry, EntrySourceType

    db_session.add(
        AccountingEntry(
            entry_number="000001",
            date=date(2025, 7, 31),
            account_number="512000",
            label="Ecriture manuelle existante",
            debit=Decimal("0"),
            credit=Decimal("25"),
            source_type=EntrySourceType.MANUAL,
            source_id=None,
        )
    )
    await db_session.commit()

    content = _make_multi_sheet_xlsx(
        {
            "Journal": (
                ["Date", "N° compte", "Libellé de l'écriture", "Débit", "Crédit"],
                [["2025-08-01", "411100", "Facture 2025-0142", 55, None]],
            ),
        }
    )

    preview_response = await client.post(
        "/api/import/excel/comptabilite/preview",
        files={"file": ("Comptabilite manuel.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert preview_response.status_code == 200
    preview_data = preview_response.json()
    assert preview_data["can_import"] is True
    assert all("auto-generees" not in error.lower() for error in preview_data["errors"])

    import_response = await client.post(
        "/api/import/excel/comptabilite",
        files={"file": ("Comptabilite manuel.xlsx", content, _XLSX_MIME)},
        headers=auth_headers,
    )

    assert import_response.status_code == 200
    import_data = import_response.json()
    assert import_data["entries_created"] == 1
    assert all("auto-generees" not in error.lower() for error in import_data["errors"])
